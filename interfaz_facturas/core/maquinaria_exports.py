"""Generación de exports de mantenimiento: Service History (PDF/Excel), Asset Passport, Certificado CAE."""
from __future__ import annotations

import hashlib
import io
import json
import os
from datetime import datetime
from pathlib import Path

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT
from reportlab.lib.utils import ImageReader
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image, PageBreak,
)

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from core import maquinaria_db

# ── Paths ──
_ASSETS_DIR = Path(__file__).resolve().parent.parent / "static" / "assets" / "presupuestos"
_EXPORTS_DIR = Path(__file__).resolve().parent.parent / "data" / "exports_maquinaria"

# Datos corporativos
EMPRESA = {
    "nombre": "HINCADO DIRECTO S.L.",
    "cif": "B-88261458",
    "direccion": "Calle Francisco Luján nº2, Badajoz, 06004",
    "telefono": "+34 637 70 54 33",
    "email": "direccion@hincadodirecto.com",
    "web": "www.hincadodirecto.com",
    "registro": "Tomo 38499, Folio 174, Inscripción 1ª, Hoja M-684725",
    "firmante_nombre": "Sergio Garcia Cascallana",
    "firmante_cargo": "Administrador",
}

# Colores corporativos
_GRIS_OSCURO = colors.HexColor("#1E293B")
_GRIS_MEDIO = colors.HexColor("#475569")
_GRIS_CLARO = colors.HexColor("#F1F5F9")
_AZUL = colors.HexColor("#2563EB")
_VERDE = colors.HexColor("#16A34A")
_ROJO = colors.HexColor("#DC2626")
_AMARILLO = colors.HexColor("#CA8A04")


def _ensure_exports_dir():
    _EXPORTS_DIR.mkdir(parents=True, exist_ok=True)
    return _EXPORTS_DIR


def _find_logo():
    for name in ("logo.jpg", "logo.png", "logo_hincado_directo.png"):
        p = _ASSETS_DIR / name
        if p.exists():
            return p
    return None


def _hash_bytes(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


def _register_doc(maquina_id, tipo, titulo, filename, filepath, data_bytes, generado_por=None, metadata=None):
    """Registra un documento generado en la DB."""
    mime_map = {".pdf": "application/pdf", ".xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"}
    ext = Path(filename).suffix.lower()
    return maquinaria_db.registrar_documento({
        "maquina_id": maquina_id,
        "tipo": tipo,
        "titulo": titulo,
        "filename": filename,
        "filepath": filepath,
        "mime_type": mime_map.get(ext, "application/octet-stream"),
        "size_bytes": len(data_bytes),
        "hash_sha256": _hash_bytes(data_bytes),
        "generado_por": generado_por,
        "metadata": metadata,
    })


# ══════════════════════════════════════════════════════════════════════════════
# ██  SERVICE HISTORY — PDF                                                 ██
# ══════════════════════════════════════════════════════════════════════════════


def generar_service_history_pdf(maquina_id: int, desde: str = None, hasta: str = None,
                                 generado_por: str = None) -> tuple[bytes, dict]:
    """Genera PDF de historial de servicio. Retorna (bytes_pdf, doc_record)."""

    historial = maquinaria_db.obtener_historial_servicio(maquina_id, desde, hasta)
    if not historial:
        raise ValueError("Máquina no encontrada")

    maq = historial["maquina"]
    revisiones = historial["revisiones"]
    checks = historial["checks"]
    incidencias = historial["incidencias"]

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=A4,
        leftMargin=18 * mm, rightMargin=18 * mm,
        topMargin=20 * mm, bottomMargin=20 * mm,
    )

    styles = getSampleStyleSheet()
    s_title = ParagraphStyle("SHTitle", parent=styles["Heading1"], fontSize=16, textColor=_GRIS_OSCURO,
                              spaceAfter=2 * mm, fontName="Helvetica-Bold")
    s_subtitle = ParagraphStyle("SHSub", parent=styles["Normal"], fontSize=10, textColor=_GRIS_MEDIO,
                                 spaceAfter=4 * mm)
    s_section = ParagraphStyle("SHSection", parent=styles["Heading2"], fontSize=13, textColor=_AZUL,
                                spaceBefore=6 * mm, spaceAfter=3 * mm, fontName="Helvetica-Bold")
    s_normal = ParagraphStyle("SHNormal", parent=styles["Normal"], fontSize=9, textColor=_GRIS_OSCURO,
                               leading=12)
    s_small = ParagraphStyle("SHSmall", parent=styles["Normal"], fontSize=8, textColor=_GRIS_MEDIO,
                              leading=10)
    s_footer = ParagraphStyle("SHFooter", parent=styles["Normal"], fontSize=7, textColor=_GRIS_MEDIO,
                               alignment=TA_CENTER)

    elements = []

    # ── CABECERA ──
    logo = _find_logo()
    if logo:
        try:
            img_reader = ImageReader(str(logo))
            iw, ih = img_reader.getSize()
            ratio = min(45 * mm / iw, 20 * mm / ih)
            elements.append(Image(str(logo), width=iw * ratio, height=ih * ratio))
            elements.append(Spacer(1, 3 * mm))
        except Exception:
            pass

    elements.append(Paragraph(f"Historial de Servicio — {maq['nombre']}", s_title))

    filtro = ""
    if desde or hasta:
        filtro = f" | Filtro: {desde or '...'} a {hasta or '...'}"
    elements.append(Paragraph(
        f"{maq['internal_id']} · {maq['modelo']}"
        f"{' · S/N: ' + maq['numero_serie'] if maq.get('numero_serie') else ''}"
        f" · Horómetro: {maq['horometro_actual']:,.0f}h{filtro}", s_subtitle))

    elements.append(Paragraph(
        f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')} · {EMPRESA['nombre']}", s_small))
    elements.append(Spacer(1, 4 * mm))

    # ── RESUMEN ──
    elements.append(Paragraph("Resumen", s_section))
    summary_data = [
        ["Revisiones preventivas", str(len(revisiones))],
        ["Checks semanales", str(len(checks))],
        ["Incidencias reportadas", str(len(incidencias))],
        ["Incidencias abiertas", str(sum(1 for i in incidencias if i.get("estado") != "cerrada"))],
    ]
    t = Table(summary_data, colWidths=[120 * mm, 40 * mm])
    t.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("TEXTCOLOR", (0, 0), (0, -1), _GRIS_OSCURO),
        ("TEXTCOLOR", (1, 0), (1, -1), _AZUL),
        ("FONTNAME", (1, 0), (1, -1), "Helvetica-Bold"),
        ("ALIGN", (1, 0), (1, -1), "RIGHT"),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("LINEBELOW", (0, 0), (-1, -2), 0.5, _GRIS_CLARO),
    ]))
    elements.append(t)
    elements.append(Spacer(1, 4 * mm))

    # ── REVISIONES PREVENTIVAS ──
    elements.append(Paragraph("Revisiones Preventivas", s_section))
    if revisiones:
        rev_data = [["Horómetro", "Fecha", "Tareas", "Operario"]]
        for r in revisiones:
            rev_data.append([
                f"{r['horometro']:,.0f}h",
                r["fecha"] or "—",
                str(len(r["tareas"])) + " tareas",
                r.get("operario", "—") or "—",
            ])
        t = Table(rev_data, colWidths=[30 * mm, 28 * mm, 25 * mm, 77 * mm])
        t.setStyle(TableStyle([
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("BACKGROUND", (0, 0), (-1, 0), _GRIS_CLARO),
            ("TEXTCOLOR", (0, 0), (-1, 0), _GRIS_OSCURO),
            ("TEXTCOLOR", (0, 1), (-1, -1), _GRIS_OSCURO),
            ("ALIGN", (0, 0), (0, -1), "RIGHT"),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("LINEBELOW", (0, 0), (-1, -1), 0.3, colors.HexColor("#E2E8F0")),
            ("LINEBELOW", (0, 0), (-1, 0), 1, _GRIS_OSCURO),
        ]))
        elements.append(t)
    else:
        elements.append(Paragraph("Sin revisiones registradas en el periodo seleccionado.", s_small))
    elements.append(Spacer(1, 4 * mm))

    # ── CHECKS SEMANALES ──
    elements.append(Paragraph("Checks Semanales", s_section))
    if checks:
        chk_data = [["Fecha", "Horómetro", "Estado", "Operario"]]
        for c in checks:
            chk_data.append([
                (c.get("fecha") or "")[:10],
                f"{c.get('horometro', 0):,.0f}h",
                c.get("estado", "—"),
                c.get("usuario_nombre", "—") or "—",
            ])
        t = Table(chk_data, colWidths=[30 * mm, 30 * mm, 25 * mm, 75 * mm])
        t.setStyle(TableStyle([
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("BACKGROUND", (0, 0), (-1, 0), _GRIS_CLARO),
            ("TEXTCOLOR", (0, 0), (-1, -1), _GRIS_OSCURO),
            ("ALIGN", (1, 0), (1, -1), "RIGHT"),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("LINEBELOW", (0, 0), (-1, -1), 0.3, colors.HexColor("#E2E8F0")),
            ("LINEBELOW", (0, 0), (-1, 0), 1, _GRIS_OSCURO),
        ]))
        elements.append(t)
    else:
        elements.append(Paragraph("Sin checks registrados en el periodo seleccionado.", s_small))
    elements.append(Spacer(1, 4 * mm))

    # ── INCIDENCIAS ──
    elements.append(Paragraph("Incidencias", s_section))
    if incidencias:
        inc_data = [["Fecha", "Severidad", "Estado", "Descripción"]]
        for i in incidencias:
            desc = (i.get("descripcion") or "")[:80]
            if len(i.get("descripcion", "")) > 80:
                desc += "..."
            inc_data.append([
                (i.get("fecha") or "")[:10],
                i.get("severidad", "—"),
                i.get("estado", "—"),
                Paragraph(desc, s_small),
            ])
        t = Table(inc_data, colWidths=[25 * mm, 22 * mm, 20 * mm, 93 * mm])
        t.setStyle(TableStyle([
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("BACKGROUND", (0, 0), (-1, 0), _GRIS_CLARO),
            ("TEXTCOLOR", (0, 0), (-1, -1), _GRIS_OSCURO),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("LINEBELOW", (0, 0), (-1, -1), 0.3, colors.HexColor("#E2E8F0")),
            ("LINEBELOW", (0, 0), (-1, 0), 1, _GRIS_OSCURO),
        ]))
        elements.append(t)
    else:
        elements.append(Paragraph("Sin incidencias registradas en el periodo seleccionado.", s_small))

    # ── PIE ──
    elements.append(Spacer(1, 10 * mm))
    elements.append(Paragraph(
        f"{EMPRESA['nombre']} — C.I.F.: {EMPRESA['cif']} — {EMPRESA['registro']}", s_footer))

    doc.build(elements)
    pdf_bytes = buffer.getvalue()

    # Guardar y registrar
    _ensure_exports_dir()
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    nombre_limpio = maq["nombre"].replace(" ", "_")
    filename = f"ServiceHistory_{nombre_limpio}_{ts}.pdf"
    filepath = str(_EXPORTS_DIR / filename)
    with open(filepath, "wb") as f:
        f.write(pdf_bytes)

    doc_record = _register_doc(
        maquina_id, "service_history_pdf",
        f"Historial de Servicio — {maq['nombre']}",
        filename, filepath, pdf_bytes,
        generado_por=generado_por,
        metadata={"desde": desde, "hasta": hasta, "n_revisiones": len(revisiones),
                   "n_checks": len(checks), "n_incidencias": len(incidencias)},
    )
    return pdf_bytes, doc_record


# ══════════════════════════════════════════════════════════════════════════════
# ██  SERVICE HISTORY — EXCEL                                               ██
# ══════════════════════════════════════════════════════════════════════════════


def generar_service_history_xlsx(maquina_id: int, desde: str = None, hasta: str = None,
                                  generado_por: str = None) -> tuple[bytes, dict]:
    """Genera Excel de historial de servicio. Retorna (bytes_xlsx, doc_record)."""

    historial = maquinaria_db.obtener_historial_servicio(maquina_id, desde, hasta)
    if not historial:
        raise ValueError("Máquina no encontrada")

    maq = historial["maquina"]
    revisiones = historial["revisiones"]
    checks = historial["checks"]
    incidencias = historial["incidencias"]

    wb = openpyxl.Workbook()

    # Estilos
    header_font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    normal_font = Font(name="Arial", size=9)
    thin_border = Border(
        bottom=Side(style="thin", color="E2E8F0"),
    )
    header_align = Alignment(horizontal="center", vertical="center")

    def _setup_sheet(ws, title, headers):
        ws.title = title
        for col_idx, (header, width) in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width
        return ws

    # ── Hoja Resumen ──
    ws = wb.active
    ws.title = "Resumen"
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 40

    resumen_data = [
        ("Máquina", maq["nombre"]),
        ("ID Interno", maq["internal_id"]),
        ("Modelo", maq["modelo"]),
        ("Nº Serie", maq.get("numero_serie") or "—"),
        ("Horómetro actual", f"{maq['horometro_actual']:,.0f}h"),
        ("Fecha comisión", maq.get("fecha_comision") or "—"),
        ("Estado", maq.get("estado", "—")),
        ("", ""),
        ("Periodo", f"{desde or 'Todo'} — {hasta or 'Todo'}"),
        ("Revisiones preventivas", len(revisiones)),
        ("Checks semanales", len(checks)),
        ("Incidencias totales", len(incidencias)),
        ("Incidencias abiertas", sum(1 for i in incidencias if i.get("estado") != "cerrada")),
        ("", ""),
        ("Generado", datetime.now().strftime("%d/%m/%Y %H:%M")),
        ("Empresa", EMPRESA["nombre"]),
    ]
    for row_idx, (label, value) in enumerate(resumen_data, 1):
        ws.cell(row=row_idx, column=1, value=label).font = Font(name="Arial", bold=True, size=10)
        ws.cell(row=row_idx, column=2, value=value).font = normal_font

    # ── Hoja Revisiones ──
    ws_rev = wb.create_sheet()
    _setup_sheet(ws_rev, "Revisiones", [
        ("Horómetro", 15), ("Fecha", 15), ("Nº Tareas", 12),
        ("Operario", 25), ("Observaciones", 50),
    ])
    for i, r in enumerate(revisiones, 2):
        ws_rev.cell(row=i, column=1, value=r["horometro"]).font = normal_font
        ws_rev.cell(row=i, column=2, value=r["fecha"]).font = normal_font
        ws_rev.cell(row=i, column=3, value=len(r["tareas"])).font = normal_font
        ws_rev.cell(row=i, column=4, value=r.get("operario") or "—").font = normal_font
        ws_rev.cell(row=i, column=5, value=r.get("observaciones") or "").font = normal_font
        for col in range(1, 6):
            ws_rev.cell(row=i, column=col).border = thin_border

    # ── Hoja Checks ──
    ws_chk = wb.create_sheet()
    _setup_sheet(ws_chk, "Checks Semanales", [
        ("Fecha", 15), ("Horómetro", 15), ("Estado", 12),
        ("Operario", 25), ("Observaciones", 50),
    ])
    for i, c in enumerate(checks, 2):
        ws_chk.cell(row=i, column=1, value=(c.get("fecha") or "")[:10]).font = normal_font
        ws_chk.cell(row=i, column=2, value=c.get("horometro", 0)).font = normal_font
        ws_chk.cell(row=i, column=3, value=c.get("estado", "—")).font = normal_font
        ws_chk.cell(row=i, column=4, value=c.get("usuario_nombre") or "—").font = normal_font
        ws_chk.cell(row=i, column=5, value=c.get("observaciones") or "").font = normal_font
        for col in range(1, 6):
            ws_chk.cell(row=i, column=col).border = thin_border

    # ── Hoja Incidencias ──
    ws_inc = wb.create_sheet()
    _setup_sheet(ws_inc, "Incidencias", [
        ("Fecha", 15), ("Severidad", 12), ("Estado", 12),
        ("Descripción", 50), ("Resolución", 50), ("Reportado por", 20),
    ])
    for i, inc in enumerate(incidencias, 2):
        ws_inc.cell(row=i, column=1, value=(inc.get("fecha") or "")[:10]).font = normal_font
        ws_inc.cell(row=i, column=2, value=inc.get("severidad", "—")).font = normal_font
        ws_inc.cell(row=i, column=3, value=inc.get("estado", "—")).font = normal_font
        ws_inc.cell(row=i, column=4, value=inc.get("descripcion") or "").font = normal_font
        ws_inc.cell(row=i, column=5, value=inc.get("resolucion") or "").font = normal_font
        ws_inc.cell(row=i, column=6, value=inc.get("usuario_nombre") or "—").font = normal_font
        for col in range(1, 7):
            ws_inc.cell(row=i, column=col).border = thin_border

    # Guardar
    xlsx_buffer = io.BytesIO()
    wb.save(xlsx_buffer)
    xlsx_bytes = xlsx_buffer.getvalue()

    _ensure_exports_dir()
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    nombre_limpio = maq["nombre"].replace(" ", "_")
    filename = f"ServiceHistory_{nombre_limpio}_{ts}.xlsx"
    filepath = str(_EXPORTS_DIR / filename)
    with open(filepath, "wb") as f:
        f.write(xlsx_bytes)

    doc_record = _register_doc(
        maquina_id, "service_history_xlsx",
        f"Historial de Servicio — {maq['nombre']}",
        filename, filepath, xlsx_bytes,
        generado_por=generado_por,
        metadata={"desde": desde, "hasta": hasta, "n_revisiones": len(revisiones),
                   "n_checks": len(checks), "n_incidencias": len(incidencias)},
    )
    return xlsx_bytes, doc_record


# ══════════════════════════════════════════════════════════════════════════════
# ██  CERTIFICADO CAE/PRL — PDF (formato idéntico al histórico)             ██
# ══════════════════════════════════════════════════════════════════════════════

# Meses en español para fechas
_MESES = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
          "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]


def _fecha_es(dt: datetime) -> str:
    """Formatea fecha como '10 de Marzo 2025'."""
    return f"{dt.day} de {_MESES[dt.month - 1]} {dt.year}"


def generar_certificado_cae(
    maquina_id: int,
    modo: str = "ultima",
    hito_horas: int = None,
    lugar: str = "Badajoz",
    firmante_nombre: str = None,
    firmante_cargo: str = None,
    generado_por: str = None,
) -> tuple[bytes, dict]:
    """Genera certificado CAE/PRL en PDF, replicando el formato histórico.

    Modos:
      - "ultima": certifica la última revisión realizada (la más reciente en maintenance_logs)
      - "hito": certifica la revisión al hito especificado (hito_horas)

    Retorna (bytes_pdf, doc_record).
    """
    from reportlab.platypus import BaseDocTemplate, Frame, PageTemplate
    from reportlab.lib.pagesizes import A4

    maquinaria_db.init_maquinaria_db()

    firmante_nombre = firmante_nombre or EMPRESA["firmante_nombre"]
    firmante_cargo = firmante_cargo or EMPRESA["firmante_cargo"]

    # ── Obtener datos de la máquina ──
    from core.db import conectar as _db_conectar
    with _db_conectar() as conn:
        maq = conn.execute("SELECT * FROM maquinas WHERE id = ?", [maquina_id]).fetchone()
        if not maq:
            raise ValueError("Máquina no encontrada")
        maq = dict(maq)

        if modo == "hito" and hito_horas:
            # Buscar revisión al hito concreto
            rev = conn.execute(
                "SELECT horometro_at, completed_at FROM maquinaria_maintenance_logs "
                "WHERE maquina_id = ? AND due_hours = ? ORDER BY completed_at DESC LIMIT 1",
                [maquina_id, hito_horas],
            ).fetchone()
            if not rev:
                raise ValueError(f"No se encontró revisión al hito {hito_horas}h")
            rev_horas = rev[0]
            rev_fecha = rev[1]
        else:
            # Última revisión (mayor horometro_at)
            rev = conn.execute(
                "SELECT horometro_at, completed_at FROM maquinaria_maintenance_logs "
                "WHERE maquina_id = ? ORDER BY horometro_at DESC LIMIT 1",
                [maquina_id],
            ).fetchone()
            if not rev:
                raise ValueError("No hay revisiones registradas para esta máquina")
            rev_horas = rev[0]
            rev_fecha = rev[1]

    # Parsear fecha de revisión
    try:
        dt_rev = datetime.fromisoformat(rev_fecha.replace("Z", "+00:00"))
    except Exception:
        dt_rev = datetime.now()

    # Descripción de la máquina según modelo
    modelo_desc = "Hincadora de Postes Fotovoltaicos"
    if "MHPW" in maq.get("modelo", ""):
        modelo_desc = "Máquina Hincapostes"

    serie_txt = maq.get("numero_serie") or "—"

    # ── Generar PDF con ReportLab (bajo nivel para control preciso) ──
    buffer = io.BytesIO()
    from reportlab.pdfgen import canvas as rl_canvas

    c = rl_canvas.Canvas(buffer, pagesize=A4)
    width, height = A4  # 595.27 x 841.89 pts

    # Márgenes
    ml = 55  # left
    mr = 55  # right
    content_w = width - ml - mr

    # ── CABECERA: Logo a la izquierda, datos empresa a la derecha ──
    logo_path = _ASSETS_DIR / "logo_hincado_directo.png"
    if logo_path.exists():
        try:
            img = ImageReader(str(logo_path))
            iw, ih = img.getSize()
            target_w = 180
            ratio = target_w / iw
            target_h = ih * ratio
            c.drawImage(str(logo_path), ml, height - 40 - target_h,
                        width=target_w, height=target_h, preserveAspectRatio=True, mask="auto")
        except Exception:
            pass

    # Datos empresa alineados a la derecha
    c.setFont("Helvetica", 9)
    c.setFillColor(colors.HexColor("#1E293B"))
    right_x = width - mr
    y_emp = height - 50
    emp_lines = [
        EMPRESA["nombre"],
        "",
        "Francisco Luján nº2,",
        "Badajoz, Badajoz, 06004",
        f"T  {EMPRESA['telefono']}",
        "+34 686 27 09 37",
        "",
        EMPRESA["email"],
        EMPRESA["web"],
    ]
    for line in emp_lines:
        if line:
            c.drawRightString(right_x, y_emp, line)
        y_emp -= 13

    # ── CUERPO ──
    y = height - 210

    # Línea: La empresa HINCADO DIRECTO S.L. ... CERTIFICA:
    c.setFont("Helvetica", 11)
    c.setFillColor(colors.black)

    from reportlab.platypus import Paragraph as _Para
    from reportlab.lib.styles import ParagraphStyle as _PS

    body_style = _PS("body", fontName="Helvetica", fontSize=11, leading=16,
                      textColor=colors.black)
    bold_style = _PS("bodyBold", fontName="Helvetica-Bold", fontSize=11, leading=16,
                      textColor=colors.black)

    # Primer párrafo
    text1 = (
        f'La empresa <b>HINCADO DIRECTO S.L.</b> con domicilio en calle Francisco Luján nº2, '
        f'Badajoz, 06004 con <b>CIF {EMPRESA["cif"]}</b>, CERTIFICA:'
    )
    p1 = _Para(text1, body_style)
    pw, ph = p1.wrap(content_w, 200)
    p1.drawOn(c, ml, y - ph)
    y -= ph + 20

    # Segundo párrafo: certificación
    horas_fmt = f"{int(rev_horas):,}".replace(",", ".")
    fecha_rev_fmt = _fecha_es(dt_rev)

    text2 = (
        f'Que se ha realizado la revisión a las {horas_fmt} horas de trabajo a la '
        f'{modelo_desc} con número de serie {serie_txt}, '
        f'con fecha {fecha_rev_fmt} tal y como establece el Manual de Uso y mantenimiento, '
        f'emitido por el fabricante Orteco {maq["modelo"].replace("ORTECO ", "")}, '
        f'ejecutándose todos los trabajos que en el mismo se recomiendan. '
        f'El resultado del mantenimiento es <b>Favorable</b>'
    )
    p2 = _Para(text2, body_style)
    pw, ph = p2.wrap(content_w, 300)
    p2.drawOn(c, ml, y - ph)
    y -= ph + 25

    # "Para que así conste..."
    text3 = "Para que así conste y pueda presentarse ante la autoridad competente,"
    p3 = _Para(text3, body_style)
    pw, ph = p3.wrap(content_w, 100)
    p3.drawOn(c, ml, y - ph)

    # ── FIRMA (parte inferior) ──
    y_firma = 260

    c.setFont("Helvetica", 11)
    c.setFillColor(colors.black)
    c.drawString(ml, y_firma, "Sin otro particular,")

    y_firma -= 22
    fecha_hoy = _fecha_es(datetime.now())
    c.drawString(ml, y_firma, f"En {lugar}, a {fecha_hoy}")

    # Bloque de firma: "Por el HINCADO DIRECTO" + "Fdo." + firma + línea + sello
    firma_path = _ASSETS_DIR / "firma_sg.png"
    sello_path = _ASSETS_DIR / "sello_hincado.png"

    # Centrar bloque de firma
    firma_cx = ml + content_w / 2  # centro horizontal

    y_firma -= 24
    c.setFont("Helvetica-Bold", 11)
    c.drawCentredString(firma_cx, y_firma, "Hincado Directo")

    y_firma -= 18
    c.setFont("Helvetica-Bold", 11)
    c.drawCentredString(firma_cx, y_firma, f"Fdo. {firmante_nombre}")

    # Firma manuscrita (encima de la línea)
    y_firma -= 4
    if firma_path.exists():
        try:
            img_f = ImageReader(str(firma_path))
            fw, fh = img_f.getSize()
            target_fw = 90
            ratio_f = target_fw / fw
            target_fh = fh * ratio_f
            c.drawImage(str(firma_path), firma_cx - target_fw / 2,
                        y_firma - target_fh,
                        width=target_fw, height=target_fh,
                        preserveAspectRatio=True, mask="auto")
            y_firma -= target_fh
        except Exception:
            y_firma -= 50

    # Línea horizontal separadora
    y_firma -= 1
    line_w = 140
    c.setStrokeColor(colors.HexColor("#1E293B"))
    c.setLineWidth(0.8)
    c.line(firma_cx - line_w / 2, y_firma, firma_cx + line_w / 2, y_firma)

    # Sello (debajo de la línea)
    y_firma -= 2
    if sello_path.exists():
        try:
            img_s = ImageReader(str(sello_path))
            sw, sh = img_s.getSize()
            target_sw = 120
            ratio_s = target_sw / sw
            target_sh = sh * ratio_s
            c.drawImage(str(sello_path), firma_cx - target_sw / 2,
                        y_firma - target_sh,
                        width=target_sw, height=target_sh,
                        preserveAspectRatio=True, mask="auto")
            y_firma -= target_sh
        except Exception:
            pass

    # ── PIE ──
    c.setFont("Helvetica", 7)
    c.setFillColor(colors.HexColor("#64748B"))
    pie = (f"{EMPRESA['nombre']} - C.I.F.: {EMPRESA['cif']} - "
           f"Reg: {EMPRESA['registro']}")
    c.drawCentredString(width / 2, 35, pie)

    # Línea separadora pie
    c.setStrokeColor(colors.HexColor("#E2E8F0"))
    c.setLineWidth(0.5)
    c.line(ml, 50, width - mr, 50)

    c.save()
    pdf_bytes = buffer.getvalue()

    # ── Guardar y registrar ──
    _ensure_exports_dir()
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    nombre_limpio = maq["nombre"].replace(" ", "_")
    horas_tag = f"{int(rev_horas)}h"
    filename = f"Certificado_{nombre_limpio}_{horas_tag}_{ts}.pdf"
    filepath = str(_EXPORTS_DIR / filename)
    with open(filepath, "wb") as f:
        f.write(pdf_bytes)

    titulo = f"Certificado CAE — {maq['nombre']} — {horas_tag}"
    doc_record = _register_doc(
        maquina_id, "certificado_cae", titulo, filename, filepath, pdf_bytes,
        generado_por=generado_por,
        metadata={"modo": modo, "hito_horas": int(rev_horas),
                   "fecha_revision": rev_fecha, "firmante": firmante_nombre},
    )
    return pdf_bytes, doc_record


# ══════════════════════════════════════════════════════════════════════════════
# ██  ASSET PASSPORT — Charts helper (page 2)                               ██
# ══════════════════════════════════════════════════════════════════════════════

def _draw_passport_charts(c, w, h, ml, mr, cw, maq, maquina_id,
                          _GO, _GM, _GC, _AZ, horo, horo_ini):
    """Adds page 2 to the Asset Passport: hourometer charts + summary stats.

    Uses matplotlib (Agg backend) to render two charts as PNG images:
      1. Cumulative hourometer line chart with area fill
      2. Biweekly consumption bar chart with average line
    Then embeds them in the ReportLab canvas via ImageReader.
    """
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    import matplotlib.dates as mdates
    from matplotlib.ticker import MaxNLocator
    from datetime import timedelta

    _VE_HEX = "#16A34A"
    _AZ_HEX = "#2563EB"
    _GO_HEX = "#1E293B"

    # ── Query hourometer readings from DB ──
    # Historical maintenance logs store horometro_at as the hourometer reading
    # at the time of each maintenance.  For PDF-imported history these are
    # milestone values (500, 1000 …) that may NOT be date-monotonic (a 500h
    # service dated after a 1000h one because it was done retroactively).
    #
    # Strategy:
    #   1. Collect (date, horo) from checks + maintenance_logs
    #   2. Do NOT anchor with horometro_inicial — that is the value when the
    #      machine was registered in the ERP, which can be higher than all
    #      imported milestones and would flatten the chart.
    #   3. Add (today, horometro_actual) as final anchor.
    #   4. Sort by date, keep max horo per date, enforce monotonicity.
    from core.db import conectar as _db_conectar
    with _db_conectar() as conn:
        # Source 1: maquinaria_checks (fecha, horometro)
        rows_checks = conn.execute(
            "SELECT fecha, horometro FROM maquinaria_checks "
            "WHERE maquina_id = ? AND horometro IS NOT NULL AND horometro > 0 "
            "AND estado != 'enmendado' "
            "ORDER BY fecha",
            [maquina_id],
        ).fetchall()

        # Source 2: maquinaria_maintenance_logs — use MAX(horometro_at) per
        # completed_at date to collapse multiple task rows into one reading
        rows_logs = conn.execute(
            "SELECT completed_at, MAX(horometro_at) as horo "
            "FROM maquinaria_maintenance_logs "
            "WHERE maquina_id = ? AND horometro_at IS NOT NULL AND horometro_at > 0 "
            "GROUP BY completed_at "
            "ORDER BY completed_at",
            [maquina_id],
        ).fetchall()

    # Combine into (date, horo) pairs
    combined = []
    for r in rows_checks:
        try:
            d = datetime.fromisoformat(r["fecha"][:10])
            combined.append((d, float(r["horometro"])))
        except (ValueError, TypeError):
            pass
    for r in rows_logs:
        try:
            d_str = r["completed_at"] or ""
            d = datetime.fromisoformat(d_str[:10])
            combined.append((d, float(r["horo"])))
        except (ValueError, TypeError):
            pass

    # Anchor: today → horometro_actual (always a reliable endpoint)
    if horo > 0:
        combined.append((datetime.now().replace(hour=0, minute=0, second=0, microsecond=0),
                         float(horo)))

    # Sort by date, deduplicate keeping max horo per date
    combined.sort(key=lambda x: (x[0], x[1]))
    deduped = {}
    for d, hr in combined:
        if d not in deduped or hr > deduped[d]:
            deduped[d] = hr
    by_date = sorted(deduped.items(), key=lambda x: x[0])

    # Enforce strict monotonicity: keep only readings where horo >= running max
    sorted_readings = []
    running_max = -1.0
    for d, hr in by_date:
        if hr >= running_max:
            sorted_readings.append((d, hr))
            running_max = hr

    # Need at least 2 readings to draw charts
    if len(sorted_readings) < 2:
        # Not enough data — add a simple text page instead
        c.showPage()
        bar_h = 50
        bar_y = h - 40 - bar_h
        c.setFillColor(_GO)
        c.rect(0, bar_y, w, bar_h, fill=1, stroke=0)
        logo_path = _ASSETS_DIR / "logo_hincado_directo.png"
        if logo_path.exists():
            try:
                c.drawImage(str(logo_path), ml, bar_y + 8, width=100, height=32,
                            preserveAspectRatio=True, mask="auto")
            except Exception:
                pass
        c.setFillColor(colors.white)
        c.setFont("Helvetica-Bold", 16)
        c.drawRightString(w - mr, bar_y + 28, f"ASSET PASSPORT — {maq['nombre']}")
        c.setFont("Helvetica", 9)
        c.drawRightString(w - mr, bar_y + 12, "Análisis de consumo de horas")

        c.setFillColor(_GM)
        c.setFont("Helvetica", 12)
        c.drawCentredString(w / 2, h / 2,
                            "Datos insuficientes para generar gráficos de consumo.")
        c.setFont("Helvetica", 9)
        c.drawCentredString(w / 2, h / 2 - 20,
                            "Se necesitan al menos 2 lecturas de horómetro con fecha.")
        # Footer
        c.setFont("Helvetica", 7)
        c.setFillColor(_GM)
        c.drawCentredString(w / 2, 25,
                            f"{EMPRESA['nombre']} — Documento interno de análisis operativo")
        c.setStrokeColor(colors.HexColor("#E2E8F0"))
        c.setLineWidth(0.5)
        c.line(ml, 35, w - mr, 35)
        return

    dates = [x[0] for x in sorted_readings]
    horos_list = [x[1] for x in sorted_readings]

    # ═══ CHART 1: Cumulative hours (line chart) ═══
    fig1, ax1 = plt.subplots(figsize=(7.2, 2.8))
    ax1.fill_between(dates, horos_list, alpha=0.15, color=_AZ_HEX)
    ax1.plot(dates, horos_list, color=_AZ_HEX, linewidth=2, marker="o", markersize=3)
    ax1.set_ylabel("Horómetro (h)", fontsize=9, color=_GO_HEX)
    ax1.set_title("Evolución del horómetro", fontsize=11, fontweight="bold",
                  color=_GO_HEX, pad=10)
    ax1.xaxis.set_major_formatter(mdates.DateFormatter("%b %Y"))
    # Smart tick interval: aim for ~8-12 ticks regardless of date span
    span_months = max(1, (dates[-1] - dates[0]).days // 30)
    tick_interval = max(1, span_months // 10)
    ax1.xaxis.set_major_locator(mdates.MonthLocator(interval=tick_interval))
    plt.xticks(fontsize=7, rotation=30)
    plt.yticks(fontsize=8)
    ax1.grid(axis="y", alpha=0.3)
    ax1.spines["top"].set_visible(False)
    ax1.spines["right"].set_visible(False)
    fig1.tight_layout()
    buf1 = io.BytesIO()
    fig1.savefig(buf1, format="png", dpi=150, bbox_inches="tight")
    plt.close(fig1)
    buf1.seek(0)

    # ═══ CHART 2: Biweekly consumption (bar chart) ═══
    period_days = 14
    start, end = dates[0], dates[-1]
    periods = []
    current = start
    while current <= end:
        periods.append(current)
        current += timedelta(days=period_days)

    # Interpolate hourometer at each period boundary
    interp_horos = []
    for p in periods:
        before_h, after_h = 0, horos_list[-1]
        before_d, after_d = dates[0], dates[-1]
        for i in range(len(dates)):
            if dates[i] <= p:
                before_h = horos_list[i]
                before_d = dates[i]
            if dates[i] >= p:
                after_h = horos_list[i]
                after_d = dates[i]
                break
        if before_d == after_d:
            interp_horos.append(before_h)
        else:
            ratio = (p - before_d).total_seconds() / (after_d - before_d).total_seconds()
            interp_horos.append(before_h + (after_h - before_h) * ratio)

    # Calculate consumption per period
    consumptions = []
    period_labels = []
    for i in range(1, len(periods)):
        delta = interp_horos[i] - interp_horos[i - 1]
        consumptions.append(max(0, delta))
        period_labels.append(periods[i])

    # Only show last 26 periods (~1 year)
    max_bars = 26
    if len(consumptions) > max_bars:
        consumptions = consumptions[-max_bars:]
        period_labels = period_labels[-max_bars:]

    fig2, ax2 = plt.subplots(figsize=(7.2, 2.8))
    bar_colors = [_AZ_HEX if c_val > 0 else "#E2E8F0" for c_val in consumptions]
    ax2.bar(range(len(consumptions)), consumptions, color=bar_colors, alpha=0.8, width=0.7)

    if consumptions:
        avg = sum(consumptions) / len(consumptions)
        ax2.axhline(y=avg, color=_VE_HEX, linestyle="--", linewidth=1.5, alpha=0.7,
                    label=f"Media: {avg:.0f}h")
        ax2.legend(fontsize=8, loc="upper right")

    ax2.set_ylabel("Horas consumidas", fontsize=9, color=_GO_HEX)
    ax2.set_title("Consumo bisemanal de horas", fontsize=11, fontweight="bold",
                  color=_GO_HEX, pad=10)

    tick_positions = list(range(0, len(period_labels), 4))
    tick_labels_str = [period_labels[i].strftime("%d/%m/%y") for i in tick_positions]
    ax2.set_xticks(tick_positions)
    ax2.set_xticklabels(tick_labels_str, fontsize=7, rotation=30)
    plt.yticks(fontsize=8)
    ax2.yaxis.set_major_locator(MaxNLocator(integer=True))
    ax2.grid(axis="y", alpha=0.3)
    ax2.spines["top"].set_visible(False)
    ax2.spines["right"].set_visible(False)
    fig2.tight_layout()
    buf2 = io.BytesIO()
    fig2.savefig(buf2, format="png", dpi=150, bbox_inches="tight")
    plt.close(fig2)
    buf2.seek(0)

    # ═══ BUILD PDF PAGE 2 ═══
    c.showPage()

    # Header bar
    bar_h = 50
    bar_y = h - 40 - bar_h
    c.setFillColor(_GO)
    c.rect(0, bar_y, w, bar_h, fill=1, stroke=0)
    logo_path = _ASSETS_DIR / "logo_hincado_directo.png"
    if logo_path.exists():
        try:
            c.drawImage(str(logo_path), ml, bar_y + 8, width=100, height=32,
                        preserveAspectRatio=True, mask="auto")
        except Exception:
            pass
    c.setFillColor(colors.white)
    c.setFont("Helvetica-Bold", 16)
    c.drawRightString(w - mr, bar_y + 28, f"ASSET PASSPORT — {maq['nombre']}")
    c.setFont("Helvetica", 9)
    c.drawRightString(w - mr, bar_y + 12, "Análisis de consumo de horas")

    y = bar_y - 25

    # Chart 1: Cumulative
    img1 = ImageReader(buf1)
    chart_w = cw
    chart_h = 200
    c.drawImage(img1, ml, y - chart_h, width=chart_w, height=chart_h,
                preserveAspectRatio=True, mask="auto")
    y -= chart_h + 25

    # Chart 2: Biweekly consumption
    img2 = ImageReader(buf2)
    c.drawImage(img2, ml, y - chart_h, width=chart_w, height=chart_h,
                preserveAspectRatio=True, mask="auto")
    y -= chart_h + 20

    # Summary stats box
    c.setFillColor(_GC)
    c.roundRect(ml, y - 60, cw, 60, 4, fill=1, stroke=0)

    c.setFillColor(_GO)
    c.setFont("Helvetica-Bold", 10)
    c.drawString(ml + 12, y - 16, "Resumen de actividad")

    c.setFont("Helvetica", 9)
    c.setFillColor(_GM)
    total_hours = horos_list[-1] - horos_list[0]
    total_days = max((dates[-1] - dates[0]).days, 1)
    avg_daily = total_hours / total_days
    avg_weekly = avg_daily * 7
    avg_monthly = avg_daily * 30

    stats_text = (
        f"Período analizado: {dates[0].strftime('%d/%m/%Y')} — {dates[-1].strftime('%d/%m/%Y')}  ·  "
        f"Horas totales operadas: {total_hours:,.0f}h  ·  "
        f"Media semanal: {avg_weekly:,.1f}h  ·  "
        f"Media mensual: {avg_monthly:,.0f}h"
    )
    c.drawString(ml + 12, y - 36, stats_text.replace(",", "."))

    # Utilización (assuming 10h/day, 5 days/week = 50h/week potential)
    utilizacion = min(100, avg_weekly / 50 * 100)
    c.drawString(ml + 12, y - 52,
                 f"Utilización estimada: {utilizacion:.0f}% (sobre 50h/semana potencial)")

    # Footer
    c.setFont("Helvetica", 7)
    c.setFillColor(_GM)
    c.drawCentredString(w / 2, 25,
                        f"{EMPRESA['nombre']} — Documento interno de análisis operativo")
    c.setStrokeColor(colors.HexColor("#E2E8F0"))
    c.setLineWidth(0.5)
    c.line(ml, 35, w - mr, 35)


# ══════════════════════════════════════════════════════════════════════════════
# ██  ASSET PASSPORT — 1-page executive summary                             ██
# ══════════════════════════════════════════════════════════════════════════════

def generar_asset_passport(maquina_id: int, generado_por: str = None) -> tuple[bytes, dict]:
    """Genera un Asset Passport PDF de 1 página con resumen ejecutivo de la máquina."""
    from reportlab.pdfgen import canvas as rl_canvas

    maquinaria_db.init_maquinaria_db()

    # ── Obtener todos los datos ──
    historial = maquinaria_db.obtener_historial_servicio(maquina_id)
    if not historial:
        raise ValueError("Máquina no encontrada")

    maq = historial["maquina"]
    revisiones = historial["revisiones"]
    checks = historial["checks"]
    incidencias = historial["incidencias"]

    from core.db import conectar as _db_conectar
    with _db_conectar() as conn:
        from core.maquinaria_db import _calcular_revisiones_pendientes
        rev_pendientes = _calcular_revisiones_pendientes(
            conn, maquina_id, maq["horometro_actual"])
        inc_total = conn.execute(
            "SELECT COUNT(*) FROM maquinaria_incidencias WHERE maquina_id = ?",
            [maquina_id]).fetchone()[0]

    horo = maq.get("horometro_actual") or 0
    horo_ini = maq.get("horometro_inicial") or 0
    horas_op = horo - horo_ini
    n_revs = len(revisiones)
    n_checks = len(checks)
    n_pend = len(rev_pendientes)
    n_inc_abiertas = len(incidencias)
    tiene_urgente = any(r.get("urgente") for r in rev_pendientes)

    if n_pend == 0:
        compliance = 100
    elif tiene_urgente:
        compliance = max(30, 100 - n_pend * 20)
    else:
        compliance = max(60, 100 - n_pend * 10)

    # ── Build PDF ──
    buffer = io.BytesIO()
    c = rl_canvas.Canvas(buffer, pagesize=A4)
    w, h = A4
    ml, mr = 40, 40
    cw = w - ml - mr

    _AZ = colors.HexColor("#2563EB")
    _VE = colors.HexColor("#16A34A")
    _RO = colors.HexColor("#DC2626")
    _AM = colors.HexColor("#CA8A04")
    _GO = _GRIS_OSCURO
    _GM = _GRIS_MEDIO
    _GC = _GRIS_CLARO

    def _fc(s):
        return s[:10] if s else "—"

    # ═══ HEADER BAR ═══
    bar_h = 65
    bar_y = h - 40 - bar_h
    c.setFillColor(_GO)
    c.rect(0, bar_y, w, bar_h, fill=1, stroke=0)
    logo_path = _ASSETS_DIR / "logo_hincado_directo.png"
    if logo_path.exists():
        try:
            c.drawImage(str(logo_path), ml, bar_y + 12, width=120, height=40,
                        preserveAspectRatio=True, mask="auto")
        except Exception:
            pass
    c.setFillColor(colors.white)
    c.setFont("Helvetica-Bold", 20)
    c.drawRightString(w - mr, bar_y + 38, "ASSET PASSPORT")
    c.setFont("Helvetica", 9)
    c.drawRightString(w - mr, bar_y + 22, f"Generado: {_fecha_es(datetime.now())}")
    c.setFont("Helvetica", 8)
    c.drawRightString(w - mr, bar_y + 10, f"Documento interno — {EMPRESA['nombre']}")

    y = bar_y - 20

    # ═══ MACHINE IDENTITY ═══
    c.setFillColor(_AZ)
    c.rect(ml, y - 4, cw, 4, fill=1, stroke=0)
    y -= 18
    c.setFillColor(_GO)
    c.setFont("Helvetica-Bold", 18)
    c.drawString(ml, y, maq["nombre"])

    estado_colors = {"disponible": _VE, "en_proyecto": _AZ, "en_taller": _AM, "baja": _RO}
    estado_labels = {"disponible": "DISPONIBLE", "en_proyecto": "EN PROYECTO",
                     "en_taller": "EN TALLER", "baja": "DE BAJA"}
    est = maq.get("estado_computado") or maq.get("estado", "disponible")
    ec = estado_colors.get(est, _GM)
    el = estado_labels.get(est, est.upper())
    bw = c.stringWidth(el, "Helvetica-Bold", 9) + 16
    bx = w - mr - bw
    c.setFillColor(ec)
    c.roundRect(bx, y - 2, bw, 16, 3, fill=1, stroke=0)
    c.setFillColor(colors.white)
    c.setFont("Helvetica-Bold", 9)
    c.drawCentredString(bx + bw / 2, y + 2, el)

    y -= 18
    c.setFillColor(_GM)
    c.setFont("Helvetica", 10)
    parts = [maq.get("internal_id", "")]
    if maq.get("modelo"): parts.append(maq["modelo"])
    if maq.get("numero_serie"): parts.append(f"S/N: {maq['numero_serie']}")
    if maq.get("proyecto_nombre"): parts.append(f"Proyecto: {maq['proyecto_nombre']}")
    c.drawString(ml, y, " · ".join(parts))
    y -= 30

    # ═══ KPI BOXES ═══
    def draw_kpi(x, yy, kw, kh, label, value, color, sublabel=None):
        c.setStrokeColor(colors.HexColor("#E2E8F0")); c.setLineWidth(0.5)
        c.setFillColor(colors.white)
        c.roundRect(x, yy, kw, kh, 4, fill=1, stroke=1)
        c.setFillColor(color)
        c.rect(x, yy + kh - 3, kw, 3, fill=1, stroke=0)
        c.setFillColor(_GM); c.setFont("Helvetica", 8)
        c.drawString(x + 10, yy + kh - 18, label)
        c.setFillColor(color); c.setFont("Helvetica-Bold", 20)
        c.drawString(x + 10, yy + 12, str(value))
        if sublabel:
            c.setFillColor(_GM); c.setFont("Helvetica", 7)
            c.drawString(x + 10, yy + 4, sublabel)

    kh = 58; kg = 10; kw = (cw - 3 * kg) / 4
    draw_kpi(ml, y - kh, kw, kh, "HORÓMETRO ACTUAL",
             f"{horo:,.0f}h".replace(",", "."), _AZ,
             f"Inicial: {horo_ini:,.0f}h · Operadas: {horas_op:,.0f}h".replace(",", "."))
    draw_kpi(ml + kw + kg, y - kh, kw, kh, "REVISIONES REALIZADAS",
             str(n_revs), _VE,
             f"Última: {_fc(revisiones[0]['fecha']) if revisiones else '—'}")
    draw_kpi(ml + 2*(kw+kg), y - kh, kw, kh, "CHECKS SEMANALES",
             str(n_checks), _GO,
             f"Último: {_fc(checks[0]['fecha']) if checks else '—'}")
    pc = _VE if n_pend == 0 else (_RO if tiene_urgente else _AM)
    pv = "✓ OK" if n_pend == 0 else str(n_pend)
    ps = "Todas al día" if n_pend == 0 else \
        f"Próx: {rev_pendientes[0]['proximo_hito']:,.0f}h".replace(",",".")
    draw_kpi(ml + 3*(kw+kg), y - kh, kw, kh, "REVISIONES PENDIENTES", pv, pc, ps)
    y -= kh + 20

    # ═══ TWO COLUMNS ═══
    col_w = (cw - 14) / 2
    cl = ml; cr = ml + col_w + 14

    def sec_hdr(x, yy, title):
        c.setFillColor(_GC)
        c.roundRect(x, yy - 16, col_w, 18, 3, fill=1, stroke=0)
        c.setFillColor(_GO); c.setFont("Helvetica-Bold", 10)
        c.drawString(x + 8, yy - 12, title)
        return yy - 22

    # LEFT: Historial
    yl = sec_hdr(cl, y, "HISTORIAL DE REVISIONES")
    mx = min(len(revisiones), 8)
    for i in range(mx):
        r = revisiones[i]; ry = yl - 4
        c.setFillColor(_AZ); c.setFont("Helvetica-Bold", 9)
        c.drawString(cl+6, ry, f"{r['horometro']:,.0f}h".replace(",","."))
        c.setFillColor(_GM); c.setFont("Helvetica", 8)
        c.drawString(cl+60, ry, _fc(r["fecha"]))
        c.drawString(cl+130, ry, f"{len(r['tareas'])} tareas")
        c.drawString(cl+190, ry, r.get("operario") or "—")
        yl -= 14
        c.setStrokeColor(colors.HexColor("#E2E8F0")); c.setLineWidth(0.3)
        c.line(cl+4, yl+2, cl+col_w-4, yl+2)
    if len(revisiones) > mx:
        c.setFillColor(_GM); c.setFont("Helvetica-Oblique", 8)
        c.drawString(cl+6, yl-4, f"... y {len(revisiones)-mx} más"); yl -= 14

    # LEFT: Incidencias
    yl -= 10
    yl = sec_hdr(cl, yl, "INCIDENCIAS")
    if n_inc_abiertas == 0:
        c.setFillColor(_VE); c.setFont("Helvetica-Bold", 9)
        c.drawString(cl+6, yl-4, "✓ Sin incidencias abiertas")
        if inc_total:
            c.setFillColor(_GM); c.setFont("Helvetica", 8)
            c.drawString(cl+150, yl-4, f"({inc_total} históricas cerradas)")
    else:
        c.setFillColor(_RO); c.setFont("Helvetica-Bold", 9)
        c.drawString(cl+6, yl-4, f"{n_inc_abiertas} incidencia(s) abierta(s)")

    # RIGHT: Identificación
    yr = sec_hdr(cr, y, "IDENTIFICACIÓN DEL ACTIVO")
    def irow(x, yy, lbl, val):
        c.setFillColor(_GM); c.setFont("Helvetica", 8)
        c.drawString(x+6, yy, lbl)
        c.setFillColor(_GO); c.setFont("Helvetica-Bold", 9)
        c.drawString(x+110, yy, str(val) if val else "—")
        return yy - 16
    yr = irow(cr, yr-4, "Nombre:", maq["nombre"])
    yr = irow(cr, yr, "ID interno:", maq.get("internal_id"))
    yr = irow(cr, yr, "Modelo:", maq.get("modelo"))
    yr = irow(cr, yr, "Nº Serie:", maq.get("numero_serie"))
    yr = irow(cr, yr, "Fecha comisión:", _fc(maq.get("fecha_comision")))
    yr = irow(cr, yr, "Ubicación:", maq.get("ubicacion") or maq.get("proyecto_nombre"))

    # RIGHT: Pendientes
    yr -= 10
    yr = sec_hdr(cr, yr, "REVISIONES PENDIENTES")
    if n_pend == 0:
        c.setFillColor(_VE); c.setFont("Helvetica-Bold", 9)
        c.drawString(cr+6, yr-4, "✓ Todas las revisiones al día")
    else:
        for rp in rev_pendientes[:5]:
            c.setFillColor(_RO if rp.get("urgente") else _AM)
            c.setFont("Helvetica-Bold", 9)
            c.drawString(cr+6, yr-4, f"● {rp['proximo_hito']:,.0f}h".replace(",","."))
            c.setFillColor(_GM); c.setFont("Helvetica", 8)
            c.drawString(cr+80, yr-4, f"Intervalo {rp['tipo']}")
            if rp.get("urgente"):
                c.setFillColor(_RO); c.setFont("Helvetica-Bold", 8)
                c.drawString(cr+170, yr-4, "¡ATRASADA!")
            yr -= 16

    # RIGHT: Cumplimiento
    yr -= 10
    yr = sec_hdr(cr, yr, "CUMPLIMIENTO")
    bbx = cr + 6; bbw = col_w - 20; bbh = 14; bby = yr - 8
    c.setFillColor(colors.HexColor("#E2E8F0"))
    c.roundRect(bbx, bby, bbw, bbh, 3, fill=1, stroke=0)
    cc = _VE if compliance >= 90 else (_AM if compliance >= 70 else _RO)
    c.setFillColor(cc)
    c.roundRect(bbx, bby, max(bbw * compliance / 100, 20), bbh, 3, fill=1, stroke=0)
    c.setFillColor(colors.white if compliance > 30 else _GO)
    c.setFont("Helvetica-Bold", 9)
    c.drawCentredString(bbx + bbw/2, bby+3, f"{compliance}% cumplimiento")

    # ═══ QR Code — link to Auditor View ═══
    try:
        from core.db import conectar as _db_qr
        with _db_qr() as conn:
            qr_row = conn.execute(
                "SELECT token FROM maquinaria_auditor_links "
                "WHERE maquina_id = ? AND revocado = 0 AND expires_at > ? "
                "ORDER BY created_at DESC LIMIT 1",
                [maquina_id, datetime.now().isoformat()],
            ).fetchone()
        if qr_row:
            from reportlab.graphics.barcode import qr as qr_mod
            from reportlab.graphics.shapes import Drawing
            from reportlab.graphics import renderPDF
            audit_url = f"/audit/{qr_row['token']}"
            qr_code = qr_mod.QrCodeWidget(audit_url)
            qr_code.barWidth = 60
            qr_code.barHeight = 60
            qr_code.qrVersion = 4
            d = Drawing(60, 60)
            d.add(qr_code)
            renderPDF.draw(d, c, w - mr - 65, 40)
            c.setFont("Helvetica", 6)
            c.setFillColor(_GM)
            c.drawCentredString(w - mr - 35, 36, "Auditor View")
    except Exception:
        pass  # QR is optional, don't break passport generation

    # ═══ FOOTER ═══
    c.setFont("Helvetica", 7); c.setFillColor(_GM)
    c.drawCentredString(w/2, 25,
        f"{EMPRESA['nombre']} — C.I.F.: {EMPRESA['cif']} — {EMPRESA['registro']}")
    c.drawCentredString(w/2, 15,
        "Este documento es un resumen interno y no constituye certificación oficial. "
        f"Generado automáticamente el {_fecha_es(datetime.now())}.")
    c.setStrokeColor(colors.HexColor("#E2E8F0")); c.setLineWidth(0.5)
    c.line(ml, 35, w - mr, 35)

    # ═══ PAGE 2: Charts ═══
    _draw_passport_charts(c, w, h, ml, mr, cw, maq, maquina_id,
                          _GO, _GM, _GC, _AZ, horo, horo_ini)

    c.save()
    pdf_bytes = buffer.getvalue()

    # ── Guardar y registrar ──
    _ensure_exports_dir()
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    nombre_limpio = maq["nombre"].replace(" ", "_")
    filename = f"AssetPassport_{nombre_limpio}_{ts}.pdf"
    filepath = str(_EXPORTS_DIR / filename)
    with open(filepath, "wb") as f:
        f.write(pdf_bytes)

    titulo = f"Asset Passport — {maq['nombre']}"
    doc_record = _register_doc(
        maquina_id, "asset_passport", titulo, filename, filepath, pdf_bytes,
        generado_por=generado_por,
        metadata={"horometro": horo, "revisiones": n_revs,
                  "pendientes": n_pend, "compliance": compliance},
    )
    return pdf_bytes, doc_record


# ══════════════════════════════════════════════════════════════════════════════
# ██  INFORME DE DISPONIBILIDAD — PDF                                       ██
# ══════════════════════════════════════════════════════════════════════════════


def generar_informe_disponibilidad(maquina_id: int, dias: int = 90, generado_por: str = None) -> tuple[bytes, dict]:
    """Genera un Informe de Disponibilidad PDF de 1 página con KPIs de uptime/downtime."""
    from reportlab.pdfgen import canvas as rl_canvas

    maquinaria_db.init_maquinaria_db()

    # ── Obtener datos ──
    maq = maquinaria_db.obtener_maquina(maquina_id)
    if not maq:
        raise ValueError("Máquina no encontrada")

    disp = maquinaria_db.calcular_disponibilidad(maquina_id, dias)
    incidencias = maquinaria_db.listar_incidencias(maquina_id=maquina_id, limit=10)

    # ── Colores locales ──
    _AZ = _AZUL
    _VE = _VERDE
    _RO = _ROJO
    _AM = _AMARILLO
    _GO = _GRIS_OSCURO
    _GM = _GRIS_MEDIO
    _GC = _GRIS_CLARO

    # ── Build PDF ──
    buffer = io.BytesIO()
    c = rl_canvas.Canvas(buffer, pagesize=A4)
    w, h = A4
    ml, mr = 40, 40
    cw = w - ml - mr

    # ═══ HEADER BAR ═══
    bar_h = 65
    bar_y = h - 40 - bar_h
    c.setFillColor(_GO)
    c.rect(0, bar_y, w, bar_h, fill=1, stroke=0)
    logo_path = _find_logo()
    if logo_path and logo_path.exists():
        try:
            c.drawImage(str(logo_path), ml, bar_y + 12, width=120, height=40,
                        preserveAspectRatio=True, mask="auto")
        except Exception:
            pass
    c.setFillColor(colors.white)
    c.setFont("Helvetica-Bold", 18)
    c.drawRightString(w - mr, bar_y + 38, "INFORME DE DISPONIBILIDAD")
    c.setFont("Helvetica", 9)
    c.drawRightString(w - mr, bar_y + 22, f"Generado: {_fecha_es(datetime.now())}")
    c.setFont("Helvetica", 8)
    c.drawRightString(w - mr, bar_y + 10, f"Documento interno — {EMPRESA['nombre']}")

    y = bar_y - 20

    # ═══ MACHINE IDENTITY ═══
    c.setFillColor(_AZ)
    c.rect(ml, y - 4, cw, 4, fill=1, stroke=0)
    y -= 18
    c.setFillColor(_GO)
    c.setFont("Helvetica-Bold", 16)
    c.drawString(ml, y, maq["nombre"])
    y -= 16
    c.setFillColor(_GM)
    c.setFont("Helvetica", 10)
    parts = []
    if maq.get("modelo"):
        parts.append(maq["modelo"])
    if maq.get("matricula"):
        parts.append(f"Matrícula: {maq['matricula']}")
    horo = maq.get("horometro_actual") or 0
    parts.append(f"Horómetro: {horo:,.0f}h".replace(",", "."))
    c.drawString(ml, y, " · ".join(parts))
    y -= 25

    # ═══ PERIOD INFO ═══
    c.setFillColor(_GO)
    c.setFont("Helvetica-Bold", 11)
    c.drawString(ml, y, f"Periodo analizado: Últimos {dias} días")
    y -= 8
    c.setStrokeColor(colors.HexColor("#E2E8F0"))
    c.setLineWidth(0.5)
    c.line(ml, y, w - mr, y)
    y -= 15

    # ═══ KPI BOXES (4 boxes) ═══
    def draw_kpi(x, yy, kw, kh, label, value, color, sublabel=None):
        c.setStrokeColor(colors.HexColor("#E2E8F0"))
        c.setLineWidth(0.5)
        c.setFillColor(colors.white)
        c.roundRect(x, yy, kw, kh, 4, fill=1, stroke=1)
        c.setFillColor(color)
        c.rect(x, yy + kh - 3, kw, 3, fill=1, stroke=0)
        c.setFillColor(_GM)
        c.setFont("Helvetica", 8)
        c.drawString(x + 10, yy + kh - 18, label)
        c.setFillColor(color)
        c.setFont("Helvetica-Bold", 20)
        c.drawString(x + 10, yy + 12, str(value))
        if sublabel:
            c.setFillColor(_GM)
            c.setFont("Helvetica", 7)
            c.drawString(x + 10, yy + 4, sublabel)

    kh = 58
    kg = 10
    kw = (cw - 3 * kg) / 4
    draw_kpi(ml, y - kh, kw, kh, "HORAS DOWNTIME",
             f"{disp['horas_downtime']:.1f}h",
             _RO if disp['horas_downtime'] > 0 else _VE,
             f"en los últimos {dias} días")
    draw_kpi(ml + kw + kg, y - kh, kw, kh, "DÍAS PARADOS",
             f"{disp['dias_parados']:.1f}",
             _RO if disp['dias_parados'] > 0 else _VE,
             "jornadas de 8h equiv.")
    draw_kpi(ml + 2 * (kw + kg), y - kh, kw, kh, "MTTR (HORAS)",
             f"{disp['mttr_horas']:.1f}",
             _AM if disp['mttr_horas'] > 0 else _VE,
             "tiempo medio reparación")
    draw_kpi(ml + 3 * (kw + kg), y - kh, kw, kh, "INCID. / 100h",
             f"{disp['incidencias_por_100h']:.2f}",
             _AM if disp['incidencias_por_100h'] > 0 else _VE,
             f"total histórico")

    y = y - kh - 20

    # ═══ RESUMEN ECONÓMICO ═══
    c.setFillColor(_GC)
    c.roundRect(ml, y - 32, cw, 32, 4, fill=1, stroke=0)
    c.setFillColor(_GO)
    c.setFont("Helvetica-Bold", 10)
    c.drawString(ml + 12, y - 14, "COSTE ACUMULADO (periodo)")
    c.setFillColor(_RO if disp['coste_acumulado'] > 0 else _VE)
    c.setFont("Helvetica-Bold", 14)
    c.drawRightString(w - mr - 12, y - 16,
                      f"{disp['coste_acumulado']:,.2f} €".replace(",", "X").replace(".", ",").replace("X", "."))
    y -= 52

    # ═══ TABLA INCIDENCIAS RECIENTES ═══
    c.setFillColor(_GO)
    c.setFont("Helvetica-Bold", 11)
    c.drawString(ml, y, "Últimas incidencias (máx. 10)")
    y -= 15

    # Table header
    col_widths = [65, 200, 65, 65, 70]  # fecha, descripción, severidad, estado, downtime
    headers = ["Fecha", "Descripción", "Severidad", "Estado", "Downtime (h)"]
    row_h = 18
    # Header row
    c.setFillColor(_GO)
    c.rect(ml, y - row_h, cw, row_h, fill=1, stroke=0)
    c.setFillColor(colors.white)
    c.setFont("Helvetica-Bold", 8)
    cx = ml
    for i, hdr in enumerate(headers):
        c.drawString(cx + 4, y - row_h + 5, hdr)
        cx += col_widths[i]

    y -= row_h

    if not incidencias:
        c.setFillColor(_GM)
        c.setFont("Helvetica", 9)
        c.drawString(ml + 4, y - 14, "Sin incidencias registradas para esta máquina.")
    else:
        sev_colors = {"baja": _VE, "media": _AM, "alta": _RO, "seguridad": _RO}
        for idx, inc in enumerate(incidencias[:10]):
            bg = colors.white if idx % 2 == 0 else _GC
            c.setFillColor(bg)
            c.rect(ml, y - row_h, cw, row_h, fill=1, stroke=0)
            c.setFillColor(_GO)
            c.setFont("Helvetica", 8)

            fecha = (inc.get("fecha") or "—")[:10]
            desc = inc.get("descripcion") or "—"
            if len(desc) > 50:
                desc = desc[:47] + "..."
            sev = inc.get("severidad") or "—"
            estado = inc.get("estado") or "—"
            dt_h = inc.get("horas_downtime")
            dt_str = f"{dt_h:.1f}" if dt_h else "—"

            cx = ml
            c.drawString(cx + 4, y - row_h + 5, fecha)
            cx += col_widths[0]
            c.drawString(cx + 4, y - row_h + 5, desc)
            cx += col_widths[1]
            # Severidad with color
            sc = sev_colors.get(sev, _GM)
            c.setFillColor(sc)
            c.setFont("Helvetica-Bold", 8)
            c.drawString(cx + 4, y - row_h + 5, sev.upper())
            cx += col_widths[2]
            c.setFillColor(_GO)
            c.setFont("Helvetica", 8)
            c.drawString(cx + 4, y - row_h + 5, estado.replace("_", " "))
            cx += col_widths[3]
            c.drawString(cx + 4, y - row_h + 5, dt_str)

            y -= row_h

    # ═══ FOOTER ═══
    c.setFont("Helvetica", 7)
    c.setFillColor(_GM)
    c.drawCentredString(w / 2, 25,
        f"{EMPRESA['nombre']} — C.I.F.: {EMPRESA['cif']} — {EMPRESA['registro']}")
    c.drawCentredString(w / 2, 15,
        "Este documento es un resumen interno y no constituye certificación oficial. "
        f"Generado automáticamente el {_fecha_es(datetime.now())}.")
    c.setStrokeColor(colors.HexColor("#E2E8F0"))
    c.setLineWidth(0.5)
    c.line(ml, 35, w - mr, 35)

    c.save()
    pdf_bytes = buffer.getvalue()

    # ── Guardar y registrar ──
    _ensure_exports_dir()
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    nombre_limpio = maq["nombre"].replace(" ", "_")
    filename = f"InformeDisponibilidad_{nombre_limpio}_{ts}.pdf"
    filepath = str(_EXPORTS_DIR / filename)
    with open(filepath, "wb") as f:
        f.write(pdf_bytes)

    titulo = f"Informe de Disponibilidad — {maq['nombre']} ({dias}d)"
    try:
        doc_record = _register_doc(
            maquina_id, "informe_disponibilidad_pdf", titulo, filename, filepath, pdf_bytes,
            generado_por=generado_por,
            metadata={"dias": dias, "horas_downtime": disp["horas_downtime"],
                      "dias_parados": disp["dias_parados"], "mttr": disp["mttr_horas"],
                      "coste_acumulado": disp["coste_acumulado"]},
        )
    except Exception:
        # DB CHECK constraint may not include this tipo yet (old schema) — return synthetic record
        doc_record = {"filename": filename, "filepath": filepath, "titulo": titulo,
                      "tipo": "informe_disponibilidad_pdf", "maquina_id": maquina_id}
    return pdf_bytes, doc_record
