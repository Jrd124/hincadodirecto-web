# -*- coding: utf-8 -*-
"""Estados Financieros — base de datos, parser e importador."""
from __future__ import annotations

import logging
import os
import re
import shutil
import tempfile
from datetime import datetime

logger = logging.getLogger("erp")


# ── DDL ──────────────────────────────────────────────────────────────────────

EEFF_DDL = """
CREATE TABLE IF NOT EXISTS eeff_periodos (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    sociedad TEXT NOT NULL,
    tipo TEXT NOT NULL CHECK(tipo IN ('balance','pyg','sumas_saldos')),
    periodo TEXT NOT NULL,
    fecha_desde TEXT NOT NULL,
    fecha_hasta TEXT NOT NULL,
    año INTEGER NOT NULL,
    trimestre INTEGER,
    fuente TEXT DEFAULT 'gestoria',
    fichero_origen TEXT,
    created_at TEXT NOT NULL
);

CREATE TABLE IF NOT EXISTS eeff_lineas (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    periodo_id INTEGER NOT NULL REFERENCES eeff_periodos(id) ON DELETE CASCADE,
    codigo_cuenta TEXT,
    descripcion TEXT NOT NULL,
    nivel INTEGER DEFAULT 0,
    tipo_nivel TEXT,
    importe_actual REAL DEFAULT 0,
    importe_anterior REAL,
    debe_periodo REAL,
    haber_periodo REAL,
    debe_acumulado REAL,
    haber_acumulado REAL,
    debe_saldo REAL,
    haber_saldo REAL,
    saldo_inicial REAL,
    orden INTEGER DEFAULT 0,
    created_at TEXT NOT NULL
);

CREATE UNIQUE INDEX IF NOT EXISTS idx_eeff_periodo
    ON eeff_periodos(sociedad, tipo, fecha_desde, fecha_hasta);

CREATE TABLE IF NOT EXISTS eeff_plan_cuentas (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    codigo TEXT NOT NULL UNIQUE,
    nombre TEXT NOT NULL,
    nivel1 TEXT NOT NULL,
    nivel2 TEXT,
    nivel3 TEXT,
    signo INTEGER DEFAULT 1,
    activo INTEGER DEFAULT 1,
    notas TEXT,
    created_at TEXT NOT NULL,
    updated_at TEXT
);

CREATE TABLE IF NOT EXISTS eeff_formulas (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    nombre TEXT NOT NULL UNIQUE,
    descripcion TEXT,
    formula TEXT NOT NULL,
    formato TEXT DEFAULT 'EUR',
    orden INTEGER DEFAULT 0,
    grupo TEXT DEFAULT 'metricas',
    activo INTEGER DEFAULT 1,
    created_at TEXT NOT NULL
);
"""


def crear_tablas(conn):
    conn.executescript(EEFF_DDL)


# ── Helpers ──────────────────────────────────────────────────────────────────

def _float(v):
    """Convierte valor de celda a float, tolerando None y strings."""
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return round(float(v), 2)
    s = str(v).strip().replace("\u20ac", "").replace(" ", "")
    if not s:
        return 0.0
    s = s.replace(".", "").replace(",", ".")
    try:
        return round(float(s), 2)
    except ValueError:
        return 0.0


def _detectar_nivel(codigo_str):
    """Devuelve (nivel, tipo_nivel) según el patrón del código PGC."""
    if not codigo_str:
        return (4, "cuenta")
    s = str(codigo_str)
    # Cuenta numérica (4-8 dígitos)
    if re.match(r"^\d{4,8}$", s.strip()):
        return (4, "cuenta")
    # A), B), C) — sección principal
    if re.match(r"^[A-Z]\)", s.strip()):
        return (0, "seccion")
    # A.1), A.2) etc — resultado agrupado
    if re.match(r"^[A-Z]\.\d\)", s.strip()):
        return (0, "resultado")
    # I. II. III. — romano
    if re.match(r"^\s*[IVX]+\.", s):
        return (1, "grupo")
    # 1. 2. 3. — numérico
    if re.match(r"^\s*\d+\.", s):
        return (2, "subgrupo")
    # a) b) c) — letra
    if re.match(r"^\s*[a-z]\)", s):
        return (3, "partida")
    return (2, "subgrupo")


def _parse_fecha(s):
    """Convierte '01/01/2024' o '2024-01-01' a 'YYYY-MM-DD'."""
    if not s:
        return ""
    s = str(s).strip()
    m = re.search(r"(\d{2})/(\d{2})/(\d{4})", s)
    if m:
        return f"{m.group(3)}-{m.group(2)}-{m.group(1)}"
    m = re.search(r"(\d{4})-(\d{2})-(\d{2})", s)
    if m:
        return m.group(0)
    return s


def _trimestre_de_fecha(fecha_hasta):
    """Devuelve 1-4 según el mes de fecha_hasta, o None."""
    try:
        m = int(fecha_hasta.split("-")[1])
        return (m - 1) // 3 + 1
    except Exception:
        return None


def _str(v):
    """Celda a string limpio."""
    if v is None:
        return ""
    return str(v).strip()


# ── Parser ───────────────────────────────────────────────────────────────────

def _abrir_workbook(filepath):
    """Abre fichero Excel (.xlsx o .xls-que-es-xlsx)."""
    import openpyxl
    try:
        return openpyxl.load_workbook(filepath, data_only=True)
    except Exception:
        # .XLS que internamente es .xlsx
        tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        tmp.close()
        shutil.copy(filepath, tmp.name)
        try:
            return openpyxl.load_workbook(tmp.name, data_only=True)
        finally:
            try:
                os.unlink(tmp.name)
            except OSError:
                pass


def _detectar_tipo_hoja(ws):
    """Lee filas 1-2 y devuelve tipo: balance, pyg, sumas_saldos, diario, None."""
    for r in range(1, 3):
        for c in range(1, 12):
            v = _str(ws.cell(r, c).value).upper()
            if "SUMAS Y SALDOS" in v or "SUMAS" in v and "SALDOS" in v:
                return "sumas_saldos"
            if "PERDIDAS Y GANANCIAS" in v:
                return "pyg"
            if "LIBRO DIARIO" in v:
                return "diario"
            if "BALANCE" in v:
                return "balance"
    return None


def _extraer_meta(ws):
    """Extrae empresa, desde, hasta de filas 2-4."""
    empresa, desde, hasta = "", "", ""
    for r in range(1, 6):
        v = _str(ws.cell(r, 1).value)
        if "EMPRESA:" in v.upper():
            empresa = v.split(":", 1)[1].strip()
        elif "DESDE:" in v.upper():
            desde = _parse_fecha(v.split(":", 1)[1].strip())
        elif "HASTA:" in v.upper():
            hasta = _parse_fecha(v.split(":", 1)[1].strip())
    return empresa, desde, hasta


def _detectar_columnas_balance(ws):
    """Detecta layout de columnas según formato.

    Returns: dict con keys desc_col, actual_col, anterior_col (1-based).
    """
    # Fila 5 contiene las cabeceras.
    row5 = [_str(ws.cell(5, c).value) for c in range(1, 12)]
    # Formato 2025: "EJERCICIO 20XX" en col 8 y 10
    for c in range(7, 11):
        if "EJERCICIO" in row5[c - 1].upper():
            # 2025 format: desc in col 3, actual in col 8, anterior in col 10
            ant = None
            for c2 in range(c + 1, 11):
                if "EJERCICIO" in row5[c2 - 1].upper():
                    ant = c2
                    break
            return {"desc_col": 3, "actual_col": c, "anterior_col": ant}

    # Formato 2021-2024: check cols 3 and 4
    if "EJERCICIO" in row5[2].upper():
        ant = 4 if "EJERCICIO" in row5[3].upper() else None
        # desc in col B(2) — check row 6
        desc_test = _str(ws.cell(6, 2).value)
        if desc_test:
            return {"desc_col": 2, "actual_col": 3, "anterior_col": ant}
        return {"desc_col": 3, "actual_col": 3, "anterior_col": ant}

    # Formato 2020: col D only
    if "EJERCICIO" in row5[3].upper():
        # desc in col C (2020) or col B
        desc_test = _str(ws.cell(6, 3).value)
        if desc_test:
            return {"desc_col": 3, "actual_col": 4, "anterior_col": None}
        return {"desc_col": 2, "actual_col": 4, "anterior_col": None}

    # Fallback: look for first numeric value in row 6
    for c in range(3, 11):
        v = ws.cell(6, c).value
        if isinstance(v, (int, float)):
            return {"desc_col": 2, "actual_col": c, "anterior_col": c + 1 if ws.cell(6, c + 1).value else None}

    return {"desc_col": 2, "actual_col": 3, "anterior_col": 4}


def _parsear_balance_pyg(ws, tipo):
    """Parsea hoja de Balance o P&G."""
    empresa, desde, hasta = _extraer_meta(ws)
    cols = _detectar_columnas_balance(ws)
    dc = cols["desc_col"]
    ac = cols["actual_col"]
    an = cols["anterior_col"]

    lineas = []
    orden = 0
    for r in range(6, ws.max_row + 1):
        codigo_raw = _str(ws.cell(r, 1).value)
        desc = _str(ws.cell(r, dc).value)

        # Skip empty or header-repeat rows
        if not codigo_raw and not desc:
            continue
        # Skip PASIVO/ACTIVO header repeats (e.g. "EJERCICIO 2024" standalone)
        if desc and re.match(r"^EJERCICIO\s+\d{4}$", desc.strip(), re.IGNORECASE):
            continue

        # Determine description — if desc is empty try the other column
        if not desc and dc == 2:
            desc = _str(ws.cell(r, 3).value)
        elif not desc and dc == 3:
            desc = _str(ws.cell(r, 2).value)
        if not desc:
            desc = codigo_raw

        # Detect level
        nivel, tipo_nivel = _detectar_nivel(codigo_raw)

        # For TOTAL rows, use the description text
        if "TOTAL" in desc.upper():
            nivel, tipo_nivel = 0, "total"

        actual = _float(ws.cell(r, ac).value)
        anterior = _float(ws.cell(r, an).value) if an else None

        lineas.append({
            "codigo_cuenta": codigo_raw.strip() if codigo_raw else None,
            "descripcion": desc.strip(),
            "nivel": nivel,
            "tipo_nivel": tipo_nivel,
            "importe_actual": actual,
            "importe_anterior": anterior,
            "orden": orden,
        })
        orden += 1

    return {
        "tipo": tipo,
        "empresa": empresa,
        "desde": desde,
        "hasta": hasta,
        "lineas": lineas,
    }


def _parsear_sumas_saldos(ws):
    """Parsea hoja de Sumas y Saldos."""
    empresa, desde, hasta = _extraer_meta(ws)

    # Cabecera fija en fila 5: CUENTA, NOMBRE, APERTURA/SALDO_INICIAL,
    # DEBE_PERIODO, HABER_PERIODO, DEBE_ACUM, HABER_ACUM, DEBE_SALDO, HABER_SALDO
    lineas = []
    orden = 0
    for r in range(6, ws.max_row + 1):
        cuenta = ws.cell(r, 1).value
        nombre = _str(ws.cell(r, 2).value)
        if cuenta is None and not nombre:
            continue
        # Skip totals row
        if nombre and "TOTAL" in nombre.upper():
            continue

        lineas.append({
            "codigo_cuenta": str(int(cuenta)) if isinstance(cuenta, (int, float)) else _str(cuenta),
            "descripcion": nombre,
            "nivel": 4,
            "tipo_nivel": "cuenta",
            "saldo_inicial": _float(ws.cell(r, 3).value),
            "debe_periodo": _float(ws.cell(r, 4).value),
            "haber_periodo": _float(ws.cell(r, 5).value),
            "debe_acumulado": _float(ws.cell(r, 6).value),
            "haber_acumulado": _float(ws.cell(r, 7).value),
            "debe_saldo": _float(ws.cell(r, 8).value),
            "haber_saldo": _float(ws.cell(r, 9).value),
            "orden": orden,
        })
        orden += 1

    return {
        "tipo": "sumas_saldos",
        "empresa": empresa,
        "desde": desde,
        "hasta": hasta,
        "lineas": lineas,
    }


def parse_fichero_eeff(filepath):
    """Parsea un fichero Excel de EEFF y devuelve lista de informes."""
    wb = _abrir_workbook(filepath)
    informes = []
    for sn in wb.sheetnames:
        ws = wb[sn]
        tipo = _detectar_tipo_hoja(ws)
        if tipo == "diario" or tipo is None:
            continue
        if tipo in ("balance", "pyg"):
            informe = _parsear_balance_pyg(ws, tipo)
        else:
            informe = _parsear_sumas_saldos(ws)
        if informe and informe["lineas"]:
            informe["hoja"] = sn
            informes.append(informe)
    wb.close()
    return informes


# ── Importador ───────────────────────────────────────────────────────────────

def _generar_nombre_periodo(desde, hasta, año, trimestre, tipo, meses_es):
    """Genera nombre legible para el periodo: 'Febrero 2025', 'Q1 2025', 'Anual 2024'."""
    if desde and hasta:
        mes_desde = desde[5:7] if len(desde) >= 7 else ""
        mes_hasta = hasta[5:7] if len(hasta) >= 7 else ""
        # Mensual: mismo mes en desde y hasta
        if mes_desde == mes_hasta and mes_desde:
            try:
                idx = int(mes_hasta) - 1
                return f"{meses_es[idx]} {año}"
            except (ValueError, IndexError):
                pass
        # Anual: enero a diciembre
        if mes_desde == "01" and mes_hasta == "12":
            return f"Anual {año}"
    # Trimestral
    if trimestre:
        return f"Q{trimestre} {año}"
    return str(año)


def importar_eeff(filepath, conn):
    """Parsea el fichero y guarda en BD. Si el periodo ya existe, lo reemplaza."""
    crear_tablas(conn)
    informes = parse_fichero_eeff(filepath)
    now = datetime.now().isoformat()
    filename = os.path.basename(filepath)
    resultados = []

    _MESES_ES = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
                  "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]

    for inf in informes:
        sociedad = inf["empresa"]
        tipo = inf["tipo"]
        desde = inf["desde"]
        hasta = inf["hasta"]
        año = int(hasta[:4]) if hasta else 0
        trimestre = _trimestre_de_fecha(hasta)
        periodo = _generar_nombre_periodo(desde, hasta, año, trimestre, tipo, _MESES_ES)

        # Delete existing period if any (upsert)
        reemplazado = False
        existing = conn.execute(
            "SELECT id FROM eeff_periodos WHERE sociedad = ? AND tipo = ? AND fecha_desde = ? AND fecha_hasta = ?",
            (sociedad, tipo, desde, hasta),
        ).fetchone()
        if existing:
            conn.execute("DELETE FROM eeff_lineas WHERE periodo_id = ?", (existing["id"],))
            conn.execute("DELETE FROM eeff_periodos WHERE id = ?", (existing["id"],))
            reemplazado = True

        # Insert period
        conn.execute(
            "INSERT INTO eeff_periodos (sociedad, tipo, periodo, fecha_desde, fecha_hasta, año, trimestre, fuente, fichero_origen, created_at)"
            " VALUES (?, ?, ?, ?, ?, ?, ?, 'gestoria', ?, ?)",
            (sociedad, tipo, periodo, desde, hasta, año, trimestre, filename, now),
        )
        periodo_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]

        # Insert lines
        for ln in inf["lineas"]:
            conn.execute(
                "INSERT INTO eeff_lineas (periodo_id, codigo_cuenta, descripcion, nivel, tipo_nivel,"
                " importe_actual, importe_anterior, debe_periodo, haber_periodo,"
                " debe_acumulado, haber_acumulado, debe_saldo, haber_saldo, saldo_inicial, orden, created_at)"
                " VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
                (
                    periodo_id,
                    ln.get("codigo_cuenta"),
                    ln["descripcion"],
                    ln.get("nivel", 0),
                    ln.get("tipo_nivel"),
                    ln.get("importe_actual", 0),
                    ln.get("importe_anterior"),
                    ln.get("debe_periodo"),
                    ln.get("haber_periodo"),
                    ln.get("debe_acumulado"),
                    ln.get("haber_acumulado"),
                    ln.get("debe_saldo"),
                    ln.get("haber_saldo"),
                    ln.get("saldo_inicial"),
                    ln.get("orden", 0),
                    now,
                ),
            )

        resultados.append({
            "tipo": tipo,
            "periodo": periodo,
            "empresa": sociedad,
            "lineas": len(inf["lineas"]),
            "hoja": inf.get("hoja", ""),
            "reemplazado": reemplazado,
        })

    conn.commit()
    # Fix names of existing periods
    _corregir_nombres_periodos(conn)
    return {"importados": len(resultados), "detalle": resultados}


def _corregir_nombres_periodos(conn):
    """Corrige nombres de periodos existentes (ej: '2025 Q1' → 'Febrero 2025' si es mensual)."""
    _MESES = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
              "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]
    rows = conn.execute("SELECT id, fecha_desde, fecha_hasta, año, trimestre, tipo, periodo FROM eeff_periodos").fetchall()
    for r in rows:
        nuevo = _generar_nombre_periodo(r["fecha_desde"], r["fecha_hasta"], r["año"], r["trimestre"], r["tipo"], _MESES)
        if nuevo != r["periodo"]:
            conn.execute("UPDATE eeff_periodos SET periodo = ? WHERE id = ?", (nuevo, r["id"]))
    conn.commit()


# ── Consultas ────────────────────────────────────────────────────────────────

def listar_periodos(conn, sociedad=None, tipo=None, año=None):
    sql = "SELECT * FROM eeff_periodos WHERE 1=1"
    params = []
    if sociedad:
        sql += " AND sociedad = ?"
        params.append(sociedad)
    if tipo:
        sql += " AND tipo = ?"
        params.append(tipo)
    if año:
        sql += " AND año = ?"
        params.append(int(año))
    sql += " ORDER BY año DESC, fecha_hasta DESC, tipo"
    return [dict(r) for r in conn.execute(sql, params).fetchall()]


def obtener_lineas(conn, periodo_id):
    rows = conn.execute(
        "SELECT * FROM eeff_lineas WHERE periodo_id = ? ORDER BY orden",
        (periodo_id,),
    ).fetchall()
    return [dict(r) for r in rows]


def eliminar_periodo(conn, periodo_id):
    conn.execute("DELETE FROM eeff_lineas WHERE periodo_id = ?", (periodo_id,))
    conn.execute("DELETE FROM eeff_periodos WHERE id = ?", (periodo_id,))
    conn.commit()


def calcular_metricas(conn, sociedad, año):
    """Calcula KPIs financieros a partir de Balance + P&G del año."""
    # Buscar balance anual
    bal = conn.execute(
        "SELECT id FROM eeff_periodos WHERE sociedad = ? AND tipo = 'balance' AND año = ? ORDER BY fecha_hasta DESC LIMIT 1",
        (sociedad, int(año)),
    ).fetchone()
    pyg = conn.execute(
        "SELECT id FROM eeff_periodos WHERE sociedad = ? AND tipo = 'pyg' AND año = ? ORDER BY fecha_hasta DESC LIMIT 1",
        (sociedad, int(año)),
    ).fetchone()

    metricas = {}
    if not bal and not pyg:
        return metricas

    def _normalizar(s):
        """Quita tildes para comparación robusta."""
        import unicodedata
        return "".join(
            c for c in unicodedata.normalize("NFD", s)
            if unicodedata.category(c) != "Mn"
        ).upper()

    def _buscar(periodo_id, *terms, last=False):
        """Busca línea cuya descripción contenga todos los terms (normalizado).

        By default returns the first match. If last=True, returns the last.
        """
        rows_full = conn.execute(
            "SELECT descripcion, importe_actual, tipo_nivel FROM eeff_lineas WHERE periodo_id = ? ORDER BY orden",
            (periodo_id,),
        ).fetchall()
        terms_norm = [_normalizar(t) for t in terms]
        resultado = 0
        for r in rows_full:
            desc_norm = _normalizar(r["descripcion"])
            if all(t in desc_norm for t in terms_norm):
                resultado = r["importe_actual"] or 0
                if not last:
                    return resultado
        return resultado

    if pyg:
        pid = pyg["id"]
        ingresos = _buscar(pid, "IMPORTE NETO", "CIFRA DE NEGOCIOS")
        resultado_expl = _buscar(pid, "RESULTADO DE EXPLOTACION")
        amortizacion = abs(_buscar(pid, "AMORTIZACION", "INMOVILIZADO"))
        resultado_ej = _buscar(pid, "RESULTADO DEL EJERCICIO", last=True)

        ebitda = resultado_expl + amortizacion
        metricas["ingresos"] = ingresos
        metricas["resultado_explotacion"] = resultado_expl
        metricas["amortizacion"] = amortizacion
        metricas["ebitda"] = ebitda
        metricas["margen_ebitda"] = round(ebitda / ingresos * 100, 1) if ingresos else 0
        metricas["resultado_ejercicio"] = resultado_ej
        metricas["resultado_antes_impuestos"] = _buscar(pid, "RESULTADO ANTES DE IMPUESTOS")

    if bal:
        bid = bal["id"]
        activo_corriente = _buscar(bid, "ACTIVO CORRIENTE")
        pasivo_corriente = _buscar(bid, "PASIVO CORRIENTE")
        pn = _buscar(bid, "PATRIMONIO NETO")
        total_activo = _buscar(bid, "TOTAL ACTIVO")

        metricas["activo_corriente"] = activo_corriente
        metricas["pasivo_corriente"] = pasivo_corriente
        metricas["working_capital"] = activo_corriente - pasivo_corriente
        metricas["patrimonio_neto"] = pn
        metricas["total_activo"] = total_activo

        # Deuda: buscar pasivo no corriente (deuda LP) + pasivo corriente financiero
        pasivo_no_corriente = _buscar(bid, "PASIVO NO CORRIENTE")
        # Efectivo
        efectivo = _buscar(bid, "EFECTIVO")
        deuda_neta = pasivo_no_corriente + pasivo_corriente - efectivo
        metricas["deuda_neta"] = deuda_neta

        ebitda = metricas.get("ebitda", 0)
        metricas["deuda_neta_ebitda"] = round(deuda_neta / ebitda, 2) if ebitda else None

        # ROE
        resultado = metricas.get("resultado_ejercicio", 0)
        metricas["roe"] = round(resultado / pn * 100, 1) if pn else None

        # DSO — clientes / facturación * 365
        clientes = _buscar(bid, "DEUDORES COMERCIALES")
        facturacion = metricas.get("ingresos", 0)
        metricas["dso"] = round(clientes / facturacion * 365, 0) if facturacion else None

    return metricas


# ══════════════════════════════════════════════════════════════════════════════
# MOTOR FINANCIERO v2 — Plan de cuentas, fórmulas, informes desde SS
# ══════════════════════════════════════════════════════════════════════════════

# Mapeo PGC estándar: rango inicio-fin → (nivel1, nivel2, nivel3, signo)
_PGC_MAPEO = [
    # ACTIVO NO CORRIENTE
    ("2030", "2090", "Activo", "Activo No Corriente", "Inmovilizado intangible", 1),
    ("2130", "2199", "Activo", "Activo No Corriente", "Inmovilizado material", 1),
    ("2300", "2399", "Activo", "Activo No Corriente", "Inmovilizado en curso", 1),
    ("2400", "2519", "Activo", "Activo No Corriente", "Inversiones financieras LP", 1),
    ("2600", "2609", "Activo", "Activo No Corriente", "Fianzas LP", 1),
    ("2800", "2819", "Activo", "Activo No Corriente", "Amortización acumulada", 1),
    ("2900", "2979", "Activo", "Activo No Corriente", "Deterioro inmovilizado", 1),
    # ACTIVO CORRIENTE
    ("3000", "3999", "Activo", "Activo Corriente", "Existencias", 1),
    ("4300", "4369", "Activo", "Activo Corriente", "Clientes", 1),
    ("4400", "4409", "Activo", "Activo Corriente", "Deudores varios", 1),
    ("4700", "4709", "Activo", "Activo Corriente", "HP deudora", 1),
    ("4720", "4749", "Activo", "Activo Corriente", "HP deudora", 1),
    ("4780", "4789", "Activo", "Activo Corriente", "Otros deudores", 1),
    ("4800", "4809", "Activo", "Activo Corriente", "Gastos anticipados", 1),
    ("5400", "5419", "Activo", "Activo Corriente", "Inversiones financieras CP", 1),
    ("5650", "5669", "Activo", "Activo Corriente", "Fianzas CP", 1),
    ("5700", "5729", "Activo", "Activo Corriente", "Bancos", 1),
    # PATRIMONIO NETO
    ("1000", "1099", "Pasivo y PN", "Patrimonio Neto", "Capital", -1),
    ("1120", "1139", "Pasivo y PN", "Patrimonio Neto", "Reservas", -1),
    ("1200", "1209", "Pasivo y PN", "Patrimonio Neto", "Resultado ejercicios anteriores", -1),
    ("1290", "1299", "Pasivo y PN", "Patrimonio Neto", "Resultado del ejercicio", -1),
    # PASIVO NO CORRIENTE
    ("1700", "1739", "Pasivo y PN", "Pasivo No Corriente", "Deudas LP", -1),
    ("1800", "1809", "Pasivo y PN", "Pasivo No Corriente", "Fianzas recibidas LP", -1),
    ("4790", "4799", "Pasivo y PN", "Pasivo No Corriente", "Pasivos diferidos", -1),
    # PASIVO CORRIENTE
    ("4000", "4009", "Pasivo y PN", "Pasivo Corriente", "Proveedores", -1),
    ("4100", "4119", "Pasivo y PN", "Pasivo Corriente", "Acreedores", -1),
    ("4650", "4659", "Pasivo y PN", "Pasivo Corriente", "Remuneraciones pendientes", -1),
    ("4750", "4769", "Pasivo y PN", "Pasivo Corriente", "HP acreedora", -1),
    ("4770", "4779", "Pasivo y PN", "Pasivo Corriente", "HP acreedora", -1),
    ("4994", "4999", "Pasivo y PN", "Pasivo Corriente", "Provisiones CP", -1),
    ("5200", "5269", "Pasivo y PN", "Pasivo Corriente", "Deudas CP", -1),
    ("5500", "5559", "Pasivo y PN", "Pasivo Corriente", "Otras deudas CP", -1),
    # P&G INGRESOS
    ("7000", "7009", "P&G", "Ingresos explotación", "Ventas", -1),
    ("7050", "7059", "P&G", "Ingresos explotación", "Prestación de servicios", -1),
    ("7400", "7409", "P&G", "Otros ingresos explotación", "Subvenciones", -1),
    ("7520", "7529", "P&G", "Otros ingresos explotación", "Otros ingresos", -1),
    ("7600", "7699", "P&G", "Ingresos financieros", "Ingresos financieros", -1),
    ("7780", "7789", "P&G", "Otros ingresos explotación", "Ingresos excepcionales", -1),
    ("7940", "7949", "P&G", "Otros ingresos explotación", "Reversión deterioro", -1),
    # P&G GASTOS (signo -1: convierte saldo deudor positivo a negativo en presentación)
    ("6000", "6009", "P&G", "Aprovisionamientos", "Compras", -1),
    ("6020", "6029", "P&G", "Aprovisionamientos", "Otros aprovisionamientos", -1),
    ("6070", "6079", "P&G", "Aprovisionamientos", "Trabajos otras empresas", -1),
    ("6100", "6109", "P&G", "Variación existencias", "Variación existencias", -1),
    ("6210", "6299", "P&G", "Otros gastos explotación", "Servicios exteriores", -1),
    ("6300", "6309", "P&G", "Impuesto sobre beneficios", "Impuesto sociedades", -1),
    ("6310", "6319", "P&G", "Otros gastos explotación", "Otros tributos", -1),
    ("6400", "6409", "P&G", "Gastos personal", "Sueldos y salarios", -1),
    ("6420", "6429", "P&G", "Gastos personal", "Seguridad Social empresa", -1),
    ("6490", "6499", "P&G", "Gastos personal", "Otros gastos sociales", -1),
    ("6500", "6509", "P&G", "Otros gastos explotación", "Otros gastos gestión", -1),
    ("6620", "6699", "P&G", "Gastos financieros", "Gastos financieros", -1),
    ("6780", "6789", "P&G", "Gastos excepcionales", "Gastos excepcionales", -1),
    ("6810", "6819", "P&G", "Amortización", "Amortización inmovilizado", -1),
    ("6940", "6949", "P&G", "Deterioro", "Deterioro créditos", -1),
]


def _clasificar_cuenta(codigo4):
    """Devuelve (nivel1, nivel2, nivel3, signo) para un código de 4 dígitos."""
    for rango_ini, rango_fin, n1, n2, n3, signo in _PGC_MAPEO:
        if rango_ini <= codigo4 <= rango_fin:
            return n1, n2, n3, signo
    return "Sin clasificar", "Sin clasificar", "Sin clasificar", 1


def seed_plan_cuentas(conn):
    """Genera plan de cuentas desde las subcuentas existentes en eeff_lineas."""
    crear_tablas(conn)
    now = datetime.now().isoformat()

    # Get all distinct 4-digit prefixes from sumas_saldos
    rows = conn.execute(
        "SELECT DISTINCT SUBSTR(l.codigo_cuenta, 1, 4) as cod4,"
        " MIN(l.descripcion) as nombre"
        " FROM eeff_lineas l JOIN eeff_periodos p ON l.periodo_id = p.id"
        " WHERE p.tipo = 'sumas_saldos' AND l.codigo_cuenta GLOB '[0-9]*'"
        " GROUP BY cod4 ORDER BY cod4"
    ).fetchall()

    inserted = 0
    for r in rows:
        cod4 = r["cod4"]
        nombre = r["nombre"] or cod4
        n1, n2, n3, signo = _clasificar_cuenta(cod4)

        # Use a generic name for the 4-digit group
        existing = conn.execute("SELECT id FROM eeff_plan_cuentas WHERE codigo = ?", (cod4,)).fetchone()
        if existing:
            continue

        conn.execute(
            "INSERT INTO eeff_plan_cuentas (codigo, nombre, nivel1, nivel2, nivel3, signo, activo, created_at)"
            " VALUES (?, ?, ?, ?, ?, ?, 1, ?)",
            (cod4, nombre, n1, n2, n3, signo, now),
        )
        inserted += 1

    conn.commit()
    return inserted


_FORMULAS_SEED = [
    # Balance
    ("Total Activo No Corriente", "Suma activo no corriente", "SUM_NIVEL2:Activo No Corriente", "EUR", 1, "balance"),
    ("Total Activo Corriente", "Suma activo corriente", "SUM_NIVEL2:Activo Corriente", "EUR", 2, "balance"),
    ("Total Activo", "Total activo", "{Total Activo No Corriente} + {Total Activo Corriente}", "EUR", 3, "balance"),
    ("Total Patrimonio Neto", "Suma patrimonio neto", "SUM_NIVEL2:Patrimonio Neto", "EUR", 4, "balance"),
    ("Total Pasivo No Corriente", "Suma pasivo no corriente", "SUM_NIVEL2:Pasivo No Corriente", "EUR", 5, "balance"),
    ("Total Pasivo Corriente", "Suma pasivo corriente", "SUM_NIVEL2:Pasivo Corriente", "EUR", 6, "balance"),
    ("Total Pasivo y PN", "Total pasivo y patrimonio neto", "{Total Patrimonio Neto} + {Total Pasivo No Corriente} + {Total Pasivo Corriente}", "EUR", 7, "balance"),
    # P&G
    ("Ingresos", "Total ingresos de explotación", "SUM_NIVEL2:Ingresos explotación", "EUR", 10, "pyg"),
    ("Otros Ingresos", "Otros ingresos de explotación", "SUM_NIVEL2:Otros ingresos explotación", "EUR", 11, "pyg"),
    ("Aprovisionamientos", "Total aprovisionamientos", "SUM_NIVEL2:Aprovisionamientos", "EUR", 12, "pyg"),
    ("Margen Bruto", "Ingresos - Aprovisionamientos", "{Ingresos} + {Otros Ingresos} + {Aprovisionamientos}", "EUR", 13, "pyg"),
    ("Gastos Personal", "Total gastos de personal", "SUM_NIVEL2:Gastos personal", "EUR", 14, "pyg"),
    ("Otros Gastos", "Otros gastos de explotación", "SUM_NIVEL2:Otros gastos explotación", "EUR", 15, "pyg"),
    ("EBITDA", "Resultado antes de amortización", "{Margen Bruto} + {Gastos Personal} + {Otros Gastos}", "EUR", 16, "pyg"),
    ("Amortización", "Amortización del inmovilizado", "SUM_NIVEL2:Amortización", "EUR", 17, "pyg"),
    ("Deterioro", "Deterioro de créditos", "SUM_NIVEL2:Deterioro", "EUR", 17, "pyg"),
    ("Gastos Excepcionales", "Gastos excepcionales", "SUM_NIVEL2:Gastos excepcionales", "EUR", 17, "pyg"),
    ("EBIT", "Resultado de explotación", "{EBITDA} + {Amortización} + {Deterioro} + {Gastos Excepcionales}", "EUR", 18, "pyg"),
    ("Ingresos Financieros", "Ingresos financieros", "SUM_NIVEL2:Ingresos financieros", "EUR", 19, "pyg"),
    ("Gastos Financieros", "Gastos financieros", "SUM_NIVEL2:Gastos financieros", "EUR", 20, "pyg"),
    ("Resultado Financiero", "Ingresos - gastos financieros", "{Ingresos Financieros} + {Gastos Financieros}", "EUR", 21, "pyg"),
    ("Resultado Antes Impuestos", "EBIT + resultado financiero", "{EBIT} + {Resultado Financiero}", "EUR", 22, "pyg"),
    ("Impuesto Sociedades", "Impuesto sobre beneficios", "SUM_NIVEL2:Impuesto sobre beneficios", "EUR", 23, "pyg"),
    ("Resultado Neto", "Resultado después de impuestos", "{Resultado Antes Impuestos} + {Impuesto Sociedades}", "EUR", 24, "pyg"),
    # Métricas
    ("Margen EBITDA %", "EBITDA / Ingresos x 100", "{EBITDA} / {Ingresos} * 100", "PCT", 30, "metricas"),
    ("Working Capital", "Activo Corriente - Pasivo Corriente", "{Total Activo Corriente} - {Total Pasivo Corriente}", "EUR", 31, "metricas"),
    ("Deuda Neta", "Deuda LP + CP - Bancos", "{Total Pasivo No Corriente} + {Total Pasivo Corriente} - SUM_NIVEL3:Bancos", "EUR", 32, "metricas"),
    ("Deuda Neta / EBITDA", "Ratio de apalancamiento", "{Deuda Neta} / {EBITDA}", "RATIO", 33, "metricas"),
    ("ROE %", "Resultado / Patrimonio Neto x 100", "{Resultado Neto} / {Total Patrimonio Neto} * 100", "PCT", 34, "metricas"),
    ("DSO", "Días de cobro medio", "SUM_NIVEL3:Clientes / {Ingresos} * 365", "DIAS", 35, "metricas"),
]


def seed_formulas(conn):
    """Inserta fórmulas estándar si no existen."""
    crear_tablas(conn)
    now = datetime.now().isoformat()
    inserted = 0
    for nombre, desc, formula, fmt, orden, grupo in _FORMULAS_SEED:
        existing = conn.execute("SELECT id FROM eeff_formulas WHERE nombre = ?", (nombre,)).fetchone()
        if existing:
            continue
        conn.execute(
            "INSERT INTO eeff_formulas (nombre, descripcion, formula, formato, orden, grupo, activo, created_at)"
            " VALUES (?, ?, ?, ?, ?, ?, 1, ?)",
            (nombre, desc, formula, fmt, orden, grupo, now),
        )
        inserted += 1
    conn.commit()
    return inserted


# ── Motor de cálculo v2 ─────────────────────────────────────────────────────

def calcular_informe(conn, periodo_id):
    """Genera Balance + P&G + Métricas desde sumas y saldos + plan de cuentas."""
    crear_tablas(conn)

    # 1. Load all SS lines for this period
    lineas = conn.execute(
        "SELECT codigo_cuenta, descripcion, debe_saldo, haber_saldo"
        " FROM eeff_lineas WHERE periodo_id = ? AND codigo_cuenta GLOB '[0-9]*'"
        " ORDER BY codigo_cuenta",
        (periodo_id,),
    ).fetchall()

    # 2. Load plan de cuentas
    plan = {}
    for r in conn.execute("SELECT codigo, nombre, nivel1, nivel2, nivel3, signo FROM eeff_plan_cuentas WHERE activo = 1").fetchall():
        plan[r["codigo"]] = dict(r)

    # 3. Classify each subcuenta
    agrupado = {}  # key = (nivel1, nivel2, nivel3) → sum of signed saldos
    sin_clasificar = []
    detalle_cuentas = []  # all accounts with their classification

    for ln in lineas:
        cod = str(ln["codigo_cuenta"]).strip()
        cod4 = cod[:4]
        saldo_deudor = ln["debe_saldo"] or 0
        saldo_acreedor = ln["haber_saldo"] or 0
        saldo = saldo_deudor - saldo_acreedor

        mapping = plan.get(cod4)
        if not mapping or mapping["nivel1"] == "Sin clasificar":
            sin_clasificar.append({
                "codigo": cod4,
                "subcuenta": cod,
                "nombre": ln["descripcion"],
                "saldo": round(saldo, 2),
            })
            continue

        n1 = mapping["nivel1"]
        n2 = mapping["nivel2"]
        n3 = mapping["nivel3"]
        signo = mapping["signo"]
        saldo_signed = round(saldo * signo, 2)

        key = (n1, n2, n3)
        agrupado[key] = round(agrupado.get(key, 0) + saldo_signed, 2)
        detalle_cuentas.append({
            "codigo": cod, "nombre": ln["descripcion"],
            "saldo": round(saldo, 2), "saldo_signed": saldo_signed,
            "nivel1": n1, "nivel2": n2, "nivel3": n3,
        })

    # 4. Build aggregations by nivel2 and nivel3
    by_nivel2 = {}
    by_nivel3 = {}
    for (n1, n2, n3), val in agrupado.items():
        by_nivel2[n2] = round(by_nivel2.get(n2, 0) + val, 2)
        by_nivel3[n3] = round(by_nivel3.get(n3, 0) + val, 2)

    # 5. Evaluate formulas
    formulas = conn.execute(
        "SELECT nombre, formula, formato, orden, grupo FROM eeff_formulas WHERE activo = 1 ORDER BY orden"
    ).fetchall()

    formula_cache = {}

    def _eval_formula(formula_str):
        """Evalúa una fórmula recursivamente."""
        s = formula_str.strip()

        # SUM_NIVEL2:X
        s = re.sub(
            r"SUM_NIVEL2:([^+\-*/{}]+)",
            lambda m: str(by_nivel2.get(m.group(1).strip(), 0)),
            s,
        )
        # SUM_NIVEL3:X
        s = re.sub(
            r"SUM_NIVEL3:([^+\-*/{}]+)",
            lambda m: str(by_nivel3.get(m.group(1).strip(), 0)),
            s,
        )
        # {FormulaRef}
        def _resolve_ref(m):
            name = m.group(1).strip()
            if name in formula_cache:
                return str(formula_cache[name])
            return "0"

        s = re.sub(r"\{([^}]+)\}", _resolve_ref, s)

        # Evaluate arithmetic
        try:
            result = eval(s, {"__builtins__": {}}, {})
            return round(float(result), 2)
        except (ZeroDivisionError, TypeError, ValueError):
            return 0
        except Exception:
            return 0

    formula_results = []
    for f in formulas:
        val = _eval_formula(f["formula"])
        formula_cache[f["nombre"]] = val
        formula_results.append({
            "nombre": f["nombre"],
            "valor": val,
            "formato": f["formato"],
            "grupo": f["grupo"],
            "orden": f["orden"],
        })

    # 6. Build structured balance
    balance = {"Activo": {}, "Pasivo y PN": {}}
    for (n1, n2, n3), val in sorted(agrupado.items()):
        if n1 in balance:
            if n2 not in balance[n1]:
                balance[n1][n2] = {"total": 0, "detalle": {}}
            balance[n1][n2]["detalle"][n3] = round(balance[n1][n2]["detalle"].get(n3, 0) + val, 2)
            balance[n1][n2]["total"] = round(balance[n1][n2]["total"] + val, 2)

    # 7. Build P&G lines from formulas
    pyg_lines = [f for f in formula_results if f["grupo"] == "pyg"]
    metricas = [f for f in formula_results if f["grupo"] == "metricas"]
    balance_totals = [f for f in formula_results if f["grupo"] == "balance"]

    # Deduplicate sin_clasificar by cod4
    seen_sc = set()
    sc_unique = []
    for sc in sin_clasificar:
        if sc["codigo"] not in seen_sc:
            seen_sc.add(sc["codigo"])
            sc_unique.append(sc)

    return {
        "balance": balance,
        "balance_totals": balance_totals,
        "pyg": pyg_lines,
        "metricas": metricas,
        "sin_clasificar": sc_unique,
        "agrupado": {f"{k[0]}|{k[1]}|{k[2]}": v for k, v in agrupado.items()},
    }


# ── P&G multi-periodo (comparativas) ──────────────────────────────────────
#
# Los SS de la gestoría son ACUMULADOS desde enero:
#   - debe_saldo / haber_saldo = acumulado YTD (unambiguamente)
#   - debe_periodo / haber_periodo = puede ser mensual o acumulado según fichero
#
# Para evitar ambigüedad usamos siempre debe_saldo/haber_saldo y derivamos
# el mensual por sustracción: P&G(mes M) = YTD(M) - YTD(M-1).

def _evaluar_formulas_pyg(conn, by_nivel2, by_nivel3):
    """Evalúa las fórmulas P&G sobre los dicts de nivel2/nivel3 ya calculados."""
    formulas = conn.execute(
        "SELECT nombre, formula, formato, orden, grupo FROM eeff_formulas"
        " WHERE activo = 1 AND grupo = 'pyg' ORDER BY orden"
    ).fetchall()

    formula_cache = {}

    def _eval(formula_str):
        s = formula_str.strip()
        s = re.sub(
            r"SUM_NIVEL2:([^+\-*/{}]+)",
            lambda m: str(by_nivel2.get(m.group(1).strip(), 0)), s,
        )
        s = re.sub(
            r"SUM_NIVEL3:([^+\-*/{}]+)",
            lambda m: str(by_nivel3.get(m.group(1).strip(), 0)), s,
        )
        s = re.sub(
            r"\{([^}]+)\}",
            lambda m: str(formula_cache.get(m.group(1).strip(), 0)), s,
        )
        try:
            return round(float(eval(s, {"__builtins__": {}}, {})), 2)
        except Exception:
            return 0

    result = []
    for f in formulas:
        val = _eval(f["formula"])
        formula_cache[f["nombre"]] = val
        result.append({"nombre": f["nombre"], "valor": val, "formato": f["formato"], "orden": f["orden"]})
    return result


def _agregar_saldos_ytd(conn, periodo_id):
    """Lee debe_saldo/haber_saldo de un periodo SS y devuelve (by_nivel2, by_nivel3).

    Estos saldos son ACUMULADOS desde inicio de año — exactamente lo que
    necesitamos para YTD.
    """
    plan = {}
    for r in conn.execute(
        "SELECT codigo, nivel1, nivel2, nivel3, signo FROM eeff_plan_cuentas WHERE activo = 1"
    ).fetchall():
        plan[r["codigo"]] = dict(r)

    lineas = conn.execute(
        "SELECT codigo_cuenta, COALESCE(debe_saldo,0) as debe, COALESCE(haber_saldo,0) as haber"
        " FROM eeff_lineas WHERE periodo_id = ? AND codigo_cuenta GLOB '[0-9]*'",
        (periodo_id,),
    ).fetchall()

    by_nivel2 = {}
    by_nivel3 = {}
    for ln in lineas:
        cod4 = str(ln["codigo_cuenta"]).strip()[:4]
        mapping = plan.get(cod4)
        if not mapping or mapping["nivel1"] == "Sin clasificar":
            continue
        saldo = (ln["debe"] or 0) - (ln["haber"] or 0)
        saldo_signed = round(saldo * mapping["signo"], 2)
        n2, n3 = mapping["nivel2"], mapping["nivel3"]
        by_nivel2[n2] = round(by_nivel2.get(n2, 0) + saldo_signed, 2)
        by_nivel3[n3] = round(by_nivel3.get(n3, 0) + saldo_signed, 2)
    return by_nivel2, by_nivel3


def _restar_niveles(a_n2, a_n3, b_n2, b_n3):
    """Devuelve a - b para los dicts de nivel2 y nivel3."""
    all_n2 = set(a_n2) | set(b_n2)
    all_n3 = set(a_n3) | set(b_n3)
    r_n2 = {k: round(a_n2.get(k, 0) - b_n2.get(k, 0), 2) for k in all_n2}
    r_n3 = {k: round(a_n3.get(k, 0) - b_n3.get(k, 0), 2) for k in all_n3}
    return r_n2, r_n3


def calcular_pyg_ytd(conn, periodo_id):
    """P&G acumulada YTD usando debe_saldo/haber_saldo de un solo periodo."""
    crear_tablas(conn)
    n2, n3 = _agregar_saldos_ytd(conn, periodo_id)
    return _evaluar_formulas_pyg(conn, n2, n3)


def calcular_pyg_mensual(conn, periodo_id, periodo_id_anterior=None):
    """P&G de un solo mes = YTD(mes) - YTD(mes-1).

    Para enero, periodo_id_anterior es None y se usa el YTD directamente.
    """
    crear_tablas(conn)
    n2_m, n3_m = _agregar_saldos_ytd(conn, periodo_id)
    if periodo_id_anterior:
        n2_prev, n3_prev = _agregar_saldos_ytd(conn, periodo_id_anterior)
        n2_m, n3_m = _restar_niveles(n2_m, n3_m, n2_prev, n3_prev)
    return _evaluar_formulas_pyg(conn, n2_m, n3_m)


def calcular_pyg_ltm(conn, periodo_id_mes_actual, periodo_id_dic_anterior, periodo_id_mes_anterior=None):
    """P&G últimos 12 meses = YTD(mes M, año Y) + Anual(Y-1) - YTD(mes M, año Y-1).

    Si no hay YTD del mismo mes del año anterior, se usa solo YTD actual + Dic anterior.
    """
    crear_tablas(conn)
    # YTD actual (ene..M del año Y)
    n2_ytd, n3_ytd = _agregar_saldos_ytd(conn, periodo_id_mes_actual)
    # Anual año anterior (ene..dic Y-1)
    n2_dic, n3_dic = _agregar_saldos_ytd(conn, periodo_id_dic_anterior)
    # Sumar YTD + Dic anterior
    all_n2 = set(n2_ytd) | set(n2_dic)
    all_n3 = set(n3_ytd) | set(n3_dic)
    n2_sum = {k: round(n2_ytd.get(k, 0) + n2_dic.get(k, 0), 2) for k in all_n2}
    n3_sum = {k: round(n3_ytd.get(k, 0) + n3_dic.get(k, 0), 2) for k in all_n3}
    # Restar YTD del mismo mes del año anterior (para no contar doble ene..M de Y-1)
    if periodo_id_mes_anterior:
        n2_prev, n3_prev = _agregar_saldos_ytd(conn, periodo_id_mes_anterior)
        n2_sum, n3_sum = _restar_niveles(n2_sum, n3_sum, n2_prev, n3_prev)
    return _evaluar_formulas_pyg(conn, n2_sum, n3_sum)


# Keep old name for backwards compatibility
def calcular_pyg_multiples_periodos(conn, periodo_ids):
    """Deprecated — usa calcular_pyg_ytd / calcular_pyg_mensual / calcular_pyg_ltm."""
    return calcular_pyg_ytd(conn, periodo_ids[0]) if periodo_ids else []


# ── Plan de cuentas CRUD ────────────────────────────────────────────────────

def obtener_plan_cuentas(conn):
    rows = conn.execute("SELECT * FROM eeff_plan_cuentas ORDER BY codigo").fetchall()
    return [dict(r) for r in rows]


def actualizar_cuenta(conn, cuenta_id, nivel1, nivel2, nivel3, signo):
    conn.execute(
        "UPDATE eeff_plan_cuentas SET nivel1=?, nivel2=?, nivel3=?, signo=?, updated_at=? WHERE id=?",
        (nivel1, nivel2, nivel3, signo, datetime.now().isoformat(), cuenta_id),
    )
    conn.commit()


def crear_cuenta(conn, codigo, nombre, nivel1, nivel2, nivel3, signo):
    now = datetime.now().isoformat()
    conn.execute(
        "INSERT OR REPLACE INTO eeff_plan_cuentas (codigo, nombre, nivel1, nivel2, nivel3, signo, activo, created_at)"
        " VALUES (?, ?, ?, ?, ?, ?, 1, ?)",
        (codigo, nombre, nivel1, nivel2, nivel3, signo, now),
    )
    conn.commit()


def obtener_formulas(conn):
    rows = conn.execute("SELECT * FROM eeff_formulas WHERE activo = 1 ORDER BY orden").fetchall()
    return [dict(r) for r in rows]
