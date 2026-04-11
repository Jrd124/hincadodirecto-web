"""
Importador de transacciones Moeve/Cepsa desde Excel.
Hoja 'data', header en fila 3 (header=2 en pandas/openpyxl).
Idempotente: UPSERT por clave única (origen, tarjeta_completa, fecha_hora_original, concepto, litros).
Netea descuentos Billed=2 automáticamente.
"""
from __future__ import annotations

import logging
import uuid
from io import BytesIO
from collections import defaultdict

import openpyxl

from core.db import get_conn

logger = logging.getLogger("erp")

# ── Ensure tables ────────────────────────────────────────────────────────

_tables_ok = False


def _ensure_tables():
    global _tables_ok
    if _tables_ok:
        return
    conn = get_conn()
    try:
        conn.execute("""CREATE TABLE IF NOT EXISTS combustible_transacciones (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            origen TEXT NOT NULL DEFAULT 'moeve',
            tarjeta TEXT,
            tarjeta_completa TEXT,
            matricula TEXT,
            estacion TEXT,
            localidad TEXT,
            pais TEXT,
            fecha TEXT NOT NULL,
            hora TEXT,
            fecha_hora_original TEXT,
            concepto TEXT,
            litros REAL DEFAULT 0,
            importe REAL DEFAULT 0,
            iva_porcentaje REAL DEFAULT 0,
            descuento REAL DEFAULT 0,
            factura TEXT,
            billed INTEGER,
            operacion TEXT,
            latitud REAL,
            longitud REAL,
            geo_confidence TEXT,
            proyecto_id INTEGER,
            imputacion_tipo TEXT,
            imputacion_confianza TEXT,
            imputacion_notas TEXT,
            importacion_id TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (proyecto_id) REFERENCES proyectos(id),
            UNIQUE(origen, tarjeta_completa, fecha_hora_original, concepto, litros)
        )""")
        conn.execute("CREATE INDEX IF NOT EXISTS ix_comb_fecha ON combustible_transacciones(fecha)")
        conn.execute("CREATE INDEX IF NOT EXISTS ix_comb_matricula ON combustible_transacciones(matricula)")
        conn.execute("CREATE INDEX IF NOT EXISTS ix_comb_proyecto ON combustible_transacciones(proyecto_id)")

        conn.execute("""CREATE TABLE IF NOT EXISTS moeve_estaciones_geo (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            estacion TEXT UNIQUE NOT NULL,
            localidad_extraida TEXT,
            latitud REAL,
            longitud REAL,
            municipio TEXT,
            provincia TEXT,
            geo_source TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )""")

        conn.execute("""CREATE TABLE IF NOT EXISTS moeve_vehiculos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            matricula TEXT UNIQUE NOT NULL,
            tipo TEXT,
            descripcion TEXT,
            empleado_id INTEGER,
            maquina_id INTEGER,
            activo INTEGER DEFAULT 1,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (empleado_id) REFERENCES empleados(id),
            FOREIGN KEY (maquina_id) REFERENCES maquinas(id)
        )""")

        conn.commit()
        _tables_ok = True
    finally:
        conn.close()


# ── Helpers ──────────────────────────────────────────────────────────────

def _parse_fecha(dt_str):
    """'07/04/2026 07:53:16' → ('2026-04-07', '07:53:16')"""
    if not dt_str or not isinstance(dt_str, str):
        return None, None
    dt_str = dt_str.strip()
    parts = dt_str.split(" ", 1)
    date_part = parts[0]
    time_part = parts[1] if len(parts) > 1 else ""
    # DD/MM/YYYY
    dp = date_part.split("/")
    if len(dp) == 3:
        fecha = f"{dp[2]}-{dp[1].zfill(2)}-{dp[0].zfill(2)}"
    else:
        return None, None
    return fecha, time_part


def _extraer_localidad(location):
    """Extrae localidad de Location field.
    'TRUJILLO-II TRUJILLO-II' → 'TRUJILLO'
    'LA MAYA MONTEJO DE SALVATIERRA' → 'MONTEJO DE SALVATIERRA'
    '- -' → ''
    'VASCO-ARAGONESA .' → ''
    """
    if not location or location.strip() in ("- -", "-", ".", ""):
        return ""
    loc = location.strip()
    # Si termina en "." es peaje, sin localidad útil
    if loc.endswith(" .") or loc.endswith("."):
        return ""
    # Split: la localidad suele ser la segunda mitad
    # Muchas estaciones repiten: "NOMBRE NOMBRE" → tomar una
    words = loc.split()
    if len(words) <= 1:
        return loc
    # Buscar si la segunda mitad es repetición de la primera
    mid = len(words) // 2
    first_half = " ".join(words[:mid])
    second_half = " ".join(words[mid:])
    if first_half == second_half:
        # Repetición: "TRUJILLO-II TRUJILLO-II" → "TRUJILLO-II"
        # Limpiar sufijos tipo "-II", "-I"
        clean = first_half
        for suffix in ["-II", "-I", "-III", "-IV", "-V"]:
            if clean.endswith(suffix):
                clean = clean[:-len(suffix)].strip()
                break
        return clean if clean else first_half
    # No es repetición: tomar segunda mitad como localidad
    # Pero si es solo un par de palabras, quizá todo es el nombre
    return second_half


def _normalizar_matricula(mat):
    """Quita guiones y espacios de matrícula."""
    if not mat or mat.strip() == "-":
        return ""
    return mat.replace("-", "").replace(" ", "").upper().strip()


# ── Main import ──────────────────────────────────────────────────────────

def importar_moeve(excel_path=None, excel_bytes=None):
    """Importa transacciones Moeve/Cepsa desde Excel.
    Returns dict con resumen.
    """
    _ensure_tables()

    if excel_bytes:
        wb = openpyxl.load_workbook(BytesIO(excel_bytes), read_only=True, data_only=True)
    elif excel_path:
        wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    else:
        raise ValueError("Se requiere excel_path o excel_bytes")

    ws = wb["data"]

    # Read all data rows (skip rows 0,1,2 = title, blank, header)
    raw_rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 3:
            continue
        if not row[7]:  # No date → skip
            continue
        raw_rows.append(row)
    wb.close()

    # ── Fase 0: Detectar meses con doble conteo (B1 + B2 simultáneo) ──
    # Cuando un mes tiene tanto B1 como B2, las transacciones B1 son duplicados
    # pre-facturación de las B2 (facturadas). Conservar solo B2, pero enriquecer
    # su Location con el valor de B1 (B2 suele tener "- -").
    mes_b1 = defaultdict(list)
    mes_b2 = defaultdict(list)
    for row in raw_rows:
        dt = str(row[7] or "").strip()
        dp = dt.split(" ")[0].split("/")
        if len(dp) != 3:
            continue
        mes = f"{dp[2]}-{dp[1].zfill(2)}"
        billed = str(row[19] or "").strip()
        if billed == "1":
            mes_b1[mes].append(row)
        else:
            mes_b2[mes].append(row)

    meses_overlap = set(mes_b1.keys()) & set(mes_b2.keys())

    # Para meses con overlap: construir lookup B1 para enriquecer ubicación de B2
    b1_location_lookup = {}  # (tarjeta, dia, concepto, litros_round) → location
    b1_descartados = 0
    if meses_overlap:
        for mes in meses_overlap:
            for row in mes_b1[mes]:
                tarjeta = str(row[0] or "").strip()
                dt = str(row[7] or "").strip()
                dia = dt.split(" ")[0]  # DD/MM/YYYY
                concepto = str(row[10] or "").strip()
                litros = round(abs(float(row[11] or 0)), 1)
                location = str(row[4] or "").strip()
                if location and location not in ("- -", "-"):
                    key = (tarjeta, dia, concepto, litros)
                    b1_location_lookup[key] = location
            b1_descartados += len(mes_b1[mes])

    # Filtrar raw_rows: descartar B1 de meses con overlap
    filtered_rows = []
    for row in raw_rows:
        dt = str(row[7] or "").strip()
        dp = dt.split(" ")[0].split("/")
        if len(dp) != 3:
            filtered_rows.append(row)
            continue
        mes = f"{dp[2]}-{dp[1].zfill(2)}"
        billed = str(row[19] or "").strip()
        if billed == "1" and mes in meses_overlap:
            continue  # Descartar B1 duplicado
        # Si es B2 de mes overlap y no tiene location, enriquecer desde B1
        if billed == "2" and mes in meses_overlap:
            location = str(row[4] or "").strip()
            if not location or location in ("- -", "-"):
                tarjeta = str(row[0] or "").strip()
                dia = dt.split(" ")[0]
                concepto = str(row[10] or "").strip()
                litros = round(abs(float(row[11] or 0)), 1)
                key = (tarjeta, dia, concepto, litros)
                b1_loc = b1_location_lookup.get(key, "")
                if b1_loc:
                    row = list(row)
                    row[4] = b1_loc  # Enriquecer Location
                    row = tuple(row)
        filtered_rows.append(row)

    # ── Fase 1: Netear descuentos Billed=2 ──
    # Group by (tarjeta, fecha_hora, concepto, litros) to find pairs
    groups = defaultdict(list)
    for row in filtered_rows:
        tarjeta = str(row[0] or "").strip()
        fecha_hora = str(row[7] or "").strip()
        concepto = str(row[10] or "").strip()
        litros = float(row[11] or 0)
        key = (tarjeta, fecha_hora, concepto, round(abs(litros), 4))
        groups[key].append(row)

    # Build final rows: netear pares
    final_rows = []
    neteados = 0
    for key, rows in groups.items():
        if len(rows) == 1:
            final_rows.append(rows[0])
        elif len(rows) == 2:
            # Par: sum importes (positive + negative = net)
            imp1 = float(rows[0][12] or 0)
            imp2 = float(rows[1][12] or 0)
            desc1 = float(rows[0][18] or 0)
            desc2 = float(rows[1][18] or 0)
            # Keep the positive one, adjust importe
            base = rows[0] if imp1 >= imp2 else rows[1]
            importe_neto = imp1 + imp2
            descuento_neto = desc1 + desc2
            # Create modified tuple
            base_list = list(base)
            base_list[12] = importe_neto
            base_list[18] = descuento_neto
            final_rows.append(tuple(base_list))
            neteados += 1
        else:
            # More than 2: keep all (rare edge case)
            final_rows.extend(rows)

    # ── Insert into DB ──
    import_id = str(uuid.uuid4())[:8]
    conn = get_conn()
    stats = {
        "registros_excel": len(raw_rows),
        "b1_descartados": b1_descartados,
        "meses_overlap": sorted(meses_overlap),
        "b1_locations_enriquecidas": len(b1_location_lookup),
        "pares_neteados": neteados,
        "registros_finales": len(final_rows),
        "insertados": 0,
        "duplicados": 0,
        "errores": 0,
        "vehiculos_creados": 0,
        "estaciones_creadas": 0,
    }

    try:
        matriculas_vistas = set()
        estaciones_vistas = set()

        for row in final_rows:
            try:
                tarjeta_completa = str(row[0] or "").strip()
                tarjeta = tarjeta_completa[-6:] if len(tarjeta_completa) >= 6 else tarjeta_completa
                matricula_raw = str(row[2] or "").strip()
                matricula = _normalizar_matricula(matricula_raw)
                estacion = str(row[4] or "").strip()
                localidad = _extraer_localidad(estacion)
                pais = str(row[5] or "").strip()
                fecha_hora_orig = str(row[7] or "").strip()
                fecha, hora = _parse_fecha(fecha_hora_orig)
                if not fecha:
                    stats["errores"] += 1
                    continue
                concepto = str(row[10] or "").strip()
                litros = float(row[11] or 0)
                importe = float(row[12] or 0)
                iva = float(row[13] or 0)
                factura = str(row[6] or "").strip()
                if factura == "-":
                    factura = ""
                billed = int(row[19]) if row[19] else None
                operacion = str(row[9] or "").strip()
                descuento = float(row[18] or 0)

                conn.execute("""
                    INSERT OR IGNORE INTO combustible_transacciones (
                        origen, tarjeta, tarjeta_completa, matricula,
                        estacion, localidad, pais, fecha, hora, fecha_hora_original,
                        concepto, litros, importe, iva_porcentaje, descuento,
                        factura, billed, operacion, importacion_id
                    ) VALUES (
                        'moeve', ?, ?, ?,
                        ?, ?, ?, ?, ?, ?,
                        ?, ?, ?, ?, ?,
                        ?, ?, ?, ?
                    )
                """, (
                    tarjeta, tarjeta_completa, matricula,
                    estacion, localidad, pais, fecha, hora, fecha_hora_orig,
                    concepto, litros, importe, iva, descuento,
                    factura, billed, operacion, import_id,
                ))
                if conn.total_changes:
                    stats["insertados"] += 1
                else:
                    stats["duplicados"] += 1

                # Track vehicles
                if matricula and matricula not in matriculas_vistas:
                    matriculas_vistas.add(matricula)

                # Track stations
                if estacion and estacion not in ("- -", "-") and estacion not in estaciones_vistas:
                    estaciones_vistas.add(estacion)

            except Exception as e:
                stats["errores"] += 1
                logger.debug("Error importando fila: %s", e)

        conn.commit()

        # ── Create vehicles ──
        for mat in matriculas_vistas:
            try:
                conn.execute(
                    "INSERT OR IGNORE INTO moeve_vehiculos (matricula) VALUES (?)",
                    (mat,),
                )
                if conn.total_changes:
                    stats["vehiculos_creados"] += 1
            except Exception:
                pass
        conn.commit()

        # ── Create stations ──
        for est in estaciones_vistas:
            loc = _extraer_localidad(est)
            try:
                conn.execute(
                    "INSERT OR IGNORE INTO moeve_estaciones_geo (estacion, localidad_extraida) VALUES (?, ?)",
                    (est, loc),
                )
                if conn.total_changes:
                    stats["estaciones_creadas"] += 1
            except Exception:
                pass
        conn.commit()

        # Fix insertados count (total_changes is cumulative)
        stats["insertados"] = conn.execute(
            "SELECT COUNT(*) FROM combustible_transacciones WHERE importacion_id = ?",
            (import_id,),
        ).fetchone()[0]
        stats["duplicados"] = stats["registros_finales"] - stats["insertados"] - stats["errores"]

    finally:
        conn.close()

    logger.info(
        "Moeve import: %d rows, %d netted, %d inserted, %d dupes, %d errors",
        stats["registros_excel"], stats["pares_neteados"],
        stats["insertados"], stats["duplicados"], stats["errores"],
    )
    return stats
