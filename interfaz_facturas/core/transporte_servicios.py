"""
Servicio de búsqueda de ruta y proveedores de transporte de maquinaria.
Usa OpenRouteService para geocoding y direcciones; filtra proveedores cercanos a la ruta.
Acepta datos desde proveedores_transporte.xlsx (prioridad) o proveedores_transporte.csv.
"""
from __future__ import annotations

import csv
import json
import math
import unicodedata
import urllib.parse
import urllib.request
from pathlib import Path
from typing import Any


def _ruta_proveedores_csv(base_dir: Path) -> Path:
  return base_dir / "data" / "proveedores_transporte.csv"


def _ruta_proveedores_xlsx(base_dir: Path) -> Path:
  return base_dir / "data" / "proveedores_transporte.xlsx"


def _normalizar_header(s: str) -> str:
  """Minúsculas, sin espacios extra, sin acentos para comparar nombres de columna."""
  if not s or not isinstance(s, str):
    return ""
  s = " ".join(str(s).strip().lower().split())
  s = unicodedata.normalize("NFD", s)
  s = "".join(c for c in s if unicodedata.category(c) != "Mn")
  return s


def _valor_celda(val: Any) -> str:
  """Convierte el valor de una celda Excel a string limpio."""
  if val is None:
    return ""
  if isinstance(val, (int, float)):
    if isinstance(val, float) and val == int(val):
      return str(int(val))
    return str(val).strip().replace(",", ".")
  return str(val).strip()


def decode_polyline(encoded: str) -> list[tuple[float, float]]:
  """Decodifica una polyline codificada (formato Google/ORS) a lista de (lat, lon)."""
  if not encoded or not encoded.strip():
    return []
  result: list[tuple[float, float]] = []
  prev_lat, prev_lon = 0.0, 0.0
  i = 0
  while i < len(encoded):
    shift = 0
    b = 0
    while True:
      c = ord(encoded[i]) - 63
      i += 1
      b |= (c & 0x1F) << shift
      shift += 5
      if c < 0x20:
        break
    dlat = (b & 1) and ~(b >> 1) or (b >> 1)
    prev_lat += dlat / 1e5
    shift = 0
    b = 0
    while i < len(encoded):
      c = ord(encoded[i]) - 63
      i += 1
      b |= (c & 0x1F) << shift
      shift += 5
      if c < 0x20:
        break
    dlon = (b & 1) and ~(b >> 1) or (b >> 1)
    prev_lon += dlon / 1e5
    result.append((prev_lat, prev_lon))
  return result


def haversine_km(lat1: float, lon1: float, lat2: float, lon2: float) -> float:
  """Distancia en km entre dos puntos (aproximación esférica)."""
  R = 6371.0
  phi1 = math.radians(lat1)
  phi2 = math.radians(lat2)
  dphi = math.radians(lat2 - lat1)
  dlam = math.radians(lon2 - lon1)
  a = math.sin(dphi / 2) ** 2 + math.cos(phi1) * math.cos(phi2) * math.sin(dlam / 2) ** 2
  c = 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))
  return R * c


def _distancia_punto_segmento_km(
  plat: float, plon: float,
  lat1: float, lon1: float, lat2: float, lon2: float,
) -> float:
  """Distancia mínima en km del punto al segmento."""
  dx = (lon2 - lon1) * 111.0 * max(0.01, math.cos(math.radians((lat1 + lat2) / 2)))
  dy = (lat2 - lat1) * 111.0
  seg_len = math.sqrt(dx * dx + dy * dy)
  if seg_len < 1e-9:
    return haversine_km(plat, plon, lat1, lon1)
  px = (plon - lon1) * 111.0 * math.cos(math.radians(plat))
  py = plat - lat1
  t = (px * dx + py * dy) / (seg_len * seg_len)
  t = max(0.0, min(1.0, t))
  qlat = lat1 + t * (lat2 - lat1)
  qlon = lon1 + t * (lon2 - lon1)
  return haversine_km(plat, plon, qlat, qlon)


def distancia_punto_polilinea_km(
  lat: float, lon: float, puntos: list[tuple[float, float]]
) -> float:
  """Distancia mínima del punto a la polilínea (lista de lat, lon)."""
  if not puntos:
    return float("inf")
  if len(puntos) == 1:
    return haversine_km(lat, lon, puntos[0][0], puntos[0][1])
  min_d = float("inf")
  for i in range(len(puntos) - 1):
    d = _distancia_punto_segmento_km(
      lat, lon,
      puntos[i][0], puntos[i][1],
      puntos[i + 1][0], puntos[i + 1][1],
    )
    min_d = min(min_d, d)
  return min_d


def geocode_ors(texto: str, api_key: str) -> tuple[float, float] | None:
  """Geocodifica una dirección/localidad con ORS. Devuelve (lat, lon) o None."""
  if not texto or not api_key:
    return None
  q = urllib.parse.quote(texto.strip())
  url = f"https://api.openrouteservice.org/geocode/search?api_key={api_key}&text={q}&size=1"
  try:
    req = urllib.request.Request(url, headers={"Authorization": api_key})
    with urllib.request.urlopen(req, timeout=10) as resp:
      data = json.loads(resp.read().decode("utf-8"))
  except Exception:
    return None
  features = data.get("features") or []
  if not features:
    return None
  coords = features[0].get("geometry", {}).get("coordinates")
  if not coords or len(coords) < 2:
    return None
  return (float(coords[1]), float(coords[0]))


def directions_ors(
  origen_lat: float, origen_lon: float,
  destino_lat: float, destino_lon: float,
  api_key: str,
) -> dict | None:
  """Obtiene la ruta entre dos puntos. Devuelve dict con distance_m, duration_s, geometry_encoded o coordinates."""
  return directions_ors_multi(
    [(origen_lat, origen_lon), (destino_lat, destino_lon)],
    api_key,
  )


def directions_ors_multi(puntos: list[tuple[float, float]], api_key: str) -> dict | None:
  """Ruta que pasa por todos los puntos en orden. puntos = [(lat, lon), ...]. Devuelve distance_m, duration_s, coordinates."""
  if not api_key or len(puntos) < 2:
    return None
  url = "https://api.openrouteservice.org/v2/directions/driving-car"
  coords_geojson = [[float(lon), float(lat)] for lat, lon in puntos]
  body = {
    "coordinates": coords_geojson,
    "instructions": False,
  }
  try:
    req = urllib.request.Request(
      url,
      data=json.dumps(body).encode("utf-8"),
      headers={"Content-Type": "application/json", "Authorization": api_key},
      method="POST",
    )
    with urllib.request.urlopen(req, timeout=30) as resp:
      data = json.loads(resp.read().decode("utf-8"))
  except Exception:
    return None
  routes = data.get("routes") or []
  if not routes:
    return None
  r = routes[0]
  summary = r.get("summary") or {}
  geometry = r.get("geometry")
  enc = None
  coords = None
  if isinstance(geometry, str):
    enc = geometry
  elif isinstance(geometry, list):
    coords = geometry
  elif isinstance(geometry, dict) and geometry.get("type") == "LineString":
    coords = geometry.get("coordinates") or []
  return {
    "distance_m": summary.get("distance") or 0,
    "duration_s": summary.get("duration") or 0,
    "geometry_encoded": enc,
    "coordinates": coords,
  }


def cargar_proveedores_transporte(base_dir: Path) -> list[dict[str, Any]]:
  """Carga proveedores desde .xlsx (si existe) o desde .csv. Devuelve lista con nombre, telefono, email, localidad, lat, lon."""
  xlsx_path = _ruta_proveedores_xlsx(base_dir)
  if xlsx_path.exists():
    return _cargar_proveedores_desde_xlsx(base_dir)
  return _cargar_proveedores_desde_csv(base_dir)


def _cargar_proveedores_desde_xlsx(base_dir: Path) -> list[dict[str, Any]]:
  """Lee la primera hoja de proveedores_transporte.xlsx y mapea columnas al formato interno."""
  try:
    import openpyxl
  except ImportError:
    return _cargar_proveedores_desde_csv(base_dir)

  ruta = _ruta_proveedores_xlsx(base_dir)
  if not ruta.exists():
    return []

  proveedores: list[dict[str, Any]] = []
  wb = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
  ws = wb.active
  if not ws:
    wb.close()
    return []

  # Cabeceras: primera fila
  headers: list[str] = []
  for cell in ws[1]:
    headers.append(_normalizar_header(_valor_celda(cell.value)))

  # Nombres de columna -> índice (varios alias por campo)
  def col_index(aliases: list[str]) -> int | None:
    for a in aliases:
      for i, h in enumerate(headers):
        if a in h or h in a or (a.replace(" ", "") in h.replace(" ", "")):
          return i
    return None

  idx_nombre = col_index(["nombre de empresa", "nombre", "empresa"])
  idx_localidad = col_index(["localidad"])
  idx_provincia = col_index(["provincia"])
  idx_cp = col_index(["codigo postal", "cp", "c postal"])
  idx_direccion = col_index(["direccion"])
  idx_email = col_index(["email", "correo", "e-mail"])
  idx_web = col_index(["web", "pagina web", "url"])
  idx_lat = col_index(["lat", "latitud"])
  idx_lon = col_index(["lon", "longitud", "lng"])
  idx_fijo = col_index(["fijo"])
  idx_movil = col_index(["movil", "móvil"])

  # Columnas de teléfono: todas las que contengan "telefono", "tel", "tlf", "fijo", "movil"
  idx_telefonos: list[int] = []
  for i, h in enumerate(headers):
    if not h:
      continue
    if "telefono" in h or "tel " in h or "tlf" in h or "telef" in h or h == "tel" or h == "fijo" or h == "movil":
      idx_telefonos.append(i)

  def get(row: tuple, idx: int | None) -> str:
    if idx is None or idx >= len(row):
      return ""
    return _valor_celda(row[idx])

  for row in ws.iter_rows(min_row=2, values_only=True):
    if not row:
      continue
    row = tuple(row) if not isinstance(row, tuple) else row

    nombre = get(row, idx_nombre) if idx_nombre is not None else ""
    localidad = get(row, idx_localidad)
    provincia = get(row, idx_provincia)
    cp = get(row, idx_cp)
    direccion = get(row, idx_direccion)

    telefonos = []
    for i in idx_telefonos:
      t = get(row, i)
      if t:
        telefonos.append(t)
    telefono = " / ".join(telefonos) if telefonos else ""

    tel_fijo = get(row, idx_fijo)
    tel_movil = get(row, idx_movil)

    email = get(row, idx_email)
    web = get(row, idx_web)

    lat_s = get(row, idx_lat).replace(",", ".")
    lon_s = get(row, idx_lon).replace(",", ".")
    lat, lon = None, None
    if lat_s and lon_s:
      try:
        lat = float(lat_s)
        lon = float(lon_s)
      except ValueError:
        pass

    texto_geocode = localidad or ""
    if direccion or cp or provincia:
      partes = [p for p in [direccion, cp, localidad, provincia] if p]
      texto_geocode = ", ".join(partes) if partes else texto_geocode

    p: dict[str, Any] = {
      "nombre": nombre,
      "telefono": telefono,
    "telefono_fijo": tel_fijo,
    "telefono_movil": tel_movil,
      "email": email,
      "localidad": texto_geocode or localidad,
      "lat": lat,
      "lon": lon,
    }
    if web:
      p["web"] = web
    if provincia:
      p["provincia"] = provincia
    if direccion:
      p["direccion"] = direccion
    if cp:
      p["codigo_postal"] = cp

    if p["nombre"] or p["localidad"]:
      proveedores.append(p)

  wb.close()
  return proveedores


def _cargar_proveedores_desde_csv(base_dir: Path) -> list[dict[str, Any]]:
  """Carga el CSV de proveedores. Columnas: nombre, telefono, email, localidad, lat, lon."""
  ruta = _ruta_proveedores_csv(base_dir)
  if not ruta.exists():
    return []
  proveedores: list[dict[str, Any]] = []
  with open(ruta, newline="", encoding="utf-8-sig") as f:
    reader = csv.DictReader(f)
    for row in reader:
      p = {
        "nombre": (row.get("nombre") or "").strip(),
        "telefono": (row.get("telefono") or "").strip(),
        "email": (row.get("email") or "").strip(),
        "localidad": (row.get("localidad") or "").strip(),
        "lat": None,
        "lon": None,
      }
      lat_s = (row.get("lat") or "").strip().replace(",", ".")
      lon_s = (row.get("lon") or "").strip().replace(",", ".")
      if lat_s and lon_s:
        try:
          p["lat"] = float(lat_s)
          p["lon"] = float(lon_s)
        except ValueError:
          pass
      if p["nombre"] or p["localidad"]:
        proveedores.append(p)
  return proveedores


def buscar_ruta_y_proveedores(
  origen_texto: str,
  destino_texto: str,
  base_dir: Path,
  api_key: str,
  radio_km: float = 50.0,
  paradas: list[str] | None = None,
) -> dict[str, Any]:
  """Calcula la ruta (origen → paradas → destino) y devuelve proveedores a menos de radio_km de la ruta."""
  resultado: dict[str, Any] = {"ruta": {}, "proveedores": []}
  if not api_key:
    return resultado

  origen = (origen_texto or "").strip()
  destino = (destino_texto or "").strip()
  if not origen or not destino:
    return resultado

  coord_origen = geocode_ors(origen, api_key)
  coord_destino = geocode_ors(destino, api_key)
  if not coord_origen or not coord_destino:
    return resultado

  paradas_texto = [p for p in (paradas or []) if (p or "").strip()]
  coords_paradas: list[tuple[float, float]] = []
  paradas_coords_info: list[dict[str, Any]] = []

  for i, texto in enumerate(paradas_texto):
    texto = (texto or "").strip()
    if not texto:
      continue
    geo = geocode_ors(texto, api_key)
    if geo:
      coords_paradas.append(geo)
      paradas_coords_info.append({
        "nombre": texto,
        "numero": i + 1,
        "lat": round(geo[0], 5),
        "lon": round(geo[1], 5),
      })

  puntos_ordenados: list[tuple[float, float]] = [coord_origen] + coords_paradas + [coord_destino]
  dirs = directions_ors_multi(puntos_ordenados, api_key)
  if not dirs:
    return resultado

  dist_m = dirs.get("distance_m") or 0
  dur_s = dirs.get("duration_s") or 0
  resultado["ruta"] = {
    "distancia_km": round(dist_m / 1000.0, 2),
    "duracion_min": round(dur_s / 60.0, 1),
    "paradas_coords": paradas_coords_info,
  }

  puntos_ruta: list[tuple[float, float]] = []
  enc = dirs.get("geometry_encoded")
  coords = dirs.get("coordinates")
  if enc and isinstance(enc, str):
    puntos_ruta = decode_polyline(enc)
  if coords and isinstance(coords, list):
    for c in coords:
      if len(c) >= 2:
        puntos_ruta.append((float(c[1]), float(c[0])))

  if puntos_ruta:
    resultado["ruta"]["coordenadas_ruta"] = [[round(p[0], 5), round(p[1], 5)] for p in puntos_ruta]

  if not puntos_ruta:
    return resultado

  proveedores_raw = cargar_proveedores_transporte(base_dir)
  for p in proveedores_raw:
    lat = p.get("lat")
    lon = p.get("lon")
    if lat is None or lon is None:
      if p.get("localidad"):
        geo = geocode_ors(p["localidad"], api_key)
        if geo:
          lat, lon = geo
          p["lat"], p["lon"] = lat, lon
    if lat is None or lon is None:
      continue
    dist_km = distancia_punto_polilinea_km(lat, lon, puntos_ruta)
    if dist_km <= radio_km:
      resultado["proveedores"].append({
        "nombre": p.get("nombre") or "—",
        "telefono": p.get("telefono") or "",
        "telefono_fijo": p.get("telefono_fijo") or "",
        "telefono_movil": p.get("telefono_movil") or "",
        "email": p.get("email") or "",
        "web": p.get("web") or "",
        "localidad": p.get("localidad") or "—",
        "distancia_km": round(dist_km, 2),
        "lat": round(lat, 5),
        "lon": round(lon, 5),
      })

  resultado["proveedores"].sort(key=lambda x: x["distancia_km"])
  return resultado
