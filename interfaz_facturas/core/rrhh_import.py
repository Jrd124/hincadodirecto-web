"""
Importador de nóminas desde el Excel de análisis.
Lee la hoja 'Base de Datos' del Excel y carga empleados + nóminas en SQLite.
Idempotente: ejecutar múltiples veces no duplica datos.
"""
from __future__ import annotations

import logging
from io import BytesIO
from datetime import datetime

import openpyxl

from core.db import conectar as _conectar, now_iso as _now
from core import empleados_db

logger = logging.getLogger("erp")

# ── Normalización de categorías ──────────────────────────────────────────
_CAT_MAP = {
    "peones": "Peones",
    "peón": "Peones",
    "peon": "Peones",
    "peones ord": "Peones",
    "peones ord.": "Peones",
    "oficial de 1ª": "Oficial de 1ª",
    "oficial de 1a": "Oficial de 1ª",
    "jefes de 1ª": "Jefes de 1ª",
    "jefes de 1a": "Jefes de 1ª",
}


def _normalizar_categoria(cat: str | None) -> str:
    if not cat:
        return ""
    return _CAT_MAP.get(cat.strip().lower(), cat.strip())


def importar_nominas(excel_path: str = None, excel_bytes: bytes = None) -> dict:
    """Importa nóminas desde archivo Excel o bytes.

    Returns dict con resumen: empleados_creados, empleados_actualizados,
    nominas_importadas, finiquitos_importados, errores.
    """
    empleados_db.init_empleados_db()

    # Abrir Excel
    if excel_bytes:
        wb = openpyxl.load_workbook(BytesIO(excel_bytes), read_only=True, data_only=True)
    elif excel_path:
        wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    else:
        raise ValueError("Se requiere excel_path o excel_bytes")

    ws = wb["Base de Datos"]

    # Leer todas las filas (skip header)
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue  # header
        if not row[0]:
            continue  # fila vacía
        rows.append(row)
    wb.close()

    stats = {
        "empleados_creados": 0,
        "empleados_actualizados": 0,
        "nominas_importadas": 0,
        "finiquitos_importados": 0,
        "errores": [],
    }

    now = _now()

    with _conectar() as conn:
        # ── Paso 1: Crear/actualizar empleados por DNI ──────────────
        # Recopilar info única por DNI
        dni_info = {}  # dni -> {nombre, categoria, antiguedad, periodos, tiene_finiquito, ultimo_periodo_nomina}
        for row in rows:
            periodo = str(row[0]).strip()
            nombre_raw = str(row[1]).strip()
            dni = str(row[2]).strip()
            categoria = str(row[3]).strip() if row[3] else ""
            antiguedad = str(row[4]).strip() if row[4] else ""
            tipo = str(row[5]).strip().upper()

            if dni not in dni_info:
                dni_info[dni] = {
                    "nombre_raw": nombre_raw,
                    "categoria": categoria,
                    "antiguedad": antiguedad,
                    "periodos": [],
                    "tiene_finiquito": False,
                    "ultimo_periodo_nomina": None,
                    "ultimo_periodo_finiquito": None,
                }

            info = dni_info[dni]
            info["periodos"].append(periodo)

            # Preferir nombre de NOMINA sobre FINIQUITO (sin sufijo "(Finiq.)")
            if tipo == "NOMINA":
                info["nombre_raw"] = nombre_raw
                info["ultimo_periodo_nomina"] = periodo
            elif tipo == "FINIQUITO":
                info["tiene_finiquito"] = True
                info["ultimo_periodo_finiquito"] = periodo

            # Actualizar categoría si tenemos una más reciente
            if categoria:
                info["categoria"] = categoria

        # Crear/actualizar cada empleado
        dni_to_id = {}
        for dni, info in dni_info.items():
            # Parsear nombre "Apellidos, Nombre"
            nombre_raw = info["nombre_raw"]
            # Limpiar sufijo de finiquito
            nombre_raw = nombre_raw.replace(" (Finiq.)", "").strip()

            if "," in nombre_raw:
                parts = nombre_raw.split(",", 1)
                apellidos = parts[0].strip()
                nombre = parts[1].strip()
            else:
                nombre = nombre_raw
                apellidos = ""

            categoria = _normalizar_categoria(info["categoria"])
            fecha_alta = min(info["periodos"]) + "-01"  # primer periodo

            # Determinar estado
            if info["tiene_finiquito"]:
                # Si finiquito es posterior a última nómina → exempleado
                ultimo_nom = info["ultimo_periodo_nomina"] or ""
                ultimo_fin = info["ultimo_periodo_finiquito"] or ""
                if ultimo_fin >= ultimo_nom:
                    estado = "exempleado"
                    fecha_baja = ultimo_fin + "-28"  # aprox fin de mes
                else:
                    estado = "activo"
                    fecha_baja = None
            else:
                estado = "activo"
                fecha_baja = None

            # Buscar si ya existe por DNI
            existing = conn.execute(
                "SELECT id FROM empleados WHERE dni = ?", (dni,)
            ).fetchone()

            if existing:
                emp_id = existing[0]
                conn.execute(
                    "UPDATE empleados SET nombre=?, apellidos=?, categoria=?, "
                    "fecha_antiguedad=?, estado=?, fecha_baja=?, updated_at=? "
                    "WHERE id=?",
                    (nombre, apellidos, categoria, info["antiguedad"],
                     estado, fecha_baja, now, emp_id),
                )
                stats["empleados_actualizados"] += 1
            else:
                cur = conn.execute(
                    "INSERT INTO empleados (nombre, apellidos, dni, categoria, "
                    "fecha_alta, fecha_baja, estado, fecha_antiguedad, created_at, updated_at) "
                    "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                    (nombre, apellidos, dni, categoria,
                     fecha_alta, fecha_baja, estado, info["antiguedad"], now, now),
                )
                emp_id = cur.lastrowid
                stats["empleados_creados"] += 1

            dni_to_id[dni] = emp_id

        # ── Paso 2: Importar nóminas ────────────────────────────────
        for row in rows:
            try:
                periodo = str(row[0]).strip()
                dni = str(row[2]).strip()
                tipo = str(row[5]).strip().upper()
                emp_id = dni_to_id.get(dni)
                if not emp_id:
                    stats["errores"].append(f"DNI {dni} no encontrado para periodo {periodo}")
                    continue

                dias = int(row[6] or 0)
                salario_base = float(row[7] or 0)
                antiguedad_euros = float(row[8] or 0)
                plus_asistencia = float(row[9] or 0)
                extra_mes = float(row[10] or 0)
                mejora_voluntaria = float(row[11] or 0)
                a_cuenta_convenio = float(row[12] or 0)
                dietas = float(row[13] or 0)
                cot_cc = float(row[14] or 0)
                cot_mei = float(row[15] or 0)
                cot_fp = float(row[16] or 0)
                cot_desempleo = float(row[17] or 0)
                irpf_porcentaje = float(row[18] or 0)
                irpf_euros = float(row[19] or 0)
                embargo = float(row[20] or 0)
                indemnizacion = float(row[21] or 0)
                vacaciones_prop = float(row[22] or 0)
                rem_total = float(row[23] or 0)
                base_ss = float(row[24] or 0)
                total_devengado = float(row[25] or 0)
                total_deducir = float(row[26] or 0)
                liquido = float(row[27] or 0)
                coste_empresa = float(row[28] or 0)

                coste_dia = round(coste_empresa / dias, 2) if dias > 0 else 0
                ss_empresa = round(coste_empresa - total_devengado, 2)

                conn.execute("""
                    INSERT OR REPLACE INTO nominas (
                        empleado_id, periodo, tipo, dias,
                        salario_base, antiguedad_euros, plus_asistencia, extra_mes,
                        mejora_voluntaria, a_cuenta_convenio, dietas,
                        indemnizacion, vacaciones_proporcionales,
                        cot_cc, cot_mei, cot_fp, cot_desempleo,
                        irpf_porcentaje, irpf_euros, embargo,
                        rem_total, base_ss, total_devengado, total_deducir,
                        liquido, coste_empresa, coste_dia, ss_empresa, created_at
                    ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                """, (
                    emp_id, periodo, tipo, dias,
                    salario_base, antiguedad_euros, plus_asistencia, extra_mes,
                    mejora_voluntaria, a_cuenta_convenio, dietas,
                    indemnizacion, vacaciones_prop,
                    cot_cc, cot_mei, cot_fp, cot_desempleo,
                    irpf_porcentaje, irpf_euros, embargo,
                    rem_total, base_ss, total_devengado, total_deducir,
                    liquido, coste_empresa, coste_dia, ss_empresa, now,
                ))

                if tipo == "FINIQUITO":
                    stats["finiquitos_importados"] += 1
                else:
                    stats["nominas_importadas"] += 1

            except Exception as e:
                stats["errores"].append(f"Fila periodo={row[0]}, dni={row[2]}: {e}")

    logger.info(
        "Importación nóminas: %d empleados creados, %d actualizados, %d nóminas, %d finiquitos, %d errores",
        stats["empleados_creados"], stats["empleados_actualizados"],
        stats["nominas_importadas"], stats["finiquitos_importados"],
        len(stats["errores"]),
    )
    return stats
