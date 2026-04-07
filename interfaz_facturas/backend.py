from __future__ import annotations

import csv
import difflib
import hashlib
import io
import logging
from logging.handlers import RotatingFileHandler
import os
import re
import shutil
import sqlite3
import time
import unicodedata
import zipfile

# ─── Logging con rotación ────────────────────────────────────────────────────
_log_dir = os.path.join(os.path.dirname(__file__), "data", "logs")
os.makedirs(_log_dir, exist_ok=True)
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        RotatingFileHandler(
            os.path.join(_log_dir, "erp.log"),
            maxBytes=5 * 1024 * 1024,
            backupCount=5,
            encoding="utf-8",
        ),
        logging.StreamHandler(),
    ],
)
logger = logging.getLogger("erp")
from concurrent.futures import ThreadPoolExecutor
from datetime import datetime
from pathlib import Path

from flask import Flask, Blueprint, jsonify, redirect, request, send_file, send_from_directory, Response, url_for
from flask_login import LoginManager, UserMixin, login_user, logout_user, login_required, current_user

from config import (
  ADMIN_PASSWORD,
  ADMIN_USER,
  BASE_DIR,
  BANCOS_DIR,
  DATOS_DIR,
  EMPRESAS_CLIENTE,
  EMPRESAS_DIR,
  FACTURAS_EMITIDAS_DIR,
  FACTURAS_RECIBIDAS_DIR,
  GESTION_DB,
  MOVIMIENTOS_DB,
  OPENROUTESERVICE_API_KEY,
  SECRET_KEY,
  SUBIDAS_DIR,
  client,
)
from core.facturas_servicios import filtrar_filas_csv as _filtrar_filas_csv
from core.transporte_servicios import (
  buscar_ruta_y_proveedores as _buscar_ruta_y_proveedores,
  parsear_xlsx_proveedores_desde_stream as _parsear_xlsx_proveedores_stream,
)
from core.proveedores_transporte_db import (
  alta_proveedor as _alta_proveedor_transporte,
  listar_proveedores_para_admin as _listar_proveedores_transporte_admin,
  obtener_proveedor as _obtener_proveedor_transporte,
  actualizar_proveedor as _actualizar_proveedor_transporte,
  insertar_desde_lista as _insertar_proveedores_transporte_lista,
)
from core import terceros_db, facturas_db, tarjetas_db, facturas_cliente_db, crm_db
from core.facturas_db import CAMPOS_FACTURAS_PROVEEDOR
from core.facturas_cliente_db import CAMPOS_FACTURAS_CLIENTE
from core.ocr import leer_texto_factura as _leer_texto_factura
from core.geocoding import (
  obtener_pais_desde_localidad as _obtener_pais_desde_localidad,
  enriquecer_pais_desde_localidad as _enriquecer_pais_desde_localidad,
)
from core.revisor import (
  revisor_basico as _revisor_basico,
  revisor_basico_clientes as _revisor_basico_clientes,
)
from core.proveedores import (
  normalizar_texto_proveedor as _normalizar_texto_proveedor,
  normalizar_nif as _normalizar_nif,
  cargar_proveedores_maestros as _cargar_proveedores_maestros,
  listar_proveedores_para_selector as _listar_proveedores_para_selector,
  listar_proveedores_con_facturas as _listar_proveedores_con_facturas,
  guardar_proveedores_maestros as _guardar_proveedores_maestros,
  sincronizar_proveedores_desde_facturas as _sincronizar_proveedores_desde_facturas,
  similitud_nombres as _similitud_nombres,
  buscar_o_crear_proveedor as _buscar_o_crear_proveedor,
  homogeneizar_proveedores as _homogeneizar_proveedores,
)
from core.llm import (
  limpiar_json_respuesta as _limpiar_json_respuesta,
  extraer_campos_llm as _extraer_campos_llm,
  extraer_campos_vision as _extraer_campos_vision,
  registrar_vision_control as _registrar_vision_control,
  sugerencias_llm as _sugerencias_llm,
  CLAVES_FACTURA_CLIENTE as _CLAVES_FACTURA_CLIENTE,
)
from core.archivador import (
  hash_archivo as _hash_archivo,
  normalizar_fecha_factura_clave as _normalizar_fecha_factura_clave,
  clave_logica_factura_proveedor as _clave_logica_factura_proveedor,
  añadir_hashes_tabla as _añadir_hashes_tabla,
  archivar_por_fecha as _archivar_por_fecha,
)
from core.parser import (
  normalizar_texto as _normalizar_texto,
  normalizar_importe_str as _normalizar_importe_str,
  extraer_ultimo_importe_linea as _extraer_ultimo_importe_linea,
  normalizar_fecha_a_iso as _normalizar_fecha_a_iso,
  buscar_primera_fecha as _buscar_primera_fecha,
  buscar_nif_cif as _buscar_nif_cif,
  buscar_numero_factura as _buscar_numero_factura,
  buscar_proveedor_y_localizacion as _buscar_proveedor_y_localizacion,
  buscar_concepto as _buscar_concepto,
  buscar_importes as _buscar_importes,
)

# ─── Validación de entrada centralizada (respuestas 400 consistentes) ─────────

def _bad_request(mensaje: str):
  """Devuelve respuesta 400 con formato consistente: { \"error\": \"mensaje\" }."""
  return jsonify({"error": mensaje}), 400


def _validar_empresa_id_requerido(val) -> tuple:
  """
  Valida que empresa_id esté presente y no vacío tras strip().
  Devuelve (empresa_id_limpio, None) si es válido, o (None, (response, 400)) si no.
  """
  if val is None:
    return None, _bad_request("Falta empresa_id")
  empresa_id = (val if isinstance(val, str) else str(val or "")).strip()
  if not empresa_id:
    return None, _bad_request("Falta empresa_id")
  return empresa_id, None


# ─── Configuración de rutas (Blueprints por dominio) ───────────────────────────

facturas_proveedores_bp = Blueprint("facturas_proveedores", __name__)
proveedores_bp = Blueprint("proveedores", __name__)
facturas_clientes_bp = Blueprint("facturas_clientes", __name__)
archivo_bp = Blueprint("archivo", __name__)
control_calidad_bp = Blueprint("control_calidad", __name__)
bancos_bp = Blueprint("bancos", __name__)

# Workers para procesamiento en paralelo del pipeline de facturas (OpenAI/vision)
_MAX_WORKERS_EXTRACTOR_LLM = 4


def ensure_dirs() -> None:
  """Crea directorios e inicializa todas las bases de datos."""
  DATOS_DIR.mkdir(exist_ok=True)
  SUBIDAS_DIR.mkdir(exist_ok=True)
  FACTURAS_RECIBIDAS_DIR.mkdir(exist_ok=True)
  FACTURAS_EMITIDAS_DIR.mkdir(exist_ok=True)
  EMPRESAS_DIR.mkdir(exist_ok=True)
  BANCOS_DIR.mkdir(parents=True, exist_ok=True)
  # Inicializar todas las BDs al arranque (idempotente gracias a flags internos)
  facturas_db.init_facturas_db()
  facturas_cliente_db.init_facturas_cliente_db()
  terceros_db.init_terceros_db()
  tarjetas_db.init_tarjetas_db()
  crm_db.init_crm_db()
  _init_movimientos_db()


def _recolector(carpeta: Path) -> list[Path]:
  """
  Recolector: devuelve los archivos de la carpeta de entrada.
  """
  if not carpeta.exists() or not carpeta.is_dir():
    return []
  return [p for p in carpeta.iterdir() if p.is_file()]


def _extraer_una_factura_llm_proveedor(ruta: Path, empresa_id: str) -> dict | None:
  """
  Extrae los campos de una sola factura (proveedor) con visión y/o LLM.
  Devuelve la fila de datos o None si no se pudo extraer (equivalente a continue).
  Usado por _extractor_llm para ejecución en paralelo.
  """
  usar_vision = False
  datos: dict = {}
  suf = (ruta.suffix or "").lower()
  es_imagen = suf in (".jpg", ".jpeg", ".png", ".webp", ".gif")

  if es_imagen:
    datos = _extraer_campos_vision(ruta, empresa_id)
    if datos:
      usar_vision = True
      _registrar_vision_control(empresa_id, ruta.name, str(ruta))

  texto = ""

  if not datos:
    texto = _leer_texto_factura(ruta)
    texto = _normalizar_texto(texto)
    datos = _extraer_campos_llm(texto, empresa_id, tipo="proveedor")

  if not datos:
    return None

  claves_clave = ["proveedor", "cif_nif", "fecha_factura", "numero_factura", "total_a_pagar", "total_factura"]

  def _es_casi_sin_datos(d: dict) -> bool:
    num_no_vacias = sum(
      1
      for c in claves_clave
      if str(d.get(c, "") or "").strip()
    )
    bases_raw_local = d.get("bases") or []
    rets_raw_local = d.get("retenciones") or []
    return num_no_vacias <= 2 and not bases_raw_local and not rets_raw_local

  if _es_casi_sin_datos(datos):
    if es_imagen and usar_vision:
      if not texto:
        texto = _leer_texto_factura(ruta)
        texto = _normalizar_texto(texto)
      datos_texto = _extraer_campos_llm(texto, empresa_id, tipo="proveedor")
      if datos_texto and not _es_casi_sin_datos(datos_texto):
        datos = datos_texto
        usar_vision = False
    elif es_imagen and not usar_vision:
      datos_vision = _extraer_campos_vision(ruta, empresa_id)
      if datos_vision and not _es_casi_sin_datos(datos_vision):
        datos = datos_vision
        usar_vision = True
        _registrar_vision_control(empresa_id, ruta.name, str(ruta))

  if _es_casi_sin_datos(datos):
    return None

  bases_raw = datos.get("bases") or []
  rets_raw = datos.get("retenciones") or []
  bases = bases_raw
  retenciones = rets_raw

  def _suma_lista(lst: list[dict], campo: str) -> float:
    total = 0.0
    for item in lst:
      val = _normalizar_importe_str(str(item.get(campo, "")))
      if val is not None:
        total += val
    return total

  base_total = _suma_lista(bases, "base")
  iva_total = _suma_lista(bases, "cuota")
  ret_total = _suma_lista(retenciones, "importe")

  def fmt(v: float | None) -> str:
    return "" if v is None else f"{v:.2f}"

  base_str = fmt(base_total)
  iva_str = fmt(iva_total)
  ret_str = fmt(ret_total)

  total_factura = str(datos.get("total_factura") or "").strip()
  total_a_pagar = str(datos.get("total_a_pagar") or "").strip() or total_factura

  return {
    "ruta_archivo": str(ruta),
    "empresa_id": empresa_id,
    "fecha_factura": str(datos.get("fecha_factura") or "").strip(),
    "proveedor": str(datos.get("proveedor") or "").strip(),
    "nif_proveedor": str(datos.get("cif_nif") or "").strip(),
    "pais_proveedor": str(datos.get("pais_proveedor") or "").strip(),
    "localidad_proveedor": str(datos.get("localidad_proveedor") or "").strip(),
    "resumen_concepto": str(datos.get("resumen_concepto") or "").strip(),
    "numero_factura": str(datos.get("numero_factura") or "").strip(),
    "base_imponible_total": base_str,
    "base_imponible_detalle": "; ".join(
      f"{(b.get('base') or '').strip()}@{(b.get('tipo_iva') or '').strip()}" for b in bases
    ).strip(),
    "iva_cuota_total": iva_str,
    "iva_cuota_detalle": "; ".join(
      f"{(b.get('cuota') or '').strip()}@{(b.get('tipo_iva') or '').strip()}" for b in bases
    ).strip(),
    "retenciones_total": ret_str,
    "retenciones_detalle": "; ".join(
      f"{(r.get('importe') or '').strip()}@{(r.get('tipo') or '').strip()}" for r in retenciones
    ).strip(),
    "total_factura": total_factura,
    "total_a_pagar": total_a_pagar,
    "base_imponible": base_str,
    "iva": iva_str,
    "total": total_a_pagar,
    "categoria": "",
    "extraccion_vision": "1" if usar_vision else "",
  }


def _extractor_llm(rutas: list[Path], empresa_id: str) -> list[dict]:
  """
  Extractor principal: usa visión para imágenes y OCR+texto para el resto.
  Procesa facturas en paralelo con ThreadPoolExecutor (respeta límites de tasa razonables).
  """
  filas: list[dict] = []
  if not rutas:
    return filas
  workers = min(_MAX_WORKERS_EXTRACTOR_LLM, len(rutas))
  with ThreadPoolExecutor(max_workers=workers) as executor:
    futures = [executor.submit(_extraer_una_factura_llm_proveedor, ruta, empresa_id) for ruta in rutas]
    for fut in futures:
      try:
        fila = fut.result()
        if fila is not None:
          filas.append(fila)
      except Exception as e:
        logger.warning("Error procesando factura en paralelo: %s", e)
  return filas


def _extractor_basico(rutas: list[Path], empresa_id: str) -> list[dict]:
  """
  Extractor basado en reglas (backup). Intenta leer el texto de la factura y
  rellenar los campos principales sin usar LLM.
  """
  filas: list[dict] = []
  for ruta in rutas:
    texto = _leer_texto_factura(ruta)
    texto = _normalizar_texto(texto)
    num_factura = _buscar_numero_factura(texto)
    proveedor, pais, localidad = _buscar_proveedor_y_localizacion(texto, num_factura)
    if proveedor and not num_factura and re.match(r"^[A-Z0-9\/\-\.]+\s*$", proveedor.strip()):
      num_factura = proveedor[:60]
      proveedor, pais, localidad = _buscar_proveedor_y_localizacion(texto, num_factura)
    nif = _buscar_nif_cif(texto)
    fecha = _buscar_primera_fecha(texto)
    concepto = _buscar_concepto(texto, proveedor)
    if concepto and proveedor and concepto.lower().strip() == proveedor.lower().strip():
      concepto = ""
    importes = _buscar_importes(texto)

    base_total = importes["base_imponible_total"]
    iva_total = importes["iva_cuota_total"]
    ret_total = importes["retenciones_total"]

    # Convertir a strings para almacenamiento
    def fmt(v: float | None) -> str:
      return "" if v is None else f"{v:.2f}"

    base_str = fmt(base_total)
    iva_str = fmt(iva_total)
    ret_str = fmt(ret_total)

    total_factura = importes["total_factura"]
    total_a_pagar = importes["total_a_pagar"] or total_factura

    fila = {
      "ruta_archivo": str(ruta),
      "empresa_id": empresa_id,
      "fecha_factura": fecha,
      "proveedor": proveedor,
      "nif_proveedor": nif,
      "pais_proveedor": pais,
      "localidad_proveedor": localidad,
      "resumen_concepto": concepto,
      "numero_factura": num_factura,
      # Campos agregados
      "base_imponible_total": base_str,
      "base_imponible_detalle": importes["base_imponible_detalle"],
      "iva_cuota_total": iva_str,
      "iva_cuota_detalle": importes["iva_cuota_detalle"],
      "retenciones_total": ret_str,
      "retenciones_detalle": importes["retenciones_detalle"],
      "total_factura": total_factura,
      "total_a_pagar": total_a_pagar,
      # Campos históricos sencillos
      "base_imponible": base_str,
      "iva": iva_str,
      "total": total_a_pagar,
      "categoria": "",
      "extraccion_vision": "",
    }
    filas.append(fila)
  return filas


def _analizar_fila_proveedor(fila: dict) -> list[str]:
  """
  Aplica las mismas reglas que _revisor_basico pero sin modificar la fila.
  Devuelve una lista de mensajes de error (vacía si no hay problemas).
  Usado por Control de calidad para listar facturas con problemas.
  """
  errores: list[str] = []
  fecha = (fila.get("fecha_factura") or "").strip()
  if not fecha:
    errores.append("Sin fecha de factura.")
  else:
    try:
      fecha_dt = datetime.fromisoformat(fecha[:10]).date()
      if fecha_dt > datetime.now().date():
        errores.append(f"Fecha futura ({fecha[:10]}).")
    except Exception as e:
      logger.debug("No se pudo parsear fecha '%s': %s", fecha[:10], e)

  base_str = (fila.get("base_imponible") or fila.get("base_imponible_total") or "").strip()
  iva_str = (fila.get("iva") or fila.get("iva_cuota_total") or "").strip()
  ret_str = (fila.get("retenciones_total") or "").strip()
  total_pagar_str = (fila.get("total_a_pagar") or "").strip()

  base_val = _normalizar_importe_str(base_str) or 0.0
  iva_val = _normalizar_importe_str(iva_str) or 0.0
  ret_val = _normalizar_importe_str(ret_str) or 0.0
  total_pagar_val = _normalizar_importe_str(total_pagar_str)
  esperado_pagar = base_val + iva_val - ret_val

  if total_pagar_val is not None and esperado_pagar > 0:
    if abs(esperado_pagar - total_pagar_val) > 0.05:
      errores.append(
        f"Descuadre: base({base_val:.2f}) + iva({iva_val:.2f}) - ret({ret_val:.2f}) != total_pagar({total_pagar_val:.2f})."
      )
  return errores


def _analizar_fila_cliente(fila: dict) -> list[str]:
  """
  Aplica las mismas reglas que _revisor_basico_clientes sin modificar la fila.
  Devuelve lista de mensajes de error. Usado por Control de calidad.
  """
  errores: list[str] = []
  fecha = (fila.get("fecha_factura") or "").strip()
  if not fecha:
    errores.append("Sin fecha de factura.")
  else:
    try:
      fecha_dt = datetime.fromisoformat(fecha[:10]).date()
      if fecha_dt > datetime.now().date():
        errores.append(f"Fecha futura ({fecha[:10]}).")
    except Exception as e:
      logger.debug("No se pudo parsear fecha '%s': %s", fecha[:10], e)

  iva_val = _normalizar_importe_str(fila.get("iva") or "")
  total_val = _normalizar_importe_str(fila.get("total_a_pagar") or "")
  pricing_sum = 0.0
  for campo_p in ("pricing_servicio", "pricing_transporte"):
    v = _normalizar_importe_str(fila.get(campo_p) or "")
    if v is not None:
      pricing_sum += v

  if pricing_sum > 0 and iva_val is not None and total_val is not None:
    esperado = pricing_sum + iva_val
    if abs(esperado - total_val) > 0.05:
      errores.append(
        f"Descuadre: pricing({pricing_sum:.2f}) + iva({iva_val:.2f}) != total({total_val:.2f})"
      )
  return errores


def _analizar_facturas_proveedores(filas: list[dict]) -> list[dict]:
  """
  Recorre las filas de facturas de proveedores y devuelve las que tienen al menos un error,
  con la lista de mensajes y la fila completa. No modifica los datos.
  """
  resultado: list[dict] = []
  for i, fila in enumerate(filas):
    errores = _analizar_fila_proveedor(fila)
    if not errores:
      continue
    ruta = (fila.get("ruta_destino") or fila.get("ruta_archivo") or "").strip()
    resultado.append({
      "indice": i,
      "ruta_archivo": ruta,
      "errores": errores,
      "fila": fila,
    })
  return resultado


def _analizar_facturas_clientes(filas: list[dict]) -> list[dict]:
  """
  Recorre las filas de facturas de clientes y devuelve las que tienen al menos un error.
  """
  resultado: list[dict] = []
  for i, fila in enumerate(filas):
    errores = _analizar_fila_cliente(fila)
    if not errores:
      continue
    ruta = (fila.get("ruta_archivo") or "").strip()
    resultado.append({
      "indice": i,
      "ruta_archivo": ruta,
      "errores": errores,
      "fila": fila,
    })
  return resultado


def _sugerencias_heuristicas(fila: dict, errores: list[str], tipo: str) -> list[dict]:
  """
  Genera sugerencias a partir de los mensajes de error y la fila (reglas heurísticas).
  tipo: "proveedores" | "clientes".
  Devuelve list[dict] con { "campo", "valor_actual", "valor_sugerido", "motivo" }.
  """
  sugerencias: list[dict] = []
  fila = fila or {}

  for texto in (errores or []):
    texto_lower = texto.lower()
    # Descuadre proveedores: base + iva - ret != total_pagar → sugerir total_a_pagar
    if "descuadre" in texto_lower and "base" in texto_lower and tipo == "proveedores":
      base_val = _normalizar_importe_str(fila.get("base_imponible") or fila.get("base_imponible_total") or "") or 0.0
      iva_val = _normalizar_importe_str(fila.get("iva") or fila.get("iva_cuota_total") or "") or 0.0
      ret_val = _normalizar_importe_str(fila.get("retenciones_total") or "") or 0.0
      total_sugerido = base_val + iva_val - ret_val
      valor_actual = (fila.get("total_a_pagar") or "").strip()
      sugerencias.append({
        "campo": "total_a_pagar",
        "valor_actual": valor_actual or "—",
        "valor_sugerido": f"{total_sugerido:.2f}",
        "motivo": "Corregir total para que cuadre con base + IVA − retenciones.",
      })
      continue
    # Descuadre clientes: pricing + iva != total → sugerir total_a_pagar
    if "descuadre" in texto_lower and "pricing" in texto_lower and tipo == "clientes":
      pricing_sum = 0.0
      for c in ("pricing_servicio", "pricing_transporte"):
        v = _normalizar_importe_str(fila.get(c) or "")
        if v is not None:
          pricing_sum += v
      iva_val = _normalizar_importe_str(fila.get("iva") or "") or 0.0
      total_sugerido = pricing_sum + iva_val
      valor_actual = (fila.get("total_a_pagar") or "").strip()
      sugerencias.append({
        "campo": "total_a_pagar",
        "valor_actual": valor_actual or "—",
        "valor_sugerido": f"{total_sugerido:.2f}",
        "motivo": "Corregir total para que cuadre con pricing + IVA.",
      })
      continue
    # Sin fecha: no sugerir valor concreto
    if "sin fecha" in texto_lower:
      sugerencias.append({
        "campo": "fecha_factura",
        "valor_actual": (fila.get("fecha_factura") or "").strip() or "—",
        "valor_sugerido": "",
        "motivo": "Requiere revisión manual.",
      })
      continue
    # Fecha futura: sugerir revisar / vacío
    if "fecha futura" in texto_lower:
      valor_actual = (fila.get("fecha_factura") or "").strip() or "—"
      sugerencias.append({
        "campo": "fecha_factura",
        "valor_actual": valor_actual,
        "valor_sugerido": "",
        "motivo": "Revisar fecha.",
      })

  return sugerencias


def _base_maestra_csv(filas: list[dict], empresa_id: str) -> dict:
  """
  Guarda las filas en la base maestra de facturas (SQLite).
  Las filas recibidas ya están filtradas (sin duplicados por hash).
  """
  resultado = facturas_db.insert_facturas(empresa_id, filas)
  ruta_csv = EMPRESAS_DIR / empresa_id / "base_maestra_facturas.csv"
  return {
    "ruta_base_maestra": str(ruta_csv),
    "filas_añadidas": resultado["insertados"],
    "ids_insertados": resultado["ids"],
  }


def _subir_lote_a_sharepoint(tabla: list[dict], tipo: str = "Facturas Recibidas") -> None:
  """Sube PDFs procesados a SharePoint en background. No falla si OneDrive no está configurado."""
  if not os.environ.get("MICROSOFT_CLIENT_ID"):
    return
  try:
    from core.onedrive_db import get_sharepoint_client
    client = get_sharepoint_client()
  except Exception as exc:
    logger.warning("No se pudo conectar a SharePoint para subida: %s", exc)
    return

  meses_en = [
    "January", "February", "March", "April", "May", "June",
    "July", "August", "September", "October", "November", "December",
  ]
  for fila in tabla:
    ruta_local = fila.get("ruta_destino") or fila.get("ruta_archivo") or ""
    if not ruta_local or not Path(ruta_local).is_file():
      continue
    empresa = fila.get("empresa_id", "")
    fecha = fila.get("fecha_factura", "")
    try:
      from datetime import datetime as _dt
      dt = _dt.strptime(fecha[:10], "%Y-%m-%d")
    except Exception:
      from datetime import datetime as _dt
      dt = _dt.now()
    mes_carpeta = f"{dt.month:02d}. {meses_en[dt.month - 1]}"
    nombre = Path(ruta_local).name
    sp_path = f"{tipo}/{empresa}/{dt.year}/{mes_carpeta}/{nombre}"
    try:
      with open(ruta_local, "rb") as f:
        contenido = f.read()
      resultado = client.subir_archivo(sp_path, contenido)
      if resultado:
        logger.info("Factura subida a SharePoint: %s", sp_path)
      else:
        logger.warning("Error subiendo factura a SharePoint: %s", sp_path)
    except Exception as exc:
      logger.warning("Error subiendo %s a SharePoint: %s", sp_path, exc)


def procesar_lote(empresa_id: str, carpeta: Path, tarjeta_id: str | None = None) -> dict:
  """
  Orquestador interno: aplica Recolector → Extractor → Revisor → Archivador → Base de datos.
  """
  archivos = _recolector(carpeta)
  if not archivos:
    return {
      "procesado": False,
      "motivo": "No se han encontrado archivos en la carpeta de entrada.",
      "empresa_id": empresa_id,
      "carpeta_entrada": str(carpeta),
    }
  logger.info("Procesando %d archivos para empresa '%s'", len(archivos), empresa_id)

  # Primero intentamos con el extractor LLM (OpenAI). Si falla o no devuelve filas, usamos el backup.
  tabla = _extractor_llm(archivos, empresa_id)
  if not tabla:
    logger.info("Extractor LLM sin resultados, usando extractor básico")
    tabla = _extractor_basico(archivos, empresa_id)
  else:
    logger.info("Extractor LLM: %d filas extraídas", len(tabla))

  tabla = _revisor_basico(tabla)
  _añadir_hashes_tabla(tabla)
  hashes_existentes = facturas_db.get_hashes_empresa_proveedor(empresa_id)
  claves_existentes = {
    _clave_logica_factura_proveedor(num, prov, fec)
    for num, prov, fec in facturas_db.get_claves_facturas_proveedor(empresa_id)
  }
  tabla_sin_duplicados = []
  for f in tabla:
    h = (f.get("hash_archivo") or "").strip()
    if h and h in hashes_existentes:
      continue
    clave = _clave_logica_factura_proveedor(
      f.get("numero_factura"), f.get("proveedor"), f.get("fecha_factura")
    )
    if clave in claves_existentes:
      continue
    tabla_sin_duplicados.append(f)
  duplicados_omitidos = len(tabla) - len(tabla_sin_duplicados)
  if duplicados_omitidos:
    logger.info("Duplicados omitidos: %d", duplicados_omitidos)
  tabla = _archivar_por_fecha(tabla_sin_duplicados, FACTURAS_RECIBIDAS_DIR)
  _subir_lote_a_sharepoint(tabla, "Facturas Recibidas")
  tabla = _homogeneizar_proveedores(tabla, empresa_id)
  tabla = _enriquecer_pais_desde_localidad(tabla)

  # Si se ha indicado tarjeta en el procesado, marcamos todas las facturas
  # de este lote como pagadas con esa tarjeta.
  tarjeta_int: int | None = None
  if tarjeta_id:
    try:
      tarjeta_int = int(str(tarjeta_id).strip())
    except ValueError:
      tarjeta_int = None
  if tarjeta_int:
    for fila in tabla:
      fila["tarjeta_id"] = tarjeta_int
      if not (fila.get("estado_pago") or "").strip():
        fila["estado_pago"] = "pagada"
  resumen_bd = _base_maestra_csv(tabla, empresa_id)

  # Post-insert: vincular facturas recien insertadas a terceros (por CIF/nombre)
  try:
    crm_db.vincular_facturas_a_terceros()
  except Exception as e:
    logger.warning("Error al vincular facturas a terceros post-procesamiento: %s", e)

  facturas_con_vision = sum(1 for f in tabla if (str(f.get("extraccion_vision") or "").strip() == "1"))
  logger.info(
    "Lote completado: %d facturas procesadas, %d con visión, %d añadidas a BD",
    len(tabla_sin_duplicados), facturas_con_vision, resumen_bd["filas_añadidas"],
  )

  return {
    "procesado": True,
    "empresa_id": empresa_id,
    "carpeta_entrada": str(carpeta),
    "facturas_procesadas": len(tabla_sin_duplicados),
    "facturas_omitidas_duplicadas": duplicados_omitidos,
    "facturas_con_vision": facturas_con_vision,
    "ruta_base_maestra": resumen_bd["ruta_base_maestra"],
    "filas_añadidas": resumen_bd["filas_añadidas"],
    "ids_insertados": resumen_bd.get("ids_insertados", []),
  }


app = Flask(__name__, static_folder=".", static_url_path="")
app.secret_key = SECRET_KEY

# ─── Autenticación (Flask-Login + BD) ────────────────────────────────────────

from core.usuarios_db import init_usuarios_db, verificar_credenciales
init_usuarios_db()

from core import proyectos_db
proyectos_db.init_proyectos_db()

from core import maquinaria_db
maquinaria_db.init_maquinaria_db()

from core import empleados_db
empleados_db.init_empleados_db()

from core import vehiculos_db
vehiculos_db.init_vehiculos_db()

from core import cae_db
cae_db.init_cae_db()

login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = "usuarios.login_page"


class _User(UserMixin):
  def __init__(self, uid, username, nombre="", rol="admin"):
    self.id = str(uid)
    self.username = username
    self.nombre = nombre
    self.rol = rol


@login_manager.user_loader
def _load_user(user_id: str):
  # Intentar cargar desde BD
  from core.usuarios_db import obtener_usuario
  u = obtener_usuario(int(user_id)) if user_id.isdigit() else None
  if u:
    return _User(u["id"], u["username"], u["nombre"], u["rol"])
  # Fallback legacy: el user_id es el username del .env
  if user_id == ADMIN_USER:
    return _User(0, ADMIN_USER, "Admin (.env)", "admin")
  return None


def requiere_rol(*roles_permitidos):
  """Decorador para proteger endpoints por rol."""
  from functools import wraps
  def decorator(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
      rol = getattr(current_user, "rol", "admin") if current_user.is_authenticated else ""
      if rol not in roles_permitidos:
        return jsonify({"error": "Sin permisos"}), 403
      return f(*args, **kwargs)
    return wrapper
  return decorator


@app.before_request
def _require_login():
  """Protege todas las rutas excepto login, estáticos y acceso operario por token."""
  rutas_publicas = ("usuarios.login_page", "usuarios.login_post", "static", "api_general.api_health")
  if request.endpoint in rutas_publicas:
    return
  # Rutas públicas de operario via token (sin login)
  if request.path.startswith("/m/") or request.path.startswith("/api/m/"):
    return
  # Rutas públicas de mantenimiento por tarea via token (sin login)
  if request.path.startswith("/w/") or request.path.startswith("/api/w/"):
    return
  # Servir fotos de maquinaria sin auth (protegidas por filename opaco)
  if request.path.startswith("/fotos_maquinaria/"):
    return
  if not current_user.is_authenticated:
    if request.path.startswith("/api/"):
      return jsonify({"error": "No autenticado"}), 401
    return redirect(url_for("usuarios.login_page"))


@app.after_request
def _log_api_request(response):
  if request.path.startswith("/api/"):
    user = current_user.username if current_user.is_authenticated else "anon"
    logger.info("%s %s -> %s [%s]", request.method, request.path, response.status_code, user)
    # Prevent browser caching of API responses
    response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    response.headers["Pragma"] = "no-cache"
  # Seguridad: rutas /w/ y /m/ no indexables
  if request.path.startswith("/w/") or request.path.startswith("/m/"):
    response.headers["X-Robots-Tag"] = "noindex, nofollow"
  return response


@app.route("/robots.txt")
def robots_txt():
  return Response("User-agent: *\nDisallow: /m/\nDisallow: /w/\nDisallow: /api/\n",
                  mimetype="text/plain")


# Auth + usuarios endpoints → routes/usuarios.py


# Maquinaria endpoints → routes/maquinaria.py


# index() + health + dashboard + finanzas + empresas → routes/api_general.py


@facturas_proveedores_bp.post("/api/procesar")
def procesar():
  """
  Backend para la interfaz de facturas.

  - Recibe empresa_id + archivos.
  - Guarda los archivos en data/subidas/{empresa_id}/{timestamp}/.
  - Lanza el flujo interno (Recolector → Extractor → Revisor → Archivador → Base de datos).
  """
  ensure_dirs()

  empresa_id = request.form.get("empresa_id")
  empresa_id, err = _validar_empresa_id_requerido(empresa_id)
  if err:
    return err[0], err[1]

  tarjeta_id = (request.form.get("tarjeta_id") or "").strip() or None

  files = request.files.getlist("archivos")
  if not files:
    return _bad_request("No se han recibido archivos")

  timestamp = int(time.time())
  destino = SUBIDAS_DIR / empresa_id / str(timestamp)
  destino.mkdir(parents=True, exist_ok=True)

  nombres_guardados = []
  for f in files:
    nombre = f.filename or f"factura_{len(nombres_guardados) + 1}.dat"
    ruta = destino / os.path.basename(nombre)
    f.save(ruta)
    nombres_guardados.append(str(ruta))

  resumen = procesar_lote(empresa_id, destino, tarjeta_id=tarjeta_id)
  _invalidar_cache_listado_proveedores(empresa_id)

  mensaje = "Facturas procesadas correctamente."
  if not resumen.get("procesado"):
    mensaje = resumen.get("motivo", "No se han podido procesar las facturas.")

  return jsonify(
    {
      "mensaje": mensaje,
      "empresa_id": empresa_id,
      "carpeta_entrada": str(destino),
      "archivos_entrada": nombres_guardados,
      "resumen_proceso": resumen,
    }
  )


@facturas_proveedores_bp.get("/api/facturas")
def listar_facturas():
  """
  Devuelve el listado de facturas de una empresa (base maestra en CSV).
  Usa cache en memoria por empresa; se invalida al editar/eliminar/procesar.
  """
  empresa_id = request.args.get("empresa_id")
  empresa_id, err = _validar_empresa_id_requerido(empresa_id)
  if err:
    return jsonify({"facturas": [], "error": "Falta empresa_id"}), 400

  if empresa_id in _cache_listado_facturas_proveedores:
    facturas = _cache_listado_facturas_proveedores[empresa_id]
  else:
    facturas = facturas_db.get_facturas_empresa(empresa_id)
    _cache_listado_facturas_proveedores[empresa_id] = facturas

  proveedor_filtro = (request.args.get("proveedor") or "").strip()
  if proveedor_filtro:
    facturas = [f for f in facturas if (f.get("proveedor") or "").strip() == proveedor_filtro]

  return jsonify({"facturas": facturas, "empresa_id": empresa_id})


@proveedores_bp.get("/api/proveedores")
def listar_proveedores():
  """
  Devuelve el listado de proveedores para la empresa.
  Por defecto: maestro + proveedores únicos que aparecen en facturas (para el desplegable de edición).
  Si solo_con_facturas=1 o true: solo proveedores que tienen al menos una factura (para el listado único, evita ver duplicados sin facturas).
  """
  empresa_id = request.args.get("empresa_id")
  empresa_id, err = _validar_empresa_id_requerido(empresa_id)
  if err:
    return jsonify({"proveedores": [], "error": "Falta empresa_id"}), 400

  solo_con_facturas = request.args.get("solo_con_facturas", "").strip().lower() in ("1", "true", "yes")
  lista = (
    _listar_proveedores_con_facturas(empresa_id)
    if solo_con_facturas
    else _listar_proveedores_para_selector(empresa_id)
  )
  return jsonify({"proveedores": lista, "empresa_id": empresa_id})


@proveedores_bp.get("/api/empresas/<empresa_id>/proveedores")
def listar_proveedores_por_empresa(empresa_id: str):
  """Listado de proveedores de una empresa. Query solo_con_facturas=1 para listado único (solo con facturas)."""
  empresa_id, err = _validar_empresa_id_requerido(empresa_id)
  if err:
    return err[0], err[1]
  solo_con_facturas = request.args.get("solo_con_facturas", "").strip().lower() in ("1", "true", "yes")
  lista = (
    _listar_proveedores_con_facturas(empresa_id)
    if solo_con_facturas
    else _listar_proveedores_para_selector(empresa_id)
  )
  return jsonify({"proveedores": lista, "empresa_id": empresa_id})


@proveedores_bp.get("/api/empresas/<empresa_id>/terceros")
def listar_terceros_por_empresa(empresa_id: str):
  """
  Listado unificado de terceros (proveedores y/o clientes) de una empresa.
  Query: rol=proveedor|cliente|ambos (por defecto ambos).
  Cada ítem incluye "rol" ("proveedor" o "cliente") y campos comunes: nif, nombre_canonico, direccion, localidad, pais, email, telefono; proveedores añaden centro_coste; clientes añaden proyecto.
  """
  empresa_id, err = _validar_empresa_id_requerido(empresa_id)
  if err:
    return err[0], err[1]
  rol = (request.args.get("rol") or "ambos").strip().lower()
  if rol not in ("proveedor", "cliente", "ambos"):
    return _bad_request('rol debe ser "proveedor", "cliente" o "ambos"')

  terceros: list[dict] = []
  if rol in ("proveedor", "ambos"):
    proveedores = _cargar_proveedores_maestros(empresa_id)
    for p in proveedores:
      terceros.append({
        "rol": "proveedor",
        "nif": (p.get("nif") or "").strip(),
        "nombre_canonico": (p.get("nombre_canonico") or "").strip(),
        "direccion": (p.get("direccion") or "").strip(),
        "localidad": (p.get("localidad") or "").strip(),
        "pais": (p.get("pais") or "").strip(),
        "email": (p.get("email") or "").strip(),
        "telefono": (p.get("telefono") or "").strip(),
        "centro_coste": (p.get("centro_coste") or "").strip(),
      })
  if rol in ("cliente", "ambos"):
    if terceros_db.hay_clientes_en_bd():
      clientes_bd = terceros_db.get_clientes_empresa(empresa_id)
    else:
      clientes_bd = []
    clientes_agg = _get_clientes_unicos_empresa(empresa_id)
    seen_cliente: set[tuple[str, str]] = set()
    for c in clientes_bd:
      nombre = (c.get("cliente") or "").strip()
      nif = (c.get("cif_nif") or "").strip()
      key = (nombre, nif)
      if key not in seen_cliente:
        seen_cliente.add(key)
        terceros.append({
          "rol": "cliente",
          "nif": nif,
          "nombre_canonico": nombre,
          "direccion": (c.get("direccion") or "").strip(),
          "localidad": (c.get("localidad") or "").strip(),
          "pais": (c.get("pais") or "").strip(),
          "email": (c.get("email") or "").strip(),
          "telefono": (c.get("telefono") or "").strip(),
          "proyecto": (c.get("proyecto") or "").strip(),
        })
    for c in clientes_agg:
      nombre = (c.get("cliente") or "").strip()
      nif = (c.get("cif_nif") or "").strip()
      key = (nombre, nif)
      if key not in seen_cliente:
        seen_cliente.add(key)
        terceros.append({
          "rol": "cliente",
          "nif": nif,
          "nombre_canonico": nombre,
          "direccion": (c.get("direccion") or "").strip(),
          "localidad": (c.get("localidad") or "").strip(),
          "pais": (c.get("pais") or "").strip(),
          "email": (c.get("email") or "").strip(),
          "telefono": (c.get("telefono") or "").strip(),
          "proyecto": (c.get("proyecto") or "").strip(),
        })

  return jsonify({"terceros": terceros, "empresa_id": empresa_id, "rol": rol})


@bancos_bp.get("/api/empresas/<empresa_id>/tarjetas")
def listar_tarjetas_por_empresa(empresa_id: str):
  """
  Listado de tarjetas de banco de una empresa.
  Query opcional: solo_activas=true|false (por defecto true).
  """
  empresa_id, err = _validar_empresa_id_requerido(empresa_id)
  if err:
    return err[0], err[1]
  solo_activas_raw = (request.args.get("solo_activas") or "true").strip().lower()
  solo_activas = solo_activas_raw not in ("false", "0", "no")
  tarjetas = tarjetas_db.get_tarjetas_empresa(empresa_id, solo_activas=solo_activas)
  return jsonify({"tarjetas": tarjetas, "empresa_id": empresa_id, "solo_activas": solo_activas})


@bancos_bp.post("/api/tarjetas")
def crear_tarjeta():
  """
  Alta de tarjeta de banco en el maestro de la empresa.
  JSON: empresa_id (obligatorio), banco (obligatorio), persona (obligatorio), ultimos4?, alias?, activa?.
  """
  data = request.get_json(silent=True) or {}
  empresa_id, err = _validar_empresa_id_requerido(data.get("empresa_id"))
  if err:
    return err[0], err[1]
  banco = (data.get("banco") or "").strip()
  persona = (data.get("persona") or "").strip()
  if not banco:
    return _bad_request("El banco de la tarjeta es obligatorio")
  if not persona:
    return _bad_request("La persona titular de la tarjeta es obligatoria")
  try:
    tarjeta = tarjetas_db.crear_tarjeta(empresa_id, data)
  except Exception as e:
    return jsonify({"error": str(e)}), 500
  return jsonify({"ok": True, "tarjeta": tarjeta, "empresa_id": empresa_id}), 201


@bancos_bp.put("/api/tarjetas/<int:tarjeta_id>")
def actualizar_tarjeta(tarjeta_id: int):
  """
  Edición de una tarjeta de banco.
  JSON: empresa_id (obligatorio) y campos a actualizar: banco?, persona?, ultimos4?, alias?, activa?.
  No permite cambiar empresa_id.
  """
  data = request.get_json(silent=True) or {}
  empresa_id, err = _validar_empresa_id_requerido(data.get("empresa_id"))
  if err:
    return err[0], err[1]
  try:
    tarjeta = tarjetas_db.actualizar_tarjeta(tarjeta_id, empresa_id, data)
  except ValueError as e:
    return _bad_request(str(e))
  except Exception as e:
    return jsonify({"error": str(e)}), 500
  if not tarjeta:
    return jsonify({"error": "Tarjeta no encontrada"}), 404
  return jsonify({"ok": True, "tarjeta": tarjeta, "empresa_id": empresa_id}), 200


@bancos_bp.get("/api/empresas/<empresa_id>/tarjetas/liquidaciones-resumen")
def resumen_liquidaciones_tarjetas(empresa_id: str):
  """
  Resumen de \"liquidaciones\" por tarjeta y periodo (YYYY-MM) calculado a partir de las facturas
  de proveedores que tienen tarjeta_id asignada. Incluye total_movimiento (suma del importe de
  movimientos bancarios vinculados a ese tarjeta_id+periodo) para ver pendiente de incorporar.

  Periodo de la factura: liquidacion_periodo si está rellenado, si no el mes de fecha_factura.
  """
  empresa_id, err = _validar_empresa_id_requerido(empresa_id)
  if err:
    return err[0], err[1]

  facturas_db.init_facturas_db()
  conn = sqlite3.connect(str(GESTION_DB))
  conn.row_factory = sqlite3.Row
  try:
    cur = conn.execute(
      """
      SELECT
        t.id AS tarjeta_id,
        t.banco,
        t.persona,
        t.alias,
        sub.periodo,
        sub.num_facturas,
        sub.total_facturas
      FROM (
        SELECT
          tarjeta_id,
          COALESCE(
            NULLIF(TRIM(liquidacion_periodo), ''),
            substr(COALESCE(NULLIF(fecha_factura, ''), '0000-00-00'), 1, 7)
          ) AS periodo,
          COUNT(*) AS num_facturas,
          SUM(
            CAST(
              REPLACE(
                COALESCE(
                  NULLIF(total_a_pagar, ''),
                  NULLIF(total_factura, ''),
                  '0'
                ),
                ',', '.'
              ) AS REAL
            )
          ) AS total_facturas
        FROM facturas_proveedor
        WHERE empresa_id = ?
          AND tarjeta_id IS NOT NULL
          AND TRIM(CAST(tarjeta_id AS TEXT)) <> ''
        GROUP BY tarjeta_id, periodo
      ) AS sub
      JOIN tarjetas t
        ON t.id = sub.tarjeta_id
       AND t.empresa_id = ?
      ORDER BY sub.periodo DESC, t.banco, t.persona
      """,
      (empresa_id, empresa_id),
    )
    filas = []
    for row in cur.fetchall():
      total_facturas = row["total_facturas"] or 0.0
      try:
        total_facturas = float(total_facturas)
      except Exception as e:
        logger.debug("No se pudo convertir total_facturas a float: %s", e)
        total_facturas = 0.0
      porcentaje = 100.0 if total_facturas > 0 else 0.0
      filas.append(
        {
          "tarjeta_id": row["tarjeta_id"],
          "tarjeta_banco": row["banco"],
          "tarjeta_persona": row["persona"],
          "tarjeta_alias": row["alias"],
          "periodo": row["periodo"],
          "num_facturas": row["num_facturas"],
          "total_facturas": total_facturas,
          "estado": "pendiente",
          "porcentaje_facturas": porcentaje,
          "total_movimiento": None,
        }
      )
  finally:
    conn.close()

  # Enriquecer con importe de movimientos vinculados (tarjeta_id + liquidacion_periodo) en movimientos.db
  _init_movimientos_db()
  conn_bancos = _get_bancos_db()
  conn_bancos.row_factory = sqlite3.Row
  try:
    # Build a set of (tarjeta_id, periodo) already present from facturas
    existing_keys = {(f["tarjeta_id"], f.get("periodo") or "") for f in filas}

    # Find movimientos linked to tarjetas of this empresa that have no matching facturas row
    mov_rows = conn_bancos.execute(
      """
      SELECT tarjeta_id, liquidacion_periodo,
             COALESCE(SUM(CAST(importe AS REAL)), 0) AS total
      FROM movimientos
      WHERE tarjeta_id IS NOT NULL AND tarjeta_id != 0
        AND liquidacion_periodo IS NOT NULL AND TRIM(liquidacion_periodo) != ''
      GROUP BY tarjeta_id, liquidacion_periodo
      """,
    ).fetchall()

    # Fetch tarjetas of this empresa for resolving names
    tarjetas_map = {}
    conn_gest2 = sqlite3.connect(str(GESTION_DB))
    conn_gest2.row_factory = sqlite3.Row
    try:
      for t in conn_gest2.execute(
        "SELECT id, banco, persona, alias FROM tarjetas WHERE empresa_id = ?",
        (empresa_id,),
      ):
        tarjetas_map[t["id"]] = dict(t)
    finally:
      conn_gest2.close()

    # Add missing (tarjeta, periodo) combinations from movimientos
    for mr in mov_rows:
      tid = mr["tarjeta_id"]
      per = mr["liquidacion_periodo"]
      if tid not in tarjetas_map:
        continue  # tarjeta belongs to another empresa
      if (tid, per) not in existing_keys:
        t_info = tarjetas_map[tid]
        filas.append(
          {
            "tarjeta_id": tid,
            "tarjeta_banco": t_info.get("banco"),
            "tarjeta_persona": t_info.get("persona"),
            "tarjeta_alias": t_info.get("alias"),
            "periodo": per,
            "num_facturas": 0,
            "total_facturas": 0.0,
            "estado": "pendiente",
            "porcentaje_facturas": 0.0,
            "total_movimiento": None,
          }
        )

    for fila in filas:
      tid = fila["tarjeta_id"]
      per = fila.get("periodo") or ""
      if not per or len(per) != 7:
        continue
      row_sum = conn_bancos.execute(
        """
        SELECT COALESCE(SUM(CAST(importe AS REAL)), 0) AS total
        FROM movimientos
        WHERE tarjeta_id = ? AND liquidacion_periodo = ?
        """,
        (tid, per),
      ).fetchone()
      fila["total_movimiento"] = float(row_sum[0]) if row_sum else 0.0
    # Pendiente = total_facturas + total_movimiento (movimientos negativos; suma = saldo pendiente)
    # Estado: pendiente si no hay movimiento; conciliado si hay movimiento y pendiente ≈ 0; cargo recibido si hay movimiento pero no cuadra
    for fila in filas:
      total_fac = fila.get("total_facturas") or 0.0
      total_mov = fila.get("total_movimiento") or 0.0
      fila["pendiente_facturas"] = round(total_fac + total_mov, 2)
      if total_mov == 0:
        fila["estado"] = "pendiente"
      elif abs(fila["pendiente_facturas"]) < 0.01:
        fila["estado"] = "conciliado"
      else:
        fila["estado"] = "cargo recibido"

    # Sort: most recent period first, then by banco/persona
    filas.sort(key=lambda f: (f.get("periodo") or "", f.get("tarjeta_banco") or "", f.get("tarjeta_persona") or ""), reverse=True)
  finally:
    conn_bancos.close()

  return jsonify({"empresa_id": empresa_id, "liquidaciones": filas})


@bancos_bp.get("/api/empresas/<empresa_id>/tarjetas/extracto-export")
def exportar_extracto_tarjeta(empresa_id: str):
  """
  Exporta la conciliación de un extracto (tarjeta + periodo) a Excel.
  Query: tarjeta_id (int), periodo (YYYY-MM), tipo=excel|facturas.
  - tipo=excel: Excel con hojas "Facturas" y "Movimientos" (conciliación completa).
  - tipo=facturas: Excel solo con hoja "Facturas".
  """
  empresa_id, err = _validar_empresa_id_requerido(empresa_id)
  if err:
    return err[0], err[1]
  try:
    tarjeta_id = int(request.args.get("tarjeta_id", 0))
  except (TypeError, ValueError):
    return _bad_request("tarjeta_id debe ser un entero")
  periodo = (request.args.get("periodo") or "").strip()
  if not periodo or len(periodo) != 7 or periodo[4] != "-":
    return _bad_request("periodo debe tener formato YYYY-MM")
  tipo = (request.args.get("tipo") or "excel").strip().lower()
  if tipo not in ("excel", "facturas"):
    tipo = "excel"

  facturas_db.init_facturas_db()
  conn_gest = sqlite3.connect(str(GESTION_DB))
  conn_gest.row_factory = sqlite3.Row
  try:
    # Facturas del extracto: tarjeta_id y periodo (liquidacion_periodo o mes de fecha_factura)
    cur_fac = conn_gest.execute(
      """
      SELECT id, empresa_id, fecha_factura, proveedor, nif_proveedor, pais_proveedor, localidad_proveedor,
             resumen_concepto, numero_factura, base_imponible, iva, retenciones_total, total_factura,
             total_a_pagar, categoria, ruta_archivo, ruta_destino, estado_pago, tarjeta_id, liquidacion_periodo
      FROM facturas_proveedor
      WHERE empresa_id = ? AND tarjeta_id = ?
        AND (
          (TRIM(COALESCE(liquidacion_periodo, '')) = ?)
          OR (TRIM(COALESCE(liquidacion_periodo, '')) = '' AND substr(COALESCE(NULLIF(fecha_factura, ''), '0000-00-00'), 1, 7) = ?)
        )
      ORDER BY fecha_factura, id
      """,
      (empresa_id, tarjeta_id, periodo, periodo),
    )
    filas_fac = [dict(row) for row in cur_fac.fetchall()]
  finally:
    conn_gest.close()

  movimientos_rows = []
  if tipo == "excel":
    _init_movimientos_db()
    conn_bancos = _get_bancos_db()
    try:
      cur_mov = conn_bancos.execute(
        """
        SELECT id, fecha_operacion, fecha_valor, concepto, importe, divisa, saldo,
               banco, codigo, numero_documento, referencia_1, referencia_2, empresa_id,
               tarjeta_id, liquidacion_periodo, created_at
        FROM movimientos
        WHERE tarjeta_id = ? AND liquidacion_periodo = ?
        ORDER BY fecha_operacion, id
        """,
        (tarjeta_id, periodo),
      )
      movimientos_rows = list(cur_mov.fetchall())
    finally:
      conn_bancos.close()

  try:
    import openpyxl
    from openpyxl import Workbook
  except ImportError:
    return jsonify({"error": "openpyxl no instalado. pip install openpyxl"}), 500

  wb = Workbook()
  ws_fac = wb.active
  ws_fac.title = "Facturas"
  cols_fac = [
    "id", "fecha_factura", "proveedor", "nif_proveedor", "numero_factura", "resumen_concepto",
    "base_imponible", "iva", "retenciones_total", "total_factura", "total_a_pagar", "categoria", "estado_pago",
  ]
  ws_fac.append(cols_fac)
  for f in filas_fac:
    ws_fac.append([f.get(c) or "" for c in cols_fac])

  if tipo == "excel" and movimientos_rows:
    ws_mov = wb.create_sheet("Movimientos")
    cols_mov = [
      "id", "fecha_operacion", "fecha_valor", "concepto", "importe", "divisa", "saldo",
      "banco", "codigo", "numero_documento", "referencia_1", "referencia_2", "empresa_id",
      "tarjeta_id", "liquidacion_periodo", "created_at",
    ]
    ws_mov.append(cols_mov)
    for r in movimientos_rows:
      ws_mov.append(list(r))

  output = io.BytesIO()
  wb.save(output)
  output.seek(0)
  safe_per = periodo.replace("/", "-")
  nombre = f"extracto_tarjeta_{tarjeta_id}_{safe_per}_{tipo}.xlsx"
  return send_file(
    output,
    mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    as_attachment=True,
    download_name=nombre,
  )


@bancos_bp.post("/api/bancos/tarjetas/conciliar-movimiento")
def conciliar_movimiento_con_liquidacion_tarjeta():
  """
  Vincula un movimiento bancario (cargo de tarjeta) con una \"liquidación\" (tarjeta_id + periodo).

  Body JSON:
    {
      "empresa_id": "...",
      "movimiento_id": 123,
      "tarjeta_id": 5,
      "periodo": "2025-02"
    }
  """
  data = request.get_json(silent=True) or {}
  empresa_id_body, err = _validar_empresa_id_requerido(data.get("empresa_id"))
  if err:
    return err[0], err[1]

  try:
    mov_id = int(data.get("movimiento_id"))
    tarjeta_id = int(data.get("tarjeta_id"))
  except (TypeError, ValueError):
    return _bad_request("movimiento_id y tarjeta_id deben ser enteros válidos")
  periodo = (data.get("periodo") or "").strip()
  if not periodo or len(periodo) != 7 or periodo[4] != "-":
    return _bad_request("El periodo debe tener formato YYYY-MM")

  # Validar tarjeta pertenece a empresa
  facturas_db.init_facturas_db()
  conn_gest = sqlite3.connect(str(GESTION_DB))
  try:
    row_tar = conn_gest.execute(
      "SELECT id, empresa_id FROM tarjetas WHERE id = ?",
      (tarjeta_id,),
    ).fetchone()
    if not row_tar or (row_tar[1] or "").strip() != empresa_id_body:
      return jsonify({"error": "Tarjeta no encontrada o no pertenece a la empresa indicada"}), 404
  finally:
    conn_gest.close()

  # Actualizar movimiento en BD de bancos
  _init_movimientos_db()
  conn_bancos = _get_bancos_db()
  try:
    row_mov = conn_bancos.execute(
      "SELECT id, empresa_id, importe FROM movimientos WHERE id = ?",
      (mov_id,),
    ).fetchone()
    if not row_mov:
      return jsonify({"error": "Movimiento no encontrado"}), 404
    empresa_mov = (row_mov[1] or "").strip()
    if empresa_mov and empresa_mov != empresa_id_body:
      return jsonify({"error": "El movimiento pertenece a otra empresa"}), 400
    conn_bancos.execute(
      """
      UPDATE movimientos
         SET tarjeta_id = ?, liquidacion_periodo = ?
       WHERE id = ?
      """,
      (tarjeta_id, periodo, mov_id),
    )
    conn_bancos.commit()
  finally:
    conn_bancos.close()

  return jsonify(
    {
      "ok": True,
      "mensaje": "Movimiento vinculado a la liquidación de tarjeta.",
      "movimiento_id": mov_id,
      "tarjeta_id": tarjeta_id,
      "periodo": periodo,
      "empresa_id": empresa_id_body,
    }
  )


@bancos_bp.post("/api/bancos/tarjetas/desvincular-movimiento")
def desvincular_movimiento_de_liquidacion_tarjeta():
  """
  Quita la vinculación de un movimiento con la liquidación de tarjeta (tarjeta_id y liquidacion_periodo a NULL).
  Body JSON: { "empresa_id": "...", "movimiento_id": 123 }
  """
  data = request.get_json(silent=True) or {}
  empresa_id_body, err = _validar_empresa_id_requerido(data.get("empresa_id"))
  if err:
    return err[0], err[1]
  try:
    mov_id = int(data.get("movimiento_id"))
  except (TypeError, ValueError):
    return _bad_request("movimiento_id debe ser un entero válido")
  _init_movimientos_db()
  conn_bancos = _get_bancos_db()
  try:
    row = conn_bancos.execute(
      "SELECT id, empresa_id FROM movimientos WHERE id = ?",
      (mov_id,),
    ).fetchone()
    if not row:
      return jsonify({"error": "Movimiento no encontrado"}), 404
    empresa_mov = (row[1] or "").strip()
    if empresa_mov and empresa_mov != empresa_id_body:
      return jsonify({"error": "El movimiento pertenece a otra empresa"}), 400
    conn_bancos.execute(
      "UPDATE movimientos SET tarjeta_id = NULL, liquidacion_periodo = NULL WHERE id = ?",
      (mov_id,),
    )
    conn_bancos.commit()
  finally:
    conn_bancos.close()
  return jsonify({"ok": True, "mensaje": "Movimiento desvinculado del extracto de tarjeta.", "movimiento_id": mov_id})


# Frases que identifican movimientos bancarios de liquidación/recibo de tarjeta.
# Orden: frases largas primero para match más específico, genéricas al final.
_KEYWORDS_TARJETA = [
  "recibo mensual tarjeta", "adeudo mensual de tarjeta", "adeudo mensual tarjeta",
  "recibo tarjeta", "liquidacion tarjeta", "liquidacion de las tarjetas",
  "liquidacion contrato", "pago tarjeta", "cargo tarjeta",
]


def _es_movimiento_tarjeta(concepto: str) -> bool:
  """Devuelve True si el concepto del movimiento bancario indica un cargo/liquidación de tarjeta."""
  c = (concepto or "").lower()
  return any(kw in c for kw in _KEYWORDS_TARJETA)


def _extraer_ultimos4_de_concepto(concepto: str) -> str | None:
  """Intenta extraer los últimos 4 dígitos de tarjeta del concepto (ej: '5478240009522305' → '2305')."""
  import re
  # Buscar secuencias de 16 dígitos (número de tarjeta completo)
  m = re.search(r"\b(\d{16})\b", concepto or "")
  if m:
    return m.group(1)[-4:]
  # Buscar secuencias de 4 dígitos precedidas de * o espacio (ej: *2305, XXXX2305)
  m = re.search(r"[*xX]+(\d{4})\b", concepto or "")
  if m:
    return m.group(1)
  return None


@bancos_bp.get("/api/bancos/tarjetas/sugerencias-conciliacion")
def sugerencias_conciliacion_tarjeta():
  """
  Devuelve sugerencias de conciliación: movimientos de tipo tarjeta sin conciliar vs tarjetas + periodos.

  Detecta movimientos cuyo concepto contiene palabras clave de tarjeta (recibo, liquidación, etc.),
  intenta extraer los últimos 4 dígitos del número de tarjeta del concepto, y sugiere la tarjeta
  y periodo (YYYY-MM basado en fecha del movimiento) correspondientes.

  Query: empresa_id (obligatorio).
  """
  empresa_id = (request.args.get("empresa_id") or "").strip()
  if not empresa_id:
    return _bad_request("Falta empresa_id")

  _init_movimientos_db()
  facturas_db.init_facturas_db()

  # Obtener tarjetas de la empresa
  conn_gest = sqlite3.connect(str(GESTION_DB))
  conn_gest.row_factory = sqlite3.Row
  try:
    tarjetas = conn_gest.execute(
      "SELECT id, banco, persona, ultimos4, alias FROM tarjetas WHERE empresa_id = ? AND activa = 1",
      (empresa_id,),
    ).fetchall()
    tarjetas_list = [dict(t) for t in tarjetas]
  finally:
    conn_gest.close()

  if not tarjetas_list:
    return jsonify({"sugerencias": [], "mensaje": "No hay tarjetas activas para esta empresa"})

  # Indexar tarjetas por últimos 4 dígitos
  tarjetas_por_u4: dict[str, list[dict]] = {}
  for t in tarjetas_list:
    u4 = (t.get("ultimos4") or "").strip()
    if u4:
      tarjetas_por_u4.setdefault(u4, []).append(t)

  # Obtener movimientos sin conciliar con tarjeta (tarjeta_id IS NULL, importe negativo)
  conn_bancos = _get_bancos_db()
  try:
    movimientos = conn_bancos.execute(
      """
      SELECT id, fecha_operacion, concepto, importe
      FROM movimientos
      WHERE empresa_id = ?
        AND (tarjeta_id IS NULL OR tarjeta_id = 0)
        AND (factura_proveedor_id IS NULL)
        AND (conciliado_at IS NULL OR conciliado_at = '')
        AND importe < 0
      ORDER BY fecha_operacion DESC
      LIMIT 500
      """,
      (empresa_id,),
    ).fetchall()
  finally:
    conn_bancos.close()

  sugerencias = []
  for m in movimientos:
    mov_id, fecha_op, concepto, importe = m[0], m[1], m[2], m[3]
    concepto_str = (concepto or "").strip()

    if not _es_movimiento_tarjeta(concepto_str):
      continue

    # Extraer últimos 4 dígitos del concepto
    u4 = _extraer_ultimos4_de_concepto(concepto_str)

    # Calcular periodo sugerido (mes del movimiento)
    periodo_sugerido = (fecha_op or "")[:7] if fecha_op and len(fecha_op) >= 7 else None

    # Buscar tarjeta candidata por últimos 4 dígitos
    tarjeta_match = None
    confianza = "baja"
    if u4 and u4 in tarjetas_por_u4:
      candidatas = tarjetas_por_u4[u4]
      if len(candidatas) == 1:
        tarjeta_match = candidatas[0]
        confianza = "alta"
      else:
        tarjeta_match = candidatas[0]
        confianza = "media"
    elif len(tarjetas_list) == 1:
      # Solo hay una tarjeta, sugerir esa
      tarjeta_match = tarjetas_list[0]
      confianza = "media"

    sugerencias.append({
      "movimiento_id": mov_id,
      "movimiento_fecha": fecha_op,
      "movimiento_concepto": concepto_str,
      "movimiento_importe": importe,
      "ultimos4_detectados": u4,
      "tarjeta_sugerida": {
        "tarjeta_id": tarjeta_match["id"],
        "banco": tarjeta_match.get("banco"),
        "persona": tarjeta_match.get("persona"),
        "ultimos4": tarjeta_match.get("ultimos4"),
        "alias": tarjeta_match.get("alias"),
      } if tarjeta_match else None,
      "periodo_sugerido": periodo_sugerido,
      "confianza": confianza,
    })

  return jsonify({
    "sugerencias": sugerencias,
    "total": len(sugerencias),
    "tarjetas_disponibles": [
      {"id": t["id"], "banco": t.get("banco"), "persona": t.get("persona"),
       "ultimos4": t.get("ultimos4"), "alias": t.get("alias")}
      for t in tarjetas_list
    ],
  })


def _normalizar_proveedor_body(data: dict) -> dict:
  """Extrae y normaliza campos de proveedor desde el body (CAMPOS_PROVEEDORES_MAESTROS)."""
  return {
    "nombre_canonico": (data.get("nombre_canonico") or "").strip(),
    "nif": (data.get("nif") or "").strip(),
    "direccion": (data.get("direccion") or "").strip(),
    "localidad": (data.get("localidad") or "").strip(),
    "pais": (data.get("pais") or "").strip(),
    "email": (data.get("email") or "").strip(),
    "telefono": (data.get("telefono") or "").strip(),
    "centro_coste": (data.get("centro_coste") or "").strip(),
  }


@proveedores_bp.post("/api/proveedores")
def crear_proveedor():
  """
  Alta de proveedor en el maestro de la empresa.
  JSON: empresa_id (obligatorio), nombre_canonico, nif (mínimos), direccion, localidad, pais, email, telefono, centro_coste.
  Usa el gateway unificado crear_o_vincular_tercero para deduplicacion.
  """
  from core.terceros_db import crear_o_vincular_tercero
  data = request.get_json(silent=True) or {}
  empresa_id, err = _validar_empresa_id_requerido(data.get("empresa_id"))
  if err:
    return err[0], err[1]
  p = _normalizar_proveedor_body(data)
  nombre = p["nombre_canonico"]
  nif = p["nif"]
  if not nombre:
    return _bad_request("El nombre del proveedor es obligatorio")
  if not nif:
    return _bad_request("El NIF/CIF del proveedor es obligatorio")

  # Verificar via gateway unificado si ya existe
  resultado = crear_o_vincular_tercero(
    nombre, nif,
    datos_extra={k: p.get(k) for k in ("pais", "localidad", "direccion", "email", "telefono")},
    rol="proveedor", origen="manual",
  )
  if resultado["accion"] != "creado":
    warning = None
    if resultado["accion"] == "vinculado_similar" and resultado.get("requiere_revision"):
      warning = f"Vinculado a tercero similar: {resultado['nombre_match']} (similitud {resultado['similitud']:.0%})"
    return jsonify({
      "error": f"Ya existe un proveedor equivalente: {resultado['nombre_match']} (#{resultado['id']})",
      "tercero_existente": resultado,
      "warning": warning,
    }), 409

  lista = _cargar_proveedores_maestros(empresa_id)
  lista.append(p)
  _guardar_proveedores_maestros(empresa_id, lista)
  return jsonify({"ok": True, "proveedores": lista, "empresa_id": empresa_id, "tercero_id": resultado["id"]}), 201


@proveedores_bp.put("/api/proveedores")
def actualizar_proveedor():
  """
  Edición de un proveedor del maestro. Se identifica por old_nombre_canonico y old_nif.
  JSON: empresa_id, old_nombre_canonico, old_nif, y el resto de campos (nombre_canonico, nif, direccion, ...).
  """
  data = request.get_json(silent=True) or {}
  empresa_id, err = _validar_empresa_id_requerido(data.get("empresa_id"))
  if err:
    return err[0], err[1]
  old_nombre = (data.get("old_nombre_canonico") or "").strip()
  old_nif = (data.get("old_nif") or "").strip()
  if not old_nombre and not old_nif:
    return _bad_request("Faltan old_nombre_canonico y old_nif para identificar el proveedor")
  p = _normalizar_proveedor_body(data)
  if not p["nombre_canonico"]:
    return _bad_request("El nombre del proveedor es obligatorio")
  if not p["nif"]:
    return _bad_request("El NIF/CIF del proveedor es obligatorio")

  lista = _cargar_proveedores_maestros(empresa_id)
  idx = -1
  for i, existente in enumerate(lista):
    if (
      (existente.get("nombre_canonico") or "").strip() == old_nombre
      and (existente.get("nif") or "").strip() == old_nif
    ):
      idx = i
      break
  if idx < 0:
    lista.append(p)
    _guardar_proveedores_maestros(empresa_id, lista)
    facturas_db.update_facturas_datos_proveedor(
      empresa_id,
      old_nombre,
      old_nif,
      p["nombre_canonico"],
      p["nif"],
      p.get("pais") or None,
      p.get("localidad") or None,
    )
    _invalidar_cache_listado_proveedores(empresa_id)
    return jsonify({
      "ok": True,
      "proveedores": lista,
      "empresa_id": empresa_id,
      "mensaje": "Proveedor añadido al maestro con los datos indicados (no estaba en el maestro).",
    })
  lista[idx] = p
  _guardar_proveedores_maestros(empresa_id, lista)
  facturas_db.update_facturas_datos_proveedor(
    empresa_id,
    old_nombre,
    old_nif,
    p["nombre_canonico"],
    p["nif"],
    p.get("pais") or None,
    p.get("localidad") or None,
  )
  _invalidar_cache_listado_proveedores(empresa_id)
  return jsonify({"ok": True, "proveedores": lista, "empresa_id": empresa_id})


@proveedores_bp.delete("/api/proveedores")
def eliminar_proveedor():
  """
  Elimina un proveedor del maestro. Solo si está en el maestro.
  Body: { "empresa_id": "...", "nombre_canonico": "...", "nif": "..." }.
  """
  data = request.get_json(silent=True) or {}
  empresa_id, err = _validar_empresa_id_requerido(data.get("empresa_id"))
  if err:
    return err[0], err[1]
  nombre = (data.get("nombre_canonico") or "").strip()
  nif = (data.get("nif") or "").strip()
  if not nombre and not nif:
    return _bad_request("Faltan nombre_canonico y nif para identificar el proveedor")
  lista = _cargar_proveedores_maestros(empresa_id)
  nueva_lista = [
    p for p in lista
    if not (
      (p.get("nombre_canonico") or "").strip() == nombre
      and (p.get("nif") or "").strip() == nif
    )
  ]
  if len(nueva_lista) == len(lista):
    return jsonify({"error": "Proveedor no encontrado en el maestro"}), 404
  _guardar_proveedores_maestros(empresa_id, nueva_lista)
  _invalidar_cache_listado_proveedores(empresa_id)
  return jsonify({"ok": True, "proveedores": nueva_lista, "empresa_id": empresa_id, "mensaje": "Proveedor eliminado del maestro."})


@proveedores_bp.post("/api/proveedores/sincronizar-facturas")
def sincronizar_facturas_con_proveedores():
  """
  Actualiza las facturas existentes con los datos del maestro de proveedores.
  Para cada factura se busca el proveedor en el maestro (por NIF o por similitud de nombre)
  y se actualizan proveedor, nif_proveedor, pais_proveedor y localidad_proveedor.
  Body opcional: { "empresa_id": "..." } para sincronizar solo una empresa; si se omite, se procesan todas.
  """
  data = request.get_json(silent=True) or {}
  empresa_filtro = (data.get("empresa_id") or "").strip()
  empresas = (
    [empresa_filtro]
    if empresa_filtro
    else list(EMPRESAS_CLIENTE.keys())
  )
  if empresa_filtro and empresa_filtro not in EMPRESAS_CLIENTE:
    return _bad_request("empresa_id no válida")

  total_actualizadas = 0
  detalle = {}
  for empresa_id in empresas:
    proveedores = _cargar_proveedores_maestros(empresa_id)
    if not proveedores:
      detalle[empresa_id] = 0
      continue
    try:
      facturas = facturas_db.get_facturas_empresa(empresa_id)
    except Exception as e:
      logger.warning("Error leyendo facturas para sincronizar proveedores %s: %s", empresa_id, e)
      facturas = []
    count_empresa = 0
    vistos_factura: set[tuple[str, str]] = set()
    for f in facturas:
      prov = (f.get("proveedor") or "").strip()
      nif = (f.get("nif_proveedor") or "").strip()
      if not prov and not nif:
        continue
      key_f = (_normalizar_texto_proveedor(prov), _normalizar_nif(nif))
      if key_f in vistos_factura:
        continue
      vistos_factura.add(key_f)
      match = None
      nif_norm = _normalizar_nif(nif)
      if nif_norm:
        for p in proveedores:
          if _normalizar_nif(p.get("nif") or "") == nif_norm:
            match = p
            break
      if not match and prov:
        mejor_ratio = 0.0
        for p in proveedores:
          nombre_m = (p.get("nombre_canonico") or "").strip()
          if not nombre_m:
            continue
          r = _similitud_nombres(prov, nombre_m)
          if r >= 0.82 and r > mejor_ratio:
            mejor_ratio = r
            match = p
      if not match:
        continue
      nombre_new = (match.get("nombre_canonico") or "").strip()
      nif_new = (match.get("nif") or "").strip()
      pais_new = (match.get("pais") or "").strip()
      localidad_new = (match.get("localidad") or "").strip()
      if (
        nombre_new == prov
        and nif_new == nif
        and (not pais_new or (f.get("pais_proveedor") or "").strip() == pais_new)
        and (not localidad_new or (f.get("localidad_proveedor") or "").strip() == localidad_new)
      ):
        continue
      n = facturas_db.update_facturas_datos_proveedor(
        empresa_id, prov, nif, nombre_new, nif_new, pais_new or None, localidad_new or None,
      )
      count_empresa += n
      total_actualizadas += n
    detalle[empresa_id] = count_empresa
    if count_empresa:
      _invalidar_cache_listado_proveedores(empresa_id)

  return jsonify({
    "ok": True,
    "mensaje": f"Facturas actualizadas con datos del maestro: {total_actualizadas} fila(s).",
    "facturas_actualizadas": total_actualizadas,
    "empresas_procesadas": len(empresas),
    "detalle": detalle,
  })


@proveedores_bp.post("/api/terceros/migrar-desde-csv")
def migrar_terceros_desde_csv():
  """
  Migra los proveedores de todos los proveedores_maestros.csv a SQLite (terceros + empresa_tercero).
  Ejecutar una vez para pasar a usar la BD. Devuelve estadísticas y posibles errores.
  """
  try:
    resultado = terceros_db.migrar_proveedores_desde_csv()
    return jsonify({
      "ok": True,
      "mensaje": "Migración completada.",
      "empresas_procesadas": resultado["empresas_procesadas"],
      "terceros_totales": resultado["terceros_totales"],
      "relaciones_creadas": resultado["relaciones_creadas"],
      "errores": resultado.get("errores", []),
    })
  except Exception as e:
    return jsonify({"ok": False, "error": str(e)}), 500


@proveedores_bp.put("/api/proveedor_ceco")
def actualizar_centro_coste_proveedor():
  """
  Actualiza el centro de coste asociado a un proveedor en el listado maestro.
  JSON: { "empresa_id": "...", "proveedor": "nombre_canonico", "centro_coste": "..." }
  """
  data = request.get_json(silent=True) or {}
  empresa_id, err = _validar_empresa_id_requerido(data.get("empresa_id"))
  if err:
    return err[0], err[1]
  proveedor = (data.get("proveedor") or "").strip()
  centro_coste = (data.get("centro_coste") or "").strip()
  if not proveedor:
    return _bad_request("Faltan empresa_id o proveedor")

  lista = _cargar_proveedores_maestros(empresa_id)
  if not lista:
    return jsonify({"error": "No hay proveedores maestros para esta empresa"}), 404

  actualizado = False
  for p in lista:
    if (p.get("nombre_canonico") or "").strip() == proveedor:
      p["centro_coste"] = centro_coste
      actualizado = True
      break

  if not actualizado:
    return jsonify({"error": "Proveedor no encontrado en el maestro"}), 404

  _guardar_proveedores_maestros(empresa_id, lista)
  return jsonify({"ok": True, "mensaje": "Centro de coste actualizado"})


@facturas_proveedores_bp.get("/api/facturas_export")
def exportar_facturas():
  """
  Exporta las facturas de una empresa (y filtros opcionales de año/mes) en CSV con extensión .xlsx.
  Usa la misma fuente que el listado (BD primero, luego CSV si existe).
  """
  empresa_id = request.args.get("empresa_id")
  empresa_id, err = _validar_empresa_id_requerido(empresa_id)
  if err:
    return err[0], err[1]
  year = (request.args.get("year") or "").strip()
  month = (request.args.get("month") or "").strip()
  proveedor_filtro = (request.args.get("proveedor") or "").strip()
  estado_pago_filtro = (request.args.get("estado_pago") or "").strip()
  tarjeta_id_filtro = (request.args.get("tarjeta_id") or "").strip()

  CAMPOS_EXPORT = [
    "fecha_factura",
    "proveedor",
    "nif_proveedor",
    "pais_proveedor",
    "localidad_proveedor",
    "resumen_concepto",
    "numero_factura",
    "base_imponible",
    "iva",
    "retenciones_total",
    "total_a_pagar",
    "estado_pago",
    "tarjeta_asociada",
    "Nombre del archivo",
  ]

  ids_filtro = (request.args.get("ids") or "").strip()
  ids_set: set[int] | None = None
  if ids_filtro:
    try:
      ids_set = {int(x) for x in ids_filtro.split(",") if x.strip()}
    except ValueError:
      pass

  filas_export: list[dict] = []
  try:
    facturas_bd = facturas_db.get_facturas_empresa(empresa_id)
    if ids_set:
      facturas_bd = [f for f in facturas_bd if f.get("id") in ids_set]
    facturas_filtradas = _filtrar_facturas_en_memoria(facturas_bd, year, month, proveedor_filtro)
    facturas_filtradas = _aplicar_filtros_estado_tarjeta(facturas_filtradas, estado_pago_filtro, tarjeta_id_filtro)
    tarjeta_map: dict[int, str] = {}
    try:
      tarjetas = tarjetas_db.get_tarjetas_empresa(empresa_id, solo_activas=False)
      for t in tarjetas:
        tid = t.get("id")
        if tid is not None:
          alias = (t.get("alias") or "").strip()
          if not alias:
            alias = ((t.get("banco") or "").strip() + " " + (t.get("persona") or "").strip()).strip() or "Tarjeta"
          tarjeta_map[int(tid)] = alias
    except Exception as e:
      logger.warning("Error cargando tarjetas para export: %s", e)
    for f in facturas_filtradas:
      row = {k: str(f.get(k) or "").strip() for k in CAMPOS_EXPORT}
      tid = f.get("tarjeta_id")
      if tid is not None and str(tid).strip():
        try:
          row["tarjeta_asociada"] = tarjeta_map.get(int(tid), "") or ""
        except (TypeError, ValueError):
          row["tarjeta_asociada"] = ""
      else:
        row["tarjeta_asociada"] = ""
      if not row.get("estado_pago"):
        row["estado_pago"] = (f.get("estado_pago") or "pendiente").strip() or "pendiente"
      ruta = (f.get("ruta_destino") or f.get("ruta_archivo") or "").strip()
      row["Nombre del archivo"] = Path(ruta).name if ruta else ""
      filas_export.append(row)
  except Exception as e:
    logger.warning("Error preparando facturas para export: %s", e)
  if not filas_export:
    ruta_csv = EMPRESAS_DIR / empresa_id / "base_maestra_facturas.csv"
    if ruta_csv.exists():
      campos_para_csv = [
        "fecha_factura", "proveedor", "nif_proveedor", "pais_proveedor", "localidad_proveedor",
        "resumen_concepto", "numero_factura", "base_imponible", "iva", "retenciones_total",
        "total_a_pagar", "estado_pago", "tarjeta_asociada", "ruta_archivo", "ruta_destino",
      ]
      with ruta_csv.open("r", newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        campos_csv = list(campos_para_csv) + (["tarjeta_id"] if tarjeta_id_filtro else [])
        filas_export = _filtrar_filas_csv(
          reader,
          campos_csv,
          year=year,
          month=month,
          campo_filtro="proveedor",
          valor_filtro=proveedor_filtro,
          skip_header_por_empresa=False,
        )
        filas_export = _aplicar_filtros_estado_tarjeta(filas_export, estado_pago_filtro, tarjeta_id_filtro)
      for row in filas_export:
        ruta = (row.get("ruta_destino") or row.get("ruta_archivo") or "").strip()
        row["Nombre del archivo"] = Path(ruta).name if ruta else ""
        row.pop("ruta_archivo", None)
        row.pop("ruta_destino", None)
  if not filas_export:
    return jsonify({"error": "No hay facturas que cumplan el filtro para exportar"}), 404

  try:
    import openpyxl
    from openpyxl import Workbook
  except ImportError:
    return jsonify({"error": "openpyxl no instalado. pip install openpyxl"}), 500

  wb = Workbook()
  ws = wb.active
  if ws is None:
    return jsonify({"error": "No se pudo crear la hoja de Excel"}), 500
  ws.title = "Facturas"
  for col, key in enumerate(CAMPOS_EXPORT, start=1):
    ws.cell(row=1, column=col, value=key)
  for row_idx, fila in enumerate(filas_export, start=2):
    for col, key in enumerate(CAMPOS_EXPORT, start=1):
      val = fila.get(key)
      if val is None:
        val = ""
      ws.cell(row=row_idx, column=col, value=val)

  buf = io.BytesIO()
  wb.save(buf)
  buf.seek(0)
  nombre_empresa = EMPRESAS_CLIENTE.get(empresa_id, empresa_id)
  sufijo = (year or "todos") + "_" + (month or "todos")
  if proveedor_filtro:
    nombre_safe = re.sub(r'[^\w\s\-]', "", proveedor_filtro)[:30].strip() or "proveedor"
    nombre_safe = nombre_safe.replace(" ", "_")
    filename = f"facturas_{nombre_safe}_{sufijo}.xlsx"
  else:
    filename = f"facturas_{nombre_empresa}_{sufijo}.xlsx"
  filename = re.sub(r'[<>:"/\\|?*,]', "", filename)
  if not filename.endswith(".xlsx"):
    filename = filename.rstrip() + ".xlsx"

  return Response(
    buf.read(),
    mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    headers={"Content-Disposition": f'attachment; filename="{filename}"'},
  )


def _aplicar_filtros_estado_tarjeta(
  filas: list[dict], estado_pago: str, tarjeta_id: str,
) -> list[dict]:
  """Filtra filas por estado_pago y/o tarjeta_id cuando vienen informados."""
  if not estado_pago and not tarjeta_id:
    return filas
  resultado: list[dict] = []
  tarjeta_str = (tarjeta_id or "").strip()
  for f in filas:
    if estado_pago:
      ep = (f.get("estado_pago") or "pendiente").strip() or "pendiente"
      if ep != estado_pago:
        continue
    if tarjeta_str:
      tid = f.get("tarjeta_id")
      if tid is None or str(tid).strip() != tarjeta_str:
        continue
    resultado.append(f)
  return resultado


def _filtrar_facturas_en_memoria(
  facturas: list[dict], year: str, month: str, proveedor_filtro: str = "",
) -> list[dict]:
  """Filtra lista de facturas (dicts) por año, mes y opcionalmente proveedor."""
  resultado: list[dict] = []
  for f in facturas:
    fecha = (f.get("fecha_factura") or "").strip()
    if year and not fecha.startswith(year):
      continue
    if month and (len(fecha) < 7 or fecha[5:7] != month):
      continue
    if proveedor_filtro and (f.get("proveedor") or "").strip() != proveedor_filtro:
      continue
    resultado.append(f)
  return resultado


def _facturas_filtradas_por_fecha(
  empresa_id: str, year: str, month: str, proveedor: str = "",
) -> list[dict]:
  """Devuelve facturas de la empresa filtradas por año, mes y opcionalmente proveedor. Usa BD primero (misma fuente que el listado), luego CSV si existe."""
  try:
    facturas_bd = facturas_db.get_facturas_empresa(empresa_id)
  except Exception as e:
    logger.warning("Error leyendo facturas de BD: %s", e)
    facturas_bd = []
  filas_bd = _filtrar_facturas_en_memoria(facturas_bd, year, month, proveedor)
  if filas_bd:
    return filas_bd
  ruta_csv = EMPRESAS_DIR / empresa_id / "base_maestra_facturas.csv"
  if not ruta_csv.exists():
    return []
  CAMPOS = [
    "fecha_factura", "proveedor", "numero_factura",
    "ruta_archivo", "ruta_destino",
  ]
  with ruta_csv.open("r", newline="", encoding="utf-8") as f:
    reader = csv.DictReader(f)
    filas = _filtrar_filas_csv(
      reader,
      CAMPOS,
      year=year,
      month=month,
      campo_filtro="proveedor",
      valor_filtro=proveedor,
      skip_header_por_empresa=False,
    )
  return filas


@facturas_proveedores_bp.get("/api/facturas_zip")
def descargar_facturas_zip():
  """
  Devuelve un ZIP con todos los archivos de facturas visibles (empresa + filtros año/mes).
  """
  empresa_id = request.args.get("empresa_id")
  year = (request.args.get("year") or "").strip()
  month = (request.args.get("month") or "").strip()
  proveedor_filtro = (request.args.get("proveedor") or "").strip()
  estado_pago_filtro = (request.args.get("estado_pago") or "").strip()
  tarjeta_id_filtro = (request.args.get("tarjeta_id") or "").strip()

  empresa_id, err = _validar_empresa_id_requerido(empresa_id)
  if err:
    return err[0], err[1]

  ids_filtro = (request.args.get("ids") or "").strip()
  ids_set: set[int] | None = None
  if ids_filtro:
    try:
      ids_set = {int(x) for x in ids_filtro.split(",") if x.strip()}
    except ValueError:
      pass

  filas = _facturas_filtradas_por_fecha(empresa_id, year, month, proveedor_filtro)
  filas = _aplicar_filtros_estado_tarjeta(filas, estado_pago_filtro, tarjeta_id_filtro)
  if ids_set:
    filas = [f for f in filas if f.get("id") in ids_set]
  if not filas:
    return jsonify({"error": "No hay facturas que cumplan el filtro para descargar"}), 404

  datos_resolved = DATOS_DIR.resolve()
  buf = io.BytesIO()
  nombres_usados: dict[str, int] = {}
  added: list[str] = []

  with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
    for fila in filas:
      ruta_param = (fila.get("ruta_destino") or fila.get("ruta_archivo") or "").strip()
      if not ruta_param:
        continue
      ruta = Path(ruta_param)
      if not ruta.is_absolute():
        ruta = DATOS_DIR / ruta
      ruta = ruta.resolve()
      try:
        ruta.relative_to(datos_resolved)
      except ValueError:
        continue
      if not ruta.exists() or not ruta.is_file():
        continue
      nombre_base = ruta.name
      if nombre_base not in nombres_usados:
        nombres_usados[nombre_base] = 0
      else:
        nombres_usados[nombre_base] += 1
        stem, sufijo = ruta.stem, ruta.suffix
        nombre_base = f"{stem}_{nombres_usados[nombre_base]}{sufijo}"
      zf.write(ruta, arcname=nombre_base)
      added.append(nombre_base)

  if not added:
    return jsonify({"error": "No se encontraron archivos de facturas para incluir en el ZIP"}), 404

  buf.seek(0)
  nombre_empresa = EMPRESAS_CLIENTE.get(empresa_id, empresa_id)
  sufijo = (year or "todos") + "_" + (month or "todos")
  if proveedor_filtro:
    nombre_safe = re.sub(r'[^\w\s-]', "", proveedor_filtro)[:30].strip() or "proveedor"
    filename = f"facturas_{nombre_safe}_{sufijo}.zip"
  else:
    filename = f"facturas_{nombre_empresa}_{sufijo}.zip"

  return send_file(
    buf,
    mimetype="application/zip",
    as_attachment=True,
    download_name=filename,
  )


# Cache en memoria de listados por empresa (invalidación al editar/eliminar/procesar)
_cache_listado_facturas_proveedores: dict[str, list[dict]] = {}
_cache_listado_facturas_clientes: dict[str, list[dict]] = {}


def _invalidar_cache_listado_proveedores(empresa_id: str) -> None:
  """Invalida el cache de listado de facturas proveedores para una empresa."""
  _cache_listado_facturas_proveedores.pop(empresa_id, None)


def _invalidar_cache_listado_clientes(empresa_id: str) -> None:
  """Invalida el cache de listado de facturas clientes para una empresa."""
  _cache_listado_facturas_clientes.pop(empresa_id, None)


@facturas_proveedores_bp.post("/api/facturas/migrar-desde-csv")
def migrar_facturas_desde_csv():
  """
  Migra la base maestra de facturas de proveedores desde los CSV (base_maestra_facturas.csv)
  a SQLite (tabla facturas_proveedor). Ejecutar una vez para pasar a usar la BD.
  Devuelve estadísticas y posibles errores.
  """
  try:
    resultado = facturas_db.migrar_desde_csv()
    return jsonify({
      "ok": True,
      "mensaje": "Migración de facturas completada.",
      "empresas_procesadas": resultado["empresas_procesadas"],
      "filas_migradas": resultado["filas_migradas"],
      "errores": resultado.get("errores", []),
    })
  except Exception as e:
    return jsonify({"ok": False, "error": str(e)}), 500


@facturas_proveedores_bp.put("/api/factura")
def actualizar_factura():
  """
  Actualiza una fila de la base maestra (SQLite). La fila se identifica por ruta_destino o ruta_archivo.
  Recibe JSON: { "empresa_id": "...", "factura": { ... } }.
  """
  data = request.get_json(silent=True) or {}
  empresa_id, err = _validar_empresa_id_requerido(data.get("empresa_id"))
  if err:
    return err[0], err[1]
  factura = data.get("factura")
  if not isinstance(factura, dict):
    return _bad_request("Falta empresa_id o factura")

  ruta_identificar = (factura.get("ruta_destino") or factura.get("ruta_archivo") or "").strip()
  if not ruta_identificar:
    return _bad_request("La factura debe tener ruta_destino o ruta_archivo para identificarla")

  todas = facturas_db.get_facturas_empresa(empresa_id)
  actualizado = None
  factura_id = None
  for f in todas:
    if ((f.get("ruta_destino") or f.get("ruta_archivo")) or "").strip() == ruta_identificar:
      actualizado = {c: (f.get(c) or "") for c in CAMPOS_FACTURAS_PROVEEDOR}
      factura_id = f.get("id")
      break
  if not actualizado:
    return jsonify({"error": "No se encontró la factura con esa ruta"}), 404
  for k, v in factura.items():
    if k in CAMPOS_FACTURAS_PROVEEDOR:
      actualizado[k] = (v.strip() if isinstance(v, str) else str(v or ""))
  actualizado["empresa_id"] = empresa_id
  actualizado["flag_error"] = False
  actualizado["motivo_error"] = ""
  if actualizado.get("estado_pago") not in ("pendiente", "pagada", "parcial"):
    actualizado["estado_pago"] = "pendiente"
  _revisor_basico([actualizado])
  nueva_fecha = (actualizado.get("fecha_factura") or "").strip()
  ruta_archivo_actual = Path(actualizado.get("ruta_destino") or actualizado.get("ruta_archivo") or "")
  if nueva_fecha and ruta_archivo_actual.exists():
    ano_dest = "Sin_fecha"
    mes_dest = "Sin fecha"
    try:
      dt = datetime.fromisoformat(nueva_fecha[:10])
      ano_dest = str(dt.year)
      mes_dest = f"{dt.month:02d}. {dt.strftime('%B')}"
    except Exception as e:
      logger.debug("Fecha no parseable al reubicar factura '%s': %s", nueva_fecha, e)
    destino_dir = FACTURAS_RECIBIDAS_DIR / empresa_id / ano_dest / mes_dest
    destino_dir.mkdir(parents=True, exist_ok=True)
    destino_final = destino_dir / ruta_archivo_actual.name
    if destino_final != ruta_archivo_actual:
      cnt = 2
      while destino_final.exists():
        destino_final = destino_dir / f"{ruta_archivo_actual.stem}_{cnt}{ruta_archivo_actual.suffix}"
        cnt += 1
      shutil.move(str(ruta_archivo_actual), destino_final)
      actualizado["ruta_destino"] = str(destino_final)
  ok = facturas_db.update_factura(empresa_id, actualizado)
  if not ok:
    return jsonify({"error": "No se pudo actualizar la factura en la base de datos"}), 500

  # FIX 2: Vincular factura a tercero automáticamente
  tercero_id_resultado = None
  if factura_id:
    tercero_id_explicito = factura.get("tercero_id")
    if tercero_id_explicito not in (None, "", "null"):
      # El usuario seleccionó del maestro — usar directamente
      try:
        tercero_id_int = int(tercero_id_explicito)
        facturas_db.update_tercero_id(factura_id, tercero_id_int)
        tercero_id_resultado = tercero_id_int
      except (ValueError, TypeError):
        logger.warning("tercero_id no válido: %s", tercero_id_explicito)
    else:
      # El usuario editó nombre/NIF manualmente — vincular automáticamente
      nombre_prov = (actualizado.get("proveedor") or "").strip()
      nif_prov = (actualizado.get("nif_proveedor") or "").strip()
      if nombre_prov or nif_prov:
        try:
          from core.terceros_db import crear_o_vincular_tercero
          resultado_tercero = crear_o_vincular_tercero(
            nombre=nombre_prov,
            cif=nif_prov,
            datos_extra={
              "pais": (actualizado.get("pais_proveedor") or "").strip(),
              "localidad": (actualizado.get("localidad_proveedor") or "").strip(),
            },
            rol="proveedor",
            origen="edicion_factura",
          )
          if resultado_tercero and resultado_tercero.get("id"):
            facturas_db.update_tercero_id(factura_id, resultado_tercero["id"])
            tercero_id_resultado = resultado_tercero["id"]
        except Exception as e:
          logger.warning("Error al vincular tercero en edición de factura: %s", e)

  # Vincular factura a proyecto si se especificó
  if factura_id:
    proyecto_id_val = factura.get("proyecto_id")
    if proyecto_id_val is not None:
      try:
        from core.db import conectar as _fc_conectar
        with _fc_conectar() as _fc_conn:
          # Asegurar que la columna existe
          _fc_cols = {r[1] for r in _fc_conn.execute("PRAGMA table_info(facturas_proveedor)").fetchall()}
          if "proyecto_id" not in _fc_cols:
            _fc_conn.execute("ALTER TABLE facturas_proveedor ADD COLUMN proyecto_id INTEGER REFERENCES proyectos(id)")
          pid = int(proyecto_id_val) if proyecto_id_val not in ("", "null") else None
          _fc_conn.execute("UPDATE facturas_proveedor SET proyecto_id = ? WHERE id = ?", (pid, factura_id))
      except Exception as e:
        logger.warning("Error al vincular proyecto en factura: %s", e)

  _sincronizar_proveedores_desde_facturas(empresa_id)
  _invalidar_cache_listado_proveedores(empresa_id)
  return jsonify({"ok": True, "mensaje": "Factura actualizada en la base maestra.", "tercero_id": tercero_id_resultado})


@facturas_proveedores_bp.delete("/api/facturas")
def eliminar_facturas():
  """
  Elimina una o varias filas del maestro (SQLite) identificadas por sus rutas (ruta_destino o ruta_archivo).
  Recibe JSON: { "empresa_id": "...", "rutas": ["ruta1", "ruta2", ...] }.
  """
  data = request.get_json(silent=True) or {}
  empresa_id, err = _validar_empresa_id_requerido(data.get("empresa_id"))
  if err:
    return err[0], err[1]
  rutas_eliminar = data.get("rutas", [])
  ids_eliminar = data.get("ids", [])

  if not isinstance(rutas_eliminar, list):
    rutas_eliminar = []
  if not isinstance(ids_eliminar, list):
    ids_eliminar = []

  rutas_set = set(r.strip() for r in rutas_eliminar if isinstance(r, str) and r.strip())
  ids_set = [int(i) for i in ids_eliminar if str(i).strip().isdigit()]

  if not rutas_set and not ids_set:
    return _bad_request("Falta empresa_id, rutas o ids")

  eliminadas = 0
  if rutas_set:
    eliminadas += facturas_db.delete_facturas(empresa_id, list(rutas_set))
  if ids_set:
    eliminadas += facturas_db.delete_facturas_por_ids(empresa_id, ids_set)
  _sincronizar_proveedores_desde_facturas(empresa_id)
  _invalidar_cache_listado_proveedores(empresa_id)
  return jsonify({"ok": True, "eliminadas": eliminadas, "mensaje": f"{eliminadas} factura(s) eliminada(s)."})


@archivo_bp.get("/api/archivo")
def servir_archivo():
  """
  Sirve un archivo de factura por ruta. Solo permite rutas dentro de data/.
  Si el archivo no existe localmente y SharePoint está configurado,
  intenta descargarlo de SharePoint como fallback.
  """
  ruta_param = request.args.get("ruta")
  if not ruta_param:
    return _bad_request("Falta ruta")

  # Extraer ruta relativa a data/ (las rutas en BD pueden ser absolutas de Windows)
  ruta_rel = ruta_param
  for sep in ("\\data\\", "/data/", "\\data/", "/data\\"):
    if sep in ruta_rel:
      ruta_rel = ruta_rel.split(sep, 1)[1]
      break
  ruta_rel = ruta_rel.replace("\\", "/")

  ruta = (DATOS_DIR / ruta_rel).resolve()

  try:
    ruta.relative_to(DATOS_DIR.resolve())
  except ValueError:
    return jsonify({"error": "Ruta no permitida"}), 403

  if ruta.exists() and ruta.is_file():
    return send_file(ruta, as_attachment=False, download_name=ruta.name)

  # Fallback: intentar desde SharePoint si está configurado
  if os.environ.get("MICROSOFT_CLIENT_ID"):
    try:
      from core.onedrive_db import get_sharepoint_client

      client = get_sharepoint_client()
      contenido = client.descargar_archivo(ruta_rel)
      if contenido:
        ext = ruta.suffix.lstrip(".").lower()
        ct_map = {
            "pdf": "application/pdf", "jpg": "image/jpeg",
            "jpeg": "image/jpeg", "png": "image/png",
        }
        from flask import Response

        return Response(
            contenido,
            mimetype=ct_map.get(ext, "application/octet-stream"),
            headers={"Content-Disposition": f'inline; filename="{ruta.name}"'},
        )
    except Exception as exc:
      logger.warning("Fallback SharePoint falló para %s: %s", ruta_param, exc)

  return jsonify({"error": "Archivo no encontrado"}), 404


# ─── Pipeline de procesamiento de Facturas de Clientes (Emitidas) ─────────


def _extraer_una_factura_llm_cliente(ruta: Path, empresa_id: str) -> dict | None:
  """
  Extrae los campos de una sola factura emitida (cliente) con visión y/o LLM.
  Devuelve la fila de datos o None si no se pudo extraer.
  Usado por _extractor_llm_clientes para ejecución en paralelo.
  """
  usar_vision = False
  datos: dict = {}
  suf = (ruta.suffix or "").lower()
  es_imagen = suf in (".jpg", ".jpeg", ".png", ".webp", ".gif")

  if es_imagen:
    datos = _extraer_campos_vision(ruta, empresa_id, tipo="cliente")
    if datos:
      usar_vision = True
      _registrar_vision_control(empresa_id, ruta.name, str(ruta))

  texto = ""

  if not datos:
    texto = _leer_texto_factura(ruta)
    texto = _normalizar_texto(texto)
    datos = _extraer_campos_llm(texto, empresa_id, tipo="cliente")

  if not datos:
    return None

  claves_clave = ["cliente", "cif_nif", "fecha_factura", "numero_factura", "total_a_pagar"]

  def _es_casi_sin_datos(d: dict) -> bool:
    num_no_vacias = sum(1 for c in claves_clave if str(d.get(c, "") or "").strip())
    return num_no_vacias <= 2

  if _es_casi_sin_datos(datos):
    if es_imagen and usar_vision:
      if not texto:
        texto = _leer_texto_factura(ruta)
        texto = _normalizar_texto(texto)
      datos_texto = _extraer_campos_llm(texto, empresa_id, tipo="cliente")
      if datos_texto and not _es_casi_sin_datos(datos_texto):
        datos = datos_texto
        usar_vision = False
    elif es_imagen and not usar_vision:
      datos_vision = _extraer_campos_vision(ruta, empresa_id, tipo="cliente")
      if datos_vision and not _es_casi_sin_datos(datos_vision):
        datos = datos_vision
        usar_vision = True
        _registrar_vision_control(empresa_id, ruta.name, str(ruta))

  if _es_casi_sin_datos(datos):
    return None

  fila = {
    "ruta_archivo": str(ruta),
    "empresa_id": empresa_id,
    "extraccion_vision": "1" if usar_vision else "",
  }
  for campo in _CLAVES_FACTURA_CLIENTE:
    fila[campo] = str(datos.get(campo, "") or "").strip()

  return fila


def _extractor_llm_clientes(rutas: list[Path], empresa_id: str) -> list[dict]:
  """
  Extractor LLM para facturas emitidas a clientes.
  Misma cascada visión/texto que el de proveedores. Procesa en paralelo con ThreadPoolExecutor.
  """
  filas: list[dict] = []
  if not rutas:
    return filas
  workers = min(_MAX_WORKERS_EXTRACTOR_LLM, len(rutas))
  with ThreadPoolExecutor(max_workers=workers) as executor:
    futures = [executor.submit(_extraer_una_factura_llm_cliente, ruta, empresa_id) for ruta in rutas]
    for fut in futures:
      try:
        fila = fut.result()
        if fila is not None:
          filas.append(fila)
      except Exception as e:
        logger.warning("Error procesando factura cliente en paralelo: %s", e)
  return filas


def _extractor_basico_clientes(rutas: list[Path], empresa_id: str) -> list[dict]:
  """Extractor básico (backup) para facturas de clientes: solo extrae fecha, NIF y totales."""
  filas: list[dict] = []
  for ruta in rutas:
    texto = _leer_texto_factura(ruta)
    texto = _normalizar_texto(texto)
    fecha = _buscar_primera_fecha(texto)
    nif = _buscar_nif_cif(texto)
    importes = _buscar_importes(texto)

    def fmt(v) -> str:
      return "" if v is None else f"{v:.2f}" if isinstance(v, float) else str(v)

    fila = {
      "ruta_archivo": str(ruta),
      "empresa_id": empresa_id,
      "fecha_factura": fecha,
      "cliente": "",
      "cif_nif": nif,
      "pais": "",
      "localidad": "",
      "proyecto": "",
      "tipologia": "",
      "num_hincadoras": "",
      "num_ayudantes": "",
      "pricing_servicio": "",
      "pricing_transporte": "",
      "iva": fmt(importes.get("iva_cuota_total")),
      "total_a_pagar": fmt(importes.get("total_a_pagar") or importes.get("total_factura")),
      "numero_factura": _buscar_numero_factura(texto),
      "extraccion_vision": "",
    }
    filas.append(fila)
  return filas


def _get_hashes_csv_clientes(empresa_id: str) -> set[str]:
  """Devuelve el conjunto de hash_archivo ya presentes en la BD de clientes de la empresa."""
  return facturas_cliente_db.get_hashes_empresa_cliente(empresa_id)


def _base_maestra_csv_clientes(filas: list[dict], empresa_id: str) -> dict:
  """Guarda las facturas de clientes procesadas en la BD (tabla facturas_cliente)."""
  resultado = facturas_cliente_db.insert_facturas_clientes(empresa_id, filas)
  return {
    "ruta_base_maestra": "BD (facturas_cliente)",
    "filas_añadidas": resultado["insertados"],
    "ids_insertados": resultado["ids"],
  }


def procesar_lote_clientes(empresa_id: str, carpeta: Path) -> dict:
  """
  Orquestador para facturas emitidas a clientes:
  Recolector → Extractor → Revisor → Archivador → Base de datos.
  """
  archivos = _recolector(carpeta)
  if not archivos:
    return {
      "procesado": False,
      "motivo": "No se han encontrado archivos en la carpeta de entrada.",
      "empresa_id": empresa_id,
      "carpeta_entrada": str(carpeta),
    }

  tabla = _extractor_llm_clientes(archivos, empresa_id)
  if not tabla:
    tabla = _extractor_basico_clientes(archivos, empresa_id)

  for fila in tabla:
    pais = (fila.get("pais") or "").strip()
    localidad = (fila.get("localidad") or "").strip()
    if not pais and localidad:
      pais_detectado = _obtener_pais_desde_localidad(localidad)
      if pais_detectado:
        fila["pais"] = pais_detectado

  tabla = _revisor_basico_clientes(tabla)
  _añadir_hashes_tabla(tabla)
  hashes_existentes = _get_hashes_csv_clientes(empresa_id)
  tabla_sin_duplicados = [
    f for f in tabla
    if (f.get("hash_archivo") or "").strip() and (f.get("hash_archivo") or "").strip() not in hashes_existentes
  ]
  duplicados_omitidos = len(tabla) - len(tabla_sin_duplicados)
  tabla = _archivar_por_fecha(tabla_sin_duplicados, FACTURAS_EMITIDAS_DIR, actualizar_ruta_archivo=True)
  _subir_lote_a_sharepoint(tabla, "Facturas Emitidas")

  # Crear/vincular clientes a terceros via gateway unificado antes de insertar
  from core.terceros_db import crear_o_vincular_tercero, init_terceros_db
  init_terceros_db()
  for fila in tabla:
    cliente = (fila.get("cliente") or "").strip()
    cif = (fila.get("cif_nif") or "").strip()
    if not cliente:
      continue
    try:
      crear_o_vincular_tercero(
        cliente, cif,
        datos_extra={"pais": fila.get("pais"), "localidad": fila.get("localidad")},
        rol="cliente", origen="ocr",
      )
    except Exception as e:
      logger.warning("Error vinculando cliente '%s' a tercero: %s", cliente, e)

  resumen_bd = _base_maestra_csv_clientes(tabla, empresa_id)

  # Post-insert: vincular facturas recien insertadas a terceros (por CIF/nombre)
  try:
    crm_db.vincular_facturas_a_terceros()
  except Exception as e:
    logger.warning("Error al vincular facturas cliente a terceros: %s", e)
  facturas_con_vision = sum(1 for f in tabla if str(f.get("extraccion_vision") or "").strip() == "1")

  return {
    "procesado": True,
    "empresa_id": empresa_id,
    "carpeta_entrada": str(carpeta),
    "facturas_procesadas": len(tabla_sin_duplicados),
    "facturas_omitidas_duplicadas": duplicados_omitidos,
    "facturas_con_vision": facturas_con_vision,
    "ruta_base_maestra": resumen_bd["ruta_base_maestra"],
    "filas_añadidas": resumen_bd["filas_añadidas"],
    "ids_insertados": resumen_bd.get("ids_insertados", []),
  }


# ─── Facturas de Clientes (Facturas Emitidas) – CRUD ─────────────────────────


def _get_clientes_unicos_empresa(empresa_id: str) -> list[dict]:
  """
  Devuelve la lista de clientes únicos de la empresa a partir de las facturas de clientes (BD, tabla facturas_cliente).
  Agrupa por (cliente, cif_nif); cada elemento tiene cliente, cif_nif, pais, localidad
  (y opcionalmente proyecto) tomados de la primera factura que aparezca para ese cliente.
  Compatible con el maestro de terceros (GET /api/empresas/<id>/clientes fusiona maestro + este agregado).
  """
  facturas = facturas_cliente_db.get_facturas_cliente_empresa(empresa_id)
  vistos: dict[tuple[str, str], dict] = {}
  for f in facturas:
    cliente = (f.get("cliente") or "").strip()
    cif_nif = (f.get("cif_nif") or "").strip()
    if not cliente and not cif_nif:
      continue
    clave = (cliente, cif_nif)
    if clave not in vistos:
      vistos[clave] = {
        "cliente": cliente,
        "cif_nif": cif_nif,
        "pais": (f.get("pais") or "").strip(),
        "localidad": (f.get("localidad") or "").strip(),
        "proyecto": (f.get("proyecto") or "").strip(),
      }
  return sorted(vistos.values(), key=lambda x: (x["cliente"], x["cif_nif"]))


@facturas_clientes_bp.post("/api/procesar_clientes")
def procesar_clientes():
  """Recibe archivos de facturas emitidas a clientes y lanza el pipeline de extracción."""
  ensure_dirs()

  empresa_id = request.form.get("empresa_id")
  empresa_id, err = _validar_empresa_id_requerido(empresa_id)
  if err:
    return err[0], err[1]

  files = request.files.getlist("archivos")
  if not files:
    return _bad_request("No se han recibido archivos")

  timestamp = int(time.time())
  destino = SUBIDAS_DIR / empresa_id / f"clientes_{timestamp}"
  destino.mkdir(parents=True, exist_ok=True)

  nombres_guardados = []
  for f in files:
    nombre = f.filename or f"factura_{len(nombres_guardados) + 1}.dat"
    ruta = destino / os.path.basename(nombre)
    f.save(ruta)
    nombres_guardados.append(str(ruta))

  resumen = procesar_lote_clientes(empresa_id, destino)
  _invalidar_cache_listado_clientes(empresa_id)

  mensaje = "Facturas de clientes procesadas correctamente."
  if not resumen.get("procesado"):
    mensaje = resumen.get("motivo", "No se han podido procesar las facturas.")

  return jsonify(
    {
      "mensaje": mensaje,
      "empresa_id": empresa_id,
      "carpeta_entrada": str(destino),
      "archivos_entrada": nombres_guardados,
      "resumen_proceso": resumen,
    }
  )


@facturas_clientes_bp.get("/api/facturas_clientes")
def listar_facturas_clientes():
  """Listado de facturas emitidas a clientes; usa cache en memoria por empresa.
  Si solo_pendientes_vinculacion=1, excluye las ya vinculadas a un movimiento de caja."""
  empresa_id, err = _validar_empresa_id_requerido(request.args.get("empresa_id"))
  if err:
    return jsonify({"facturas": [], "error": "Falta empresa_id"}), 400
  filtro_cliente = (request.args.get("cliente") or "").strip()
  filtro_cobro = (request.args.get("estado_cobro") or "").strip().lower()
  solo_pendientes = request.args.get("solo_pendientes_vinculacion", "").strip().lower() in ("1", "true", "yes")
  if empresa_id in _cache_listado_facturas_clientes:
    facturas = list(_cache_listado_facturas_clientes[empresa_id])
  else:
    facturas = facturas_cliente_db.get_facturas_cliente_empresa(empresa_id)
    _cache_listado_facturas_clientes[empresa_id] = facturas
  if filtro_cliente:
    facturas = [f for f in facturas if (f.get("cliente") or "").strip() == filtro_cliente]
  if filtro_cobro:
    facturas = [f for f in facturas if (f.get("estado_cobro") or "pendiente").strip().lower() == filtro_cobro]
  if solo_pendientes and facturas:
    # Excluir las ya cobradas totalmente (para el modal de conciliación de cobros)
    facturas = [f for f in facturas if (f.get("estado_cobro") or "pendiente").strip().lower() != "cobrada"]
  return jsonify({"facturas": facturas, "empresa_id": empresa_id})


@facturas_clientes_bp.get("/api/clientes_unicos")
def clientes_unicos():
  """Lista de nombres de cliente únicos; usa cache de listado de facturas clientes."""
  empresa_id, err = _validar_empresa_id_requerido(request.args.get("empresa_id"))
  if err:
    return jsonify({"clientes": [], "error": "Falta empresa_id"}), 400
  if empresa_id in _cache_listado_facturas_clientes:
    facturas = _cache_listado_facturas_clientes[empresa_id]
  else:
    facturas = facturas_cliente_db.get_facturas_cliente_empresa(empresa_id)
    _cache_listado_facturas_clientes[empresa_id] = facturas
  nombres: set[str] = set((f.get("cliente") or "").strip() for f in facturas if (f.get("cliente") or "").strip())
  return jsonify({"clientes": sorted(nombres), "empresa_id": empresa_id})


@facturas_clientes_bp.get("/api/empresas/<empresa_id>/clientes")
def listar_clientes_por_empresa(empresa_id: str):
  """Listado de clientes únicos: maestro (BD terceros) + agregado desde facturas de clientes (BD facturas_cliente). Cada item tiene en_maestro true si está en el maestro."""
  empresa_id, err = _validar_empresa_id_requerido(empresa_id)
  if err:
    return err[0], err[1]
  desde_bd = terceros_db.get_clientes_empresa(empresa_id) if terceros_db.hay_clientes_en_bd() else []
  agregado_facturas = _get_clientes_unicos_empresa(empresa_id)
  clave_bd = {((c.get("cliente") or "").strip(), (c.get("cif_nif") or "").strip()) for c in desde_bd}
  merged: list[dict] = []
  for c in desde_bd:
    merged.append({**c, "en_maestro": True})
  for c in agregado_facturas:
    cli = (c.get("cliente") or "").strip()
    cif = (c.get("cif_nif") or "").strip()
    if (cli, cif) not in clave_bd:
      merged.append({**c, "en_maestro": False})
      clave_bd.add((cli, cif))
  merged.sort(key=lambda x: ((x.get("cliente") or "").lower(), (x.get("cif_nif") or "").lower()))
  return jsonify({"clientes": merged, "empresa_id": empresa_id})


def _normalizar_cliente_body(data: dict) -> dict:
  """Extrae y normaliza campos de cliente desde el body."""
  return {
    "cliente": (data.get("cliente") or "").strip(),
    "cif_nif": (data.get("cif_nif") or "").strip(),
    "pais": (data.get("pais") or "").strip(),
    "localidad": (data.get("localidad") or "").strip(),
    "proyecto": (data.get("proyecto") or "").strip(),
    "direccion": (data.get("direccion") or "").strip(),
    "email": (data.get("email") or "").strip(),
    "telefono": (data.get("telefono") or "").strip(),
  }


@facturas_clientes_bp.post("/api/clientes")
def crear_cliente():
  """
  Alta de cliente en el maestro de la empresa (terceros + empresa_tercero con es_cliente=1).
  JSON: empresa_id (obligatorio), cliente, cif_nif (mínimos), pais, localidad, proyecto, direccion, email, telefono.
  Usa el gateway unificado crear_o_vincular_tercero para deduplicacion.
  """
  from core.terceros_db import crear_o_vincular_tercero
  data = request.get_json(silent=True) or {}
  empresa_id, err = _validar_empresa_id_requerido(data.get("empresa_id"))
  if err:
    return err[0], err[1]
  c = _normalizar_cliente_body(data)
  if not c["cliente"]:
    return _bad_request("El nombre del cliente es obligatorio")
  if not c["cif_nif"]:
    return _bad_request("El CIF/NIF del cliente es obligatorio")

  # Verificar via gateway unificado si ya existe
  resultado = crear_o_vincular_tercero(
    c["cliente"], c["cif_nif"],
    datos_extra={k: c.get(k) for k in ("pais", "localidad", "direccion", "email", "telefono")},
    rol="cliente", origen="manual",
  )
  if resultado["accion"] != "creado":
    warning = None
    if resultado["accion"] == "vinculado_similar" and resultado.get("requiere_revision"):
      warning = f"Vinculado a tercero similar: {resultado['nombre_match']} (similitud {resultado['similitud']:.0%})"
    return jsonify({
      "error": f"Ya existe un cliente equivalente: {resultado['nombre_match']} (#{resultado['id']})",
      "tercero_existente": resultado,
      "warning": warning,
    }), 409

  lista_bd = terceros_db.get_clientes_empresa(empresa_id) if terceros_db.hay_clientes_en_bd() else []
  lista_bd.append(c)
  terceros_db.init_terceros_db()
  terceros_db.guardar_clientes_empresa(empresa_id, lista_bd)
  return jsonify({"ok": True, "clientes": terceros_db.get_clientes_empresa(empresa_id), "empresa_id": empresa_id}), 201


@facturas_clientes_bp.put("/api/clientes")
def actualizar_cliente():
  """
  Edición de un cliente del maestro. Se identifica por old_cliente y old_cif_nif.
  JSON: empresa_id, old_cliente, old_cif_nif, y el resto de campos (cliente, cif_nif, ...).
  """
  data = request.get_json(silent=True) or {}
  empresa_id, err = _validar_empresa_id_requerido(data.get("empresa_id"))
  if err:
    return err[0], err[1]
  old_cliente = (data.get("old_cliente") or "").strip()
  old_cif = (data.get("old_cif_nif") or "").strip()
  if not old_cliente and not old_cif:
    return _bad_request("Faltan old_cliente y old_cif_nif para identificar el cliente")
  c = _normalizar_cliente_body(data)
  if not c["cliente"]:
    return _bad_request("El nombre del cliente es obligatorio")
  if not c["cif_nif"]:
    return _bad_request("El CIF/NIF del cliente es obligatorio")
  lista_bd = terceros_db.get_clientes_empresa(empresa_id) if terceros_db.hay_clientes_en_bd() else []
  idx = -1
  for i, existente in enumerate(lista_bd):
    if (existente.get("cliente") or "").strip() == old_cliente and (existente.get("cif_nif") or "").strip() == old_cif:
      idx = i
      break
  if idx < 0:
    return jsonify({"error": "Cliente no encontrado en el maestro"}), 404
  lista_bd[idx] = c
  terceros_db.init_terceros_db()
  terceros_db.guardar_clientes_empresa(empresa_id, lista_bd)
  return jsonify({"ok": True, "clientes": terceros_db.get_clientes_empresa(empresa_id), "empresa_id": empresa_id})


@facturas_clientes_bp.delete("/api/clientes")
def eliminar_cliente():
  """
  Elimina un cliente del maestro. Solo si está en el maestro.
  Body: { "empresa_id": "...", "cliente": "...", "cif_nif": "..." }.
  """
  data = request.get_json(silent=True) or {}
  empresa_id, err = _validar_empresa_id_requerido(data.get("empresa_id"))
  if err:
    return err[0], err[1]
  cliente = (data.get("cliente") or "").strip()
  cif_nif = (data.get("cif_nif") or "").strip()
  if not cliente and not cif_nif:
    return _bad_request("Faltan cliente y cif_nif para identificar el cliente")
  terceros_db.init_terceros_db()
  lista_bd = terceros_db.get_clientes_empresa(empresa_id) if terceros_db.hay_clientes_en_bd() else []
  nueva_lista = [
    c for c in lista_bd
    if not (
      (c.get("cliente") or "").strip() == cliente
      and (c.get("cif_nif") or "").strip() == cif_nif
    )
  ]
  if len(nueva_lista) == len(lista_bd):
    return jsonify({"error": "Cliente no encontrado en el maestro"}), 404
  terceros_db.guardar_clientes_empresa(empresa_id, nueva_lista)
  _invalidar_cache_listado_clientes(empresa_id)
  return jsonify({"ok": True, "clientes": nueva_lista, "empresa_id": empresa_id, "mensaje": "Cliente eliminado del maestro."})


@facturas_clientes_bp.get("/api/facturas_clientes_export")
def exportar_facturas_clientes():
  """Exporta las facturas de clientes en CSV con extensión .xlsx."""
  empresa_id, err = _validar_empresa_id_requerido(request.args.get("empresa_id"))
  if err:
    return err[0], err[1]
  year = (request.args.get("year") or "").strip()
  month = (request.args.get("month") or "").strip()
  cliente_filtro = (request.args.get("cliente") or "").strip()
  ids_filtro = (request.args.get("ids") or "").strip()
  ids_set: set[int] | None = None
  if ids_filtro:
    try:
      ids_set = {int(x) for x in ids_filtro.split(",") if x.strip()}
    except ValueError:
      pass
  facturas = facturas_cliente_db.get_facturas_cliente_empresa(empresa_id)
  if ids_set:
    facturas = [f for f in facturas if f.get("id") in ids_set]
  if not facturas:
    return jsonify({"error": "No hay datos para exportar"}), 404
  CAMPOS_EXPORT_CLI = [
    "fecha_factura", "cliente", "cif_nif", "pais", "localidad",
    "numero_factura", "proyecto", "tipologia", "num_hincadoras",
    "num_ayudantes", "pricing_servicio", "pricing_transporte",
    "iva", "total_a_pagar",
  ]
  filas_export = _filtrar_filas_csv(
    iter(facturas),
    CAMPOS_EXPORT_CLI,
    year=year,
    month=month,
    campo_filtro="cliente",
    valor_filtro=cliente_filtro,
    skip_header_por_empresa=False,
  )
  if not filas_export:
    return jsonify({"error": "No hay facturas que cumplan el filtro"}), 404
  output = io.StringIO()
  writer = csv.DictWriter(output, fieldnames=CAMPOS_EXPORT_CLI)
  writer.writeheader()
  for fila in filas_export:
    writer.writerow(fila)
  output.seek(0)
  nombre_empresa = EMPRESAS_CLIENTE.get(empresa_id, empresa_id)
  sufijo = (year or "todos") + "_" + (month or "todos")
  if cliente_filtro:
    nombre_safe = re.sub(r'[^\w\s-]', "", cliente_filtro)[:30].strip() or "cliente"
    filename = f"facturas_{nombre_safe}_{sufijo}.xlsx"
  else:
    filename = f"facturas_clientes_{nombre_empresa}_{sufijo}.xlsx"
  return Response(
    output.read(),
    mimetype="text/csv",
    headers={"Content-Disposition": f'attachment; filename="{filename}"'},
  )


@facturas_clientes_bp.get("/api/facturas_clientes_zip")
def descargar_facturas_clientes_zip():
  """Devuelve un ZIP con los archivos de facturas de clientes."""
  empresa_id, err = _validar_empresa_id_requerido(request.args.get("empresa_id"))
  if err:
    return err[0], err[1]
  year = (request.args.get("year") or "").strip()
  month = (request.args.get("month") or "").strip()
  cliente_filtro = (request.args.get("cliente") or "").strip()
  ids_filtro = (request.args.get("ids") or "").strip()
  ids_set: set[int] | None = None
  if ids_filtro:
    try:
      ids_set = {int(x) for x in ids_filtro.split(",") if x.strip()}
    except ValueError:
      pass
  facturas = facturas_cliente_db.get_facturas_cliente_empresa(empresa_id)
  if ids_set:
    facturas = [f for f in facturas if f.get("id") in ids_set]
  if not facturas:
    return jsonify({"error": "No hay facturas para descargar"}), 404
  CAMPOS_ZIP_CLI = [
    "fecha_factura",
    "cliente",
    "ruta_archivo",
  ]
  filas_zip = _filtrar_filas_csv(
    iter(facturas),
    CAMPOS_ZIP_CLI,
    year=year,
    month=month,
    campo_filtro="cliente",
    valor_filtro=cliente_filtro,
    skip_header_por_empresa=False,
  )
  rutas_archivos = [(f.get("ruta_archivo") or "").strip() for f in filas_zip if (f.get("ruta_archivo") or "").strip()]
  if not rutas_archivos:
    return jsonify({"error": "No hay archivos que cumplan el filtro"}), 404
  datos_resolved = DATOS_DIR.resolve()
  buf = io.BytesIO()
  nombres_usados: dict[str, int] = {}
  added: list[str] = []
  with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
    for ruta_param in rutas_archivos:
      ruta = Path(ruta_param)
      if not ruta.is_absolute():
        ruta = DATOS_DIR / ruta
      ruta = ruta.resolve()
      try:
        ruta.relative_to(datos_resolved)
      except ValueError:
        continue
      if not ruta.exists() or not ruta.is_file():
        continue
      nombre_base = ruta.name
      if nombre_base not in nombres_usados:
        nombres_usados[nombre_base] = 0
      else:
        nombres_usados[nombre_base] += 1
        stem, suf = ruta.stem, ruta.suffix
        nombre_base = f"{stem}_{nombres_usados[nombre_base]}{suf}"
      zf.write(ruta, arcname=nombre_base)
      added.append(nombre_base)
  if not added:
    return jsonify({"error": "No se encontraron archivos para incluir en el ZIP"}), 404
  buf.seek(0)
  nombre_empresa = EMPRESAS_CLIENTE.get(empresa_id, empresa_id)
  sufijo = (year or "todos") + "_" + (month or "todos")
  if cliente_filtro:
    nombre_safe = re.sub(r'[^\w\s-]', "", cliente_filtro)[:30].strip() or "cliente"
    filename = f"facturas_{nombre_safe}_{sufijo}.zip"
  else:
    filename = f"facturas_clientes_{nombre_empresa}_{sufijo}.zip"
  return Response(
    buf.read(),
    mimetype="application/zip",
    headers={"Content-Disposition": f'attachment; filename="{filename}"'},
  )


@facturas_clientes_bp.post("/api/facturas_clientes/migrar-desde-csv")
def migrar_facturas_clientes_desde_csv():
  """
  Migra las facturas de clientes desde los CSV (facturas_clientes.csv por empresa)
  a SQLite (tabla facturas_cliente). Ejecutar una vez para rellenar la BD desde CSV existentes.
  Tras ejecutarlo, la fuente de verdad es la BD; los CSV pueden conservarse como respaldo histórico.
  Devuelve estadísticas y posibles errores.
  """
  try:
    resultado = facturas_cliente_db.migrar_desde_csv_clientes()
    return jsonify({
      "ok": True,
      "mensaje": "Migración de facturas de clientes completada.",
      "empresas_procesadas": resultado["empresas_procesadas"],
      "filas_migradas": resultado["filas_migradas"],
      "errores": resultado.get("errores", []),
    })
  except Exception as e:
    return jsonify({"ok": False, "error": str(e)}), 500


@facturas_clientes_bp.post("/api/factura_cliente")
def crear_factura_cliente():
  """Crea una nueva factura de cliente (manual)."""
  data = request.get_json(silent=True) or {}
  empresa_id, err = _validar_empresa_id_requerido(data.get("empresa_id"))
  if err:
    return err[0], err[1]
  factura = data.get("factura")
  if not isinstance(factura, dict):
    return _bad_request("Falta empresa_id o factura")
  row = {c: str(factura.get(c, "") or "").strip() for c in CAMPOS_FACTURAS_CLIENTE}
  row["empresa_id"] = empresa_id
  facturas_cliente_db.insert_factura_cliente(empresa_id, row)
  _invalidar_cache_listado_clientes(empresa_id)
  return jsonify({"ok": True, "mensaje": "Factura de cliente registrada."})


@facturas_clientes_bp.put("/api/factura_cliente")
def actualizar_factura_cliente():
  """Actualiza una factura de cliente identificada por empresa_id + numero_factura + fecha_factura + cliente."""
  data = request.get_json(silent=True) or {}
  empresa_id, err = _validar_empresa_id_requerido(data.get("empresa_id"))
  if err:
    return err[0], err[1]
  factura = data.get("factura")
  clave_original = data.get("clave_original") or {}
  if not isinstance(factura, dict):
    return _bad_request("Falta empresa_id o factura")
  id_num = (clave_original.get("numero_factura") or factura.get("numero_factura") or "").strip()
  id_fecha = (clave_original.get("fecha_factura") or factura.get("fecha_factura") or "").strip()
  id_cliente = (clave_original.get("cliente") or factura.get("cliente") or "").strip()
  if not id_num and not id_fecha and not id_cliente:
    return _bad_request("No se puede identificar la factura a actualizar")
  actualizado = {c: (factura.get(c) or "").strip() for c in CAMPOS_FACTURAS_CLIENTE}
  actualizado["empresa_id"] = empresa_id
  _revisor_basico_clientes([actualizado])
  ok = facturas_cliente_db.update_factura_cliente(empresa_id, actualizado, clave_original)
  if not ok:
    return jsonify({"error": "No se encontró la factura de cliente a actualizar"}), 404
  # Vincular factura de cliente a proyecto si se especificó
  proyecto_id_cli = factura.get("proyecto_id")
  if proyecto_id_cli is not None:
    try:
      from core.db import conectar as _fcc_conectar
      with _fcc_conectar() as _fcc_conn:
        _fcc_cols = {r[1] for r in _fcc_conn.execute("PRAGMA table_info(facturas_cliente)").fetchall()}
        if "proyecto_id" in _fcc_cols:
          id_num = (clave_original.get("numero_factura") or factura.get("numero_factura") or "").strip()
          id_fecha = (clave_original.get("fecha_factura") or factura.get("fecha_factura") or "").strip()
          id_cli = (clave_original.get("cliente") or factura.get("cliente") or "").strip()
          row_cli = _fcc_conn.execute(
            """SELECT id FROM facturas_cliente WHERE empresa_id = ?
               AND (? = '' OR numero_factura = ?) AND (? = '' OR fecha_factura = ?) AND (? = '' OR cliente = ?)
               LIMIT 1""",
            (empresa_id, id_num, id_num, id_fecha, id_fecha, id_cli, id_cli),
          ).fetchone()
          if row_cli:
            pid_cli = int(proyecto_id_cli) if proyecto_id_cli not in ("", "null") else None
            _fcc_conn.execute("UPDATE facturas_cliente SET proyecto_id = ? WHERE id = ?", (pid_cli, row_cli[0]))
    except Exception as e:
      logger.warning("Error al vincular proyecto en factura cliente: %s", e)
  _invalidar_cache_listado_clientes(empresa_id)
  return jsonify({"ok": True, "mensaje": "Factura de cliente actualizada."})


@facturas_clientes_bp.delete("/api/facturas_clientes")
def eliminar_facturas_clientes():
  """Elimina facturas de clientes por índice (posición en el listado)."""
  data = request.get_json(silent=True) or {}
  empresa_id, err = _validar_empresa_id_requerido(data.get("empresa_id"))
  if err:
    return err[0], err[1]
  indices = data.get("indices")
  if not isinstance(indices, list) or not indices:
    return _bad_request("Falta empresa_id o indices")
  eliminadas = facturas_cliente_db.delete_facturas_cliente_por_indices(empresa_id, indices)
  _invalidar_cache_listado_clientes(empresa_id)
  return jsonify({"ok": True, "eliminadas": eliminadas, "mensaje": f"{eliminadas} factura(s) de cliente eliminada(s)."})


@facturas_clientes_bp.post("/api/facturas_clientes/recalcular-cobros")
def recalcular_cobros_clientes():
  """Recalcula el estado_cobro de TODAS las facturas de cliente basándose en conciliaciones."""
  result = facturas_cliente_db.recalcular_todos_estados_cobro()
  return jsonify(result)


@control_calidad_bp.post("/api/control-calidad/analizar")
def control_calidad_analizar():
  """
  Ejecuta el análisis de calidad sobre las facturas (proveedores y/o clientes).
  Parámetros (JSON): empresa_id (obligatorio), tipo opcional: "proveedores" | "clientes" | "ambos",
  incluir_tests opcional: true para ejecutar tests unitarios y devolver resultado.
  """
  import unittest

  data = request.get_json(silent=True) or {}
  empresa_id, err = _validar_empresa_id_requerido(data.get("empresa_id"))
  if err:
    return err[0], err[1]
  tipo = (data.get("tipo") or "ambos").strip().lower()
  if tipo not in ("proveedores", "clientes", "ambos"):
    return _bad_request('tipo debe ser "proveedores", "clientes" o "ambos"')
  incluir_tests = data.get("incluir_tests") is True

  # Cargar facturas (misma lógica que listado, con cache)
  filas_proveedores: list[dict] = []
  filas_clientes: list[dict] = []
  if tipo in ("proveedores", "ambos"):
    if empresa_id in _cache_listado_facturas_proveedores:
      filas_proveedores = _cache_listado_facturas_proveedores[empresa_id]
    else:
      filas_proveedores = facturas_db.get_facturas_empresa(empresa_id)
      _cache_listado_facturas_proveedores[empresa_id] = filas_proveedores
  if tipo in ("clientes", "ambos"):
    if empresa_id in _cache_listado_facturas_clientes:
      filas_clientes = _cache_listado_facturas_clientes[empresa_id]
    else:
      filas_clientes = facturas_cliente_db.get_facturas_cliente_empresa(empresa_id)
      _cache_listado_facturas_clientes[empresa_id] = filas_clientes

  # Análisis por fila (tarea 1.1)
  facturas_proveedores = _analizar_facturas_proveedores(filas_proveedores)
  facturas_clientes = _analizar_facturas_clientes(filas_clientes)

  payload = {
    "empresa_id": empresa_id,
    "tipo": tipo,
    "facturas_proveedores": facturas_proveedores,
    "facturas_clientes": facturas_clientes,
  }

  if incluir_tests:
    import sys

    inter_dir = Path(__file__).resolve().parent
    if str(inter_dir) not in sys.path:
      sys.path.insert(0, str(inter_dir))
    loader = unittest.TestLoader()
    suite = loader.loadTestsFromName("tests.test_logica_pura")
    result = unittest.TestResult()
    suite.run(result)
    payload["unit_tests"] = {
      "ok": result.wasSuccessful(),
      "total": result.testsRun,
      "fallos": [
        {"test": str(t), "error": (err or "").strip()}
        for t, err in result.failures + result.errors
      ],
    }

  return jsonify(payload)


@control_calidad_bp.post("/api/control-calidad/sugerir")
def control_calidad_sugerir():
  """
  Devuelve sugerencias para una fila con errores: heurísticas y opcionalmente LLM.
  Body (JSON): empresa_id (obligatorio), tipo ("proveedores" | "clientes"), fila (dict),
  errores (list[str]). Opcional: ruta_archivo, indice, usar_llm (bool, default False).
  Si usar_llm es true, se envía la factura y los errores al LLM para que sugiera correcciones;
  se combinan con las heurísticas (prioridad a heurísticas en descuadres).
  """
  data = request.get_json(silent=True) or {}
  empresa_id, err = _validar_empresa_id_requerido(data.get("empresa_id"))
  if err:
    return err[0], err[1]
  tipo = (data.get("tipo") or "").strip().lower()
  if tipo not in ("proveedores", "clientes"):
    return _bad_request('tipo debe ser "proveedores" o "clientes"')
  fila = data.get("fila")
  if not isinstance(fila, dict):
    fila = {}
  errores = data.get("errores")
  if not isinstance(errores, list):
    errores = []
  usar_llm = data.get("usar_llm") is True

  sugerencias = _sugerencias_heuristicas(fila, errores, tipo)
  campos_heuristicos = {s["campo"] for s in sugerencias}
  if usar_llm and client:
    llm_sug = _sugerencias_llm(fila, errores, tipo)
    for s in llm_sug:
      if s["campo"] not in campos_heuristicos:
        sugerencias.append(s)
        campos_heuristicos.add(s["campo"])
  return jsonify({"sugerencias": sugerencias})


# ─── Bancos: movimiento de caja unificado ─────────────────────────────────────

def _get_bancos_db():
  """Abre conexión a la base de datos de movimientos de caja."""
  BANCOS_DIR.mkdir(parents=True, exist_ok=True)
  return sqlite3.connect(MOVIMIENTOS_DB)


_movimientos_db_initialized = False


def _init_movimientos_db():
  """Crea la tabla movimientos si no existe y añade columnas de conciliación si faltan. No-op tras la primera llamada."""
  global _movimientos_db_initialized
  if _movimientos_db_initialized:
    return
  conn = _get_bancos_db()
  try:
    conn.execute("""
      CREATE TABLE IF NOT EXISTS movimientos (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        fecha_operacion TEXT NOT NULL,
        fecha_valor TEXT,
        concepto TEXT,
        importe REAL NOT NULL,
        divisa TEXT DEFAULT 'EUR',
        saldo REAL,
        banco TEXT NOT NULL,
        codigo TEXT,
        numero_documento TEXT,
        referencia_1 TEXT,
        referencia_2 TEXT,
        empresa_id TEXT,
        hash_dedup TEXT,
        created_at TEXT NOT NULL
      )
    """)
    conn.execute("CREATE INDEX IF NOT EXISTS ix_movimientos_banco ON movimientos(banco)")
    conn.execute("CREATE INDEX IF NOT EXISTS ix_movimientos_fecha ON movimientos(fecha_operacion)")
    conn.execute("CREATE INDEX IF NOT EXISTS ix_movimientos_empresa ON movimientos(empresa_id)")
    conn.execute("CREATE UNIQUE INDEX IF NOT EXISTS ix_movimientos_hash_dedup ON movimientos(hash_dedup) WHERE hash_dedup IS NOT NULL")
    # Migración G.3 / G.9: columnas de conciliación
    cur = conn.execute("PRAGMA table_info(movimientos)")
    columnas_existentes = {row[1] for row in cur.fetchall()}
    for col, sql_type in [
      ("factura_proveedor_id", "INTEGER"),
      ("factura_cliente_id", "INTEGER"),
      ("factura_cliente_key", "TEXT"),
      ("conciliado_at", "TEXT"),
      # G.9: vínculo movimiento ↔ liquidación tarjeta
      ("tarjeta_id", "INTEGER"),
      ("liquidacion_periodo", "TEXT"),
      # Seguros: vínculo movimiento ↔ póliza de seguro
      ("seguro_poliza_id", "INTEGER"),
      # Albaranes: vínculo movimiento ↔ albaranes
      ("albaran_ids", "TEXT"),
    ]:
      if col not in columnas_existentes:
        conn.execute(f"ALTER TABLE movimientos ADD COLUMN {col} {sql_type}")
    conn.execute("CREATE INDEX IF NOT EXISTS ix_movimientos_factura_proveedor ON movimientos(factura_proveedor_id)")
    # Backfill: assign empresa_id to movimientos that don't have one
    conn.execute("UPDATE movimientos SET empresa_id = 'hincado_directo' WHERE empresa_id IS NULL OR empresa_id = ''")
    conn.commit()
  finally:
    conn.close()
  _movimientos_db_initialized = True


def _hash_dedup(banco: str, fecha_operacion: str, importe: float, concepto: str) -> str:
  """Genera un hash para detectar movimientos duplicados."""
  raw = f"{banco}|{fecha_operacion}|{importe}|{(concepto or '').strip()}"
  return hashlib.sha256(raw.encode("utf-8")).hexdigest()


def _normalizar_fecha_dd_mm_yyyy(val) -> str | None:
  """Convierte fecha dd/mm/yyyy, datetime o número Excel (serial) a yyyy-mm-dd."""
  if val is None:
    return None
  if hasattr(val, "strftime"):
    return val.strftime("%Y-%m-%d")
  if isinstance(val, (int, float)) and 1000 < val < 100000:
    try:
      from datetime import datetime, timedelta
      base = datetime(1899, 12, 30)
      dt = base + timedelta(days=float(val))
      return dt.strftime("%Y-%m-%d")
    except Exception as e:
      logger.debug("No se pudo convertir serial Excel %s a fecha: %s", val, e)
  s = str(val).strip()
  if not s:
    return None
  parts = re.split(r"[/\-.\s]+", s)
  if len(parts) == 3:
    try:
      d, m, y = int(parts[0]), int(parts[1]), int(parts[2])
      if y < 100:
        y += 2000
      return f"{y:04d}-{m:02d}-{d:02d}"
    except (ValueError, TypeError):
      pass
  return s


def _normalizar_nombre_columna(s: str) -> str:
  """Normaliza nombre de columna para búsqueda: minúsculas, sin acentos, espacios a _."""
  if not s or not isinstance(s, str):
    return ""
  s = s.strip().lower()
  s = unicodedata.normalize("NFD", s)
  s = "".join(c for c in s if unicodedata.category(c) != "Mn")
  return s.replace(" ", "_").replace("-", "_")


def _buscar_fila_cabecera_santander(filas: list) -> tuple[int, dict[str, int]]:
  """
  Busca en una lista de filas (cada fila es lista de celdas) la que tiene cabeceras
  (Fecha/Concepto/Importe). Devuelve (índice_0based + 1 como fila 1-based, mapa).
  """
  for row_num in range(min(50, len(filas))):
    row = filas[row_num] if row_num < len(filas) else []
    if not row:
      continue
    row = list(row) if row else []
    mapa = {}
    for idx, cell in enumerate(row):
      nombre = _normalizar_nombre_columna(str(cell) if cell is not None else "")
      if not nombre:
        continue
      if ("operacion" in nombre or ("fecha" in nombre and "valor" not in nombre)) and "fecha_operacion" not in mapa:
        mapa["fecha_operacion"] = idx
      elif "fecha" in nombre and "valor" in nombre:
        mapa["fecha_valor"] = idx
      elif "concepto" in nombre:
        mapa["concepto"] = idx
      elif "importe" in nombre and "saldo" not in nombre:
        mapa["importe"] = idx
      elif "divisa" in nombre and mapa.get("divisa") is None:
        mapa["divisa"] = idx
      elif "saldo" in nombre and "importe" not in nombre:
        mapa["saldo"] = idx
      elif "codigo" in nombre:
        mapa["codigo"] = idx
      elif "numero" in nombre and "documento" in nombre:
        mapa["numero_documento"] = idx
      elif "referencia" in nombre and "1" in nombre:
        mapa["referencia_1"] = idx
      elif "referencia" in nombre and "2" in nombre:
        mapa["referencia_2"] = idx
    if "importe" in mapa and ("fecha_operacion" in mapa or "concepto" in mapa):
      if "fecha_operacion" not in mapa and "fecha_valor" in mapa:
        mapa["fecha_operacion"] = mapa["fecha_valor"]
      if "fecha_operacion" not in mapa and len(row) > 0:
        mapa["fecha_operacion"] = 0
      return row_num + 1, mapa  # 1-based row number
  return 0, {}


def _parse_santander_excel(stream) -> list[dict]:
  """
  Lee un Excel de extracto Santander. Carga la hoja en memoria para poder detectar
  cabecera y plan B sobre los mismos datos. Primero detecta por nombre de columna;
  si falla, usa layout fijo validando fecha e importe en la primera fila de datos.
  """
  try:
    import openpyxl
  except ImportError:
    raise RuntimeError("openpyxl no instalado. Ejecuta: pip install openpyxl")
  wb = openpyxl.load_workbook(stream, read_only=False, data_only=True)
  sheet_name = None
  for name in wb.sheetnames:
    if "movimiento" in name.lower():
      sheet_name = name
      break
  if not sheet_name:
    sheet_name = "movimientos" if "movimientos" in wb.sheetnames else (wb.sheetnames[0] if wb.sheetnames else None)
  if not sheet_name:
    wb.close()
    raise ValueError("No se encontró la hoja 'movimientos' en el Excel")
  ws = wb[sheet_name]
  filas_completas = list(ws.iter_rows(min_row=1, max_row=50000, values_only=True))
  wb.close()

  header_row, col_map = _buscar_fila_cabecera_santander(filas_completas)

  # Plan B: layout fijo. Buscar primera fila que tenga fecha válida en col 0 o 2 e importe numérico en col 3 o 5
  if not header_row or not col_map:
    layouts = [
      {"fecha_operacion": 0, "fecha_valor": 1, "concepto": 2, "importe": 3, "divisa": 4, "saldo": 5, "codigo": 7, "numero_documento": 8, "referencia_1": 9, "referencia_2": 10},
      {"fecha_operacion": 2, "fecha_valor": 3, "concepto": 4, "importe": 5, "divisa": 6, "saldo": 7, "codigo": 9, "numero_documento": 10, "referencia_1": 11, "referencia_2": 12},
    ]
    for layout_fijo in layouts:
      idx_fecha = layout_fijo["fecha_operacion"]
      idx_fecha_valor = layout_fijo.get("fecha_valor", idx_fecha + 1)
      idx_importe = layout_fijo["importe"]
      for candidato in range(6, min(14, len(filas_completas))):
        # candidato = fila 1-based de la cabecera; datos en filas candidato+1, candidato+2, ...
        for di in range(1, 7):
          i = candidato - 1 + di
          if i >= len(filas_completas):
            break
          fila = filas_completas[i]
          if not fila:
            continue
          r = list(fila)
          if len(r) <= max(idx_fecha, idx_fecha_valor, idx_importe):
            continue
          fecha_ok = _normalizar_fecha_dd_mm_yyyy(r[idx_fecha]) or _normalizar_fecha_dd_mm_yyyy(r[idx_fecha_valor])
          importe_ok = False
          try:
            float(r[idx_importe])
            importe_ok = True
          except (TypeError, ValueError):
            pass
          if fecha_ok and importe_ok:
            header_row = candidato
            col_map = layout_fijo
            break
        if header_row:
          break
      if header_row:
        break

  if not header_row or not col_map:
    raise ValueError(
      "No se encontró la fila de cabecera con Fecha operación, Concepto e Importe en el Excel. "
      "Asegúrate de usar un extracto de Santander con columnas Fecha operación, Concepto e Importe."
    )
  data_start_0 = header_row  # índice 0-based de la primera fila de datos (header está en header_row - 1)
  filas = filas_completas[data_start_0:]
  resultado = []
  for row in filas:
    if row is None or all(cell is None or str(cell).strip() == "" for cell in (row or [])):
      continue
    r = list(row) if row else []
    try:
      def get(col: str):
        idx = col_map.get(col)
        if idx is None or idx >= len(r):
          return None
        return r[idx]
      fecha_op = _normalizar_fecha_dd_mm_yyyy(get("fecha_operacion"))
      fecha_valor = _normalizar_fecha_dd_mm_yyyy(get("fecha_valor"))
      concepto = (get("concepto") or "")
      if isinstance(concepto, (int, float)):
        concepto = str(concepto).strip()
      else:
        concepto = str(concepto).strip() if concepto is not None else ""
      if not concepto and not fecha_op:
        continue
      importe_val = get("importe")
      try:
        importe = float(importe_val) if importe_val is not None else 0.0
      except (TypeError, ValueError):
        importe = 0.0
      divisa = (get("divisa") or "EUR")
      if isinstance(divisa, (int, float)):
        divisa = str(divisa).strip()
      else:
        divisa = str(divisa).strip() if divisa else "EUR"
      if not divisa:
        divisa = "EUR"
      saldo_val = get("saldo")
      try:
        saldo = float(saldo_val) if saldo_val is not None else None
      except (TypeError, ValueError):
        saldo = None
      codigo = get("codigo")
      codigo = str(codigo).strip() if codigo is not None else None
      codigo = codigo or None
      num_doc = get("numero_documento")
      num_doc = str(num_doc).strip() if num_doc is not None else None
      num_doc = num_doc or None
      ref1 = get("referencia_1")
      ref1 = str(ref1).strip() if ref1 is not None else None
      ref1 = ref1 or None
      ref2 = get("referencia_2")
      ref2 = str(ref2).strip() if ref2 is not None else None
      ref2 = ref2 or None
      if not fecha_op:
        fecha_op = fecha_valor or ""
      if not fecha_op:
        continue
      if not concepto and importe == 0:
        continue
      resultado.append({
        "fecha_operacion": fecha_op,
        "fecha_valor": fecha_valor,
        "concepto": concepto,
        "importe": importe,
        "divisa": divisa,
        "saldo": saldo,
        "banco": "santander",
        "codigo": codigo,
        "numero_documento": num_doc,
        "referencia_1": ref1,
        "referencia_2": ref2,
      })
    except Exception as e:
      logger.debug("Fila Santander no parseable: %s", e)
      continue
  return resultado


def _parse_bbva_excel(stream) -> list[dict]:
  """
  Lee un Excel de extracto BBVA. Compatible con ambos formatos:
  - Formato antiguo (.xlsx): hoja 'Historico', cabecera fila 16, cols desde índice 2.
    Columnas: F. CONTABLE, F. VALOR, CÓDIGO, CONCEPTO, BENEFICIARIO/ORDENANTE, OBSERVACIONES, IMPORTE, SALDO, DIVISA
  - Formato nuevo (.xls): hoja 'HistoricoMovimientos', cabecera fila 15, cols desde índice 1.
    Columnas: Fecha contable, Fecha valor, Código, Concepto, Observaciones, Importe, Saldo, Divisa, Oficina, Remesa
  Detecta automáticamente el formato por nombre de hoja y cabeceras.
  Devuelve lista de dicts con keys del modelo unificado.
  """
  # --- Leer filas brutas según formato de archivo (.xls vs .xlsx) ---
  raw_bytes = stream.read()
  sheet_names = []
  filas_all = []

  # Intentar xlrd (.xls) primero, luego openpyxl (.xlsx)
  loaded = False
  try:
    import xlrd
    from io import BytesIO
    wb = xlrd.open_workbook(file_contents=raw_bytes)
    sheet_names = wb.sheet_names()
    # Elegir hoja
    target = None
    for candidate in ["HistoricoMovimientos", "Historico"]:
      if candidate in sheet_names:
        target = candidate
        break
    if not target:
      target = sheet_names[0] if sheet_names else None
    if not target:
      raise ValueError("No se encontró ninguna hoja en el Excel")
    ws = wb.sheet_by_name(target)
    for r in range(ws.nrows):
      filas_all.append(ws.row_values(r))
    loaded = True
  except Exception:
    pass

  if not loaded:
    try:
      import openpyxl
      from io import BytesIO
      wb = openpyxl.load_workbook(BytesIO(raw_bytes), read_only=True, data_only=True)
      sheet_names = wb.sheetnames
      target = None
      for candidate in ["HistoricoMovimientos", "Historico"]:
        if candidate in sheet_names:
          target = candidate
          break
      if not target:
        target = sheet_names[0] if sheet_names else None
      if not target:
        raise ValueError("No se encontró ninguna hoja en el Excel")
      ws = wb[target]
      filas_all = [list(row) for row in ws.iter_rows(values_only=True)]
      wb.close()
      loaded = True
    except ImportError:
      raise RuntimeError("Ni xlrd ni openpyxl están instalados. Ejecuta: pip install xlrd openpyxl")

  if not loaded:
    raise ValueError("No se pudo abrir el archivo Excel")

  if not filas_all:
    return []

  # --- Detectar fila de cabecera y formato automáticamente ---
  cabecera_idx = None
  col_map = {}  # nombre_normalizado -> índice

  BBVA_HEADERS = {
    "fecha_contable": ["f._contable", "fecha_contable", "f.contable", "f_contable"],
    "fecha_valor": ["f._valor", "fecha_valor", "f.valor", "f_valor"],
    "codigo": ["codigo", "c_digo"],
    "concepto": ["concepto"],
    "beneficiario": ["beneficiario/ordenante", "beneficiario_ordenante", "beneficiario"],
    "observaciones": ["observaciones"],
    "importe": ["importe"],
    "saldo": ["saldo"],
    "divisa": ["divisa"],
  }

  for idx, fila in enumerate(filas_all[:30]):
    nombres = [_normalizar_nombre_columna(str(c)) if c else "" for c in fila]
    matches = 0
    temp_map = {}
    for campo, aliases in BBVA_HEADERS.items():
      for col_i, nombre in enumerate(nombres):
        if nombre in aliases:
          temp_map[campo] = col_i
          matches += 1
          break
    if matches >= 4:  # al menos fecha_contable, concepto, importe, saldo
      cabecera_idx = idx
      col_map = temp_map
      break

  if cabecera_idx is None:
    raise ValueError("No se encontró la fila de cabecera BBVA en el Excel. Verifica que el archivo tiene las columnas esperadas.")

  tiene_beneficiario = "beneficiario" in col_map

  # --- Parsear datos ---
  resultado = []
  for row in filas_all[cabecera_idx + 1:]:
    if row is None or all(cell is None or str(cell).strip() == "" for cell in (row or [])):
      continue
    try:
      r = list(row) if row else []

      def _val(campo):
        i = col_map.get(campo)
        if i is None or i >= len(r):
          return None
        return r[i]

      fecha_op = _normalizar_fecha_dd_mm_yyyy(_val("fecha_contable"))
      fecha_valor = _normalizar_fecha_dd_mm_yyyy(_val("fecha_valor"))
      codigo = str(_val("codigo")).strip() if _val("codigo") is not None else None
      codigo = codigo or None
      concepto_raw = str(_val("concepto")).strip() if _val("concepto") is not None else ""
      beneficiario = str(_val("beneficiario")).strip() if tiene_beneficiario and _val("beneficiario") else ""
      observaciones = str(_val("observaciones")).strip() if _val("observaciones") else ""
      partes = [p for p in [concepto_raw, beneficiario, observaciones] if p]
      concepto = " | ".join(partes) if partes else ""
      try:
        importe = float(_val("importe")) if _val("importe") is not None else 0.0
      except (TypeError, ValueError):
        importe = 0.0
      try:
        saldo = float(_val("saldo")) if _val("saldo") is not None else None
      except (TypeError, ValueError):
        saldo = None
      divisa_val = _val("divisa")
      divisa = str(divisa_val).strip() if divisa_val else "EUR"
      if not divisa:
        divisa = "EUR"
      if not fecha_op:
        fecha_op = fecha_valor or ""
      if not fecha_op:
        continue
      if not concepto and importe == 0:
        continue
      resultado.append({
        "fecha_operacion": fecha_op,
        "fecha_valor": fecha_valor,
        "concepto": concepto,
        "importe": importe,
        "divisa": divisa,
        "saldo": saldo,
        "banco": "bbva",
        "codigo": codigo,
        "numero_documento": None,
        "referencia_1": None,
        "referencia_2": None,
      })
    except Exception as e:
      logger.debug("Fila BBVA no parseable: %s", e)
      continue
  return resultado


def _insertar_movimientos_lista(movs: list, empresa_id: str | None, omitir_duplicados: bool = True) -> tuple[int, int, list]:
  """
  Inserta una lista de movimientos en la BD. Devuelve (insertados, duplicados_omitidos, errores).
  """
  now = datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%SZ")
  conn = _get_bancos_db()
  inserted = 0
  duplicados = 0
  errores = []
  try:
    for i, m in enumerate(movs):
      if not isinstance(m, dict):
        errores.append({"indice": i, "error": "Cada elemento debe ser un objeto"})
        continue
      fecha_op = (m.get("fecha_operacion") or "").strip()
      concepto = (m.get("concepto") or "").strip()
      try:
        importe = float(m.get("importe"))
      except (TypeError, ValueError):
        errores.append({"indice": i, "error": "importe inválido"})
        continue
      banco = (m.get("banco") or "santander").strip().lower()
      if not fecha_op or not banco:
        errores.append({"indice": i, "error": "fecha_operacion y banco son obligatorios"})
        continue
      if not concepto and importe == 0:
        continue
      fecha_valor = (m.get("fecha_valor") or "").strip() or None
      divisa = (m.get("divisa") or "EUR").strip() or "EUR"
      saldo = m.get("saldo")
      if saldo is not None:
        try:
          saldo = float(saldo)
        except (TypeError, ValueError):
          saldo = None
      codigo = (m.get("codigo") or "").strip() or None
      num_doc = (m.get("numero_documento") or "").strip() or None
      ref1 = (m.get("referencia_1") or "").strip() or None
      ref2 = (m.get("referencia_2") or "").strip() or None
      hash_dedup = _hash_dedup(banco, fecha_op, importe, concepto)
      if omitir_duplicados:
        existe = conn.execute(
          "SELECT 1 FROM movimientos WHERE hash_dedup = ? LIMIT 1",
          (hash_dedup,),
        ).fetchone()
        if existe:
          duplicados += 1
          continue
      conn.execute(
        """
        INSERT INTO movimientos (
          fecha_operacion, fecha_valor, concepto, importe, divisa, saldo,
          banco, codigo, numero_documento, referencia_1, referencia_2, empresa_id, hash_dedup, created_at
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
          fecha_op, fecha_valor, concepto, importe, divisa, saldo,
          banco, codigo, num_doc, ref1, ref2, empresa_id, hash_dedup, now,
        ),
      )
      inserted += 1
    conn.commit()
  finally:
    conn.close()
  return inserted, duplicados, errores


@bancos_bp.route("/api/bancos/movimientos", methods=["GET"])
def listar_movimientos():
  """
  Lista movimientos de caja con filtros opcionales.
  Query: banco, fecha_desde, fecha_hasta, empresa_id, limit, offset.
  Incluye saldo_acumulado: suma del saldo de todos los bancos tras cada movimiento (orden cronológico).
  """
  _init_movimientos_db()
  banco = request.args.get("banco", "").strip() or None
  fecha_desde = request.args.get("fecha_desde", "").strip() or None
  fecha_hasta = request.args.get("fecha_hasta", "").strip() or None
  empresa_id = request.args.get("empresa_id", "").strip() or None
  concepto = request.args.get("concepto", "").strip() or None
  limit = request.args.get("limit", type=int) or 500
  offset = request.args.get("offset", type=int) or 0
  limit = min(max(1, limit), 5000)
  offset = max(0, offset)

  conn = _get_bancos_db()
  try:
    conditions = []
    params = []
    if banco:
      conditions.append("banco = ?")
      params.append(banco.lower())
    if fecha_desde:
      conditions.append("fecha_operacion >= ?")
      params.append(fecha_desde)
    if fecha_hasta:
      conditions.append("fecha_operacion <= ?")
      params.append(fecha_hasta)
    if empresa_id:
      conditions.append("empresa_id = ?")
      params.append(empresa_id)
    if concepto:
      conditions.append("(concepto IS NOT NULL AND concepto LIKE ?)")
      params.append("%" + concepto + "%")
    where = (" WHERE " + " AND ".join(conditions)) if conditions else ""
    params_count = params.copy()
    params.extend([limit, offset])
    rows = conn.execute(
      f"""
      SELECT id, fecha_operacion, fecha_valor, concepto, importe, divisa, saldo,
             banco, codigo, numero_documento, referencia_1, referencia_2, empresa_id, created_at,
             factura_proveedor_id, factura_cliente_id, factura_cliente_key, conciliado_at,
             tarjeta_id, liquidacion_periodo
      FROM movimientos
      {where}
      ORDER BY fecha_operacion DESC, id DESC
      LIMIT ? OFFSET ?
      """,
      params,
    ).fetchall()
    total = conn.execute(
      f"SELECT COUNT(*) FROM movimientos {where}",
      params_count,
    ).fetchone()[0]
    keys = [
      "id", "fecha_operacion", "fecha_valor", "concepto", "importe", "divisa", "saldo",
      "banco", "codigo", "numero_documento", "referencia_1", "referencia_2", "empresa_id", "created_at",
      "factura_proveedor_id", "factura_cliente_id", "factura_cliente_key", "conciliado_at",
      "tarjeta_id", "liquidacion_periodo",
    ]
    movimientos = [dict(zip(keys, r)) for r in rows]

    # Enriquecer con ruta de factura enlazada (para botón "Ver factura")
    factura_ids = [m["factura_proveedor_id"] for m in movimientos if m.get("factura_proveedor_id")]
    if factura_ids:
      conn_gest = sqlite3.connect(str(GESTION_DB))
      try:
        placeholders = ",".join("?" * len(factura_ids))
        cur_gest = conn_gest.execute(
          f"SELECT id, ruta_destino, ruta_archivo, numero_factura, proveedor FROM facturas_proveedor WHERE id IN ({placeholders})",
          factura_ids,
        )
        factura_map = {}
        for row in cur_gest.fetchall():
          ruta = (row[1] or row[2] or "").strip() or ""
          factura_map[row[0]] = {
            "factura_ruta": ruta,
            "factura_numero": (row[3] or "").strip(),
            "factura_proveedor_nombre": (row[4] or "").strip(),
          }
        for m in movimientos:
          fid = m.get("factura_proveedor_id")
          if fid and fid in factura_map:
            m.update(factura_map[fid])
      finally:
        conn_gest.close()

    # Enriquecer con ruta de factura cliente cuando está vinculado por factura_cliente_key
    movs_con_cliente = [m for m in movimientos if (m.get("factura_cliente_key") or "").strip()]
    if movs_con_cliente:
      for m in movs_con_cliente:
        empresa_id_m = (m.get("empresa_id") or "").strip()
        key = (m.get("factura_cliente_key") or "").strip()
        if not empresa_id_m or not key:
          continue
        facturas_cli = facturas_cliente_db.get_facturas_cliente_empresa(empresa_id_m)
        for fc in facturas_cli:
          n = (fc.get("numero_factura") or "").strip()
          f = (fc.get("fecha_factura") or "").strip()[:10]
          c = (fc.get("cliente") or "").strip()
          k = f"{n}|{f}|{c}"
          if k == key:
            ruta = (fc.get("ruta_archivo") or "").strip()
            if ruta:
              m["factura_ruta"] = ruta
            break

    # G.9: enriquecer con alias de tarjeta si está vinculado a liquidación
    tarjeta_ids = [m["tarjeta_id"] for m in movimientos if m.get("tarjeta_id")]
    if tarjeta_ids:
      conn_gest = sqlite3.connect(str(GESTION_DB))
      try:
        placeholders = ",".join("?" * len(tarjeta_ids))
        cur_gest = conn_gest.execute(
          f"SELECT id, alias, banco, persona FROM tarjetas WHERE id IN ({placeholders})",
          tarjeta_ids,
        )
        tarjeta_map = {}
        for row in cur_gest.fetchall():
          alias = (row[1] or "").strip() or ((row[2] or "").strip() + " " + (row[3] or "").strip()).strip() or "Tarjeta"
          tarjeta_map[row[0]] = {"tarjeta_alias": alias}
        for m in movimientos:
          tid = m.get("tarjeta_id")
          if tid and tid in tarjeta_map:
            m.update(tarjeta_map[tid])
      finally:
        conn_gest.close()

    # Enrich MULTI-conciliated movements with invoice count from conciliacion_multiple
    multi_mov_ids = [m["id"] for m in movimientos if (m.get("factura_cliente_key") or "") == "MULTI"]
    if multi_mov_ids:
      conn_gest = sqlite3.connect(str(GESTION_DB))
      conn_gest.row_factory = sqlite3.Row
      try:
        placeholders = ",".join("?" * len(multi_mov_ids))
        rows_cm = conn_gest.execute(
          f"SELECT movimiento_id, COUNT(*) as n, SUM(importe_aplicado) as total"
          f" FROM conciliacion_multiple WHERE movimiento_id IN ({placeholders}) GROUP BY movimiento_id",
          multi_mov_ids,
        ).fetchall()
        cm_map = {r["movimiento_id"]: {"multi_n_facturas": r["n"], "multi_total": r["total"]} for r in rows_cm}
        for m in movimientos:
          if m["id"] in cm_map:
            m.update(cm_map[m["id"]])
      except Exception:
        pass
      finally:
        conn_gest.close()

    # Calcular saldo acumulado
    cond_acum = []
    params_acum = []
    if fecha_desde:
      cond_acum.append("fecha_operacion >= ?")
      params_acum.append(fecha_desde)
    if fecha_hasta:
      cond_acum.append("fecha_operacion <= ?")
      params_acum.append(fecha_hasta)
    if empresa_id:
      cond_acum.append("empresa_id = ?")
      params_acum.append(empresa_id)
    where_acum = (" WHERE " + " AND ".join(cond_acum)) if cond_acum else ""
    params_acum.append(5000)
    rows_acum = conn.execute(
      f"""
      SELECT id, fecha_operacion, banco, saldo
      FROM movimientos
      {where_acum}
      ORDER BY fecha_operacion ASC, id ASC
      LIMIT ?
      """,
      params_acum,
    ).fetchall()
    last_saldo_by_bank = {}
    saldo_acumulado_by_id = {}
    for r in rows_acum:
      id_, fecha_op, banco_, saldo_ = r[0], r[1], r[2], r[3]
      if saldo_ is not None:
        try:
          last_saldo_by_bank[banco_] = float(saldo_)
        except (TypeError, ValueError):
          pass
      saldo_acumulado_by_id[id_] = sum(last_saldo_by_bank.values())
    for m in movimientos:
      m["saldo_acumulado"] = saldo_acumulado_by_id.get(m["id"])

    return jsonify({"movimientos": movimientos, "total": total})
  finally:
    conn.close()


# ─── G.3 Conciliación bancaria: sugerencias y confirmación ───────────────────

def _conciliacion_umbral_default() -> float:
  """Umbral en euros para considerar match importe movimiento vs factura (diferencias de céntimos)."""
  return 0.50


def _conciliacion_dias_ventana() -> int:
  """Ventana en días para considerar factura \"cerca\" de la fecha del movimiento (4 meses)."""
  return 4 * 30  # 4 meses


def _similitud_texto_concilacion(concepto_movimiento: str, factura: dict) -> float:
  """
  Devuelve una puntuación de similitud en [0, 1] entre el concepto del movimiento bancario
  y el texto de la factura (proveedor + resumen_concepto). Usa difflib para tolerar variaciones.
  """
  a = (concepto_movimiento or "").strip().lower()
  proveedor = (factura.get("proveedor") or "").strip()
  resumen = (factura.get("resumen_concepto") or "").strip()
  b = (proveedor + " " + resumen).strip().lower()
  if not a and not b:
    return 1.0
  if not a or not b:
    return 0.0
  a = " ".join(a.split())
  b = " ".join(b.split())
  return difflib.SequenceMatcher(None, a, b).ratio()


@bancos_bp.route("/api/bancos/conciliacion/sugerencias", methods=["GET"])
def conciliacion_sugerencias():
  """
  Devuelve sugerencias de conciliación: movimientos sin conciliar vs facturas de proveedor pendientes.

  Cálculo:
  - Se toman movimientos sin factura, sin conciliado_at y sin tarjeta (límite 1000), opcionalmente
    filtrados por fecha_desde/fecha_hasta (si no se envían, se consideran todos).
  - Se excluyen movimientos cuyo concepto contenga "Nomina"/"Nómina", "Adelanto", "Liquidacion De Las Tarjetas",
    "liquidacion tarjeta", "liquidacion contrato", "recibo mensual tarjeta", "recibo tarjeta",
    "adeudo mensual...tarjeta", "pago tarjeta", "cargo tarjeta".
  - Solo se consideran movimientos con importe negativo (pagos).
  - Para cada movimiento se busca una factura pendiente cuyo total coincida con |importe| ± umbral (default 0,50 €)
    y cuya fecha esté dentro de una ventana de días respecto a la fecha del movimiento (default 365 días).
  - Entre las facturas que cumplen importe y fecha, se elige la de mayor similitud de texto entre concepto
    del movimiento y concepto de factura (proveedor + resumen_concepto). Cada movimiento genera como máximo una sugerencia.
  - Las sugerencias se ordenan por similitud de texto (mayor primero).
  - Se devuelven paginadas (por defecto 10 por página). Parámetros: page (default 1), per_page (default 10, máx. 50).

  Query: empresa_id (obligatorio), fecha_desde, fecha_hasta (opcionales), umbral (opcional), page, per_page.
  """
  empresa_id = (request.args.get("empresa_id") or "").strip()
  if not empresa_id:
    return _bad_request("Falta empresa_id")
  fecha_desde = (request.args.get("fecha_desde") or "").strip() or None
  fecha_hasta = (request.args.get("fecha_hasta") or "").strip() or None
  try:
    umbral = float(request.args.get("umbral", _conciliacion_umbral_default()))
  except (TypeError, ValueError):
    umbral = _conciliacion_umbral_default()
  umbral = max(0, min(umbral, 100))
  try:
    page = max(1, int(request.args.get("page", 1)))
  except (TypeError, ValueError):
    page = 1
  try:
    per_page = min(50, max(1, int(request.args.get("per_page", 10))))
  except (TypeError, ValueError):
    per_page = 10

  _init_movimientos_db()
  facturas_db.init_facturas_db()

  conn_bancos = _get_bancos_db()
  try:
    cond_mov = ["(factura_proveedor_id IS NULL AND (conciliado_at IS NULL OR conciliado_at = ''))"]
    params_mov: list = []
    # Excluir movimientos ya vinculados a extracto de tarjeta (no son candidatos a conciliación con factura)
    cond_mov.append("(tarjeta_id IS NULL OR tarjeta_id = 0)")
    # Excluir nóminas, adelantos y movimientos de tarjeta (recibos, liquidaciones, adeudos, etc.)
    cond_mov.append("""(
      LOWER(COALESCE(concepto, '')) NOT LIKE '%nomina%'
      AND LOWER(COALESCE(concepto, '')) NOT LIKE '%nómina%'
      AND LOWER(COALESCE(concepto, '')) NOT LIKE '%adelanto%'
      AND LOWER(COALESCE(concepto, '')) NOT LIKE '%liquidacion de las tarjetas%'
      AND LOWER(COALESCE(concepto, '')) NOT LIKE '%liquidacion tarjeta%'
      AND LOWER(COALESCE(concepto, '')) NOT LIKE '%liquidacion contrato%'
      AND LOWER(COALESCE(concepto, '')) NOT LIKE '%recibo mensual tarjeta%'
      AND LOWER(COALESCE(concepto, '')) NOT LIKE '%recibo tarjeta%'
      AND LOWER(COALESCE(concepto, '')) NOT LIKE '%adeudo mensual%tarjeta%'
      AND LOWER(COALESCE(concepto, '')) NOT LIKE '%pago tarjeta%'
      AND LOWER(COALESCE(concepto, '')) NOT LIKE '%cargo tarjeta%'
    )""")
    if fecha_desde:
      cond_mov.append("fecha_operacion >= ?")
      params_mov.append(fecha_desde)
    if fecha_hasta:
      cond_mov.append("fecha_operacion <= ?")
      params_mov.append(fecha_hasta)
    cond_mov.append("empresa_id = ?")
    params_mov.append(empresa_id)
    where_mov = " AND ".join(cond_mov)
    movimientos = conn_bancos.execute(
      f"""
      SELECT id, fecha_operacion, concepto, importe, empresa_id
      FROM movimientos
      WHERE {where_mov}
      ORDER BY fecha_operacion DESC
      LIMIT 1000
      """,
      params_mov,
    ).fetchall()
  finally:
    conn_bancos.close()

  # Total ya conciliado por factura (varios movimientos pueden apuntar a la misma factura = pagos parciales)
  conn_bancos = _get_bancos_db()
  try:
    rows_pagado = conn_bancos.execute(
      """
      SELECT factura_proveedor_id, SUM(ABS(CAST(importe AS REAL)))
      FROM movimientos
      WHERE factura_proveedor_id IS NOT NULL
      GROUP BY factura_proveedor_id
      """
    ).fetchall()
    total_pagado_por_factura = {int(r[0]): float(r[1] or 0) for r in rows_pagado}
  finally:
    conn_bancos.close()

  facturas_pendientes = [
    f for f in facturas_db.get_facturas_empresa(empresa_id)
    if (f.get("estado_pago") or "").strip().lower() in ("pendiente", "parcial")
  ]

  def importe_factura(f: dict) -> float:
    for key in ("total_a_pagar", "total_factura", "total"):
      v = f.get(key)
      if v is None or v == "":
        continue
      try:
        s = str(v).strip().replace(",", ".")
        return float(s)
      except (ValueError, TypeError):
        continue
    return 0.0

  def parse_fecha(s: str | None):
    if not s or len(str(s).strip()) < 10:
      return None
    try:
      return datetime.strptime(str(s).strip()[:10], "%Y-%m-%d")
    except Exception as e:
      logger.debug("Fecha no parseable en conciliación '%s': %s", s, e)
      return None

  sugerencias: list[dict] = []
  dias_ventana = _conciliacion_dias_ventana()
  for m in movimientos:
    mov_id, fecha_op, concepto, importe_mov, emp_id = m[0], m[1], m[2], m[3], m[4]
    try:
      imp_mov = float(importe_mov) if importe_mov is not None else 0.0
    except (TypeError, ValueError):
      imp_mov = 0.0
    if imp_mov >= 0:
      continue
    abs_mov = abs(imp_mov)
    fecha_mov = parse_fecha(fecha_op)
    concepto_mov_str = (concepto or "").strip()
    candidatos: list[tuple[dict, float]] = []
    for f in facturas_pendientes:
      imp_fac = importe_factura(f)
      if imp_fac <= 0:
        continue
      fid = f.get("id")
      total_pagado = total_pagado_por_factura.get(fid, 0.0)
      remaining = imp_fac - total_pagado
      if remaining < 0.01:
        continue
      if abs_mov > remaining + umbral:
        continue
      fecha_fac = parse_fecha(f.get("fecha_factura"))
      if fecha_mov and fecha_fac:
        delta = (fecha_mov - fecha_fac).days
        if abs(delta) > dias_ventana:
          continue
      diferencia = abs(abs_mov - remaining) if abs_mov <= remaining else abs_mov - remaining
      es_parcial = abs_mov < remaining - 0.01 or total_pagado > 0.01
      similitud = _similitud_texto_concilacion(concepto_mov_str, f)
      sug = {
        "movimiento_id": mov_id,
        "movimiento_fecha": fecha_op,
        "movimiento_concepto": concepto_mov_str,
        "movimiento_importe": imp_mov,
        "factura_id": fid,
        "factura_numero": (f.get("numero_factura") or "").strip() or "—",
        "factura_proveedor": (f.get("proveedor") or "").strip() or "—",
        "factura_resumen_concepto": (f.get("resumen_concepto") or "").strip() or None,
        "factura_total": imp_fac,
        "factura_total_pagado": round(total_pagado, 2),
        "factura_remaining": round(remaining, 2),
        "factura_fecha": (f.get("fecha_factura") or "").strip()[:10] if f.get("fecha_factura") else None,
        "factura_ruta": (f.get("ruta_destino") or f.get("ruta_archivo") or "").strip() or None,
        "diferencia": round(diferencia, 2),
        "es_parcial": es_parcial,
        "similitud_texto": round(similitud, 2),
      }
      candidatos.append((sug, similitud))
    if candidatos:
      mejor = max(candidatos, key=lambda x: x[1])
      sugerencias.append(mejor[0])
  sugerencias.sort(key=lambda s: s["similitud_texto"], reverse=True)
  total_sugerencias = len(sugerencias)
  total_paginas = (total_sugerencias + per_page - 1) // per_page if total_sugerencias else 0
  inicio = (page - 1) * per_page
  sugerencias_pagina = sugerencias[inicio:inicio + per_page]
  return jsonify({
    "sugerencias": sugerencias_pagina,
    "total_sugerencias": total_sugerencias,
    "pagina_actual": page,
    "total_paginas": total_paginas,
    "per_page": per_page,
    "movimientos_sin_conciliar": len(movimientos),
    "facturas_pendientes": len(facturas_pendientes),
  })


@bancos_bp.route("/api/bancos/conciliacion/confirmar", methods=["POST"])
def conciliacion_confirmar():
  """
  Vincula un movimiento a una factura de proveedor. Si la suma de movimientos ya vinculados
  a esa factura (incluido este) alcanza o supera el total, marca la factura como pagada; si no, como parcial.
  Body: { "movimiento_id": int, "factura_proveedor_id": int }.
  Validamos factura y empresa ANTES de actualizar el movimiento para no dejar la factura sin actualizar estado_pago.
  """
  data = request.get_json(silent=True) or {}
  mov_id = data.get("movimiento_id")
  factura_id = data.get("factura_proveedor_id")
  if mov_id is None or factura_id is None:
    return _bad_request("Falta movimiento_id o factura_proveedor_id")
  try:
    mov_id = int(mov_id)
    factura_id = int(factura_id)
  except (TypeError, ValueError):
    return _bad_request("movimiento_id y factura_proveedor_id deben ser números")

  _init_movimientos_db()
  facturas_db.init_facturas_db()

  conn_bancos = _get_bancos_db()
  try:
    row = conn_bancos.execute(
      "SELECT id, empresa_id FROM movimientos WHERE id = ?",
      (mov_id,),
    ).fetchone()
  finally:
    conn_bancos.close()
  if not row:
    return jsonify({"error": "Movimiento no encontrado"}), 404
  mov_empresa = (row[1] or "").strip()

  conn_gest = sqlite3.connect(str(GESTION_DB))
  try:
    cur = conn_gest.execute(
      "SELECT id, empresa_id, total_a_pagar, total_factura, total FROM facturas_proveedor WHERE id = ?",
      (factura_id,),
    ).fetchone()
  finally:
    conn_gest.close()
  if not cur:
    return jsonify({"error": "Factura no encontrada"}), 404
  factura_empresa = (cur[1] or "").strip()
  if mov_empresa and mov_empresa != factura_empresa:
    return jsonify({"error": "El movimiento no pertenece a la empresa de la factura"}), 400
  empresa_id = factura_empresa or mov_empresa
  if not empresa_id:
    return jsonify({"error": "Factura sin empresa asignada"}), 400

  total_fac = 0.0
  for v in (cur[2], cur[3], cur[4]):
    if v is not None and str(v).strip():
      try:
        total_fac = float(str(v).strip().replace(",", "."))
        break
      except (ValueError, TypeError):
        pass

  now = datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%SZ")
  conn_bancos = _get_bancos_db()
  try:
    if mov_empresa:
      conn_bancos.execute(
        "UPDATE movimientos SET factura_proveedor_id = ?, conciliado_at = ? WHERE id = ?",
        (factura_id, now, mov_id),
      )
    else:
      conn_bancos.execute(
        "UPDATE movimientos SET factura_proveedor_id = ?, conciliado_at = ?, empresa_id = ? WHERE id = ?",
        (factura_id, now, empresa_id, mov_id),
      )
    conn_bancos.commit()
    row_sum = conn_bancos.execute(
      "SELECT COALESCE(SUM(ABS(CAST(importe AS REAL))), 0) FROM movimientos WHERE factura_proveedor_id = ?",
      (factura_id,),
    ).fetchone()
    total_pagado = float(row_sum[0] or 0)
  finally:
    conn_bancos.close()

  estado = "pagada" if total_pagado >= total_fac - 0.02 else "parcial"
  conn_gest = sqlite3.connect(str(GESTION_DB))
  try:
    conn_gest.execute(
      "UPDATE facturas_proveedor SET estado_pago = ? WHERE id = ?",
      (estado, factura_id),
    )
    conn_gest.commit()
  finally:
    conn_gest.close()

  _invalidar_cache_listado_proveedores(empresa_id)
  msg = "Conciliación registrada. Factura marcada como pagada." if estado == "pagada" else "Pago parcial registrado. Factura en estado parcial."
  return jsonify({"ok": True, "mensaje": msg})


def _factura_cliente_key(numero_factura: str, fecha_factura: str, cliente: str) -> str:
  """Clave única para identificar una factura de cliente en el CSV."""
  n = (numero_factura or "").strip()
  f = (fecha_factura or "").strip()[:10]
  c = (cliente or "").strip()
  return f"{n}|{f}|{c}"


def _recalcular_estado_cobro_cliente(factura_cli_id: int, factura: dict | None = None) -> str:
  """Recalcula el estado_cobro de una factura de cliente sumando movimientos vinculados."""
  from routes.helpers import _parse_importe_es
  if factura is None:
    factura = facturas_cliente_db.get_factura_cliente_por_id(factura_cli_id)
  if not factura:
    return "pendiente"

  total_fac = _parse_importe_es(factura.get("total_a_pagar"))

  # Build the composite key for this invoice (for legacy key-only movements)
  fck = _factura_cliente_key(
    factura.get("numero_factura"), factura.get("fecha_factura"), factura.get("cliente")
  )

  # Sum from direct 1:1 conciliation in movimientos.db
  # Match by factura_cliente_id OR by factura_cliente_key (legacy movements)
  _init_movimientos_db()
  conn_bancos = _get_bancos_db()
  try:
    row_sum = conn_bancos.execute(
      "SELECT COALESCE(SUM(ABS(CAST(importe AS REAL))), 0) FROM movimientos"
      " WHERE (factura_cliente_id = ? OR (factura_cliente_key = ? AND factura_cliente_key != 'MULTI'))"
      " AND conciliado_at IS NOT NULL",
      (factura_cli_id, fck),
    ).fetchone()
    total_cobrado = float(row_sum[0] or 0) if row_sum else 0.0
  finally:
    conn_bancos.close()

  # Also sum from conciliacion_multiple in gestion.db
  try:
    conn_gest = sqlite3.connect(str(GESTION_DB))
    row_cm = conn_gest.execute(
      "SELECT COALESCE(SUM(importe_aplicado), 0) FROM conciliacion_multiple WHERE factura_cliente_id = ?",
      (factura_cli_id,),
    ).fetchone()
    total_cobrado += float(row_cm[0] or 0) if row_cm else 0
    conn_gest.close()
  except Exception:
    pass

  if total_cobrado < 0.01:
    estado = "pendiente"
  elif total_fac > 0 and total_cobrado >= total_fac - 1.0:
    estado = "cobrada"
  else:
    estado = "parcial"

  facturas_cliente_db.update_estado_cobro(factura_cli_id, estado)
  return estado


@bancos_bp.route("/api/bancos/conciliacion/confirmar-cliente", methods=["POST"])
def conciliacion_confirmar_cliente():
  """
  Vincula un movimiento (entrada de caja) a una factura emitida a cliente.
  Body: { "movimiento_id": int, "empresa_id": str, "factura_cliente_id": int,
          "numero_factura": str, "fecha_factura": str, "cliente": str }.
  Acepta factura_cliente_id (preferente) o la terna numero_factura+fecha_factura+cliente (retrocompatible).
  """
  data = request.get_json(silent=True) or {}
  mov_id = data.get("movimiento_id")
  empresa_id = (data.get("empresa_id") or "").strip()
  factura_cli_id = data.get("factura_cliente_id")
  numero_factura = (data.get("numero_factura") or "").strip()
  fecha_factura = (data.get("fecha_factura") or "").strip()
  cliente = (data.get("cliente") or "").strip()
  if mov_id is None or not empresa_id:
    return _bad_request("Falta movimiento_id o empresa_id")
  try:
    mov_id = int(mov_id)
  except (TypeError, ValueError):
    return _bad_request("movimiento_id debe ser un número")

  _init_movimientos_db()
  conn_bancos = _get_bancos_db()
  try:
    row = conn_bancos.execute(
      "SELECT id, empresa_id FROM movimientos WHERE id = ?",
      (mov_id,),
    ).fetchone()
    if not row:
      return jsonify({"error": "Movimiento no encontrado"}), 404
    mov_empresa = (row[1] or "").strip()
    if mov_empresa and mov_empresa != empresa_id:
      return jsonify({"error": "El movimiento no pertenece a esa empresa"}), 400
    if not mov_empresa:
      empresa_id_mov = empresa_id
    else:
      empresa_id_mov = mov_empresa
  finally:
    conn_bancos.close()

  # Resolver la factura de cliente por id o por clave compuesta
  factura = None
  if factura_cli_id is not None:
    try:
      factura_cli_id = int(factura_cli_id)
    except (TypeError, ValueError):
      return _bad_request("factura_cliente_id debe ser un número")
    factura = facturas_cliente_db.get_factura_cliente_por_id(factura_cli_id)
    if not factura:
      return jsonify({"error": "Factura de cliente no encontrada"}), 404
    if (factura.get("empresa_id") or "").strip() != empresa_id:
      return jsonify({"error": "La factura no pertenece a esa empresa"}), 400
  else:
    facturas_cli = facturas_cliente_db.get_facturas_cliente_empresa(empresa_id)
    key = _factura_cliente_key(numero_factura, fecha_factura, cliente)
    for f in facturas_cli:
      if _factura_cliente_key(f.get("numero_factura"), f.get("fecha_factura"), f.get("cliente")) == key:
        factura = f
        factura_cli_id = f.get("id")
        break
    if not factura:
      return jsonify({"error": "Factura de cliente no encontrada en la base de la empresa"}), 404

  key = _factura_cliente_key(
    factura.get("numero_factura"), factura.get("fecha_factura"), factura.get("cliente")
  )

  conn_bancos = _get_bancos_db()
  try:
    conn_bancos.execute(
      "UPDATE movimientos SET factura_cliente_key = ?, factura_cliente_id = ?, conciliado_at = ?, empresa_id = ? WHERE id = ?",
      (key, factura_cli_id, datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%SZ"), empresa_id_mov, mov_id),
    )
    conn_bancos.commit()
  finally:
    conn_bancos.close()

  # G.11: recalcular estado_cobro
  estado_cobro = _recalcular_estado_cobro_cliente(factura_cli_id, factura)
  _invalidar_cache_listado_clientes(empresa_id)

  if estado_cobro == "cobrada":
    msg = "Entrada de caja vinculada. Factura marcada como cobrada."
  elif estado_cobro == "parcial":
    msg = "Entrada de caja vinculada. Factura marcada como cobro parcial."
  else:
    msg = "Entrada de caja vinculada a la factura de cliente."
  return jsonify({"ok": True, "mensaje": msg, "estado_cobro": estado_cobro})


@bancos_bp.route("/api/bancos/conciliacion/confirmar-cliente-multiple", methods=["POST"])
def conciliacion_confirmar_cliente_multiple():
  """
  Vincula un movimiento (entrada de caja) a MÚLTIPLES facturas de cliente.
  Body: { "movimiento_id": int, "empresa_id": str,
          "aplicaciones": [{"factura_cliente_id": int, "importe_aplicado": float}, ...] }
  """
  from core.db import conectar as _conectar_gestion
  data = request.get_json(silent=True) or {}
  mov_id = data.get("movimiento_id")
  empresa_id = (data.get("empresa_id") or "").strip()
  aplicaciones = data.get("aplicaciones") or []
  if mov_id is None or not empresa_id or not aplicaciones:
    return _bad_request("Faltan movimiento_id, empresa_id o aplicaciones")
  try:
    mov_id = int(mov_id)
  except (TypeError, ValueError):
    return _bad_request("movimiento_id inválido")

  # Validate movement
  _init_movimientos_db()
  conn_bancos = _get_bancos_db()
  try:
    mov = conn_bancos.execute("SELECT id, importe, empresa_id FROM movimientos WHERE id = ?", (mov_id,)).fetchone()
    if not mov:
      return jsonify({"error": "Movimiento no encontrado"}), 404
    mov_importe = abs(float(mov[1] or 0))
  finally:
    conn_bancos.close()

  # Validate total applied <= movement amount
  total_aplicado = sum(float(a.get("importe_aplicado") or 0) for a in aplicaciones)
  if total_aplicado > mov_importe + 0.02:
    return _bad_request(f"Total aplicado ({total_aplicado:.2f}) excede el importe del movimiento ({mov_importe:.2f})")

  # Create table if needed
  with _conectar_gestion() as conn_g:
    conn_g.execute("""
      CREATE TABLE IF NOT EXISTS conciliacion_multiple (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        movimiento_id INTEGER NOT NULL,
        movimiento_fecha TEXT,
        movimiento_importe REAL,
        factura_cliente_id INTEGER NOT NULL,
        importe_aplicado REAL NOT NULL,
        created_at TEXT NOT NULL
      )
    """)

  # Mark movement as reconciled in movimientos.db
  ahora = datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%SZ")
  conn_bancos = _get_bancos_db()
  try:
    # Get movement date for the record
    mov_row = conn_bancos.execute("SELECT fecha_operacion, importe FROM movimientos WHERE id = ?", (mov_id,)).fetchone()
    mov_fecha = mov_row[0] if mov_row else ""
    mov_imp = float(mov_row[1] or 0) if mov_row else 0

    conn_bancos.execute(
      "UPDATE movimientos SET factura_cliente_key = 'MULTI', factura_cliente_id = -1,"
      " conciliado_at = ?, empresa_id = ? WHERE id = ?",
      (ahora, empresa_id, mov_id),
    )
    conn_bancos.commit()
  finally:
    conn_bancos.close()

  # Insert records and update each invoice
  resultados = []
  with _conectar_gestion() as conn_g:
    for ap in aplicaciones:
      fid = int(ap["factura_cliente_id"])
      imp = float(ap["importe_aplicado"])
      if imp <= 0:
        continue
      conn_g.execute(
        "INSERT INTO conciliacion_multiple (movimiento_id, movimiento_fecha, movimiento_importe, factura_cliente_id, importe_aplicado, created_at)"
        " VALUES (?, ?, ?, ?, ?, ?)",
        (mov_id, mov_fecha, mov_imp, fid, imp, ahora),
      )

  # Recalculate estado_cobro for each invoice
  for ap in aplicaciones:
    fid = int(ap["factura_cliente_id"])
    imp = float(ap["importe_aplicado"])
    if imp <= 0:
      continue
    factura = facturas_cliente_db.get_factura_cliente_por_id(fid)
    if not factura:
      continue

    # Get total_a_pagar
    total_fac = 0.0
    v = factura.get("total_a_pagar")
    if v is not None and str(v).strip():
      try:
        total_fac = abs(float(str(v).strip().replace(".", "").replace(",", ".")))
      except (ValueError, TypeError):
        pass

    # Sum all applied amounts for this invoice (from conciliacion_multiple)
    with _conectar_gestion() as conn_g:
      row = conn_g.execute(
        "SELECT COALESCE(SUM(importe_aplicado), 0) FROM conciliacion_multiple WHERE factura_cliente_id = ?",
        (fid,),
      ).fetchone()
      total_aplicado_factura = float(row[0] or 0)

    # Also add amounts from direct 1:1 conciliation in movimientos.db
    conn_bancos = _get_bancos_db()
    try:
      row = conn_bancos.execute(
        "SELECT COALESCE(SUM(ABS(CAST(importe AS REAL))), 0) FROM movimientos WHERE factura_cliente_id = ?",
        (fid,),
      ).fetchone()
      total_directo = float(row[0] or 0) if row else 0
    finally:
      conn_bancos.close()

    total_cobrado = total_aplicado_factura + total_directo

    if total_cobrado < 0.01:
      estado = "pendiente"
    elif total_fac > 0 and total_cobrado >= total_fac - 0.02:
      estado = "cobrada"
    else:
      estado = "parcial"

    facturas_cliente_db.update_estado_cobro(fid, estado)
    resultados.append({"factura_cliente_id": fid, "estado_cobro": estado})

  _invalidar_cache_listado_clientes(empresa_id)

  n_cobradas = sum(1 for r in resultados if r["estado_cobro"] == "cobrada")
  n_parciales = sum(1 for r in resultados if r["estado_cobro"] == "parcial")
  msg = f"Cobro vinculado a {len(resultados)} facturas"
  if n_cobradas:
    msg += f" ({n_cobradas} cobradas)"
  if n_parciales:
    msg += f" ({n_parciales} parciales)"

  return jsonify({"ok": True, "mensaje": msg, "resultados": resultados})


@bancos_bp.route("/api/bancos/conciliacion/desvincular", methods=["POST"])
def conciliacion_desvincular():
  """
  Quita la vinculación movimiento–factura. Si la factura queda sin ningún movimiento vinculado,
  pasa a pendiente; si sigue teniendo otros pagos, se recalcula estado pagada/parcial.
  Body: { "movimiento_id": int }.
  """
  data = request.get_json(silent=True) or {}
  mov_id = data.get("movimiento_id")
  if mov_id is None:
    return _bad_request("Falta movimiento_id")
  try:
    mov_id = int(mov_id)
  except (TypeError, ValueError):
    return _bad_request("movimiento_id debe ser un número")

  _init_movimientos_db()
  conn_bancos = _get_bancos_db()
  try:
    row = conn_bancos.execute(
      "SELECT id, factura_proveedor_id, factura_cliente_key, empresa_id, factura_cliente_id FROM movimientos WHERE id = ?",
      (mov_id,),
    ).fetchone()
    if not row:
      return jsonify({"error": "Movimiento no encontrado"}), 404
    factura_id = row[1]
    factura_cliente_key = (row[2] or "").strip()
    empresa_id = (row[3] or "").strip()
    factura_cliente_id_val = row[4]

    # Handle MULTI conciliation (1 movement → N invoices)
    if factura_cliente_key == "MULTI":
      conn_bancos.execute(
        "UPDATE movimientos SET factura_cliente_key = NULL, factura_cliente_id = NULL, conciliado_at = NULL WHERE id = ?",
        (mov_id,),
      )
      conn_bancos.commit()
      conn_bancos.close()
      # Delete from conciliacion_multiple and recalculate each invoice
      from core.db import conectar as _conectar_gestion
      facturas_afectadas = []
      with _conectar_gestion() as conn_g:
        rows_cm = conn_g.execute(
          "SELECT factura_cliente_id FROM conciliacion_multiple WHERE movimiento_id = ?", (mov_id,)
        ).fetchall()
        facturas_afectadas = [r["factura_cliente_id"] for r in rows_cm]
        conn_g.execute("DELETE FROM conciliacion_multiple WHERE movimiento_id = ?", (mov_id,))
      for fid in facturas_afectadas:
        _recalcular_estado_cobro_cliente(fid)
      if empresa_id:
        _invalidar_cache_listado_clientes(empresa_id)
      return jsonify({"ok": True, "mensaje": f"Conciliación múltiple deshecha ({len(facturas_afectadas)} facturas actualizadas)."})

    if not factura_id and not factura_cliente_key:
      return jsonify({"error": "El movimiento no está conciliado con ninguna factura"}), 400
    if factura_cliente_key:
      # G.11: leer factura_cliente_id antes de limpiar
      row_cli = conn_bancos.execute(
        "SELECT factura_cliente_id FROM movimientos WHERE id = ?", (mov_id,)
      ).fetchone()
      fci_id = row_cli[0] if row_cli and row_cli[0] else None
      conn_bancos.execute(
        "UPDATE movimientos SET factura_cliente_key = NULL, factura_cliente_id = NULL, conciliado_at = NULL WHERE id = ?",
        (mov_id,),
      )
      conn_bancos.commit()
      conn_bancos.close()
      # G.11: recalcular estado_cobro
      estado_cobro = "pendiente"
      if fci_id:
        estado_cobro = _recalcular_estado_cobro_cliente(fci_id)
        _invalidar_cache_listado_clientes(empresa_id)
      return jsonify({"ok": True, "mensaje": "Vinculación con factura de cliente eliminada. Estado: " + estado_cobro + "."})
    conn_bancos.execute(
      "UPDATE movimientos SET factura_proveedor_id = NULL, conciliado_at = NULL WHERE id = ?",
      (mov_id,),
    )
    conn_bancos.commit()
  finally:
    conn_bancos.close()

  facturas_db.init_facturas_db()
  conn_bancos = _get_bancos_db()
  try:
    row_sum = conn_bancos.execute(
      "SELECT COALESCE(SUM(ABS(CAST(importe AS REAL))), 0) FROM movimientos WHERE factura_proveedor_id = ?",
      (factura_id,),
    ).fetchone()
    total_pagado = float(row_sum[0] or 0) if row_sum else 0.0
  finally:
    conn_bancos.close()

  if total_pagado < 0.01:
    estado = "pendiente"
  else:
    conn_gest = sqlite3.connect(str(GESTION_DB))
    try:
      cur = conn_gest.execute(
        "SELECT total_a_pagar, total_factura, total FROM facturas_proveedor WHERE id = ?",
        (factura_id,),
      ).fetchone()
      total_fac = 0.0
      if cur:
        for v in (cur[0], cur[1], cur[2]):
          if v is not None and str(v).strip():
            try:
              total_fac = float(str(v).strip().replace(",", "."))
              break
            except (ValueError, TypeError):
              pass
      estado = "pagada" if total_pagado >= total_fac - 0.02 else "parcial"
    finally:
      conn_gest.close()

  facturas_db.init_facturas_db()
  conn_gest = sqlite3.connect(str(GESTION_DB))
  try:
    conn_gest.execute(
      "UPDATE facturas_proveedor SET estado_pago = ? WHERE id = ?",
      (estado, factura_id),
    )
    conn_gest.commit()
  finally:
    conn_gest.close()

  _invalidar_cache_listado_proveedores(empresa_id)
  return jsonify({"ok": True, "mensaje": "Conciliación deshecha. Factura vuelve a " + estado + "."})


@bancos_bp.route("/api/bancos/conciliacion/factura-proveedor/<int:factura_id>", methods=["GET"])
def conciliacion_resumen_factura_proveedor(factura_id: int):
  """
  Devuelve el resumen de conciliación bancaria de una factura de proveedor:
  total_factura (importe a pagar), total_pagado (suma de movimientos vinculados), pendiente, y lista de movimientos.
  Si la base de movimientos no está disponible, devuelve total_pagado=0 y movimientos=[].
  """
  total_factura = 0.0
  total_pagado = 0.0
  movimientos = []

  try:
    facturas_db.init_facturas_db()
  except Exception as e:
    logger.warning("No se pudo inicializar BD facturas: %s", e)
    return jsonify({"error": "Base de datos de facturas no disponible", "total_factura": 0, "total_pagado": 0, "pendiente": 0, "movimientos": []}), 200

  try:
    conn_gest = sqlite3.connect(str(GESTION_DB))
  except Exception as e:
    return jsonify({"error": "No se pudo conectar a la base de datos", "total_factura": 0, "total_pagado": 0, "pendiente": 0, "movimientos": []}), 200

  try:
    cur = conn_gest.execute(
      "SELECT total_a_pagar, total_factura, total FROM facturas_proveedor WHERE id = ?",
      (factura_id,),
    ).fetchone()
    if not cur:
      conn_gest.close()
      return jsonify({"error": "Factura no encontrada", "total_factura": 0, "total_pagado": 0, "pendiente": 0, "movimientos": []}), 200
    for v in (cur[0], cur[1], cur[2]):
      if v is not None and str(v).strip():
        try:
          total_factura = float(str(v).strip().replace(",", "."))
          break
        except (ValueError, TypeError):
          pass
  finally:
    conn_gest.close()

  try:
    _init_movimientos_db()
    conn_bancos = _get_bancos_db()
  except Exception as e:
    logger.warning("No se pudo conectar a BD movimientos: %s", e)
    pendiente = max(0.0, total_factura - total_pagado)
    return jsonify({
      "total_factura": round(total_factura, 2),
      "total_pagado": round(total_pagado, 2),
      "pendiente": round(pendiente, 2),
      "movimientos": [],
      "error": "Base de movimientos no disponible",
    })

  try:
    rows = conn_bancos.execute(
      """SELECT id, fecha_operacion, concepto, importe
         FROM movimientos WHERE factura_proveedor_id = ?
         ORDER BY fecha_operacion, id""",
      (factura_id,),
    ).fetchall()
    movimientos = [
      {
        "id": r[0],
        "fecha_operacion": (r[1] or "").strip(),
        "concepto": (r[2] or "").strip(),
        "importe": r[3],
      }
      for r in rows
    ]
    row_sum = conn_bancos.execute(
      "SELECT COALESCE(SUM(ABS(CAST(importe AS REAL))), 0) FROM movimientos WHERE factura_proveedor_id = ?",
      (factura_id,),
    ).fetchone()
    total_pagado = float(row_sum[0] or 0) if row_sum else 0.0
  finally:
    conn_bancos.close()

  pendiente = max(0.0, total_factura - total_pagado)
  return jsonify({
    "total_factura": round(total_factura, 2),
    "total_pagado": round(total_pagado, 2),
    "pendiente": round(pendiente, 2),
    "movimientos": movimientos,
  })


@bancos_bp.route("/api/bancos/conciliacion/detalle-multi/<int:movimiento_id>", methods=["GET"])
def conciliacion_detalle_multi(movimiento_id: int):
  """Detalle de conciliación múltiple: facturas vinculadas a un movimiento."""
  from routes.helpers import _parse_importe_es

  # Get movement info from movimientos.db
  mov_info = {"fecha": "", "importe": 0}
  try:
    _init_movimientos_db()
    conn_b = _get_bancos_db()
    row = conn_b.execute("SELECT fecha_operacion, importe FROM movimientos WHERE id = ?", (movimiento_id,)).fetchone()
    if row:
      mov_info = {"fecha": (row[0] or "").strip(), "importe": abs(float(row[1] or 0))}
    conn_b.close()
  except Exception:
    pass

  # Get linked invoices from conciliacion_multiple in gestion.db
  facturas = []
  total_aplicado = 0
  try:
    conn_g = sqlite3.connect(str(GESTION_DB))
    conn_g.row_factory = sqlite3.Row
    rows = conn_g.execute(
      "SELECT cm.factura_cliente_id, cm.importe_aplicado,"
      " fc.numero_factura, fc.cliente, fc.total_a_pagar"
      " FROM conciliacion_multiple cm"
      " LEFT JOIN facturas_cliente fc ON cm.factura_cliente_id = fc.id"
      " WHERE cm.movimiento_id = ?",
      (movimiento_id,),
    ).fetchall()
    for r in rows:
      imp = float(r["importe_aplicado"] or 0)
      total_aplicado += imp
      facturas.append({
        "factura_cliente_id": r["factura_cliente_id"],
        "numero_factura": (r["numero_factura"] or "").strip(),
        "cliente": (r["cliente"] or "").strip(),
        "total_factura": _parse_importe_es(r["total_a_pagar"]),
        "importe_aplicado": round(imp, 2),
      })
    conn_g.close()
  except Exception:
    pass

  return jsonify({
    "movimiento_id": movimiento_id,
    "fecha": mov_info["fecha"],
    "importe": mov_info["importe"],
    "facturas": facturas,
    "total_aplicado": round(total_aplicado, 2),
  })


@bancos_bp.route("/api/bancos/conciliacion/factura-cliente/<int:factura_id>", methods=["GET"])
def conciliacion_resumen_factura_cliente(factura_id: int):
  """Resumen de conciliación de una factura de cliente: total, cobrado, pendiente, movimientos."""
  from routes.helpers import _parse_importe_es
  total_factura = 0.0
  total_cobrado = 0.0
  movimientos = []

  # Get factura from gestion.db
  try:
    conn_gest = sqlite3.connect(str(GESTION_DB))
    conn_gest.row_factory = sqlite3.Row
    row = conn_gest.execute("SELECT total_a_pagar FROM facturas_cliente WHERE id = ?", (factura_id,)).fetchone()
    if row:
      total_factura = _parse_importe_es(row["total_a_pagar"])
    conn_gest.close()
  except Exception:
    pass

  if total_factura == 0:
    return jsonify({"total_factura": 0, "total_cobrado": 0, "pendiente": 0, "movimientos": []})

  # Movimientos from movimientos.db
  try:
    _init_movimientos_db()
    conn_bancos = _get_bancos_db()
    rows = conn_bancos.execute(
      "SELECT id, fecha_operacion, concepto, importe FROM movimientos WHERE factura_cliente_id = ? ORDER BY fecha_operacion",
      (factura_id,),
    ).fetchall()
    for r in rows:
      movimientos.append({
        "id": r[0], "fecha": (r[1] or "").strip(), "concepto": (r[2] or "").strip(),
        "importe": abs(r[3] or 0), "origen": "banco",
      })
      total_cobrado += abs(r[3] or 0)
    conn_bancos.close()
  except Exception:
    pass

  # conciliacion_multiple from gestion.db
  try:
    conn_gest = sqlite3.connect(str(GESTION_DB))
    conn_gest.row_factory = sqlite3.Row
    rows = conn_gest.execute(
      "SELECT id, movimiento_fecha, movimiento_importe, importe_aplicado, created_at"
      " FROM conciliacion_multiple WHERE factura_cliente_id = ? ORDER BY created_at",
      (factura_id,),
    ).fetchall()
    for r in rows:
      movimientos.append({
        "id": r["id"], "fecha": (r["movimiento_fecha"] or "").strip(),
        "concepto": "Conciliación múltiple", "importe": r["importe_aplicado"],
        "origen": "multiple",
      })
      total_cobrado += r["importe_aplicado"] or 0
    conn_gest.close()
  except Exception:
    pass

  pendiente = max(0.0, total_factura - total_cobrado)
  return jsonify({
    "total_factura": round(total_factura, 2),
    "total_cobrado": round(total_cobrado, 2),
    "pendiente": round(pendiente, 2),
    "movimientos": movimientos,
  })


@bancos_bp.route("/api/bancos/movimientos_export", methods=["GET"])
def exportar_movimientos_bancos():
  """
  Exporta los movimientos de bancos a CSV (con extensión .xlsx) aplicando los mismos filtros
  que el listado: banco, fecha_desde, fecha_hasta, empresa_id.
  """
  _init_movimientos_db()
  banco = (request.args.get("banco") or "").strip() or None
  fecha_desde = (request.args.get("fecha_desde") or "").strip() or None
  fecha_hasta = (request.args.get("fecha_hasta") or "").strip() or None
  empresa_id = (request.args.get("empresa_id") or "").strip() or None
  concepto = (request.args.get("concepto") or "").strip() or None

  conn = _get_bancos_db()
  try:
    conditions = []
    params = []
    if banco:
      conditions.append("banco = ?")
      params.append(banco.lower())
    if fecha_desde:
      conditions.append("fecha_operacion >= ?")
      params.append(fecha_desde)
    if fecha_hasta:
      conditions.append("fecha_operacion <= ?")
      params.append(fecha_hasta)
    if empresa_id:
      conditions.append("empresa_id = ?")
      params.append(empresa_id)
    if concepto:
      conditions.append("(concepto IS NOT NULL AND concepto LIKE ?)")
      params.append("%" + concepto + "%")
    where = (" WHERE " + " AND ".join(conditions)) if conditions else ""
    rows = conn.execute(
      f"""
      SELECT fecha_operacion, fecha_valor, concepto, importe, divisa, saldo,
             banco, codigo, numero_documento, referencia_1, referencia_2, empresa_id, created_at
      FROM movimientos
      {where}
      ORDER BY fecha_operacion DESC, id DESC
      """,
      params,
    ).fetchall()
    if not rows:
      return jsonify({"error": "No hay movimientos que cumplan el filtro para exportar"}), 404

    campos = [
      "fecha_operacion",
      "fecha_valor",
      "concepto",
      "importe",
      "divisa",
      "saldo",
      "banco",
      "codigo",
      "numero_documento",
      "referencia_1",
      "referencia_2",
      "empresa_id",
      "created_at",
    ]

    try:
      from openpyxl import Workbook
    except ImportError:
      return jsonify({"error": "openpyxl no instalado. pip install openpyxl"}), 500

    wb = Workbook()
    ws = wb.active
    ws.title = "Movimientos"
    for col, key in enumerate(campos, start=1):
      ws.cell(row=1, column=col, value=key)
    for row_idx, r in enumerate(rows, start=2):
      for col, val in enumerate(r, start=1):
        ws.cell(row=row_idx, column=col, value=val)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    fecha_tag = datetime.utcnow().strftime("%Y%m%d")
    nombre = f"movimientos_bancos_{fecha_tag}.xlsx"
    return send_file(
      buf,
      mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
      as_attachment=True,
      download_name=nombre,
    )
  finally:
    conn.close()

@bancos_bp.route("/api/bancos/movimientos", methods=["DELETE"])
def borrar_movimientos_por_ids():
  """
  Elimina movimientos por lista de IDs. Body: { "ids": [1, 2, 3] }.
  Devuelve { "eliminados": n }.
  """
  _init_movimientos_db()
  data = request.get_json() or {}
  ids_raw = data.get("ids")
  if not isinstance(ids_raw, list):
    return _bad_request("Se espera un array 'ids' en el body")
  ids = []
  for x in ids_raw:
    try:
      ids.append(int(x))
    except (ValueError, TypeError):
      pass
  if not ids:
    return jsonify({"eliminados": 0, "mensaje": "No se indicaron IDs válidos."})
  conn = _get_bancos_db()
  try:
    placeholders = ",".join("?" * len(ids))
    cur = conn.execute(
      f"DELETE FROM movimientos WHERE id IN ({placeholders})",
      ids,
    )
    conn.commit()
    n = cur.rowcount
    return jsonify({"eliminados": n, "mensaje": f"Eliminados {n} movimiento(s)."})
  finally:
    conn.close()

@bancos_bp.route("/api/bancos/movimientos/solo-fecha", methods=["DELETE"])
def eliminar_movimientos_solo_fecha():
  """
  Elimina de la base de datos los movimientos que solo tienen fecha
  (concepto vacío o nulo e importe 0). Irreversible.
  """
  _init_movimientos_db()
  conn = _get_bancos_db()
  try:
    cur = conn.execute(
      """
      SELECT COUNT(*) FROM movimientos
      WHERE TRIM(COALESCE(concepto, '')) = '' AND COALESCE(importe, 0) = 0
      """
    )
    n = cur.fetchone()[0]
    if n == 0:
      return jsonify({"eliminados": 0, "mensaje": "No hay movimientos que solo tengan fecha."})
    conn.execute(
      """
      DELETE FROM movimientos
      WHERE TRIM(COALESCE(concepto, '')) = '' AND COALESCE(importe, 0) = 0
      """
    )
    conn.commit()
    return jsonify({"eliminados": n, "mensaje": f"Eliminados {n} movimiento(s) que solo tenían fecha."})
  finally:
    conn.close()


@bancos_bp.route("/api/bancos/movimientos", methods=["POST"])
def crear_movimientos():
  """
  Crea uno o más movimientos. Body: { movimientos: [ {...} ], empresa_id?, omitir_duplicados?: true }.
  Cada movimiento: fecha_operacion, concepto, importe, banco; opcionales: fecha_valor, divisa, saldo, codigo, numero_documento, referencia_1, referencia_2.
  Si omitir_duplicados es true, no se insertan los que tengan el mismo hash_dedup (por defecto true).
  """
  _init_movimientos_db()
  data = request.get_json() or {}
  movs = data.get("movimientos")
  if not isinstance(movs, list):
    return _bad_request("Se espera un array 'movimientos'")
  empresa_id = (data.get("empresa_id") or "").strip() or None
  omitir_duplicados = data.get("omitir_duplicados", True)
  insertados, duplicados, errores = _insertar_movimientos_lista(movs, empresa_id, omitir_duplicados)
  return jsonify({
    "insertados": insertados,
    "duplicados_omitidos": duplicados,
    "errores": errores,
  })


@bancos_bp.route("/api/bancos/importar/santander", methods=["POST"])
def importar_santander():
  """
  Recibe un Excel de extracto Santander (multipart: archivo o file).
  Opcional: empresa_id en form. Omite duplicados por defecto.
  """
  _init_movimientos_db()
  archivo = request.files.get("archivo") or request.files.get("file")
  if not archivo or not archivo.filename:
    return _bad_request("Falta el archivo Excel (campo 'archivo' o 'file')")
  if not archivo.filename.lower().endswith((".xlsx", ".xls")):
    return _bad_request("El archivo debe ser Excel (.xlsx o .xls)")
  empresa_id = (request.form.get("empresa_id") or request.args.get("empresa_id") or "").strip() or None
  omitir_duplicados = request.form.get("omitir_duplicados", "true").lower() in ("1", "true", "sí", "si")
  try:
    movs = _parse_santander_excel(archivo.stream)
  except ValueError as e:
    return _bad_request(str(e))
  except RuntimeError as e:
    return jsonify({"error": str(e)}), 500
  if not movs:
    return jsonify({"insertados": 0, "duplicados_omitidos": 0, "errores": [], "mensaje": "No se encontraron movimientos en el Excel"})
  insertados, duplicados, errores = _insertar_movimientos_lista(movs, empresa_id, omitir_duplicados)
  return jsonify({
    "insertados": insertados,
    "duplicados_omitidos": duplicados,
    "errores": errores,
    "leidos": len(movs),
  })


@bancos_bp.route("/api/bancos/importar/bbva", methods=["POST"])
def importar_bbva():
  """
  Recibe un Excel de extracto BBVA (multipart: archivo o file).
  Opcional: empresa_id en form. Omite duplicados por defecto.
  """
  _init_movimientos_db()
  archivo = request.files.get("archivo") or request.files.get("file")
  if not archivo or not archivo.filename:
    return _bad_request("Falta el archivo Excel (campo 'archivo' o 'file')")
  if not archivo.filename.lower().endswith((".xlsx", ".xls")):
    return _bad_request("El archivo debe ser Excel (.xlsx o .xls)")
  empresa_id = (request.form.get("empresa_id") or request.args.get("empresa_id") or "").strip() or None
  omitir_duplicados = request.form.get("omitir_duplicados", "true").lower() in ("1", "true", "sí", "si")
  try:
    movs = _parse_bbva_excel(archivo.stream)
  except ValueError as e:
    return _bad_request(str(e))
  except RuntimeError as e:
    return jsonify({"error": str(e)}), 500
  if not movs:
    return jsonify({"insertados": 0, "duplicados_omitidos": 0, "errores": [], "mensaje": "No se encontraron movimientos en el Excel"})
  insertados, duplicados, errores = _insertar_movimientos_lista(movs, empresa_id, omitir_duplicados)
  return jsonify({
    "insertados": insertados,
    "duplicados_omitidos": duplicados,
    "errores": errores,
    "leidos": len(movs),
  })


# Transporte buscar → routes/transporte.py


# CRM endpoints → routes/crm.py


# Transporte proveedores → routes/transporte.py


# Tesorería → routes/tesoreria.py


# Proyectos + certificaciones → routes/proyectos.py


# Presupuestos → routes/presupuestos.py


# OneDrive → routes/onedrive.py


# ─── Registrar blueprints ─────────────────────────────────────────────────────

# Rutas movidas a routes/ por Javier:
# - CRM → routes/crm.py
# - Transporte → routes/transporte.py
# - Tesorería → routes/tesoreria.py
# - Proyectos → routes/proyectos.py
# - Presupuestos → routes/presupuestos.py
# - OneDrive → routes/onedrive.py


# CAE, Vehículos, Empleados → routes/cae.py, routes/vehiculos.py, routes/empleados.py






# ─── Registrar blueprints ─────────────────────────────────────────────────────
app.register_blueprint(facturas_proveedores_bp)
app.register_blueprint(proveedores_bp)
app.register_blueprint(facturas_clientes_bp)
app.register_blueprint(archivo_bp)
app.register_blueprint(control_calidad_bp)
app.register_blueprint(bancos_bp)

# Re-exportar helpers movidos a routes/ para compatibilidad con tests
from routes.helpers import _parse_importe_es, _sum_importes  # noqa: F401

# Blueprints extraídos a routes/
from routes.usuarios import usuarios_bp
from routes.maquinaria import maquinaria_bp
from routes.api_general import api_general_bp
from routes.crm import crm_bp as crm_routes_bp
from routes.tesoreria import tesoreria_bp as tesoreria_routes_bp
from routes.proyectos import proyectos_bp as proyectos_routes_bp
from routes.presupuestos import presupuestos_bp as presupuestos_routes_bp
from routes.onedrive import onedrive_bp as onedrive_routes_bp
from routes.transporte import transporte_bp as transporte_routes_bp
from routes.cae import cae_bp as cae_routes_bp
from routes.vehiculos import vehiculos_bp as vehiculos_routes_bp
from routes.empleados import empleados_bp as empleados_routes_bp

app.register_blueprint(usuarios_bp)
app.register_blueprint(maquinaria_bp)
app.register_blueprint(api_general_bp)
app.register_blueprint(crm_routes_bp)
app.register_blueprint(tesoreria_routes_bp)
app.register_blueprint(proyectos_routes_bp)
app.register_blueprint(presupuestos_routes_bp)
app.register_blueprint(onedrive_routes_bp)
app.register_blueprint(transporte_routes_bp)
app.register_blueprint(cae_routes_bp)
app.register_blueprint(vehiculos_routes_bp)
app.register_blueprint(empleados_routes_bp)

from routes.impuestos import impuestos_bp as impuestos_routes_bp
app.register_blueprint(impuestos_routes_bp)

from routes.albaranes import albaranes_bp as albaranes_routes_bp
app.register_blueprint(albaranes_routes_bp)

from routes.seguros import seguros_bp as seguros_routes_bp
app.register_blueprint(seguros_routes_bp)

from core import impuestos_db
impuestos_db.init_impuestos_db()

from routes.eeff import eeff_bp as eeff_routes_bp
app.register_blueprint(eeff_routes_bp)

from routes.gmail import gmail_bp as gmail_routes_bp
app.register_blueprint(gmail_routes_bp)

logger.info("ERP arrancado — blueprints registrados")


if __name__ == "__main__":
  # Logging ya configurado al inicio del módulo (RotatingFileHandler)
  ensure_dirs()
  app.run(debug=True, port=8000)

