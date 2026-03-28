"""Rutas de transporte: búsqueda de rutas y proveedores de transporte."""
from __future__ import annotations

import io
import logging

from flask import Blueprint, jsonify, request, send_file

from config import BASE_DIR, OPENROUTESERVICE_API_KEY
from core.transporte_servicios import (
  buscar_ruta_y_proveedores as _buscar_ruta_y_proveedores,
  parsear_xlsx_proveedores_desde_stream as _parsear_xlsx_proveedores_stream,
)
from core.proveedores_transporte_db import (
  alta_proveedor as _alta_proveedor_transporte,
  listar_proveedores_para_admin as _listar_proveedores_transporte_admin,
  obtener_proveedor as _obtener_proveedor_transporte,
  actualizar_proveedor as _actualizar_proveedor_transporte,
  insertar_desde_lista as _insertar_proveedores_transporte_lista,
)
from routes.helpers import _bad_request

logger = logging.getLogger("erp")

transporte_bp = Blueprint("transporte", __name__)


@transporte_bp.route("/api/proyectos/transporte/buscar", methods=["POST"])
def transporte_buscar():
  """
  Recibe origen y destino (texto). Calcula la ruta con OpenRouteService y devuelve
  proveedores de transporte de maquinaria situados a menos de 50 km de la ruta.
  """
  data = request.get_json(silent=True) or {}
  origen = (data.get("origen") or "").strip()
  destino = (data.get("destino") or "").strip()
  paradas = data.get("paradas")
  if paradas is not None and not isinstance(paradas, list):
    paradas = []
  paradas = [str(p).strip() for p in (paradas or []) if str(p).strip()]
  if not origen or not destino:
    return jsonify({"error": "Faltan origen o destino"}), 400
  api_key = (OPENROUTESERVICE_API_KEY or "").strip()
  if not api_key:
    return jsonify({"error": "No está configurada la API key de OpenRouteService. Añade OPENROUTESERVICE_API_KEY en .env"}), 503
  try:
    resultado = _buscar_ruta_y_proveedores(origen, destino, BASE_DIR, api_key, radio_km=50.0, paradas=paradas)
  except Exception as e:
    return jsonify({"error": "Error al calcular ruta o proveedores: " + str(e)}), 500
  return jsonify(resultado)


@transporte_bp.route("/api/proyectos/transporte/proveedores", methods=["GET"])
def transporte_listar_proveedores():
  """Lista todos los proveedores de transporte (con id) para el modal de gestión."""
  try:
    lista = _listar_proveedores_transporte_admin()
    return jsonify({"proveedores": lista})
  except Exception as e:
    return jsonify({"error": str(e)}), 500


@transporte_bp.route("/api/proyectos/transporte/proveedores", methods=["POST"])
def transporte_alta_proveedor():
  """Añade un nuevo proveedor de transporte (incorporar transportista)."""
  data = request.get_json(silent=True) or {}
  nombre = (data.get("nombre") or "").strip()
  if not nombre:
    return jsonify({"error": "El nombre es obligatorio"}), 400
  try:
    datos = {
      "nombre": nombre,
      "telefono": (data.get("telefono") or "").strip(),
      "telefono_fijo": (data.get("telefono_fijo") or "").strip(),
      "telefono_movil": (data.get("telefono_movil") or "").strip(),
      "email": (data.get("email") or "").strip(),
      "web": (data.get("web") or "").strip(),
      "localidad": (data.get("localidad") or "").strip(),
      "provincia": (data.get("provincia") or "").strip(),
      "codigo_postal": (data.get("codigo_postal") or "").strip(),
      "direccion": (data.get("direccion") or "").strip(),
      "lat": data.get("lat"),
      "lon": data.get("lon"),
    }
    if datos["lat"] is not None:
      try:
        datos["lat"] = float(datos["lat"])
      except (TypeError, ValueError):
        datos["lat"] = None
    if datos["lon"] is not None:
      try:
        datos["lon"] = float(datos["lon"])
      except (TypeError, ValueError):
        datos["lon"] = None
    id_ = _alta_proveedor_transporte(datos)
    return jsonify({"ok": True, "id": id_})
  except Exception as e:
    return jsonify({"error": str(e)}), 500


@transporte_bp.route("/api/proyectos/transporte/proveedores/<int:proveedor_id>", methods=["GET"])
def transporte_obtener_proveedor(proveedor_id):
  """Devuelve un proveedor por id."""
  p = _obtener_proveedor_transporte(proveedor_id)
  if not p:
    return jsonify({"error": "Proveedor no encontrado"}), 404
  return jsonify(p)


@transporte_bp.route("/api/proyectos/transporte/proveedores/<int:proveedor_id>", methods=["PUT"])
def transporte_actualizar_proveedor(proveedor_id):
  """Actualiza un proveedor de transporte."""
  data = request.get_json(silent=True) or {}
  nombre = (data.get("nombre") or "").strip()
  if not nombre:
    return jsonify({"error": "El nombre es obligatorio"}), 400
  try:
    datos = {
      "nombre": nombre,
      "telefono": (data.get("telefono") or "").strip(),
      "telefono_fijo": (data.get("telefono_fijo") or "").strip(),
      "telefono_movil": (data.get("telefono_movil") or "").strip(),
      "email": (data.get("email") or "").strip(),
      "web": (data.get("web") or "").strip(),
      "localidad": (data.get("localidad") or "").strip(),
      "provincia": (data.get("provincia") or "").strip(),
      "codigo_postal": (data.get("codigo_postal") or "").strip(),
      "direccion": (data.get("direccion") or "").strip(),
      "lat": data.get("lat"),
      "lon": data.get("lon"),
    }
    for key in ("lat", "lon"):
      if datos[key] is not None:
        try:
          datos[key] = float(datos[key])
        except (TypeError, ValueError):
          datos[key] = None
    ok = _actualizar_proveedor_transporte(proveedor_id, datos)
    if not ok:
      return jsonify({"error": "Proveedor no encontrado"}), 404
    return jsonify({"ok": True})
  except Exception as e:
    return jsonify({"error": str(e)}), 500


@transporte_bp.route("/api/proyectos/transporte/proveedores/carga-masiva", methods=["POST"])
def transporte_carga_masiva():
  """Subida masiva de proveedores desde un archivo Excel (.xlsx)."""
  if "archivo" not in request.files:
    return jsonify({"error": "Falta el archivo. Usa el campo 'archivo'."}), 400
  f = request.files["archivo"]
  if not f or not f.filename:
    return jsonify({"error": "No se ha seleccionado ningún archivo"}), 400
  if not f.filename.lower().endswith((".xlsx", ".xls")):
    return jsonify({"error": "El archivo debe ser Excel (.xlsx o .xls)"}), 400
  try:
    contenido = f.read()
  except Exception as e:
    return jsonify({"error": "Error leyendo el archivo: " + str(e)}), 500
  try:
    from io import BytesIO
    lista = _parsear_xlsx_proveedores_stream(BytesIO(contenido))
  except Exception as e:
    return jsonify({"error": "Error al parsear el Excel: " + str(e)}), 500
  if not lista:
    return jsonify({"error": "No se encontraron filas válidas en el Excel", "insertados": 0}), 400
  try:
    n = _insertar_proveedores_transporte_lista(lista)
    return jsonify({"ok": True, "insertados": n, "total_filas": len(lista)})
  except Exception as e:
    return jsonify({"error": str(e)}), 500


@transporte_bp.route("/api/proyectos/transporte/proveedores/exportar-excel", methods=["POST"])
def transporte_exportar_proveedores_excel():
  """Genera un Excel con los proveedores de la ruta (listado actual) e información de contacto."""
  try:
    from openpyxl import Workbook
  except ImportError:
    return jsonify({"error": "openpyxl no instalado"}), 500
  data = request.get_json(silent=True) or {}
  proveedores = data.get("proveedores")
  if not isinstance(proveedores, list):
    return jsonify({"error": "Se espera un array 'proveedores' en el body"}), 400
  ruta_info = data.get("ruta") or {}
  wb = Workbook()
  ws = wb.active
  if ws is None:
    return jsonify({"error": "No se pudo crear la hoja"}), 500
  ws.title = "Proveedores ruta"
  headers = [
    "Nombre", "Provincia", "Localidad", "Código postal", "Dirección",
    "Tel. fijo", "Tel. móvil", "Email", "Web", "Distancia (km)",
  ]
  row = 1
  if ruta_info.get("texto") or ruta_info.get("distancia_km") is not None or ruta_info.get("duracion_min") is not None:
    ws.cell(row=row, column=1, value="Ruta: " + (ruta_info.get("texto") or ""))
    row += 1
    if ruta_info.get("distancia_km") is not None or ruta_info.get("duracion_min") is not None:
      ws.cell(row=row, column=1, value="Distancia: {} km · Duración: {} min".format(
        ruta_info.get("distancia_km") if ruta_info.get("distancia_km") is not None else "",
        ruta_info.get("duracion_min") if ruta_info.get("duracion_min") is not None else "",
      ))
      row += 1
    row += 1
  for col, h in enumerate(headers, 1):
    ws.cell(row=row, column=col, value=h)
  row += 1
  for p in proveedores:
    if not isinstance(p, dict):
      continue
    dist_km = p.get("distancia_km")
    if dist_km is not None:
      try:
        dist_km = float(dist_km)
      except (TypeError, ValueError):
        dist_km = ""
    ws.cell(row=row, column=1, value=(p.get("nombre") or ""))
    ws.cell(row=row, column=2, value=(p.get("provincia") or ""))
    ws.cell(row=row, column=3, value=(p.get("localidad") or ""))
    ws.cell(row=row, column=4, value=(p.get("codigo_postal") or ""))
    ws.cell(row=row, column=5, value=(p.get("direccion") or ""))
    ws.cell(row=row, column=6, value=(p.get("telefono_fijo") or ""))
    ws.cell(row=row, column=7, value=(p.get("telefono_movil") or ""))
    ws.cell(row=row, column=8, value=(p.get("email") or ""))
    ws.cell(row=row, column=9, value=(p.get("web") or ""))
    ws.cell(row=row, column=10, value=round(dist_km, 2) if isinstance(dist_km, (int, float)) else (dist_km if dist_km != "" else ""))
    row += 1
  output = io.BytesIO()
  wb.save(output)
  output.seek(0)
  filename = "proveedores_ruta.xlsx"
  return send_file(
    output,
    mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    as_attachment=True,
    download_name=filename,
  )
