# -*- coding: utf-8 -*-
"""
Paso 4: Generar una hoja nueva con todas las empresas de todas las provincias.
- Lee el Excel Completado (o Verificado).
- Crea la hoja "Listado único" con todas las filas de las hojas de provincia.
- La columna "Provincia_hoja" (col 1) indica la provincia de origen.
"""
import openpyxl
from pathlib import Path

EXCEL_COMPLETADO = Path(r"c:\Users\javie\Desktop\HD - Listado transportistas - Completado.xlsx")
EXCEL_SALIDA = Path(r"c:\Users\javie\Desktop\HD - Listado transportistas - Listado unico.xlsx")
NOMBRE_HOJA_UNICO = "Listado único"
HOJAS_EXCLUIDAS = {"Gemini", "Listado"}


def main():
    wb = openpyxl.load_workbook(EXCEL_COMPLETADO, data_only=False)
    provincias = [s for s in wb.sheetnames if s not in HOJAS_EXCLUIDAS]
    if not provincias:
        print("No hay hojas de provincia.")
        wb.close()
        return

    # Cabecera desde la primera hoja de provincia
    ws_primera = wb[provincias[0]]
    max_col = ws_primera.max_column
    header = [ws_primera.cell(1, c).value for c in range(1, max_col + 1)]

    # Crear hoja "Listado único" (al principio o al final)
    if NOMBRE_HOJA_UNICO in wb.sheetnames:
        del wb[NOMBRE_HOJA_UNICO]
    ws_unico = wb.create_sheet(NOMBRE_HOJA_UNICO, 0)
    ws_unico.append(header)

    total_filas = 0
    for nombre in provincias:
        ws = wb[nombre]
        for row_idx in range(2, ws.max_row + 1):
            fila = [ws.cell(row_idx, c).value for c in range(1, max_col + 1)]
            ws_unico.append(fila)
            total_filas += 1
        print(f"  {nombre}: {ws.max_row - 1} filas copiadas.")

    wb.save(EXCEL_SALIDA)
    wb.close()
    print(f"\nHoja '{NOMBRE_HOJA_UNICO}' creada con {total_filas} filas.")
    print(f"Guardado: {EXCEL_SALIDA}")


if __name__ == "__main__":
    main()
