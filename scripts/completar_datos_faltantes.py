# -*- coding: utf-8 -*-
"""
Paso 3 del plan: Buscar la información pendiente para empresas que tienen ND en
localidad, código postal, web, email o teléfono.
- Usa búsqueda web (DuckDuckGo) con "[Empresa] [Provincia] transportes contacto".
- Extrae de los snippets: email, teléfono, CP, web, localidad (regex).
- Rellena solo los campos que estaban en ND.
- Guarda en HD - Listado transportistas - Completado.xlsx
"""
import re
import time
import warnings
from pathlib import Path

import openpyxl

warnings.filterwarnings("ignore", message=".*duckduckgo_search.*renamed.*")

EXCEL_VERIFICADO = Path(r"c:\Users\javie\Desktop\HD - Listado transportistas - Verificado.xlsx")
EXCEL_COMPLETADO = Path(r"c:\Users\javie\Desktop\HD - Listado transportistas - Completado.xlsx")
HOJAS_EXCLUIDAS = {"Gemini", "Listado"}
VALOR_ND = "ND"
PAUSA_ENTRE_BUSQUEDAS = 2.0  # segundos
MAX_RESULTADOS_BUSQUEDA = 5
# Límite de filas a procesar por provincia (0 = sin límite; poner ej. 5 para pruebas)
MAX_FILAS_CON_ND_POR_PROVINCIA = 0

# Columnas en Verificado (1-based): 1=Provincia_hoja, 2=Localidad, 3=CP, 4=Direccion, 5=Telefono, 6=Email, 7=Web
COL_PROVINCIA = 1
COL_LOCALIDAD = 2
COL_CP = 3
COL_TELEFONO = 5
COL_EMAIL = 6
COL_WEB = 7

# Regex para extraer de texto
RE_EMAIL = re.compile(r"[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}")
RE_CP = re.compile(r"\b(0[1-9]|[1-4]\d|5[0-2])\d{3}\b")
RE_TELEFONO = re.compile(
    r"(?:\+34\s?)?(?:\d{3}[\s.\-]?\d{3}[\s.\-]?\d{3}|\d{2}[\s.\-]?\d{2}[\s.\-]?\d{2}[\s.\-]?\d{2}[\s.\-]?\d)"
)
RE_WEB = re.compile(
    r"(?:https?://)?(?:www\.)?([a-zA-Z0-9][-a-zA-Z0-9]*\.[a-zA-Z]{2,}(?:\.[a-zA-Z]{2,})?)"
)


def es_nd(val):
    return val is None or str(val).strip() == "" or str(val).strip() == VALOR_ND


def obtener_col_empresa(ws):
    """Devuelve el índice 1-based de la columna Empresa (primera que contiene 'Empresa')."""
    for c in range(1, ws.max_column + 1):
        v = ws.cell(1, c).value
        if v and "Empresa" in str(v) and "verif" not in str(v).lower():
            return c
    return 13  # fallback por si la estructura es la esperada


def buscar_datos_empresa(empresa, provincia):
    """Busca en DuckDuckGo y devuelve dict con email, telefono, cp, web, localidad (lo que encuentre)."""
    try:
        from duckduckgo_search import DDGS
    except ImportError:
        return {}
    query = f"{empresa} {provincia} transportes contacto teléfono email"
    texto = ""
    try:
        ddgs = DDGS()
        results = list(ddgs.text(query, max_results=MAX_RESULTADOS_BUSQUEDA))
        for r in results:
            if isinstance(r, dict):
                texto += " " + str(r.get("body", "")) + " " + str(r.get("title", "")) + " " + str(r.get("href", ""))
            else:
                texto += " " + str(r)
    except Exception:
        return {}
    return extraer_datos_de_texto(texto)


def extraer_datos_de_texto(texto):
    """Extrae email, teléfono, CP, web y (opcional) localidad de un texto."""
    if not texto:
        return {}
    out = {}
    m = RE_EMAIL.search(texto)
    if m:
        out["email"] = m.group(0)
    m = RE_CP.search(texto)
    if m:
        out["cp"] = m.group(0)
    tels = RE_TELEFONO.findall(texto)
    if tels:
        t = tels[0]
        out["telefono"] = t if isinstance(t, str) else "".join(str(x) for x in t)
    # Web: primer dominio que parezca de empresa (evitar google, facebook, etc.)
    for m in RE_WEB.finditer(texto):
        d = m.group(1).lower()
        if any(x in d for x in ["google", "facebook", "twitter", "linkedin", "wikipedia", "youtube"]):
            continue
        if len(d) > 4 and len(d) < 50:
            out["web"] = d
            break
    return out


def main():
    wb = openpyxl.load_workbook(EXCEL_VERIFICADO, data_only=False)
    provincias = [s for s in wb.sheetnames if s not in HOJAS_EXCLUIDAS]
    total_actualizados = 0
    filas_procesadas = 0

    for nombre in provincias:
        ws = wb[nombre]
        max_row = ws.max_row
        if max_row < 2:
            continue
        col_empresa = obtener_col_empresa(ws)
        contador_nd_en_provincia = 0
        for row_idx in range(2, max_row + 1):
            if MAX_FILAS_CON_ND_POR_PROVINCIA and contador_nd_en_provincia >= MAX_FILAS_CON_ND_POR_PROVINCIA:
                break
            provincia = ws.cell(row_idx, COL_PROVINCIA).value
            localidad = ws.cell(row_idx, COL_LOCALIDAD).value
            cp = ws.cell(row_idx, COL_CP).value
            telefono = ws.cell(row_idx, COL_TELEFONO).value
            email = ws.cell(row_idx, COL_EMAIL).value
            web = ws.cell(row_idx, COL_WEB).value
            empresa = ws.cell(row_idx, col_empresa).value
            if not empresa or not str(empresa).strip():
                continue
            necesita_buscar = es_nd(localidad) or es_nd(cp) or es_nd(telefono) or es_nd(email) or es_nd(web)
            if not necesita_buscar:
                continue
            contador_nd_en_provincia += 1
            filas_procesadas += 1
            datos = buscar_datos_empresa(str(empresa)[:80], provincia or nombre)
            if not datos:
                time.sleep(PAUSA_ENTRE_BUSQUEDAS)
                continue
            actualizado = 0
            if es_nd(cp) and datos.get("cp"):
                ws.cell(row_idx, COL_CP, value=datos["cp"])
                actualizado += 1
            if es_nd(telefono) and datos.get("telefono"):
                ws.cell(row_idx, COL_TELEFONO, value=datos["telefono"])
                actualizado += 1
            if es_nd(email) and datos.get("email"):
                ws.cell(row_idx, COL_EMAIL, value=datos["email"])
                actualizado += 1
            if es_nd(web) and datos.get("web"):
                ws.cell(row_idx, COL_WEB, value=datos["web"])
                actualizado += 1
            if actualizado > 0:
                total_actualizados += actualizado
            time.sleep(PAUSA_ENTRE_BUSQUEDAS)
        print(f"  {nombre}: procesado.")

    wb.save(EXCEL_COMPLETADO)
    wb.close()
    print(f"\nGuardado: {EXCEL_COMPLETADO}")
    print(f"Filas con ND que se intentaron completar: {filas_procesadas}")
    print(f"Campos rellenados en total: {total_actualizados}")


if __name__ == "__main__":
    main()
