"""
import_crm_excel.py — Importador idempotente del mini-CRM Excel a la BD del ERP.

Fuente: HincadoDirecto_MiniCRM.xlsx
  Hoja 'Clientes'  → crm_empresas + crm_interacciones (última interacción)
  Hoja 'Contactos' → crm_contactos

Uso:
  python scripts/import_crm_excel.py <ruta_excel>
  python scripts/import_crm_excel.py <ruta_excel> --dry-run
  python scripts/import_crm_excel.py <ruta_excel> --only-empresas
  python scripts/import_crm_excel.py <ruta_excel> --only-contactos

Idempotencia:
  - Empresas:  match por nombre normalizado. Actualiza si ya existe.
  - Contactos: match por email (si existe) o por nombre+empresa normalizado.
  - Interacciones: solo inserta si no existe ya una interacción con mismo
    empresa_id + fecha + tipo 'nota' + asunto == "Última interacción (import)".
"""
from __future__ import annotations

import argparse
import re
import sqlite3
import sys
import unicodedata
from datetime import datetime
from pathlib import Path

# ── Rutas ──────────────────────────────────────────────────────────────────────
BASE_DIR = Path(__file__).resolve().parents[1]
DB_PATH = BASE_DIR / "data" / "gestion.db"

try:
    import openpyxl
except ImportError:
    sys.exit("ERROR: Instala openpyxl → pip install openpyxl")


# ── Helpers ────────────────────────────────────────────────────────────────────

def _norm(s: str | None) -> str:
    """Normaliza texto para comparaciones: lowercase, sin acentos, sin espacios extra."""
    if not s:
        return ""
    s = str(s).strip()
    # Quitar marcas Unicode de dirección y similares
    s = "".join(c for c in s if unicodedata.category(c) not in ("Cf",))
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    return s.lower().strip()


def _strip_notion(s: str | None) -> str | None:
    """Quita URLs de Notion del texto (p.ej. 'Acciona (https://...)' → 'Acciona')."""
    if not s:
        return None
    s = str(s).strip()
    s = re.sub(r"\s*\(https?://[^\)]*\)", "", s)
    s = re.sub(r"https?://\S+", "", s)
    return s.strip() or None


def _clean_phone(s: str | None) -> str | None:
    """Limpia teléfonos: quita caracteres de dirección Unicode, normaliza."""
    if not s:
        return None
    s = str(s)
    s = "".join(c for c in s if unicodedata.category(c) not in ("Cf",))
    s = re.sub(r"[\xa0\u202a\u202c\u200b]", "", s)
    s = s.strip()
    return s or None


def _estado_a_tipo(estado_crm: str | None, estado_alt: str | None) -> str:
    """Mapea Estado CRM del Excel al campo tipo de crm_empresas."""
    estado = (estado_crm or estado_alt or "").strip()
    m = {
        "activo": "cliente",
        "potencial": "lead",
        "dormido": "lead",
        "perdido": "lead",
        "proveedor": "proveedor",
        "partner": "lead",
    }
    return m.get(_norm(estado), "lead")


def _extraer_dominio_email(email: str | None) -> str | None:
    if not email:
        return None
    email = str(email).strip().lower()
    if "@" in email:
        domain = email.split("@", 1)[1].strip()
        if "." in domain:
            return domain
    return None


def _now() -> str:
    return datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%S")


def _fecha_excel(val) -> str | None:
    """Convierte datetime/date/str a ISO date string."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    if hasattr(val, "strftime"):
        return val.strftime("%Y-%m-%d")
    s = str(val).strip()[:10]
    if re.match(r"\d{4}-\d{2}-\d{2}", s):
        return s
    return None


# ── Conexión DB ────────────────────────────────────────────────────────────────

def conectar() -> sqlite3.Connection:
    conn = sqlite3.connect(str(DB_PATH), timeout=30)
    conn.execute("PRAGMA foreign_keys = ON")
    conn.row_factory = sqlite3.Row
    return conn


# ── IMPORT EMPRESAS ────────────────────────────────────────────────────────────

def importar_empresas(
    ws,
    conn: sqlite3.Connection,
    dry_run: bool,
) -> dict[str, int]:
    """Lee la hoja 'Clientes' e inserta/actualiza crm_empresas.
    Columnas esperadas (por posición, 1-indexed):
      A(1)=ID, B(2)=Nombre, C(3)=Tipología, D(4)=Estado CRM,
      E(5)=Contacto principal, F(6)=Contactos, G(7)=Oportunidades,
      H(8)=Proyectos, I(9)=Última interacción, J(10)=Fecha última,
      K(11)=Reuniones, L(12)=Notas, M(13)=Estado
    """
    stats = {"insertadas": 0, "actualizadas": 0, "saltadas": 0, "interacciones": 0}
    ahora = _now()

    rows = list(ws.iter_rows(min_row=2, values_only=True))
    for row in rows:
        if not row or len(row) < 2:
            continue
        nombre_raw = row[1] if len(row) > 1 else None
        if not nombre_raw:
            continue
        nombre = _strip_notion(str(nombre_raw).strip())
        if not nombre:
            continue

        tipologia = _strip_notion(row[2]) if len(row) > 2 else None
        estado_d = row[3] if len(row) > 3 else None
        notas_raw = _strip_notion(row[11]) if len(row) > 11 else None
        estado_m = row[12] if len(row) > 12 else None
        ultima_int_texto = _strip_notion(row[8]) if len(row) > 8 else None
        ultima_int_fecha = _fecha_excel(row[9]) if len(row) > 9 else None

        tipo = _estado_a_tipo(estado_d, estado_m)
        nombre_norm = _norm(nombre)

        # Buscar empresa existente por nombre normalizado
        existing = conn.execute(
            "SELECT id, nombre, sector, notas, dominio FROM crm_empresas WHERE LOWER(nombre) = ?",
            (nombre_norm,)
        ).fetchone()

        if existing:
            emp_id = existing["id"]
            # Actualizar sector/notas si están vacíos en la BD
            updates = {}
            if tipologia and not existing["sector"]:
                updates["sector"] = tipologia
            if notas_raw and not existing["notas"]:
                updates["notas"] = notas_raw
            if updates and not dry_run:
                set_clause = ", ".join(f"{k} = ?" for k in updates)
                conn.execute(
                    f"UPDATE crm_empresas SET {set_clause} WHERE id = ?",
                    list(updates.values()) + [emp_id]
                )
            stats["actualizadas"] += 1
        else:
            # Insertar nueva empresa
            if not dry_run:
                conn.execute("""
                    INSERT INTO crm_empresas
                        (nombre, sector, tipo, notas, fecha_creacion, activo)
                    VALUES (?, ?, ?, ?, ?, 1)
                """, (nombre, tipologia, tipo, notas_raw, ahora))
                emp_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
            else:
                emp_id = None
            stats["insertadas"] += 1

        # Importar "última interacción" como crm_interaccion tipo 'nota'
        if ultima_int_texto and emp_id:
            fecha_interaccion = ultima_int_fecha or ahora[:10]
            asunto_import = "Última interacción (import)"
            # Idempotencia: no duplicar si ya existe
            ya_existe = conn.execute("""
                SELECT id FROM crm_interacciones
                WHERE empresa_id = ? AND tipo = 'nota' AND asunto = ?
                LIMIT 1
            """, (emp_id, asunto_import)).fetchone()
            if not ya_existe:
                if not dry_run:
                    conn.execute("""
                        INSERT INTO crm_interacciones
                            (empresa_id, tipo, asunto, descripcion, fecha, fecha_creacion)
                        VALUES (?, 'nota', ?, ?, ?, ?)
                    """, (emp_id, asunto_import, ultima_int_texto, fecha_interaccion, ahora))
                stats["interacciones"] += 1

    return stats


# ── IMPORT CONTACTOS ───────────────────────────────────────────────────────────

def importar_contactos(
    ws,
    conn: sqlite3.Connection,
    dry_run: bool,
) -> dict[str, int]:
    """Lee la hoja 'Contactos' e inserta/actualiza crm_contactos.
    Columnas esperadas:
      A(1)=ID, B(2)=Nombre, C(3)=Cliente, D(4)=Cargo, E(5)=Categoría,
      F(6)=Email, G(7)=Teléfono, H(8)=Estado, I(9)=Último contacto,
      J(10)=Próxima acción, K(11)=Canal, L(12)=Oportunidades,
      M(13)=Proyectos, N(14)=Reuniones, O(15)=Notas
    """
    stats = {"insertados": 0, "actualizados": 0, "saltados": 0, "sin_empresa": 0}
    ahora = _now()

    rows = list(ws.iter_rows(min_row=2, values_only=True))
    for row in rows:
        if not row or len(row) < 2:
            continue
        nombre_raw = row[1] if len(row) > 1 else None
        if not nombre_raw:
            continue
        nombre = _strip_notion(str(nombre_raw).strip())
        if not nombre:
            continue

        empresa_raw = _strip_notion(row[2]) if len(row) > 2 else None
        cargo = _strip_notion(row[3]) if len(row) > 3 else None
        categoria = (row[4] or "").strip() if len(row) > 4 else ""
        email = (row[5] or "").strip() if len(row) > 5 else None
        email = email.lower() if email else None
        email = email or None
        telefono = _clean_phone(row[6]) if len(row) > 6 else None
        estado = (row[7] or "Activo").strip() if len(row) > 7 else "Activo"
        notas = _strip_notion(row[14]) if len(row) > 14 else None

        activo = 0 if _norm(estado) in ("inactivo", "baja") else 1

        # Tipo relación
        cat_map = {
            "cliente": "cliente",
            "proveedor": "proveedor",
            "experto": "otro",
            "partner": "otro",
        }
        tipo_relacion = cat_map.get(_norm(categoria), "otro")

        # Buscar empresa vinculada
        empresa_id = None
        if empresa_raw:
            emp_row = conn.execute(
                "SELECT id FROM crm_empresas WHERE LOWER(nombre) = ?",
                (_norm(empresa_raw),)
            ).fetchone()
            if emp_row:
                empresa_id = emp_row["id"]
            else:
                stats["sin_empresa"] += 1

        # Buscar contacto existente por email o nombre+empresa
        existing = None
        if email:
            existing = conn.execute(
                "SELECT id FROM crm_contactos WHERE LOWER(email) = ?",
                (email,)
            ).fetchone()
        if not existing:
            existing = conn.execute(
                "SELECT id FROM crm_contactos WHERE LOWER(nombre) = ? AND empresa_vinculada_id IS ?",
                (_norm(nombre), empresa_id)
            ).fetchone()

        if existing:
            cont_id = existing["id"]
            # Solo actualizar campos vacíos
            updates = {}
            row_db = conn.execute("SELECT * FROM crm_contactos WHERE id = ?", (cont_id,)).fetchone()
            if cargo and not row_db["cargo"]:
                updates["cargo"] = cargo
            if email and not row_db["email"]:
                updates["email"] = email
            if telefono and not row_db["telefono"]:
                updates["telefono"] = telefono
            if empresa_id and not row_db["empresa_vinculada_id"]:
                updates["empresa_vinculada_id"] = empresa_id
            if notas and not row_db["notas"]:
                updates["notas"] = notas
            if updates and not dry_run:
                set_clause = ", ".join(f"{k} = ?" for k in updates)
                conn.execute(
                    f"UPDATE crm_contactos SET {set_clause}, fecha_actualizacion = ? WHERE id = ?",
                    list(updates.values()) + [ahora, cont_id]
                )
            stats["actualizados"] += 1
        else:
            if not dry_run:
                conn.execute("""
                    INSERT INTO crm_contactos
                        (nombre, cargo, email, telefono, empresa_vinculada_id,
                         tipo_relacion, notas, fecha_creacion, activo, creado_por)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, 'import_excel')
                """, (nombre, cargo, email, telefono, empresa_id,
                      tipo_relacion, notas, ahora, activo))
            stats["insertados"] += 1

    return stats


# ── MAIN ───────────────────────────────────────────────────────────────────────

def main() -> None:
    parser = argparse.ArgumentParser(description="Importar mini-CRM Excel → gestion.db")
    parser.add_argument("excel", help="Ruta al archivo .xlsx")
    parser.add_argument("--dry-run", action="store_true",
                        help="Solo muestra qué haría, sin modificar la BD")
    parser.add_argument("--only-empresas", action="store_true",
                        help="Solo importa la hoja Clientes (empresas)")
    parser.add_argument("--only-contactos", action="store_true",
                        help="Solo importa la hoja Contactos")
    args = parser.parse_args()

    excel_path = Path(args.excel)
    if not excel_path.exists():
        sys.exit(f"ERROR: No se encuentra el archivo: {excel_path}")
    if not DB_PATH.exists():
        sys.exit(f"ERROR: No se encuentra la base de datos: {DB_PATH}")

    print(f"\n{'[DRY-RUN] ' if args.dry_run else ''}Importando desde: {excel_path.name}")
    print(f"Base de datos: {DB_PATH}\n")

    wb = openpyxl.load_workbook(str(excel_path), read_only=True, data_only=True)

    if args.dry_run:
        # En dry-run: no abrir la BD, solo contar filas del Excel
        conn = None
    else:
        conn = conectar()

    try:
        def _run(c):
            if not args.only_contactos:
                if "Clientes" not in wb.sheetnames:
                    print("AVISO: No se encontró la hoja 'Clientes'")
                else:
                    stats_e = importar_empresas(wb["Clientes"], c, dry_run=args.dry_run)
                    print("── Empresas (hoja 'Clientes') ─────────────────────")
                    print(f"  Insertadas:           {stats_e['insertadas']}")
                    print(f"  Actualizadas:         {stats_e['actualizadas']}")
                    print(f"  Interacciones import: {stats_e['interacciones']}")
            if not args.only_empresas:
                if "Contactos" not in wb.sheetnames:
                    print("AVISO: No se encontró la hoja 'Contactos'")
                else:
                    stats_c = importar_contactos(wb["Contactos"], c, dry_run=args.dry_run)
                    print("\n── Contactos (hoja 'Contactos') ─────────────────")
                    print(f"  Insertados:           {stats_c['insertados']}")
                    print(f"  Actualizados:         {stats_c['actualizados']}")
                    print(f"  Sin empresa vinculada:{stats_c['sin_empresa']}")

        if conn:
            if args.dry_run:
                # dry-run: leer BD sin BEGIN (evita write lock)
                _run(conn)
            else:
                # real import: verificar tablas y usar transacción
                tables = {r[0] for r in conn.execute(
                    "SELECT name FROM sqlite_master WHERE type='table'"
                ).fetchall()}
                required = {"crm_empresas", "crm_contactos", "crm_interacciones"}
                missing = required - tables
                if missing:
                    sys.exit(
                        f"ERROR: Tablas CRM no encontradas: {missing}\n"
                        "Arranca el ERP al menos una vez para inicializar las tablas CRM."
                    )
                conn.execute("BEGIN")
                _run(conn)
                conn.execute("COMMIT")
        else:
            # Sin BD disponible en dry-run: contar desde Excel
            if not args.only_contactos and "Clientes" in wb.sheetnames:
                n = sum(1 for r in wb["Clientes"].iter_rows(min_row=2, values_only=True) if r and r[1])
                print(f"── Empresas Excel: {n} filas con datos")
            if not args.only_empresas and "Contactos" in wb.sheetnames:
                n = sum(1 for r in wb["Contactos"].iter_rows(min_row=2, values_only=True) if r and r[1])
                print(f"── Contactos Excel: {n} filas con datos")

        if args.dry_run:
            print("\n[DRY-RUN] No se realizaron cambios en la base de datos.")
        else:
            print("\n✓ Importación completada.")

    except Exception as exc:
        if conn and not args.dry_run:
            try: conn.execute("ROLLBACK")
            except Exception: pass
        print(f"\nERROR durante la importación: {exc}", file=sys.stderr)
        import traceback; traceback.print_exc()
        sys.exit(1)
    finally:
        if conn:
            conn.close()


if __name__ == "__main__":
    main()
