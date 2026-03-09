# -*- coding: utf-8 -*-
"""
Paso 2 del plan: Verificar la información más importante de las empresas.
- Web: comprobar si la URL responde (HEAD/GET con timeout).
- Teléfono: validar formato (9 dígitos, España).
- Email: validar formato.
- Localidad: validar que no esté vacía y tenga sentido (solo letras/espacios/guiones).
- Código postal: validar formato (5 dígitos, rango 01-52 para España).

Añade columnas: Web_verif, Telefono_verif, Email_verif, CP_verif, Localidad_verif
Valores: OK | Error | ND (si no hay dato que verificar).
"""
import re
import socket
import ssl
import urllib.request
import urllib.error
from pathlib import Path

import openpyxl

EXCEL_UNIFICADO = Path(r"c:\Users\javie\Desktop\HD - Listado transportistas - Unificado.xlsx")
EXCEL_VERIFICADO = Path(r"c:\Users\javie\Desktop\HD - Listado transportistas - Verificado.xlsx")
HOJAS_EXCLUIDAS = {"Gemini", "Listado"}
VALOR_ND = "ND"
TIMEOUT_WEB = 8  # segundos

# Índices de columnas unificadas en el Excel (1-based)
# Col 1=Provincia_hoja, 2=Localidad_unif, 3=Codigo_postal_unif, 4=Direccion_unif,
# 5=Telefono_unif, 6=Email_unif, 7=Web_unif
COL_LOCALIDAD = 2
COL_CP = 3
COL_TELEFONO = 5
COL_EMAIL = 6
COL_WEB = 7

COLUMNAS_VERIF = ["Web_verif", "Telefono_verif", "Email_verif", "CP_verif", "Localidad_verif"]
COL_INICIO_VERIF = 8

# Regex de validación
RE_EMAIL = re.compile(r"^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$")
RE_CP = re.compile(r"^(0[1-9]|[1-4]\d|5[0-2])\d{3}$")
RE_TELEFONO = re.compile(r"^(\+34\s?)?(\d[\s.\-]*){8}\d$")
RE_LOCALIDAD = re.compile(r"^[a-zA-Z0-9áéíóúüñÁÉÍÓÚÜÑ\s.\-,()/'ºª]+$")
RE_WEB_DOMINIO = re.compile(r"^[a-zA-Z0-9][-a-zA-Z0-9.]*\.[a-zA-Z]{2,}(?:\.[a-zA-Z]{2,})?$")


def validar_email(val):
    if not val or str(val).strip() == "" or val == VALOR_ND:
        return "ND"
    return "OK" if RE_EMAIL.match(str(val).strip()) else "Error"


def validar_cp(val):
    if not val or str(val).strip() == "" or val == VALOR_ND:
        return "ND"
    s = str(val).strip().replace(" ", "")
    return "OK" if RE_CP.match(s) else "Error"


def validar_telefono(val):
    if not val or str(val).strip() == "" or val == VALOR_ND:
        return "ND"
    s = str(val).strip().replace(" ", "").replace(".", "").replace("-", "").replace("\u202f", "")
    if s.startswith("+34"):
        s = s[3:].lstrip()
    if len(s) == 9 and s.isdigit():
        return "OK"
    if RE_TELEFONO.match(str(val).strip()):
        return "OK"
    return "Error"


def validar_localidad(val):
    if not val or str(val).strip() == "" or val == VALOR_ND:
        return "ND"
    s = str(val).strip()
    if len(s) < 2:
        return "Error"
    return "OK" if RE_LOCALIDAD.match(s) else "Error"


def normalizar_web_url(dom):
    if not dom or str(dom).strip() == "" or dom == VALOR_ND:
        return None
    s = str(dom).strip().lower().replace("[.]", ".").replace(" ", "")
    if not s or "indicada" in s or "corporativa" in s:
        return None
    if "http" in s:
        return s if s.startswith("http") else "https://" + s
    if "." in s and len(s) < 80:
        return "https://" + s
    return None


def verificar_web_responde(url):
    try:
        req = urllib.request.Request(
            url,
            data=None,
            headers={"User-Agent": "Mozilla/5.0 (compatible; verification-bot/1.0)"},
        )
        with urllib.request.urlopen(req, timeout=TIMEOUT_WEB) as resp:
            return "OK" if 200 <= resp.getcode() < 400 else "Error"
    except urllib.error.HTTPError as e:
        if e.code in (301, 302, 303, 307, 308):
            return "OK"
        return "Error"
    except (urllib.error.URLError, socket.timeout, OSError, ssl.SSLError, Exception):
        return "Error"


def verificar_fila(ws, row_idx):
    localidad = ws.cell(row=row_idx, column=COL_LOCALIDAD).value
    cp = ws.cell(row=row_idx, column=COL_CP).value
    telefono = ws.cell(row=row_idx, column=COL_TELEFONO).value
    email = ws.cell(row=row_idx, column=COL_EMAIL).value
    web = ws.cell(row=row_idx, column=COL_WEB).value

    localidad_verif = validar_localidad(localidad)
    cp_verif = validar_cp(cp)
    telefono_verif = validar_telefono(telefono)
    email_verif = validar_email(email)
    url = normalizar_web_url(web)
    web_verif = verificar_web_responde(url) if url else "ND"

    return (web_verif, telefono_verif, email_verif, cp_verif, localidad_verif)


def main():
    wb = openpyxl.load_workbook(EXCEL_UNIFICADO, data_only=False)
    provincias = [s for s in wb.sheetnames if s not in HOJAS_EXCLUIDAS]

    for nombre in provincias:
        ws = wb[nombre]
        max_row = ws.max_row
        if max_row < 2:
            continue
        if ws.cell(1, COL_WEB).value != "Web_unif":
            print(f"  {nombre}: estructura no esperada, omitiendo.")
            continue
        ws.insert_cols(COL_INICIO_VERIF, 5)
        for c, name in enumerate(COLUMNAS_VERIF, start=COL_INICIO_VERIF):
            ws.cell(1, c, value=name)
        ok_web = ok_tel = ok_email = ok_cp = ok_loc = 0
        err_web = err_tel = err_email = err_cp = err_loc = 0
        for row_idx in range(2, max_row + 1):
            verif = verificar_fila(ws, row_idx)
            for i, v in enumerate(verif):
                ws.cell(row_idx, COL_INICIO_VERIF + i, value=v)
            if verif[0] == "OK":
                ok_web += 1
            elif verif[0] == "Error":
                err_web += 1
            if verif[1] == "OK":
                ok_tel += 1
            elif verif[1] == "Error":
                err_tel += 1
            if verif[2] == "OK":
                ok_email += 1
            elif verif[2] == "Error":
                err_email += 1
            if verif[3] == "OK":
                ok_cp += 1
            elif verif[3] == "Error":
                err_cp += 1
            if verif[4] == "OK":
                ok_loc += 1
            elif verif[4] == "Error":
                err_loc += 1
        print(
            f"  {nombre}: {max_row - 1} filas | Web OK:{ok_web} Err:{err_web} | "
            f"Tel OK:{ok_tel} Err:{err_tel} | Email OK:{ok_email} Err:{err_email} | "
            f"CP OK:{ok_cp} Err:{err_cp} | Loc OK:{ok_loc} Err:{err_loc}"
        )

    wb.save(EXCEL_VERIFICADO)
    wb.close()
    print(f"\nGuardado: {EXCEL_VERIFICADO}")


if __name__ == "__main__":
    main()
