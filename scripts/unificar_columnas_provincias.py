# -*- coding: utf-8 -*-
"""
Unificar conceptos entre columnas de distintas provincias.
- Reconocer columnas que contienen: localidad, código postal, dirección, teléfono, email, web.
- Parsear columnas combinadas (Contacto, Ubicación, Contacto según fuente, etc.).
- Añadir columnas unificadas: Localidad_unif, Codigo_postal_unif, Direccion_unif,
  Telefono_unif, Email_unif, Web_unif (rellenadas cuando haya información).
- No se elimina ninguna columna ni dato original.
"""
import re
import openpyxl
from pathlib import Path

EXCEL_HOMOGENEIZADO = Path(r"c:\Users\javie\Desktop\HD - Listado transportistas - Homogeneizado.xlsx")
EXCEL_UNIFICADO = Path(r"c:\Users\javie\Desktop\HD - Listado transportistas - Unificado.xlsx")
HOJAS_EXCLUIDAS = {"Gemini", "Listado"}
VALOR_ND = "ND"

# Columnas unificadas que añadimos (después de Provincia_hoja)
COLUMNAS_UNIF = [
    "Localidad_unif",
    "Codigo_postal_unif",
    "Direccion_unif",
    "Telefono_unif",
    "Email_unif",
    "Web_unif",
]

# Mapeo: concepto -> substrings para identificar columnas (match con col.lower())
# Usamos substrings que aparecen en los nombres reales (sin depender de encoding)
MAPEO_COLUMNAS = {
    "localidad": [
        "ubicaci",
        "municipio (provincia)",
        "municipio / provincia",
        "municipio (cr)",
        "municipio (castell",
        "municipio (cádiz)",
        "municipio (cáceres)",
        "municipio base",
        "municipio / zona",
        "municipio / direcci",
        "localidad (guadalajara)",
        "localidad (ja",
        "base(s) / localidad",
        "base / localidad",
        "base / polígono",
        "base en alicante",
        "base en álava",
        "localidad / pol",
        "municipio)",
        "municipio base",
        "localidad",
    ],
    "direccion": [
        "direcci",
        "ubicaci",
        "municipio / direcci",
        "base en alicante",
        "contacto (según fuente)",
        "contacto publicado",
        "localidad / pol",
    ],
    "telefono": [
        "teléfono",
        "tel\u00e9fono",
        "telefono",
        "contacto",
        "contacto (tel",
        "contacto (web",
        "contacto publicado",
        "contacto (según fuente)",
        "teléfono / email",
    ],
    "email": [
        "email",
        "contacto",
        "contacto (tel",
        "contacto (web",
        "contacto publicado",
        "contacto (según fuente)",
        "teléfono / email",
    ],
    "web": [
        "web / ficha",
        "web (ofuscada)",
        "web/perfil",
        "web/fuente",
        "web",
    ],
}

# Regex
RE_EMAIL = re.compile(r"[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}")
RE_CP = re.compile(r"\b(0[1-9]|[1-4]\d|5[0-2])\d{3}\b")
# Teléfono español: 9 dígitos con separadores opcionales (ej. 974 211 414, 672 246 394)
RE_TELEFONO = re.compile(
    r"(?:\+34\s?)?(?:\d{3}[\s.\-\u202f]?\d{3}[\s.\-\u202f]?\d{3}|\d{2}[\s.\-\u202f]?\d{2}[\s.\-\u202f]?\d{2}[\s.\-\u202f]?\d{2}[\s.\-\u202f]?\d)"
)
# Web: dominio tipo algo.com o algo.es (evitar "web indicada", "web corporativa")
RE_WEB_DOMINIO = re.compile(
    r"(?:https?://)?(?:www\.)?([a-zA-Z0-9][-a-zA-Z0-9]*\.[a-zA-Z]{2,}(?:\.[a-zA-Z]{2,})?)"
)


def normalizar(s):
    """Normalizar string para comparación (lower, quitar acentos)."""
    if s is None or not isinstance(s, str):
        return ""
    s = s.lower().strip()
    for a, b in [("á", "a"), ("é", "e"), ("í", "i"), ("ó", "o"), ("ú", "u"), ("ü", "u"), ("ñ", "n")]:
        s = s.replace(a, b)
    return s


def columna_para_concepto(nombre_col, concepto):
    """True si la columna nombre_col aporta el concepto."""
    n = normalizar(nombre_col)
    for sub in MAPEO_COLUMNAS.get(concepto, []):
        if normalizar(sub) in n:
            return True
    if concepto == "web" and n.strip() == "web":
        return True
    return False


def extraer_email(texto):
    if not texto or not isinstance(texto, str):
        return None
    m = RE_EMAIL.search(texto)
    return m.group(0) if m else None


def extraer_telefonos(texto):
    if not texto or not isinstance(texto, str):
        return []
    # Normalizar espacios Unicode (ej. \u202f) a espacio normal
    texto = texto.replace("\u202f", " ").replace("\xa0", " ")
    return RE_TELEFONO.findall(texto)


def extraer_cp(texto):
    if not texto or not isinstance(texto, str):
        return None
    m = RE_CP.search(texto)
    return m.group(0) if m else None


def extraer_web(texto):
    """Extraer dominio web; si es texto tipo 'gruasmarquez[.]com' normalizar a gruasmarquez.com."""
    if not texto or not isinstance(texto, str):
        return None
    t = texto.replace("[.]", ".").replace(" ", "")
    m = RE_WEB_DOMINIO.search(t)
    if m:
        d = m.group(1)
        if "web" in d or "indicada" in d or "corporativa" in d:
            return None
        return d
    # Si es solo un dominio sin punto final (ej. "transportesjunfe.com")
    if re.match(r"^[a-zA-Z0-9][-a-zA-Z0-9]*\.[a-zA-Z]{2,}$", t):
        return t
    return None


def parsear_ubicacion(texto):
    """De 'Huesca (Pol. X, C/ Y 15)' o 'Estadilla (Cam. Cremadas s/n, 22423)' -> localidad, direccion, cp."""
    if not texto or not isinstance(texto, str):
        return None, None, None
    texto = texto.strip()
    localidad = None
    direccion = None
    cp = extraer_cp(texto)
    # Patrón: "Localidad (resto)"
    m = re.match(r"^([^(]+?)\s*\((.*)\)\s*$", texto)
    if m:
        localidad = m.group(1).strip()
        resto = m.group(2).strip()
        # Si en resto hay CP, quitarlo para la dirección
        if cp:
            resto = RE_CP.sub("", resto).strip()
            resto = re.sub(r",\s*,", ",", resto).strip().strip(",")
        direccion = resto if resto else None
    else:
        # Sin paréntesis: todo puede ser localidad o dirección
        if cp:
            direccion = texto
        else:
            localidad = texto
    return localidad, direccion, cp


def parsear_contacto(texto):
    """De 'Tel.: 974 211 414; Email: info@...' extraer teléfono y email."""
    if not texto or not isinstance(texto, str):
        return None, None, None
    email = extraer_email(texto)
    tels = extraer_telefonos(texto)
    telefono = tels[0] if tels else None
    # Dirección a veces aparece como "Dirección ..." en Contacto (según fuente)
    direccion = None
    for sep in ["Dirección", "Direcci�n", "dirección", "direccion"]:
        if sep in texto:
            idx = texto.lower().find(sep.replace("�", "ó"))
            if idx >= 0:
                resto = texto[idx + len(sep):].strip(" :;")
                # Quitar hasta el siguiente campo o fin
                resto = re.split(r"\s*;\s*(?:Tel\.|Email|Direcci)", resto, flags=re.I)[0]
                if len(resto) > 5:
                    direccion = resto.strip()
            break
    return telefono, email, direccion


def valor_nd_si_vacio(v):
    if v is None:
        return VALOR_ND
    s = str(v).strip()
    return s if s else VALOR_ND


def unificar_fila(header, fila, nombre_hoja):
    """Para una fila (lista de valores según header), devolver dict con las 6 columnas unificadas."""
    row_dict = {}
    for i, col in enumerate(header):
        if i < len(fila):
            row_dict[col] = fila[i]
        else:
            row_dict[col] = VALOR_ND

    def get_val(concepto):
        for col, val in row_dict.items():
            if val and val != VALOR_ND and columna_para_concepto(col, concepto):
                return val
        return None

    localidad = None
    cp = None
    direccion = None
    telefono = None
    email = None
    web = None

    # 1) Localidad: columnas explícitas primero; luego parsear Ubicación / Base en Alicante / etc.
    for col in header:
        if row_dict.get(col) and row_dict[col] != VALOR_ND:
            if columna_para_concepto(col, "localidad"):
                v = row_dict[col]
                if "(" in str(v) and "ubicaci" in normalizar(col):
                    loc, _, _ = parsear_ubicacion(v)
                    if loc:
                        localidad = loc
                        break
                elif "direcci" not in normalizar(col) or col.strip().lower() == "localidad":
                    localidad = v
                    break
    if not localidad:
        for col in header:
            if row_dict.get(col) and row_dict[col] != VALOR_ND:
                if "ubicaci" in normalizar(col) or "base en alicante" in normalizar(col) or "municipio / direcci" in normalizar(col):
                    loc, _, _ = parsear_ubicacion(str(row_dict[col]))
                    if loc:
                        localidad = loc
                        break

    # 2) CP: extraer de cualquier campo que parezca dirección/ubicación
    for col in header:
        if row_dict.get(col) and row_dict[col] != VALOR_ND:
            v = str(row_dict[col])
            c = extraer_cp(v)
            if c:
                cp = c
                break

    # 3) Dirección: columnas explícitas; o parsear Ubicación/Contacto
    for col in header:
        if row_dict.get(col) and row_dict[col] != VALOR_ND:
            if columna_para_concepto(col, "direccion"):
                v = str(row_dict[col])
                if "ubicaci" in normalizar(col):
                    _, dir_, _ = parsear_ubicacion(v)
                    if dir_:
                        direccion = dir_
                        break
                if "contacto" in normalizar(col):
                    _, _, dir_ = parsear_contacto(v)
                    if dir_:
                        direccion = dir_
                        break
                if not direccion:
                    direccion = v
                    break
    if not direccion:
        for col in header:
            if row_dict.get(col) and row_dict[col] != VALOR_ND:
                v = str(row_dict[col])
                if "ubicaci" in normalizar(col):
                    _, dir_, _ = parsear_ubicacion(v)
                    if dir_:
                        direccion = dir_
                        break

    # 4) Teléfono y 5) Email: columnas explícitas o parsear Contacto
    for col in header:
        if row_dict.get(col) and row_dict[col] != VALOR_ND:
            if columna_para_concepto(col, "telefono") or columna_para_concepto(col, "email"):
                v = str(row_dict[col])
                if "contacto" in normalizar(col) or "tel" in normalizar(col):
                    t, e, _ = parsear_contacto(v)
                    if t:
                        telefono = t
                    if e:
                        email = e
                elif "tel" in normalizar(col) and "email" not in normalizar(col):
                    tels = extraer_telefonos(v)
                    if tels:
                        telefono = tels[0]
                elif "email" in normalizar(col):
                    em = extraer_email(v)
                    if em:
                        email = em
    if not telefono:
        for col in header:
            if row_dict.get(col) and row_dict[col] != VALOR_ND:
                t = extraer_telefonos(str(row_dict[col]))
                if t and columna_para_concepto(col, "telefono"):
                    telefono = t[0]
                    break
    if not email:
        for col in header:
            if row_dict.get(col) and row_dict[col] != VALOR_ND:
                em = extraer_email(str(row_dict[col]))
                if em:
                    email = em
                    break

    # 6) Web: columnas Web; o extraer de texto (evitar "web indicada")
    for col in header:
        if row_dict.get(col) and row_dict[col] != VALOR_ND:
            if columna_para_concepto(col, "web"):
                v = str(row_dict[col]).strip()
                if "indicada" not in v.lower() and "corporativa" not in v.lower() and "en ficha" not in v.lower():
                    w = extraer_web(v) or (v if "." in v and len(v) < 60 else None)
                    if w:
                        web = w
                        break

    return {
        "Localidad_unif": valor_nd_si_vacio(localidad),
        "Codigo_postal_unif": valor_nd_si_vacio(cp),
        "Direccion_unif": valor_nd_si_vacio(direccion),
        "Telefono_unif": valor_nd_si_vacio(telefono),
        "Email_unif": valor_nd_si_vacio(email),
        "Web_unif": valor_nd_si_vacio(web),
    }


def main():
    wb = openpyxl.load_workbook(EXCEL_HOMOGENEIZADO, data_only=True)
    provincias = [s for s in wb.sheetnames if s not in HOJAS_EXCLUIDAS]

    for nombre in provincias:
        ws = wb[nombre]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        header = list(rows[0])
        if header[0] != "Provincia_hoja":
            continue
        # Insertar 6 columnas después de la primera (Provincia_hoja)
        ws.insert_cols(2, 6)
        # Escribir cabecera de las columnas unificadas (columnas B-G = índice 2 a 7)
        for c, name in enumerate(COLUMNAS_UNIF, start=2):
            ws.cell(row=1, column=c, value=name)
        # Para cada fila de datos, calcular unif y escribir en las 6 columnas nuevas
        for row_idx, row in enumerate(rows[1:], start=2):
            fila = list(row) if isinstance(row, tuple) else list(row)
            unif = unificar_fila(header, fila, nombre)
            for c, key in enumerate(COLUMNAS_UNIF, start=2):
                ws.cell(row=row_idx, column=c, value=unif[key])
        print(f"  {nombre}: {len(rows) - 1} filas, columnas unificadas añadidas.")

    wb.save(EXCEL_UNIFICADO)
    wb.close()
    print(f"\nGuardado: {EXCEL_UNIFICADO}")


if __name__ == "__main__":
    main()
