# -*- coding: utf-8 -*-
"""
Paso 1: Homogeneizar tablas de todas las hojas de provincia.
- Esquema común = unión de todas las columnas que aparecen en alguna provincia.
- Columnas que no existan en una provincia se rellenan con "ND".
- Se añade columna "Provincia_hoja" con el nombre de la hoja.
- No se elimina información.
"""
import openpyxl
from pathlib import Path

EXCEL_ORIGEN = Path(r"c:\Users\javie\Desktop\HD - Listado transportistas.xlsx")
EXCEL_HOMOGENEIZADO = Path(r"c:\Users\javie\Desktop\HD - Listado transportistas - Homogeneizado.xlsx")
HOJAS_EXCLUIDAS = {"Gemini", "Listado"}
VALOR_ND = "ND"
COLUMNA_PROVINCIA = "Provincia_hoja"


def main():
    wb = openpyxl.load_workbook(EXCEL_ORIGEN, read_only=True, data_only=True)
    provincias = [s for s in wb.sheetnames if s not in HOJAS_EXCLUIDAS]

    # 1) Recoger todas las columnas únicas en orden de primera aparición
    todas_columnas = []
    vistas = set()
    for nombre in provincias:
        ws = wb[nombre]
        row1 = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        header = row1 if isinstance(row1, tuple) else (row1,)
        for col in header:
            if col is not None and str(col).strip() and col not in vistas:
                todas_columnas.append(col)
                vistas.add(col)
    wb.close()

    # Esquema canónico: Provincia_hoja + todas las columnas existentes
    canonical = [COLUMNA_PROVINCIA] + todas_columnas

    # 2) Leer cada hoja de provincia y construir filas homogeneizadas
    wb_in = openpyxl.load_workbook(EXCEL_ORIGEN, read_only=True, data_only=True)
    datos_por_provincia = {}

    for nombre in provincias:
        ws = wb_in[nombre]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            datos_por_provincia[nombre] = []
            continue
        header_prov = [str(c).strip() if c is not None else "" for c in rows[0]]
        # Índice: nombre canónico -> índice en esta hoja (o None)
        idx_en_hoja = {}
        for i, col in enumerate(header_prov):
            if col:
                idx_en_hoja[col] = i
        filas = []
        for row in rows[1:]:
            row_list = list(row) if isinstance(row, tuple) else [row]
            valores = []
            for col_canon in canonical:
                if col_canon == COLUMNA_PROVINCIA:
                    valores.append(nombre)
                elif col_canon in idx_en_hoja:
                    i = idx_en_hoja[col_canon]
                    val = row_list[i] if i < len(row_list) else None
                    if val is None or (isinstance(val, str) and not val.strip()):
                        valores.append(VALOR_ND)
                    else:
                        valores.append(val)
                else:
                    valores.append(VALOR_ND)
            filas.append(valores)
        datos_por_provincia[nombre] = filas

    wb_in.close()

    # 3) Escribir nuevo Excel: copiar Gemini y Listado tal cual; provincias homogeneizadas
    wb_orig = openpyxl.load_workbook(EXCEL_ORIGEN, read_only=True, data_only=True)
    wb_out = openpyxl.Workbook()
    wb_out.remove(wb_out.active)

    # Copiar hojas que no son de provincia (Gemini, Listado)
    for sheet_name in wb_orig.sheetnames:
        if sheet_name in HOJAS_EXCLUIDAS:
            ws_in = wb_orig[sheet_name]
            ws_out = wb_out.create_sheet(sheet_name)
            for row in ws_in.iter_rows(values_only=True):
                ws_out.append(row)
    wb_orig.close()

    # Crear hojas homogeneizadas por provincia (mismo orden que en el original)
    for nombre in provincias:
        ws_out = wb_out.create_sheet(nombre)
        ws_out.append(canonical)
        for fila in datos_por_provincia.get(nombre, []):
            ws_out.append(fila)

    wb_out.save(EXCEL_HOMOGENEIZADO)
    print(f"Guardado: {EXCEL_HOMOGENEIZADO}")
    print(f"Columnas canónicas: {len(canonical)}")
    print(f"Hojas de provincia homogeneizadas: {len(provincias)}")
    total_filas = sum(len(datos_por_provincia.get(p, [])) for p in provincias)
    print(f"Total filas de datos (excl. cabeceras): {total_filas}")


if __name__ == "__main__":
    main()
